import os
import sys
import sqlite3
import pandas as pd
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import calendar
import re
import json
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import secrets
from unidecode import unidecode
import io
import zipfile
import fitz
import imaplib
import email
from email.header import decode_header
import PyPDF2
import pythoncom

# GUI Imports
import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, Listbox, filedialog, simpledialog
import threading
from difflib import get_close_matches
from PIL import Image, ImageTk

# PDF Generation Imports
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch

# Excel Writing Imports
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# Import improved signature editor
from improved_signature_editor_v2 import ImprovedSignatureEditor



# Application Constants
DB_FILE = os.path.join(os.path.expanduser("~"), "AppData", "local_operations.db")
APP_DATA_FOLDER = os.path.join(os.path.expanduser("~"), "AppData")
ORDERS_FOLDER = os.path.join(APP_DATA_FOLDER, "Orders")
CONFIRMATIONS_FOLDER = os.path.join(APP_DATA_FOLDER, "Confirmations")

# UI Constants
DEFAULT_WINDOW_WIDTH = 1800
DEFAULT_WINDOW_HEIGHT = 800
DASHBOARD_ICON_SIZE = 24
FILTER_PANEL_WIDTH = 300
MAX_RETRIES = 3
DEFAULT_TIMEOUT = 30

# --- Utility Functions ---
class CaseInsensitiveDict(dict):
    """Dictionary that ignores case for string keys."""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._normalize_keys()

    def _normalize_key(self, key):
        """Normalize key to uppercase for consistent lookup."""
        if isinstance(key, str):
            return key.strip().upper()
        return key

    def _normalize_keys(self):
        """Normalize all existing keys."""
        items = list(self.items())
        self.clear()
        for key, value in items:
            normalized_key = self._normalize_key(key)
            super().__setitem__(normalized_key, value)

    def __setitem__(self, key, value):
        normalized_key = self._normalize_key(key)
        super().__setitem__(normalized_key, value)

    def __getitem__(self, key):
        normalized_key = self._normalize_key(key)
        return super().__getitem__(normalized_key)

    def get(self, key, default=None):
        normalized_key = self._normalize_key(key)
        return super().get(normalized_key, default)

    def __contains__(self, key):
        normalized_key = self._normalize_key(key)
        return super().__contains__(normalized_key)


def add_working_days(start_date, days):
    """
    Add working days to a date (opposite of subtract_working_days)

    Args:
        start_date: Starting date (datetime object)
        days: Number of working days to add

    Returns:
        datetime: Result date after adding working days
    """
    current_date = start_date
    days_added = 0

    while days_added < days:
        current_date += timedelta(days=1)
        # Skip weekends
        if current_date.weekday() < 5:  # Monday = 0, Sunday = 6
            days_added += 1

    return current_date


def subtract_working_days(end_date: datetime, num_days_to_subtract: int) -> datetime:
    """Subtract working days from a date (excluding weekends)."""
    if not isinstance(end_date, datetime) or num_days_to_subtract <= 0:
        return end_date

    days_subtracted = 0
    current_date = end_date

    while days_subtracted < num_days_to_subtract:
        current_date -= timedelta(days=1)
        if current_date.weekday() < 5:  # Monday = 0, Sunday = 6
            days_subtracted += 1

    return current_date


def extract_supplier_name(pdf_path, known_vendors, log_callback=None):
    """Extract supplier name from PDF by matching against known vendors."""

    # Normalize function matching import.py
    def normalize(text):
        """Normalize text for matching"""
        text = unidecode(str(text)).lower()
        text = re.sub(r"[^a-z0-9 ]", " ", text)
        # Remove common company suffixes
        text = re.sub(
            r"\b(ltd|inc|gmbh|kft|oy|ab|co|llc|sarl|plc|bv|sro|sa|sas|kg|as)\b",
            "",
            text,
        )
        return " ".join(text.split())

    # Create normalized vendor mapping
    normalized_vendors = {normalize(v): v for v in known_vendors}

    try:
        with open(pdf_path, "rb") as f:
            reader = fitz.open(stream=f.read(), filetype="pdf")
        text = "".join(page.get_text() for page in reader)

        # Get lines from top of PDF
        lines = [line.strip() for line in text.splitlines() if line.strip()]
        top_lines = lines[:30]

        # Filter candidates - more comprehensive keyword list
        candidates = [
            line
            for line in top_lines
            if len(line) > 5
            and not re.search(
                r"\b(phone|fax|date|page|order|number|qty|total|amount|usd|eur|sek|net|vat|delivery|invoice|payment|terms|ref)\b",
                line.lower(),
            )
        ]

        # Try fuzzy matching on each candidate
        from difflib import get_close_matches

        for candidate in candidates:
            norm = normalize(candidate)
            matches = get_close_matches(
                norm, normalized_vendors.keys(), n=1, cutoff=0.7
            )
            if matches:
                return normalized_vendors[matches[0]]

        return None

    except Exception as e:
        if log_callback:
            log_callback(f"✗ ERROR: Could not process {pdf_path}: {str(e)}")
        return None


def normalize_supplier(name):
    """Normalize supplier name for matching."""
    return unidecode(str(name)).strip().lower()


def load_email_mapping(db_manager=None, log_callback=None):
    """Loads email mappings from the database with case-insensitive lookup."""
    email_map = CaseInsensitiveDict()

    try:
        if db_manager is None:
            conn = sqlite3.connect(DB_FILE)
            cursor = conn.cursor()
            cursor.execute("SELECT display_name, emails FROM vendors")
            rows = cursor.fetchall()
            conn.close()
        else:
            # Use DatabaseManager's execute_query method
            rows = db_manager.execute_query(
                "SELECT display_name, emails FROM vendors", fetchall=True
            )

        if not rows:
            if log_callback:
                log_callback(f"⚠ WARNING: No vendors found in database")
            return email_map

        for row in rows:
            # Handle both tuple and dict formats
            if isinstance(row, dict):
                display_name = row.get("display_name")
                emails_str = row.get("emails")
            else:
                display_name = row[0]
                emails_str = row[1]

            if not display_name or not emails_str:
                continue

            # Extract emails using regex
            email_pattern = re.compile(
                r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}"
            )
            found_emails = email_pattern.findall(str(emails_str))

            if found_emails:
                # CaseInsensitiveDict automatically handles normalization
                email_map[display_name] = list(
                    set(email.lower() for email in found_emails)
                )

        if log_callback:
            log_callback(f"ℹ️ INFO: Loaded {len(email_map)} vendor email mappings")

        return email_map

    except Exception as e:
        if log_callback:
            log_callback(f"✗ ERROR: Could not load email map: {e}")
        return CaseInsensitiveDict()


# --- Outlook Integration (Optional) ---
# Requires pywin32: pip install pywin32
try:
    import win32com.client
    import pywintypes

    OUTLOOK_AVAILABLE = True
except ImportError:
    OUTLOOK_AVAILABLE = False


# --- Configuration ---
# Get the directory where the script is located
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
APP_DATA_FOLDER = os.path.join(SCRIPT_DIR, "RemoteOperationsData")
DB_FILE = os.path.join(APP_DATA_FOLDER, "operations_data.db")
CONFIG_FILE = os.path.join(APP_DATA_FOLDER, "config.json")
DOWNLOAD_FOLDER = os.path.join(
    os.path.expanduser("~"), "Downloads"
)  # Downloads folder remains in user profile
ORDERS_FOLDER = os.path.join(APP_DATA_FOLDER, "OrdersToSend")
ORDERS_SENT_FOLDER = os.path.join(APP_DATA_FOLDER, "OrdersSent")
RESCHEDULE_OUTPUT_FOLDER = os.path.join(APP_DATA_FOLDER, "reschedule_output")

# Ensure all necessary directories exist
for folder in [
    APP_DATA_FOLDER,
    DOWNLOAD_FOLDER,
    ORDERS_FOLDER,
    ORDERS_SENT_FOLDER,
    RESCHEDULE_OUTPUT_FOLDER,
]:
    os.makedirs(folder, exist_ok=True)


# ==============================================================================
# 1. DATABASE MANAGER (SQLite)
# ==============================================================================


class DatabaseManager:
    """Handles all SQLite database operations."""

    def __init__(self, db_path):
        self.db_path = db_path
        self.setup_database()
        self.create_mrp_tables()

    def get_connection(self):
        """Creates and returns a new database connection."""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = self._dict_factory
        return conn

    @staticmethod
    def _dict_factory(cursor, row):
        """Converts query results into dictionaries."""
        d = {}
        for idx, col in enumerate(cursor.description):
            d[col[0]] = row[idx]
        return d

    def execute_query(
        self, query, params=(), commit=False, fetchone=False, fetchall=False
    ):
        """
        A generic method to execute SQL queries.

        Returns:
            - cursor object if commit=True and query is INSERT/UPDATE/DELETE (for lastrowid access)
            - fetched result if fetchone=True or fetchall=True
            - cursor.rowcount otherwise
        """
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(query, params)

            if commit:
                conn.commit()
                # Return cursor for INSERT queries so caller can access lastrowid
                query_upper = query.strip().upper()
                if query_upper.startswith("INSERT"):
                    return cursor

            if fetchone:
                return cursor.fetchone()
            if fetchall:
                return cursor.fetchall()

            return cursor.rowcount

    def setup_database(self):
        """Creates database tables if they don't exist."""
        queries = [
            """
            CREATE TABLE IF NOT EXISTS vendors (
                vendor_name TEXT PRIMARY KEY,
                display_name TEXT NOT NULL UNIQUE,
                emails TEXT,
                address TEXT,
                contact_person TEXT,
                transport_days INTEGER DEFAULT 0,
                transport_days_secondary INTEGER DEFAULT 0,
                delivery_terms TEXT,
                payment_terms TEXT,
                api_key TEXT
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS open_orders (
                po TEXT NOT NULL,
                item TEXT NOT NULL,
                material_code TEXT,
                short_text TEXT,
                vendor_name TEXT,
                requested_qty INTEGER,
                requested_del_date TEXT,
                conf_delivery_date TEXT,
                rescheduling_date TEXT,
                unit TEXT,
                unit_price REAL,
                price_per_unit INTEGER DEFAULT 1,
                total_amount REAL,
                currency TEXT,
                comments TEXT,
                exception_message TEXT,
                has_new_message INTEGER DEFAULT 0,
                pdf_status TEXT DEFAULT 'Pending',
                email_status TEXT DEFAULT 'Pending',
                status TEXT DEFAULT 'Open',
                PRIMARY KEY (po, item),
                FOREIGN KEY (vendor_name) REFERENCES vendors(vendor_name)
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS messages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                po_number TEXT NOT NULL,
                item_number TEXT NOT NULL,
                sender TEXT NOT NULL,
                message_text TEXT NOT NULL,
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS app_config (
                key TEXT PRIMARY KEY,
                value TEXT
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS forecasts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                vendor_name TEXT NOT NULL,
                material_code TEXT,
                short_text TEXT,
                forecast_date DATE NOT NULL,
                forecast_qty INTEGER NOT NULL,
                unit TEXT DEFAULT 'EA',
                unit_price REAL DEFAULT 0,
                total_amount REAL DEFAULT 0,
                currency TEXT DEFAULT 'EUR',
                week_number INTEGER,
                month_number INTEGER,
                year_number INTEGER,
                comments TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (vendor_name) REFERENCES vendors(vendor_name)
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS requisitions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                req_number TEXT NOT NULL,
                item TEXT NOT NULL,
                material_code TEXT,
                short_text TEXT,
                vendor_name TEXT,
                requested_qty INTEGER,
                requested_del_date TEXT,
                unit TEXT DEFAULT 'EA',
                unit_price REAL DEFAULT 0,
                total_amount REAL DEFAULT 0,
                currency TEXT DEFAULT 'EUR',
                status TEXT DEFAULT 'Open',
                pr_status TEXT DEFAULT 'Pending',
                comments TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(req_number, item),
                FOREIGN KEY (vendor_name) REFERENCES vendors(vendor_name)
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS forecast_vs_actuals (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                vendor_name TEXT NOT NULL,
                material_code TEXT,
                period_start DATE NOT NULL,
                period_end DATE NOT NULL,
                forecast_qty INTEGER DEFAULT 0,
                actual_qty INTEGER DEFAULT 0,
                variance_qty INTEGER DEFAULT 0,
                variance_pct REAL DEFAULT 0,
                comments TEXT,
                FOREIGN KEY (vendor_name) REFERENCES vendors(vendor_name)
            )
            """,
        ]
        for query in queries:
            self.execute_query(query, commit=True)

        # Add missing columns for existing databases
        self.add_missing_columns()

    def create_mrp_tables(self):
        """Create all MRP-related tables if they don't exist"""
        queries = [
            """
            CREATE TABLE IF NOT EXISTS mrp_runs (
                run_id INTEGER PRIMARY KEY AUTOINCREMENT,
                run_date TEXT NOT NULL,
                horizon_weeks INTEGER DEFAULT 13,
                status TEXT DEFAULT 'COMPLETED',
                parameters TEXT,
                created_by TEXT
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS mrp_results (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                run_id INTEGER,
                item_code TEXT,
                item_name TEXT,
                current_stock REAL DEFAULT 0,
                allocated_stock REAL DEFAULT 0,
                available_stock REAL DEFAULT 0,
                demand REAL DEFAULT 0,
                supply REAL DEFAULT 0,
                net_requirement REAL DEFAULT 0,
                suggested_order_qty REAL DEFAULT 0,
                reorder_point REAL DEFAULT 0,
                lead_time_days INTEGER DEFAULT 0,
                supplier TEXT,
                unit_cost REAL DEFAULT 0,
                total_cost REAL DEFAULT 0,
                priority TEXT,
                notes TEXT,
                FOREIGN KEY (run_id) REFERENCES mrp_runs(id) ON DELETE CASCADE
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS mrp_demand (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                item_code TEXT,
                demand_date DATE,
                quantity REAL,
                demand_type TEXT,
                source_reference TEXT,
                status TEXT DEFAULT 'active',
                created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS mrp_supply (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                item_code TEXT,
                supply_date DATE,
                quantity REAL,
                supply_type TEXT,
                source_reference TEXT,
                status TEXT DEFAULT 'active',
                created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS materials (
                material_code TEXT PRIMARY KEY,
                description TEXT,
                unit TEXT DEFAULT 'EA',
                standard_price REAL DEFAULT 0,
                lead_time_days INTEGER DEFAULT 0,
                safety_stock REAL DEFAULT 0,
                min_order_qty REAL DEFAULT 1,
                lot_size_rule TEXT DEFAULT 'LOT_FOR_LOT',
                fixed_lot_size REAL,
                preferred_vendor TEXT,
                abc_class TEXT DEFAULT 'C',
                created_date TEXT,
                last_updated TEXT,
                FOREIGN KEY (preferred_vendor) REFERENCES vendors(vendor_name)
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS inventory (
                material_code TEXT PRIMARY KEY,
                on_hand_qty REAL DEFAULT 0,
                reserved_qty REAL DEFAULT 0,
                available_qty REAL DEFAULT 0,
                in_transit_qty REAL DEFAULT 0,
                last_count_date TEXT,
                warehouse_location TEXT,
                FOREIGN KEY (material_code) REFERENCES materials(material_code)
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS forecast_demand (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                material_code TEXT NOT NULL,
                forecast_date TEXT NOT NULL,
                forecast_qty REAL NOT NULL,
                period_type TEXT DEFAULT 'WEEK',
                forecast_source TEXT DEFAULT 'MANUAL',
                confidence_level REAL DEFAULT 100,
                created_date TEXT,
                created_by TEXT,
                FOREIGN KEY (material_code) REFERENCES materials(material_code),
                UNIQUE(material_code, forecast_date, period_type)
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS mrp_calculations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                run_id INTEGER,
                material_code TEXT,
                period_date TEXT,
                gross_requirement REAL DEFAULT 0,
                scheduled_receipts REAL DEFAULT 0,
                on_hand_start REAL DEFAULT 0,
                on_hand_end REAL DEFAULT 0,
                net_requirement REAL DEFAULT 0,
                planned_order_qty REAL DEFAULT 0,
                planned_order_date TEXT,
                vendor_name TEXT,
                FOREIGN KEY (run_id) REFERENCES mrp_runs(run_id),
                FOREIGN KEY (material_code) REFERENCES materials(material_code)
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS vendor_lead_times (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                vendor_name TEXT,
                material_code TEXT,
                lead_time_days INTEGER,
                min_order_qty REAL DEFAULT 1,
                price REAL,
                is_preferred BOOLEAN DEFAULT 0,
                FOREIGN KEY (vendor_name) REFERENCES vendors(vendor_name),
                FOREIGN KEY (material_code) REFERENCES materials(material_code),
                UNIQUE(vendor_name, material_code)
            )
            """,
        ]

        for query in queries:
            self.execute_query(query, commit=True)

        # Enhanced Requisitions table - Add columns if they don't exist
        try:
            self.execute_query(
                "ALTER TABLE requisitions ADD COLUMN source TEXT DEFAULT 'MANUAL'",
                commit=True,
            )
        except Exception:  # TODO: Add proper error handling
            pass  # TODO: Add proper error handling
        try:
            self.execute_query(
                "ALTER TABLE requisitions ADD COLUMN mrp_run_id INTEGER", commit=True
            )
        except Exception:  # TODO: Add proper error handling
            pass  # TODO: Add proper error handling
        try:
            self.execute_query(
                "ALTER TABLE requisitions ADD COLUMN priority TEXT DEFAULT 'NORMAL'",
                commit=True,
            )
        except Exception:  # TODO: Add proper error handling
            pass  # TODO: Add proper error handling
        try:
            self.execute_query(
                "ALTER TABLE requisitions ADD COLUMN approval_status TEXT DEFAULT 'PENDING'",
                commit=True,
            )
        except Exception:  # TODO: Add proper error handling
            pass  # TODO: Add proper error handling
        try:
            self.execute_query(
                "ALTER TABLE requisitions ADD COLUMN approved_by TEXT", commit=True
            )
        except Exception:  # TODO: Add proper error handling
            pass  # TODO: Add proper error handling
        try:
            self.execute_query(
                "ALTER TABLE requisitions ADD COLUMN approved_date TEXT", commit=True
            )
        except Exception:  # TODO: Add proper error handling
            pass  # TODO: Add proper error handling
        try:
            self.execute_query(
                "ALTER TABLE requisitions ADD COLUMN notes TEXT", commit=True
            )
        except Exception:  # TODO: Add proper error handling
            pass  # TODO: Add proper error handling
        print("MRP tables created successfully")

    def add_missing_columns(self):
        """Add any missing columns that were added in updates"""
        try:
            # Use a raw connection without dict_factory for PRAGMA queries
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            # Check vendors table columns
            cursor.execute("PRAGMA table_info(vendors)")
            vendor_columns = [
                row[1] for row in cursor.fetchall()
            ]  # Column name is at index 1
            print(f"Current vendor columns: {vendor_columns}")

            if "delivery_terms" not in vendor_columns:
                try:
                    cursor.execute("ALTER TABLE vendors ADD COLUMN delivery_terms TEXT")
                    conn.commit()
                    print(" Added delivery_terms column to vendors table")
                except Exception as col_error:
                    print(f" Could not add delivery_terms: {col_error}")
            else:
                print(" delivery_terms column already exists")

            if "payment_terms" not in vendor_columns:
                try:
                    cursor.execute("ALTER TABLE vendors ADD COLUMN payment_terms TEXT")
                    conn.commit()
                    print(" Added payment_terms column to vendors table")
                except Exception as col_error:
                    print(f" Could not add payment_terms: {col_error}")
            else:
                print(" payment_terms column already exists")

            # Add the new secondary transport days column to vendors table if missing
            if "transport_days_secondary" not in vendor_columns:
                try:
                    cursor.execute(
                        "ALTER TABLE vendors ADD COLUMN transport_days_secondary INTEGER DEFAULT 0"
                    )
                    conn.commit()
                    print(" Added transport_days_secondary column to vendors table")
                except Exception as col_error:
                    print(f" Could not add transport_days_secondary: {col_error}")
            else:
                print(" transport_days_secondary column already exists")

            # Check open_orders table columns
            cursor.execute("PRAGMA table_info(open_orders)")
            order_columns = [row[1] for row in cursor.fetchall()]
            print(f"Current open_orders columns: {order_columns}")

            if "exception_message" not in order_columns:
                try:
                    cursor.execute(
                        "ALTER TABLE open_orders ADD COLUMN exception_message TEXT"
                    )
                    conn.commit()
                    print(" Added exception_message column to open_orders table")
                except Exception as col_error:
                    print(f" Could not add exception_message: {col_error}")
            else:
                print(" exception_message column already exists")

            if "rescheduling_date" not in order_columns:
                try:
                    cursor.execute(
                        "ALTER TABLE open_orders ADD COLUMN rescheduling_date TEXT"
                    )
                    conn.commit()
                    print(" Added rescheduling_date column to open_orders table")
                except Exception as col_error:
                    print(f" Could not add rescheduling_date: {col_error}")
            else:
                print(" rescheduling_date column already exists")

            if "price_per_unit" not in order_columns:
                try:
                    cursor.execute(
                        "ALTER TABLE open_orders ADD COLUMN price_per_unit INTEGER DEFAULT 1"
                    )
                    conn.commit()
                    print(" Added price_per_unit column to open_orders table")
                except Exception as col_error:
                    print(f" Could not add price_per_unit: {col_error}")
            else:
                print(" price_per_unit column already exists")

            # Check for closed_by_user column
            if "closed_by_user" not in order_columns:
                try:
                    cursor.execute(
                        "ALTER TABLE open_orders ADD COLUMN closed_by_user INTEGER DEFAULT 0"
                    )
                    conn.commit()
                    print("✓ Added closed_by_user column to open_orders table")
                except Exception as col_error:
                    print(f"✗ Could not add closed_by_user: {col_error}")
            else:
                print("✓ closed_by_user column already exists")

            # NEW: Check requisitions table columns for lead_time_days
            cursor.execute("PRAGMA table_info(requisitions)")
            req_columns = [row[1] for row in cursor.fetchall()]
            print(f"Current requisitions columns: {req_columns}")

            if "lead_time_days" not in req_columns:
                try:
                    cursor.execute(
                        "ALTER TABLE requisitions ADD COLUMN lead_time_days INTEGER DEFAULT 0"
                    )
                    conn.commit()
                    print(" Added lead_time_days column to requisitions table")
                except Exception as col_error:
                    print(f" Could not add lead_time_days: {col_error}")
            else:
                print(" lead_time_days column already exists in requisitions")

            conn.close()
            print("Database migration complete!")

            cursor.execute("PRAGMA table_info(materials)")
            material_columns = [row[1] for row in cursor.fetchall()]
            print(f"Current materials columns: {material_columns}")

            if "net_price" not in material_columns:
                try:
                    cursor.execute(
                        "ALTER TABLE materials ADD COLUMN net_price REAL DEFAULT 0"
                    )
                    conn.commit()
                    print(" Added net_price column to materials table")
                except Exception as col_error:
                    print(f" Could not add net_price: {col_error}")
            else:
                print(" net_price column already exists in materials")

            if "price_per_unit" not in material_columns:
                try:
                    cursor.execute(
                        "ALTER TABLE materials ADD COLUMN price_per_unit INTEGER DEFAULT 1"
                    )
                    conn.commit()
                    print(" Added price_per_unit column to materials table")
                except Exception as col_error:
                    print(f" Could not add price_per_unit: {col_error}")
            else:
                print(" price_per_unit column already exists in materials")

        except Exception as e:
            print(f"Error in add_missing_columns: {type(e).__name__}: {e}")


# ==============================================================================
# 2. LOCAL DATA MANAGER (Replaces the API Logic)
# ==============================================================================


class LocalDataManager:
    """Handles all business logic, interacting with the DatabaseManager."""

    def __init__(self, db_manager):
        self.db = db_manager

    # --- Vendor Management ---
    def get_all_vendors(self):
        return self.db.execute_query(
            "SELECT * FROM vendors ORDER BY display_name", fetchall=True
        )

    def create_vendor(self, data):
        data["vendor_name"] = unidecode(data["display_name"]).strip().lower()
        data["api_key"] = secrets.token_urlsafe(32)
        query = """
            INSERT INTO vendors (vendor_name, display_name, emails, address, contact_person, transport_days, transport_days_secondary, delivery_terms, payment_terms, api_key)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        params = (
            data["vendor_name"],
            data["display_name"],
            data["emails"],
            data["address"],
            data["contact_person"],
            data["transport_days"],
            data.get("transport_days_secondary", 0),
            data.get("delivery_terms", ""),
            data.get("payment_terms", ""),
            data["api_key"],
        )
        self.db.execute_query(query, params, commit=True)
        return data

    def update_vendor(self, original_name, data):
        original_vendor_name = unidecode(original_name).strip().lower()
        query = """
            UPDATE vendors SET display_name=?, emails=?, address=?, contact_person=?, transport_days=?, transport_days_secondary=?, delivery_terms=?, payment_terms=?
            WHERE vendor_name=?
        """
        params = (
            data["display_name"],
            data["emails"],
            data["address"],
            data["contact_person"],
            data["transport_days"],
            data.get("transport_days_secondary", 0),
            data.get("delivery_terms", ""),
            data.get("payment_terms", ""),
            original_vendor_name,
        )
        return self.db.execute_query(query, params, commit=True) > 0

    def delete_vendor(self, name):
        vendor_name = unidecode(name).strip().lower()
        # Check for linked orders first
        linked_orders = self.db.execute_query(
            "SELECT 1 FROM open_orders WHERE vendor_name=? LIMIT 1",
            (vendor_name,),
            fetchone=True,
        )
        if linked_orders:
            raise ValueError(
                "Cannot delete vendor. They are still linked to open orders."
            )
        return (
            self.db.execute_query(
                "DELETE FROM vendors WHERE vendor_name=?", (vendor_name,), commit=True
            )
            > 0
        )

    def generate_new_api_key(self, name):
        vendor_name = unidecode(name).strip().lower()
        new_key = secrets.token_urlsafe(32)
        self.db.execute_query(
            "UPDATE vendors SET api_key=? WHERE vendor_name=?",
            (new_key, vendor_name),
            commit=True,
        )
        return new_key

    # --- Order Management ---
    def get_all_open_orders(self):
        query = """
            SELECT oo.*, v.display_name AS name
            FROM open_orders oo
            LEFT JOIN vendors v ON oo.vendor_name = v.vendor_name
            WHERE oo.status = 'Open'
            ORDER BY oo.po, oo.item
        """
        return self.db.execute_query(query, fetchall=True)

    def close_order_lines(self, order_lines, closed_by_user=True):
        """
        Marks a list of order lines as 'Closed'.
        
        Args:
            order_lines: List of tuples (po_number, item_number)
            closed_by_user: If True, marks as manually closed (prevents reopening)
                           If False, marks as auto-closed (can be reopened on upload)
        
        Returns:
            Number of rows affected
        """
        if not order_lines:
            return 0

        # order_lines is a list of tuples, where each tuple is (po_number, item_number)
        query = """
            UPDATE open_orders 
            SET status = 'Closed', closed_by_user = ? 
            WHERE po = ? AND item = ?
        """

        with self.db.get_connection() as conn:
            cursor = conn.cursor()
            # Add the closed_by_user flag to each tuple
            params = [(1 if closed_by_user else 0, po, item) for po, item in order_lines]
            cursor.executemany(query, params)
            conn.commit()
            return cursor.rowcount

    def reopen_order_lines(self, order_lines):
        """
        Reopens order lines that were manually closed.
        
        Args:
            order_lines: List of tuples (po_number, item_number)
        
        Returns:
            Number of rows affected
        """
        if not order_lines:
            return 0

        query = """
            UPDATE open_orders 
            SET status = 'Open', closed_by_user = 0 
            WHERE po = ? AND item = ?
        """

        with self.db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.executemany(query, order_lines)
            conn.commit()
            return cursor.rowcount

    def get_po_details(self, po_number):
        """Fetches full header and line item details for a given PO."""
        # Query for line items
        lines_query = """
            SELECT * FROM open_orders 
            WHERE po = ? AND status = 'Open' 
            ORDER BY item
        """
        lines = self.db.execute_query(lines_query, (po_number,), fetchall=True)

        if not lines:
            return None

        # Get vendor name from the first line to fetch vendor details
        vendor_name_internal = lines[0].get("vendor_name")

        # Query for vendor details
        vendor_query = "SELECT * FROM vendors WHERE vendor_name = ?"
        vendor_details = self.db.execute_query(
            vendor_query, (vendor_name_internal,), fetchone=True
        )

        return {"lines": lines, "vendor": vendor_details}

    ## --- ADDED --- ##
    def update_confirmation_dates(self, updates):
        """Updates the confirmation dates for multiple order lines."""
        # updates is a list of tuples: [(conf_date, po, item), ...]
        if not updates:
            return 0

        query = (
            "UPDATE open_orders SET conf_delivery_date = ? WHERE po = ? AND item = ?"
        )
        with self.db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.executemany(query, updates)
            conn.commit()
            return cursor.rowcount

    def upload_order_book(self, file_path, auto_close_missing=True):
        """
        Uploads an Excel file with open order data.
        - Upserts order lines found in the file, ensuring their status is 'Open'.
        - Optionally marks existing open order lines NOT in the file as 'Closed'.

        Args:
            file_path: Path to the Excel file
            auto_close_missing: If True, close lines missing from file (default: True)
        """
        try:
            # Let pandas auto-detect the engine, which is more robust for xls/xlsx
            df = pd.read_excel(file_path, sheet_name="RawData", engine=None)
        except Exception as e:
            # Re-raise with a more user-friendly message
            raise ValueError(
                f"Could not read the Excel file. Please ensure it is a valid, uncorrupted .xls or .xlsx file and contains a sheet named 'RawData'. Original error: {e}"
            )

        df.columns = [str(c).strip() for c in df.columns]

        # Column mapping
        col_map = {
            "po": ["Purchasing Document", "PO Number"],
            "item": ["Item"],
            "name": ["Name", "Vendor", "Supplier"],
            "material": ["Material"],
            "short_text": ["Short Text"],
            "req_qty": ["Requested quantity", "Qty"],
            "req_del_date": ["Requested del. date", "Req. Date"],
            "conf_del_date": ["Conf. delivery date", "Conf. Date"],
            "reschedule_date": ["Rescheduling date", "Reschedule Date"],
            "unit": ["Unit"],
            "price": ["Unit Price", "Net Order Price"],
            "price_per": ["Per", "Price Per", "Price Unit", "Per Unit"],
            "amount": ["Total Amount", "Net Order Value"],
            "currency": ["Currency"],
        }

        def find_col(possible_names):
            for name in possible_names:
                if name in df.columns:
                    return name
            return None

        # Data extraction and preparation
        data_to_insert = []
        vendors_to_upsert = {}
        lines_in_excel = set()  # To track all (PO, Item) tuples found in the Excel file

        for _, row in df.iterrows():
            po_num = str(row.get(find_col(col_map["po"]), "")).strip()
            item_num_val = row.get(find_col(col_map["item"]))

            # Skip rows with no PO or no valid Item number
            if not po_num or pd.isna(item_num_val):
                continue
            item_num = str(int(item_num_val)).strip()
            if item_num == "0":
                continue

            # Add the valid line to our set for comparison later
            lines_in_excel.add((po_num, item_num))

            display_name = str(row.get(find_col(col_map["name"]), "")).strip()
            if display_name:
                vendor_name = unidecode(display_name).strip().lower()
                if vendor_name not in vendors_to_upsert:
                    vendors_to_upsert[vendor_name] = {
                        "display_name": display_name,
                        "vendor_name": vendor_name,
                    }

            def format_date(val):
                if pd.isna(val):
                    return None
                try:
                    return pd.to_datetime(val).strftime("%d.%m.%Y")
                except Exception:
                    return None

            req_date_col = find_col(col_map["req_del_date"])
            conf_date_col = find_col(col_map["conf_del_date"])
            reschedule_col = find_col(col_map["reschedule_date"])

            data_to_insert.append(
                {
                    "po": po_num,
                    "item": item_num,
                    "vendor_name": (
                        unidecode(display_name).strip().lower()
                        if display_name
                        else None
                    ),
                    "material_code": str(
                        row.get(find_col(col_map["material"]), "") or ""
                    ),
                    "short_text": str(
                        row.get(find_col(col_map["short_text"]), "") or ""
                    ),
                    "requested_qty": int(row.get(find_col(col_map["req_qty"]), 0)),
                    "requested_del_date": (
                        format_date(row.get(req_date_col)) if req_date_col else None
                    ),
                    "conf_delivery_date": (
                        format_date(row.get(conf_date_col)) if conf_date_col else None
                    ),
                    "rescheduling_date": (
                        format_date(row.get(reschedule_col)) if reschedule_col else None
                    ),
                    "unit": str(row.get(find_col(col_map["unit"]), "") or "EA"),
                    "unit_price": float(
                        str(row.get(find_col(col_map["price"]), 0)).replace(",", ".")
                    ),
                    "price_per_unit": int(
                        row.get(find_col(col_map["price_per"]), 1) or 1
                    ),
                    "total_amount": float(
                        str(row.get(find_col(col_map["amount"]), 0)).replace(",", ".")
                    ),
                    "currency": str(
                        row.get(find_col(col_map["currency"]), "") or "EUR"
                    ),
                }
            )

        # Database operations
        closed_count = 0
        with self.db.get_connection() as conn:
            cursor = conn.cursor()

            # Step 1: Get all currently open order lines from the DB (only if auto_close is enabled)
            if auto_close_missing:
                cursor.execute("SELECT po, item FROM open_orders WHERE status = 'Open'")
                fetched_rows = cursor.fetchall()
                existing_open_lines = set(
                    (row["po"], row["item"]) for row in fetched_rows
                )

            # Upsert vendors from the Excel file
            for vendor in vendors_to_upsert.values():
                cursor.execute(
                    "INSERT INTO vendors (vendor_name, display_name) VALUES (?, ?) ON CONFLICT(vendor_name) DO UPDATE SET display_name=excluded.display_name",
                    (vendor["vendor_name"], vendor["display_name"]),
                )

            # Upsert order lines from the Excel file
            for order in data_to_insert:
                cursor.execute(
                    """
                    INSERT INTO open_orders (po, item, vendor_name, material_code, short_text, requested_qty, requested_del_date,
                    conf_delivery_date, rescheduling_date, unit, unit_price, price_per_unit, total_amount, currency, status)
                    VALUES (:po, :item, :vendor_name, :material_code, :short_text, :requested_qty, :requested_del_date,
                    :conf_delivery_date, :rescheduling_date, :unit, :unit_price, :price_per_unit, :total_amount, :currency, 'Open')
                    ON CONFLICT(po, item) DO UPDATE SET
                    vendor_name=excluded.vendor_name, material_code=excluded.material_code, short_text=excluded.short_text,
                    requested_qty=excluded.requested_qty, requested_del_date=excluded.requested_del_date,
                    unit=excluded.unit, unit_price=excluded.unit_price, price_per_unit=excluded.price_per_unit, 
                    total_amount=excluded.total_amount, currency=excluded.currency,
                    conf_delivery_date=excluded.conf_delivery_date,
                    rescheduling_date=excluded.rescheduling_date,
                    status='Open'
                """,
                    order,
                )

            # Step 2: CONDITIONAL - Only close missing lines if auto_close_missing is True
            if auto_close_missing:
                # Level 3: Method body (8 spaces from class start)
                lines_to_close = existing_open_lines - lines_in_excel

                # Level 3: Check if there are lines to close (8 spaces)
                if lines_to_close:
                    # Level 4: Inside if block (12 spaces)
                    lines_to_close_list = list(lines_to_close)
                    update_query = "UPDATE open_orders SET status = 'Closed' WHERE po = ? AND item = ?"
                    cursor.executemany(update_query, lines_to_close_list)
                    closed_count = cursor.rowcount
                    print(f"✓ Auto-closed {closed_count} order lines (missing from upload)")
            else:
                # Level 3: If not auto-closing (8 spaces)
                print(f"ℹ️ INFO: Auto-close disabled - existing lines remain open.")

            conn.commit()

        # *** Batch update material prices from the uploaded order data ***
        updated_prices, created_materials = (
            self.batch_update_material_prices_from_orders()
        )

        print(
            f"ℹ️ INFO: Price update complete - Updated: {updated_prices} materials, Created: {created_materials} materials"
        )

        # Return the counts including price updates
        return (len(data_to_insert), closed_count, updated_prices, created_materials)

    def batch_update_material_prices_from_orders(self):
        """
        Batch update material prices from open_orders table.
        Updates net_price, price_per_unit, and currency for each material.
        """
        try:
            # Get unique materials with their latest pricing from open orders
            query = """
                SELECT 
                    material_code,
                    unit_price,
                    price_per_unit,
                    currency,
                    COUNT(*) as order_count
                FROM open_orders
                WHERE status = 'Open'
                AND material_code IS NOT NULL
                AND material_code != ''
                GROUP BY material_code, unit_price, price_per_unit, currency
                ORDER BY material_code
            """

            price_data = self.db.execute_query(query, fetchall=True)

            if not price_data:
                return 0, 0

            # Group by material code to handle multiple prices
            from collections import defaultdict

            material_prices = defaultdict(list)

            for row in price_data:
                material_code = row["material_code"]
                material_prices[material_code].append(
                    {
                        "unit_price": float(row["unit_price"] or 0),
                        "price_per_unit": int(row["price_per_unit"] or 1),
                        "currency": row["currency"] or "EUR",  # NEW
                        "order_count": row["order_count"],
                    }
                )

            updated_count = 0
            created_count = 0

            with self.db.get_connection() as conn:
                cursor = conn.cursor()

                for material_code, prices in material_prices.items():
                    # Use the most common price (highest order_count)
                    most_common = max(prices, key=lambda x: x["order_count"])
                    unit_price = most_common["unit_price"]
                    price_per = most_common["price_per_unit"]
                    currency = most_common["currency"]  # NEW

                    # Check if material exists
                    cursor.execute(
                        "SELECT material_code FROM materials WHERE material_code = ?",
                        (material_code,),
                    )
                    existing = cursor.fetchone()

                    if existing:
                        # Update existing material prices
                        cursor.execute(
                            """
                            UPDATE materials 
                            SET net_price = ?,
                                price_per_unit = ?,
                                currency = ?,
                                last_updated = ?
                            WHERE material_code = ?
                        """,
                            (
                                unit_price,
                                price_per,
                                currency,  # NEW
                                datetime.now().isoformat(),
                                material_code,
                            ),
                        )
                        updated_count += 1
                    else:
                        # Create new material with price data
                        # Get description from first open order
                        cursor.execute(
                            """
                            SELECT short_text, unit 
                            FROM open_orders 
                            WHERE material_code = ? 
                            AND status = 'Open'
                            LIMIT 1
                        """,
                            (material_code,),
                        )
                        order_info = cursor.fetchone()

                        description = order_info["short_text"] if order_info else ""
                        unit = order_info["unit"] if order_info else "EA"

                        cursor.execute(
                            """
                            INSERT INTO materials (
                                material_code, description, unit, net_price, price_per_unit,
                                currency, created_date, last_updated
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                            (
                                material_code,
                                description,
                                unit,
                                unit_price,
                                price_per,
                                currency,  # NEW
                                datetime.now().isoformat(),
                                datetime.now().isoformat(),
                            ),
                        )
                        created_count += 1

                conn.commit()

            return updated_count, created_count

        except Exception as e:
            print(f"Error in batch_update_material_prices_from_orders: {e}")
            import traceback

            traceback.print_exc()
            return 0, 0

    # --- PDF Generation ---
    def get_pending_po_data(self):
        """Get data for pending POs - Only orders that need PDFs created"""
        query = """
            SELECT oo.po, oo.item, oo.material_code, oo.short_text, oo.requested_qty, 
                   oo.requested_del_date, oo.conf_delivery_date, oo.rescheduling_date,
                   oo.unit, oo.unit_price, oo.total_amount, oo.currency,
                   v.display_name AS "Name", v.address as "Vendor Address", v.transport_days
            FROM open_orders oo
            JOIN vendors v ON oo.vendor_name = v.vendor_name
            WHERE oo.pdf_status = 'Pending' AND oo.status = 'Open'
            ORDER BY oo.po, oo.item
        """
        results = self.db.execute_query(query, fetchall=True)

        print(
            f"DEBUG: Found {len(results)} orders with pdf_status='Pending' for batch PDF generation"
        )
        for row in results:
            print(
                f"  PO {row.get('po')} Line {row.get('item')}: "
                f"Req={row.get('requested_del_date')}, Reschedule={row.get('rescheduling_date')}"
            )

        return results

    def mark_pos_as_created(self, po_numbers):
        if not po_numbers:
            return
        query = f"UPDATE open_orders SET pdf_status='Created', email_status='Pending' WHERE po IN ({','.join('?' for _ in po_numbers)})"
        self.db.execute_query(query, po_numbers, commit=True)

    def create_po_from_data(self, po_data):
        lines_df = pd.DataFrame(po_data["lines"])

        vendor_details = self.db.execute_query(
            "SELECT * FROM vendors WHERE display_name = ?",
            (po_data["vendor_display_name"],),
            fetchone=True,
        )
        if not vendor_details:
            raise ValueError(f"Vendor '{po_data['vendor_display_name']}' not found.")

        po_number = po_data["po_number"]

        # FIXED STEP 1: Fetch the 'currency' column along with other data
        existing_lines_query = """
            SELECT item, rescheduling_date, conf_delivery_date, price_per_unit, requested_del_date, total_amount, currency
            FROM open_orders 
            WHERE po = ?
        """
        existing_lines = self.db.execute_query(
            existing_lines_query, (po_number,), fetchall=True
        )

        # Create lookup dictionary for existing data
        existing_data = {}
        for line in existing_lines:
            # FIXED STEP 2: Store the currency in the lookup dictionary
            existing_data[str(line["item"])] = {
                "rescheduling_date": line["rescheduling_date"],
                "conf_delivery_date": line["conf_delivery_date"],
                "price_per_unit": line.get("price_per_unit", 1),
                "requested_del_date": line.get("requested_del_date", ""),
                "total_amount": line.get("total_amount", 0.0),
                "currency": line.get("currency", "EUR"),  # Added currency here
            }

        print(
            f"DEBUG: Found existing data for {len(existing_data)} lines in PO {po_number}"
        )

        # Add vendor details to each line for the PDF generator
        lines_df["Name"] = vendor_details.get("display_name", "")
        lines_df["Vendor Address"] = vendor_details.get("address", "")
        lines_df["transport_days"] = vendor_details.get("transport_days", 0)

        # REMOVED STEP 4: The incorrect hardcoded line is now gone.
        # lines_df["Currency"] = "oo.currency"

        # FIXED STEP 3: Populate the DataFrame with the correct currency data
        reschedule_dates = []
        conf_dates = []
        price_per_units = []
        requested_dates = []
        total_amounts = []
        currencies = []  # Added a list for currencies

        for _, row in lines_df.iterrows():
            item_str = str(row["item"])
            if item_str in existing_data:
                reschedule_dates.append(existing_data[item_str]["rescheduling_date"])
                conf_dates.append(existing_data[item_str]["conf_delivery_date"])
                price_per_units.append(existing_data[item_str]["price_per_unit"])
                requested_dates.append(existing_data[item_str]["requested_del_date"])
                total_amounts.append(existing_data[item_str]["total_amount"])
                currencies.append(
                    existing_data[item_str]["currency"]
                )  # Add currency from DB
            else:
                reschedule_dates.append(None)
                conf_dates.append(None)
                price_per_units.append(row.get("price_per_unit", 1))
                requested_dates.append(row.get("requested_del_date", ""))
                total_amounts.append(row.get("total_amount", 0.0))
                currencies.append(
                    row.get("currency", "EUR")
                )  # Fallback for new lines not in DB

        lines_df["rescheduling_date"] = reschedule_dates
        lines_df["conf_delivery_date"] = conf_dates
        lines_df["price_per_unit"] = price_per_units
        lines_df["total_amount"] = total_amounts
        lines_df["currency"] = currencies  # Add the new currency column

        if (
            "requested_del_date" not in lines_df.columns
            or lines_df["requested_del_date"].isna().all()
        ):
            lines_df["requested_del_date"] = requested_dates

        # Generate the PDF in memory
        pdf_buffer = io.BytesIO()
        if not self._generate_single_po_pdf(pdf_buffer, po_data["po_number"], lines_df):
            raise IOError("Failed to generate PDF buffer.")

        pdf_buffer.seek(0)
        return pdf_buffer

    def _generate_single_po_pdf(self, file_buffer, po_number, lines_df):
        """Internal PDF generation logic with company logo support"""
        try:
            if lines_df.empty:
                return False

            df = lines_df.copy()
            transport_days = int(df["transport_days"].iloc[0] or 0)

            # Convert date strings to datetime objects for calculation
            for col in ["rescheduling_date", "requested_del_date"]:
                if col in df.columns:
                    df[f"{col}_dt"] = pd.to_datetime(
                        df[col], format="%d.%m.%Y", errors="coerce"
                    )

            def get_etd_display_date(row):
                """Determines the correct ETD date to display on the PO line."""
                if pd.notna(row.get("rescheduling_date_dt")) and transport_days > 0:
                    etd_date = subtract_working_days(
                        row["rescheduling_date_dt"], transport_days
                    )
                    return etd_date
                elif pd.notna(row.get("rescheduling_date_dt")):
                    return row["rescheduling_date_dt"]
                elif pd.notna(row.get("requested_del_date_dt")):
                    return row["requested_del_date_dt"]
                return pd.NaT

            df["final_etd_date"] = df.apply(get_etd_display_date, axis=1)
            df["display_delivery_date_str"] = (
                df["final_etd_date"].dt.strftime("%d.%m.%Y").fillna("")
            )

            c = canvas.Canvas(file_buffer, pagesize=letter)
            width, height = letter
            config = self.get_config("company_config", {})

            vendor_name = df.iloc[0]["Name"]
            vendor_addr = (df.iloc[0]["Vendor Address"] or "").split("\n")
            currency = df.iloc[0].get("currency", "EUR")

            # Get vendor terms
            try:
                vendor_details = self.db.execute_query(
                    "SELECT delivery_terms, payment_terms FROM vendors v JOIN open_orders o ON v.vendor_name = o.vendor_name WHERE o.po = ? LIMIT 1",
                    (po_number,),
                    fetchone=True,
                )
                delivery_terms = (
                    vendor_details.get("delivery_terms", "") if vendor_details else ""
                )
                payment_terms = (
                    vendor_details.get("payment_terms", "") if vendor_details else ""
                )
            except Exception:
                delivery_terms = ""
                payment_terms = ""

            # ==============================================
            # MODIFIED: Company info with logo (top left)
            # ==============================================
            y = height - 0.7 * inch
            logo_path = config.get("company_logo_path", "")

            # Check if logo exists and draw it
            if logo_path and os.path.exists(logo_path):
                try:
                    from reportlab.lib.utils import ImageReader

                    # Draw logo at top left
                    logo_width = 1.5 * inch  # Adjust as needed
                    logo_height = 0.4 * inch  # Adjust as needed

                    c.drawImage(
                        logo_path,
                        0.5 * inch,
                        y,
                        width=logo_width,
                        height=logo_height,
                        preserveAspectRatio=True,
                        mask="auto",
                    )

                    # Move company name down below the logo
                    y -= logo_height + 0.1 * inch

                except Exception as e:
                    print(f"⚠ WARNING: Could not load logo: {e}")
                    # Continue without logo if there's an error

            # Company name and address (below logo if present)
            c.setFont("Helvetica-Bold", 10)
            c.drawString(0.5 * inch, y, config.get("my_company_name", ""))
            c.setFont("Helvetica", 9)
            y -= 0.15 * inch
            for line in config.get("my_company_address", "").split("\n")[:4]:
                c.drawString(0.5 * inch, y, line)
                y -= 0.12 * inch

            # Title and page number
            c.setFont("Helvetica-Bold", 18)
            c.drawString(3 * inch, height - 0.5 * inch, "PURCHASE ORDER")
            c.setFont("Helvetica", 9)
            c.drawRightString(8 * inch, height - 0.5 * inch, "Page: 1 of 1")

            # Header boxes - match SHIP TO box width exactly (4 to 8 inches)
            box_top = height - 1.8 * inch

            # Top row: PO Number and Date
            c.rect(
                4 * inch, box_top + 0.8 * inch, 2 * inch, 0.4 * inch, stroke=1, fill=0
            )
            c.rect(
                6 * inch, box_top + 0.8 * inch, 2 * inch, 0.4 * inch, stroke=1, fill=0
            )

            c.setFont("Helvetica-Bold", 7)
            c.drawString(4.1 * inch, box_top + 1.05 * inch, "PURCHASE ORDER NO.")
            c.drawString(6.1 * inch, box_top + 1.05 * inch, "PURCHASE ORDER DATE")

            c.setFont("Helvetica", 10)
            c.drawString(4.1 * inch, box_top + 0.87 * inch, str(po_number))
            c.drawString(
                6.1 * inch, box_top + 0.87 * inch, datetime.now().strftime("%d.%m.%Y")
            )

            # Second row: Buyer info
            c.rect(
                4 * inch, box_top + 0.4 * inch, 4 * inch, 0.35 * inch, stroke=1, fill=0
            )
            c.setFont("Helvetica-Bold", 7)
            c.drawString(4.1 * inch, box_top + 0.62 * inch, "BUYER")
            c.setFont("Helvetica", 9)
            buyer_name = config.get("buyer_name", "")
            buyer_email = config.get("buyer_email", "")
            c.drawString(
                4.1 * inch, box_top + 0.47 * inch, f"{buyer_name} / {buyer_email}"
            )

            # Third row: Delivery and Payment Terms
            c.rect(4 * inch, box_top, 2 * inch, 0.35 * inch, stroke=1, fill=0)
            c.rect(6 * inch, box_top, 2 * inch, 0.35 * inch, stroke=1, fill=0)

            c.setFont("Helvetica-Bold", 7)
            c.drawString(4.1 * inch, box_top + 0.22 * inch, "DELIVERY TERMS")
            c.drawString(6.1 * inch, box_top + 0.22 * inch, "PAYMENT TERMS")

            c.setFont("Helvetica", 8)
            c.drawString(
                4.1 * inch,
                box_top + 0.07 * inch,
                delivery_terms[:25] if delivery_terms else "",
            )
            c.drawString(
                6.1 * inch,
                box_top + 0.07 * inch,
                payment_terms[:25] if payment_terms else "",
            )

            # Vendor and Ship To section
            box_y = box_top - 1.5 * inch
            box_height = 1.3 * inch

            c.rect(0.5 * inch, box_y, 7.5 * inch, box_height, stroke=1, fill=0)
            c.line(4 * inch, box_y, 4 * inch, box_y + box_height)
            c.line(
                0.5 * inch,
                box_y + box_height - 0.25 * inch,
                8 * inch,
                box_y + box_height - 0.25 * inch,
            )

            # Headers
            c.setFont("Helvetica-Bold", 9)
            c.drawString(0.6 * inch, box_y + box_height - 0.17 * inch, "VEND.")
            c.drawString(4.1 * inch, box_y + box_height - 0.17 * inch, "SHIP TO")

            # Vendor info
            c.setFont("Helvetica", 9)
            y_pos = box_y + box_height - 0.42 * inch
            c.drawString(0.6 * inch, y_pos, f"Vendor: {vendor_name[:35]}")
            y_pos -= 0.15 * inch

            for line in vendor_addr[:5]:
                if y_pos > box_y + 0.1 * inch:
                    c.drawString(0.6 * inch, y_pos, line[:35])
                    y_pos -= 0.15 * inch

            # Ship To info
            y_pos = box_y + box_height - 0.42 * inch
            c.drawString(4.1 * inch, y_pos, config.get("ship_to_name", "")[:35])
            y_pos -= 0.15 * inch
            for line in config.get("ship_to_address", "").split("\n")[:5]:
                if y_pos > box_y + 0.1 * inch:
                    c.drawString(4.1 * inch, y_pos, line[:35])
                    y_pos -= 0.15 * inch

            # Line items table
            table_y = box_y - 0.4 * inch

            # Header row
            c.rect(
                0.5 * inch,
                table_y - 0.3 * inch,
                7.5 * inch,
                0.3 * inch,
                stroke=1,
                fill=0,
            )
            c.setFont("Helvetica-Bold", 7)

            c.drawString(0.6 * inch, table_y - 0.15 * inch, "LINE")
            c.drawString(1.0 * inch, table_y - 0.10 * inch, "PART")
            c.drawString(1.0 * inch, table_y - 0.21 * inch, "NUMBER")
            c.drawString(2.2 * inch, table_y - 0.15 * inch, "DESCRIPTION")
            c.drawCentredString(4.75 * inch, table_y - 0.15 * inch, "QTY")
            c.drawString(5.1 * inch, table_y - 0.15 * inch, "UNIT")
            c.drawString(5.6 * inch, table_y - 0.10 * inch, "UNIT")
            c.drawString(5.6 * inch, table_y - 0.21 * inch, "PRICE")
            c.drawString(6.7 * inch, table_y - 0.10 * inch, f"AMOUNT")
            c.drawString(6.7 * inch, table_y - 0.21 * inch, f"({currency})")
            c.drawCentredString(7.75 * inch, table_y - 0.10 * inch, "DELIVERY")
            c.drawCentredString(7.75 * inch, table_y - 0.21 * inch, "DATE")

            # Data rows
            y_pos = table_y - 0.55 * inch
            c.setFont("Helvetica", 8)
            total = 0
            currency = df.iloc[0].get("currency", "EUR") if len(df) > 0 else "EUR"

            for idx, (_, line) in enumerate(df.iterrows()):
                if y_pos < 1 * inch:  # New page if needed
                    c.showPage()
                    y_pos = height - 1 * inch
                    c.setFont("Helvetica", 9)

                c.line(0.5 * inch, y_pos - 0.05 * inch, 8 * inch, y_pos - 0.05 * inch)

                # LINE
                c.drawString(0.6 * inch, y_pos + 0.03 * inch, str(line.get("item", "")))

                # PART NUMBER
                c.drawString(
                    1.0 * inch,
                    y_pos + 0.03 * inch,
                    str(line.get("material_code", ""))[:15],
                )

                # DESCRIPTION
                c.drawString(
                    2.2 * inch,
                    y_pos + 0.03 * inch,
                    str(line.get("short_text", ""))[:30],
                )

                # QTY - Center aligned under QTY header
                c.drawCentredString(
                    4.75 * inch,
                    y_pos + 0.03 * inch,
                    f'{line.get("requested_qty", ""):,}',
                )

                # UNIT
                c.drawString(
                    5.1 * inch, y_pos + 0.03 * inch, str(line.get("unit", "EA"))
                )

                # UNIT PRICE - updated to show price per unit
                price_per = int(line.get("price_per_unit", 1))
                if price_per > 1:
                    c.drawString(
                        5.6 * inch,
                        y_pos + 0.03 * inch,
                        f'{line.get("unit_price", 0.0):.2f}/ {price_per} {line.get("unit", "EA")}',
                    )
                else:
                    c.drawString(
                        5.6 * inch,
                        y_pos + 0.03 * inch,
                        f'{line.get("unit_price", 0.0):.2f}/ {line.get("unit", "EA")}',
                    )

                # AMOUNT - Right aligned under AMOUNT header
                total_amount = line.get("total_amount", 0.0)
                print(
                    f"DEBUG PDF LINE {line.get('item', 'N/A')}: total_amount from DB = {total_amount}, type = {type(total_amount)}"
                )
                c.drawRightString(
                    7.2 * inch, y_pos + 0.03 * inch, f"{total_amount:,.2f}"
                )

                # DELIVERY DATE - Center aligned under DELIVERY DATE header
                delivery_date_str = str(line.get("display_delivery_date_str", ""))
                c.drawCentredString(7.75 * inch, y_pos + 0.03 * inch, delivery_date_str)

                # Add to total - USE DATABASE VALUE DIRECTLY
                total += line.get("total_amount", 0.0)
                y_pos -= 0.25 * inch

            # Bottom line
            c.line(0.5 * inch, y_pos + 0.20 * inch, 8 * inch, y_pos + 0.20 * inch)

            # Total
            y_pos -= 0.15 * inch
            c.setFont("Helvetica-Bold", 11)
            c.drawRightString(6.5 * inch, y_pos, f"TOTAL({currency}):")
            c.drawRightString(7.4 * inch, y_pos, f"{total:,.2f}")

            # Instructions / Terms and Conditions
            y_pos -= 0.4 * inch
            c.setFont("Helvetica", 8)

            # Get terms and conditions from config
            terms_and_conditions = config.get("terms_and_conditions", "")

            if terms_and_conditions:
                # Replace placeholders
                terms_and_conditions = terms_and_conditions.replace(
                    "{buyer_email}",
                    buyer_email if buyer_email else "purchasing@company.com",
                )
                terms_and_conditions = terms_and_conditions.replace(
                    "{po_number}", str(po_number)
                )
                terms_and_conditions = terms_and_conditions.replace(
                    "{buyer_name}", buyer_name if buyer_name else "Purchasing Team"
                )

                # Use custom terms and conditions
                terms_lines = terms_and_conditions.split("\n")
                for line in terms_lines:
                    if y_pos > 0.5 * inch:  # Check if we have space
                        c.drawString(0.5 * inch, y_pos, line)
                        y_pos -= 0.15 * inch
            else:
                # Use default instructions if no custom terms set
                instructions = [
                    "Please confirm this purchase order to above reference within 3 working days",
                    f"Please send the invoice to e-mail: {buyer_email if buyer_email else 'purchasing@company.com'}",
                    "Please always attach the delivery note with corresponding PO and item number to the box",
                ]

                for instr in instructions:
                    c.drawString(0.5 * inch, y_pos, instr)
                    y_pos -= 0.15 * inch

            # Footer
            y_pos -= 0.2 * inch
            c.setFont("Helvetica", 7)
            footer_text = f"Postal address: {config.get('my_company_name', '')} | {config.get('my_company_address', '').split(chr(10))[0] if config.get('my_company_address') else ''}"
            c.drawString(0.5 * inch, y_pos, footer_text[:120])

            c.save()
            return True
        except Exception as e:
            print(f"PDF generation error: {e}")
            import traceback

            traceback.print_exc()
            return False

    # --- Reschedule Management ---
    def generate_reschedule_files(self, filters):
        """Generate reschedule files with comprehensive filtering logic"""
        query = """
            SELECT oo.*, v.display_name as "Name", v.transport_days, v.transport_days_secondary
            FROM open_orders oo
            JOIN vendors v ON oo.vendor_name = v.vendor_name
            WHERE oo.status = 'Open'
        """

        all_data = self.db.execute_query(query, fetchall=True)
        if not all_data:
            return 0

        df = pd.DataFrame(all_data)

        # ÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚Â
        # STEP 1: DATA PREPARATION & VALIDATION
        # ÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚Â
        df["requested_del_date_dt"] = pd.to_datetime(
            df["requested_del_date"], format="%d.%m.%Y", errors="coerce"
        )
        df["rescheduling_date_dt"] = pd.to_datetime(
            df["rescheduling_date"], format="%d.%m.%Y", errors="coerce"
        )
        df["conf_delivery_date_dt"] = pd.to_datetime(
            df["conf_delivery_date"], format="%d.%m.%Y", errors="coerce"
        )
        df["total_amount"] = pd.to_numeric(df["total_amount"], errors="coerce").fillna(0)
        df["transport_days"] = pd.to_numeric(df["transport_days"], errors="coerce").fillna(0)
        df["transport_days_secondary"] = pd.to_numeric(
            df["transport_days_secondary"], errors="coerce"
        ).fillna(0)

        if "exception_message" not in df.columns:
            df["exception_message"] = ""

        # Check date presence
        df["has_conf_date"] = df["conf_delivery_date"].notna() & (
            df["conf_delivery_date"].str.strip() != ""
        )
        df["has_reschedule_date"] = df["rescheduling_date_dt"].notna()

        # ÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚Â
        # STEP 2: DETERMINE RESCHEDULE STATUS
        # ÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚Â
        df["reschedule_status"] = ""

        for idx, row in df.iterrows():
            if pd.notna(row["rescheduling_date_dt"]):
                transport_days = int(row.get("transport_days", 0) or 0)

                # Calculate NEW ETD from reschedule date (which is ETA)
                if transport_days > 0:
                    new_etd = subtract_working_days(row["rescheduling_date_dt"], transport_days)
                else:
                    new_etd = row["rescheduling_date_dt"]

                # Get ORIGINAL ETD
                original_etd = None
                if pd.notna(row["conf_delivery_date_dt"]):
                    original_etd = row["conf_delivery_date_dt"]
                elif pd.notna(row["requested_del_date_dt"]):
                    original_etd = row["requested_del_date_dt"]

                # Compare ETDs to determine status
                if original_etd:
                    if new_etd < original_etd:
                        df.at[idx, "reschedule_status"] = "Reschedule In"
                    elif new_etd > original_etd:
                        df.at[idx, "reschedule_status"] = "Reschedule Out"
                    else:
                        df.at[idx, "reschedule_status"] = "No Change"
            else:
                # Orders WITHOUT reschedule dates
                if row["has_conf_date"]:
                    df.at[idx, "reschedule_status"] = "No Reschedule Date"
                else:
                    df.at[idx, "reschedule_status"] = "Unconfirmed"

        # ÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚Â
        # STEP 3: GET FILTER SETTINGS
        # ÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚Â
        include_all_open = filters.get(
            "include_all_open_orders", filters.get("include_unconfirmed_orders", False)
        )
        use_pdf_format = filters.get("use_pdf_format", False)

        # ÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚Â
        # STEP 4: APPLY SUPPLIER FILTERS (applies to all orders)
        # ÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚Â
        if filters.get("suppliers"):
            df = df[df["Name"].isin(filters["suppliers"])]

        # ÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚Â
        # STEP 5: SEPARATE INTO CATEGORIES
        # ÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚Â
        df_in = df[df["reschedule_status"] == "Reschedule In"].copy()
        df_out = df[df["reschedule_status"] == "Reschedule Out"].copy()
        df_no_change = df[df["reschedule_status"] == "No Change"].copy()
        df_no_reschedule = df[df["reschedule_status"] == "No Reschedule Date"].copy()
        df_unconfirmed = df[df["reschedule_status"] == "Unconfirmed"].copy()

        # ÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚Â
        # STEP 6: APPLY CATEGORY-SPECIFIC FILTERS
        # ÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚Â
        
        # Filter Reschedule IN orders by date range
        if filters.get("reschedule_in_req_date_start"):
            start_date = pd.to_datetime(filters["reschedule_in_req_date_start"])
            df_in = df_in[df_in["requested_del_date_dt"] >= start_date]
        if filters.get("reschedule_in_req_date_end"):
            end_date = pd.to_datetime(filters["reschedule_in_req_date_end"])
            df_in = df_in[df_in["requested_del_date_dt"] <= end_date]

        # Filter Reschedule OUT orders by date range AND value
        if filters.get("reschedule_out_req_date_start"):
            start_date = pd.to_datetime(filters["reschedule_out_req_date_start"])
            df_out = df_out[df_out["requested_del_date_dt"] >= start_date]
        if filters.get("reschedule_out_req_date_end"):
            end_date = pd.to_datetime(filters["reschedule_out_req_date_end"])
            df_out = df_out[df_out["requested_del_date_dt"] <= end_date]
        if filters.get("reschedule_out_value_min") is not None:
            df_out = df_out[df_out["total_amount"] >= filters["reschedule_out_value_min"]]
        if filters.get("reschedule_out_value_max") is not None:
            df_out = df_out[df_out["total_amount"] <= filters["reschedule_out_value_max"]]

        # ÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚Â
        # STEP 7: COMBINE FINAL DATAFRAME BASED ON INCLUDE_ALL_OPEN SETTING
        # ÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚Â
        if include_all_open:
            # Include ALL categories
            df_final = pd.concat(
                [df_in, df_out, df_no_change, df_no_reschedule, df_unconfirmed], 
                ignore_index=True
            )
        else:
            # Only include reschedule IN and OUT
            df_final = pd.concat([df_in, df_out], ignore_index=True)

        if df_final.empty:
            return 0

        # ÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚Â
        # STEP 8: APPLY MANUAL RESCHEDULE DATE OVERRIDE (if specified)
        # ÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚Â
        manual_date = filters.get("manual_reschedule_date")
        if manual_date and not df_final.empty:
            manual_date_str = manual_date.strftime("%d.%m.%Y")

            # Set the manual date directly as the ETD
            for idx, row in df_final.iterrows():
                df_final.at[idx, "rescheduling_date_dt"] = manual_date
                df_final.at[idx, "rescheduling_date"] = manual_date_str
                df_final.at[idx, "manual_etd_override"] = True

            # RECALCULATE RESCHEDULE STATUS with manual ETD
            df_final["reschedule_status"] = ""

            for idx, row in df_final.iterrows():
                transport_days = int(row.get("transport_days", 0) or 0)

                # Calculate original ETD for comparison
                original_etd = None
                if pd.notna(row.get("conf_delivery_date_dt")):
                    original_etd = (
                        subtract_working_days(row["conf_delivery_date_dt"], transport_days)
                        if transport_days > 0
                        else row["conf_delivery_date_dt"]
                    )
                elif pd.notna(row.get("requested_del_date_dt")):
                    original_etd = (
                        subtract_working_days(row["requested_del_date_dt"], transport_days)
                        if transport_days > 0
                        else row["requested_del_date_dt"]
                    )

                # Compare manual ETD to original ETD
                if original_etd:
                    if manual_date < original_etd:
                        df_final.at[idx, "reschedule_status"] = "Reschedule In"
                    elif manual_date > original_etd:
                        df_final.at[idx, "reschedule_status"] = "Reschedule Out"
                    else:
                        df_final.at[idx, "reschedule_status"] = "No Change"

            # Add note to exception_message
            df_final["exception_message"] = df_final["exception_message"].fillna("").astype(str)
            override_note = f" [Manual override: ETD {manual_date_str}]"

            mask_needs_note = ~df_final["exception_message"].str.contains(
                "Manual override", na=False
            )
            df_final.loc[mask_needs_note, "exception_message"] = (
                df_final.loc[mask_needs_note, "exception_message"] + override_note
            )
        else:
            # No manual override
            df_final["manual_etd_override"] = False

        df_final["exception_message"] = df_final["exception_message"].fillna("").astype(str)

        # ÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚Â
        # STEP 9: CREATE OUTPUT FILES
        # ÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚Â
        supplier_groups = df_final.groupby("Name")
        files_created_count = 0
        file_extension = ".pdf" if use_pdf_format else ".xlsx"

        if len(supplier_groups) > 1:
            # Multiple suppliers - create ZIP
            zip_path = os.path.join(
                RESCHEDULE_OUTPUT_FOLDER,
                f"Reschedule_Files_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.zip",
            )
            with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
                for supplier_name, lines_df in supplier_groups:
                    safe_name = re.sub(r'[\\/*?:"<>|]', "_", supplier_name)

                    if use_pdf_format:
                        # Create PDF in temp location
                        temp_pdf = os.path.join(
                            RESCHEDULE_OUTPUT_FOLDER, f"temp_{safe_name}.pdf"
                        )
                        if self._create_reschedule_pdf(temp_pdf, supplier_name, lines_df):
                            with open(temp_pdf, "rb") as f:
                                zf.writestr(f"{safe_name}.pdf", f.read())
                            os.remove(temp_pdf)
                            files_created_count += 1
                    else:
                        # Create Excel
                        temp_excel = os.path.join(
                            RESCHEDULE_OUTPUT_FOLDER, f"temp_{safe_name}.xlsx"
                        )
                        self._create_reschedule_excel(temp_excel, supplier_name, lines_df)
                        with open(temp_excel, "rb") as f:
                            zf.writestr(f"{safe_name}.xlsx", f.read())
                        os.remove(temp_excel)
                        files_created_count += 1

            messagebox.showinfo("✅ Success",
                f"Created ZIP file with {files_created_count} reschedule files\n{zip_path}",
            )
        else:
            # Single supplier - create single file
            for supplier_name, lines_df in supplier_groups:
                safe_name = re.sub(r'[\\/*?:"<>|]', "_", supplier_name)
                timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                file_path = os.path.join(
                    RESCHEDULE_OUTPUT_FOLDER, f"{safe_name}_Reschedule_{timestamp}{file_extension}"
                )

                if use_pdf_format:
                    if self._create_reschedule_pdf(file_path, supplier_name, lines_df):
                        files_created_count += 1
                else:
                    self._create_reschedule_excel(file_path, supplier_name, lines_df)
                    files_created_count += 1

            if files_created_count > 0:
                messagebox.showinfo("✅ Success", f"Created reschedule file\n{file_path}"
                )

        return files_created_count

    def _create_reschedule_excel(self, file_path, supplier_name, lines_df):
        """Create Excel file for reschedule data"""
        try:
            group_df = lines_df.copy()
            today = datetime.now()  # For comparing ETD dates

            # Ensure all text fields are properly converted to strings and handle None/NaN values
            group_df["comments"] = group_df["comments"].fillna("").astype(str)
            group_df["exception_message"] = (
                group_df["exception_message"].fillna("").astype(str)
            )
            group_df["material_code"] = group_df["material_code"].fillna("").astype(str)
            group_df["short_text"] = group_df["short_text"].fillna("").astype(str)
            group_df["conf_delivery_date"] = (
                group_df["conf_delivery_date"].fillna("").astype(str)
            )

            # Get transport days for this supplier group
            primary_days = int(group_df["transport_days"].iloc[0] or 0)
            secondary_days = int(group_df["transport_days_secondary"].iloc[0] or 0)

            def calculate_etd_and_notes(row):
                """Calculates ETD date and generates exception notes."""
                reschedule_dt = row["rescheduling_date_dt"]
                current_exception_notes = str(row.get("exception_message") or "")

                if pd.isna(reschedule_dt):
                    return pd.NaT, current_exception_notes

                # CHECK MANUAL OVERRIDE FLAG - if True, use date directly
                if row.get("manual_etd_override", False):
                    return reschedule_dt, current_exception_notes

                # Normal calculation
                final_etd = subtract_working_days(reschedule_dt, primary_days)
                note = f"ETD calculated using primary transport ({primary_days} days)."

                if final_etd < today and secondary_days > 0:
                    final_etd = subtract_working_days(reschedule_dt, secondary_days)
                    note = f"Used secondary transport ({secondary_days} days)."
                elif primary_days == 0:
                    note = "No transport days configured - ETD not calculated."

                if (
                    current_exception_notes
                    and note
                    and note not in current_exception_notes
                ):
                    updated_notes = f"{current_exception_notes}; {note}"
                else:
                    updated_notes = current_exception_notes or note

                return final_etd, updated_notes

            # Apply the function to each row
            results = group_df.apply(calculate_etd_and_notes, axis=1)
            group_df[["etd_date", "exception_message"]] = pd.DataFrame(
                results.tolist(), index=group_df.index
            )
            group_df["etd_date_str"] = (
                group_df["etd_date"].dt.strftime("%d.%m.%Y").fillna("")
            )

            # Prepare output DataFrame - removed 'rescheduling_date' column
            output_df = group_df[
                [
                    "po",
                    "item",
                    "material_code",
                    "short_text",
                    "requested_qty",
                    "total_amount",
                    "currency",
                    "requested_del_date",
                    "conf_delivery_date",
                    "etd_date_str",
                    "reschedule_status",
                    "comments",
                    "exception_message",
                ]
            ].copy()

            # Modified column names
            output_df.rename(
                columns={
                    "po": "PO Number",
                    "item": "Item",
                    "material_code": "Material",
                    "short_text": "Description",
                    "requested_qty": "Quantity",
                    "total_amount": "Value",
                    "currency": "Currency",
                    "requested_del_date": "Requested Date",
                    "conf_delivery_date": "Confirmed Date",
                    "etd_date_str": "Reschedule To (ETD)",
                    "reschedule_status": "Status",
                    "comments": "Supplier Comments",
                    "exception_message": "Exception Notes",
                },
                inplace=True,
            )

            # Create Excel file and write to file_path
            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                output_df.to_excel(writer, sheet_name="Reschedule", index=False)

                # Get the worksheet to apply formatting
                worksheet = writer.sheets["Reschedule"]

                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except Exception:  # TODO: Add proper error handling
                            pass  # TODO: Add proper error handling
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                # Add table formatting
                if len(output_df) > 0:
                    table = Table(
                        displayName="RescheduleTable",
                        ref=f"A1:{get_column_letter(len(output_df.columns))}{len(output_df) + 1}",
                    )
                    style = TableStyleInfo(
                        name="TableStyleMedium9",
                        showFirstColumn=False,
                        showLastColumn=False,
                        showRowStripes=True,
                        showColumnStripes=True,
                    )
                    table.tableStyleInfo = style
                    worksheet.add_table(table)

            return True

        except Exception as e:
            print(f"Error creating Excel for {supplier_name}: {e}")
            return False

    def _create_reschedule_pdf(self, pdf_path, vendor_name, lines_df):
        """Create a PDF reschedule file for a vendor"""
        try:
            c = canvas.Canvas(pdf_path, pagesize=letter)
            width, height = letter
            config = self.get_config("company_config", {})

            # Get transport days for ETD calculation
            transport_days = (
                int(lines_df["transport_days"].iloc[0] or 0) if len(lines_df) > 0 else 0
            )
            transport_days_secondary = (
                int(lines_df["transport_days_secondary"].iloc[0] or 0)
                if len(lines_df) > 0
                else 0
            )

            # Header
            c.setFont("Helvetica-Bold", 18)
            c.drawString(0.5 * inch, height - 0.5 * inch, "DELIVERY SCHEDULE UPDATE")

            c.setFont("Helvetica", 11)
            c.drawString(
                0.5 * inch,
                height - 0.8 * inch,
                f"Date: {datetime.now().strftime('%d.%m.%Y')}",
            )
            c.drawString(0.5 * inch, height - 1.0 * inch, f"Supplier: {vendor_name}")

            if transport_days > 0:
                c.setFont("Helvetica", 10)
                c.drawString(
                    0.5 * inch,
                    height - 1.2 * inch,
                    f"Transport Days: {transport_days} working days",
                )
                if transport_days_secondary > 0:
                    c.drawString(
                        0.5 * inch,
                        height - 1.35 * inch,
                        f"Fast Transport: {transport_days_secondary} working days",
                    )

            # Message
            y_pos = (
                height - 1.6 * inch
                if transport_days == 0
                else (
                    height - 1.8 * inch
                    if transport_days_secondary == 0
                    else height - 2.0 * inch
                )
            )
            c.setFont("Helvetica", 10)
            message_lines = [
                "Dear Supplier,",
                "",
                "We are requesting a reschedule for the following purchase order line items.",
                "Please review the proposed delivery dates and confirm your ability to meet them.",
                "",
            ]

            for line in message_lines:
                c.drawString(0.5 * inch, y_pos, line)
                y_pos -= 0.2 * inch

            # Group by PO and status
            po_groups = {}
            for _, row in lines_df.iterrows():
                po_num = row.get("po", "Unknown")
                status = row.get("reschedule_status", "Unknown")

                if po_num not in po_groups:
                    po_groups[po_num] = {
                        "lines": [],
                        "currency": row.get("currency", "EUR"),
                        "total": 0,
                    }
                po_groups[po_num]["lines"].append(row)
                po_groups[po_num]["total"] += float(row.get("total_amount", 0))

            # List all POs and their lines
            y_pos -= 0.2 * inch

            for po_num, po_data in sorted(po_groups.items()):
                lines = po_data["lines"]
                currency = po_data["currency"]

                # Check if we need a new page
                if y_pos < 2 * inch:
                    c.showPage()
                    y_pos = height - 1 * inch

                # PO Header
                c.setFont("Helvetica-Bold", 12)
                c.drawString(0.5 * inch, y_pos, f"PO {po_num}")
                y_pos -= 0.25 * inch

                # Line items
                c.setFont("Helvetica", 9)

                for line in lines:
                    # Check page space
                    if y_pos < 1 * inch:
                        c.showPage()
                        y_pos = height - 1 * inch
                        c.setFont("Helvetica", 9)

                    # Determine original date (confirmed if exists, otherwise requested)
                    original_date = ""
                    original_label = "Original"
                    if (
                        line.get("conf_delivery_date")
                        and str(line.get("conf_delivery_date")).strip()
                    ):
                        original_date = line.get("conf_delivery_date")
                        original_label = "Confirmed"
                    else:
                        original_date = line.get("requested_del_date", "N/A")
                        original_label = "Requested"

                    # Calculate proposed ETD from rescheduling date
                    proposed_date = ""
                    etd_label = "Proposed ETD"
                    reschedule_date_str = line.get("rescheduling_date")

                    if reschedule_date_str and reschedule_date_str.strip():
                        try:
                            reschedule_date = datetime.strptime(
                                reschedule_date_str, "%d.%m.%Y"
                            )

                            # CHECK IF MANUAL OVERRIDE - if so, don't subtract transport days
                            if line.get("manual_etd_override", False):
                                etd_date = reschedule_date
                                date_label = "ETD (Manual)"
                            elif transport_days > 0:
                                etd_date = subtract_working_days(
                                    reschedule_date, transport_days
                                )
                                date_label = "ETD"
                            else:
                                etd_date = reschedule_date
                                date_label = "ETD"

                            display_date = etd_date.strftime("%d.%m.%Y")

                            # Calculate ETD with primary transport days
                            etd_date = subtract_working_days(
                                reschedule_date, transport_days
                            )

                            # If primary ETD is in the past and secondary exists, use secondary
                            if etd_date < today and transport_days_secondary > 0:
                                etd_date = subtract_working_days(
                                    reschedule_date, transport_days_secondary
                                )
                                etd_label = "Proposed ETD (Fast)"

                            proposed_date = etd_date.strftime("%d.%m.%Y")
                        except Exception:
                            proposed_date = reschedule_date_str

                    # Status indicator
                    status = line.get("reschedule_status", "")
                    status_symbol = ""
                    if status == "Reschedule In":
                        status_symbol = " EARLIER"
                    elif status == "Reschedule Out":
                        status_symbol = " LATER"
                    elif status == "Unconfirmed":
                        status_symbol = " UNCONFIRMED"

                    # Bullet point with line details
                    line_text = f" Line {line.get('item', 'N/A')}  {line.get('material_code', 'N/A')}  {line.get('short_text', 'N/A')[:40]}"
                    c.drawString(0.75 * inch, y_pos, line_text)
                    y_pos -= 0.15 * inch

                    # Quantity and value on second line
                    detail_text = f"  Qty: {line.get('requested_qty', 'N/A')} {line.get('unit', 'EA')} | "
                    detail_text += (
                        f"Value: {float(line.get('total_amount', 0)):.2f} {currency}"
                    )
                    c.drawString(0.75 * inch, y_pos, detail_text)
                    y_pos -= 0.15 * inch

                    # Date change line with arrow
                    if proposed_date:
                        date_change = f"  {original_label} Date: {original_date}  {etd_label}: {proposed_date}"
                        if status_symbol:
                            date_change += f"  [{status_symbol}]"
                        c.drawString(0.75 * inch, y_pos, date_change)
                        y_pos -= 0.2 * inch
                    else:
                        date_line = f"  {original_label} Date: {original_date}"
                        if status_symbol:
                            date_line += f"  [{status_symbol}]"
                        c.drawString(0.75 * inch, y_pos, date_line)
                        y_pos -= 0.2 * inch

                    # Exception message if exists
                    exception_msg = line.get("exception_message", "")
                    if exception_msg and str(exception_msg).strip():
                        c.setFont("Helvetica-Oblique", 8)
                        c.drawString(
                            0.75 * inch, y_pos, f"  Note: {exception_msg[:80]}"
                        )
                        y_pos -= 0.15 * inch
                        c.setFont("Helvetica", 9)

                    y_pos -= 0.1 * inch

                # PO Total
                c.setFont("Helvetica-Bold", 10)
                c.drawString(
                    0.75 * inch, y_pos, f"PO Total: {po_data['total']:,.2f} {currency}"
                )
                y_pos -= 0.4 * inch

            # Calculate grand total
            grand_total = sum(po_data["total"] for po_data in po_groups.values())
            all_currency = (
                list(po_groups.values())[0]["currency"] if po_groups else "EUR"
            )

            # Grand total
            if y_pos < 1.5 * inch:
                c.showPage()
                y_pos = height - 1 * inch

            y_pos -= 0.2 * inch
            c.setFont("Helvetica-Bold", 12)
            c.drawString(
                0.5 * inch, y_pos, f"GRAND TOTAL: {grand_total:,.2f} {all_currency}"
            )

            # Footer
            y_pos -= 0.6 * inch
            c.setFont("Helvetica", 10)
            buyer_email = config.get("buyer_email", "purchasing@company.com")
            footer_lines = [
                "",
                "Please review the proposed delivery dates and confirm your acceptance.",
                "If you cannot meet these dates, please provide alternative delivery dates.",
                f"Contact: {buyer_email}",
                "",
            ]

            if transport_days > 0:
                footer_lines.append(
                    f"Note: ETD dates are calculated by subtracting {transport_days} working days from the requested arrival date."
                )
                if transport_days_secondary > 0:
                    footer_lines.append(
                        f"      Fast transport option ({transport_days_secondary} days) is used when standard ETD would be in the past."
                    )
                footer_lines.append("")

            footer_lines.append("Thank you for your cooperation.")

            for line in footer_lines:
                if y_pos > 0.5 * inch:
                    c.drawString(0.5 * inch, y_pos, line)
                    y_pos -= 0.2 * inch
                else:
                    c.showPage()
                    y_pos = height - 1 * inch
                    c.setFont("Helvetica", 10)
                    c.drawString(0.5 * inch, y_pos, line)
                    y_pos -= 0.2 * inch

            c.save()
            return True

        except Exception as e:
            print(f"✗ ERROR: Failed to create reschedule PDF: {e}")
            import traceback

            traceback.print_exc()
            return False

    # --- Email Management ---

    def get_pending_pos_with_portal_info(self):
        query = """
            SELECT oo.po, v.display_name as supplier, v.emails, v.api_key
            FROM open_orders oo
            JOIN vendors v ON oo.vendor_name = v.vendor_name
            WHERE oo.pdf_status = 'Created' AND oo.email_status = 'Pending' AND oo.status = 'Open'
            GROUP BY oo.po, v.display_name, v.emails, v.api_key
        """
        return self.db.execute_query(query, fetchall=True)

    def mark_email_sent(self, po_number):
        return (
            self.db.execute_query(
                "UPDATE open_orders SET email_status='Sent' WHERE po=?",
                (po_number,),
                commit=True,
            )
            > 0
        )

    # --- Configuration Management ---
    def get_config(self, key, default=None):
        result = self.db.execute_query(
            "SELECT value FROM app_config WHERE key=?", (key,), fetchone=True
        )
        return json.loads(result["value"]) if result else default

    def save_config(self, key, value):
        self.db.execute_query(
            "INSERT INTO app_config (key, value) VALUES (?, ?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
            (key, json.dumps(value)),
            commit=True,
        )

    def get_signature(self):
        """Get email signature HTML from database"""
        try:
            # Ensure table exists
            self.db.execute_query(
                """CREATE TABLE IF NOT EXISTS email_signatures (
                    id INTEGER PRIMARY KEY,
                    html_content TEXT,
                    last_updated TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )""",
                ()
            )
            
            # Get signature
            result = self.db.execute_query(
                "SELECT html_content FROM email_signatures WHERE id = 1",
                (),
                fetchall=False
            )
            
            return result['html_content'] if result and result.get('html_content') else ""
            
        except Exception as e:
            print(f"Error getting signature: {e}")
            return ""


class ForecastDataManager:
    """Handles all forecast-related business logic"""

    def __init__(self, db_manager):
        self.db = db_manager

    def upload_forecast_file(self, file_path, vendor_name=None):
        """Upload vendor forecast from Excel file"""
        try:
            df = pd.read_excel(file_path, engine=None)
            df.columns = [str(c).strip() for c in df.columns]

            material_col = self._find_column(
                df, ["Material", "Material Code", "Part Number", "Material Number"]
            )
            desc_col = self._find_column(
                df, ["Description", "Short Text", "Material Description", "Short text"]
            )
            vendor_col = self._find_column(
                df,
                ["Vendor", "Supplier", "Vendor Name", "Name of Vendor", "Fixed Vendor"],
            )
            price_col = self._find_column(
                df, ["Unit Price", "Price", "Price/Unit", "Net Order Price"]
            )
            unit_col = self._find_column(
                df, ["Unit", "UoM", "Unit of Measure", "Unit of Measure"]
            )

            date_columns = [
                col
                for col in df.columns
                if "week" in col.lower()
                or "month" in col.lower()
                or any(month in col.lower() for month in calendar.month_abbr[1:])
            ]

            if not material_col or not date_columns:
                raise ValueError(
                    "Could not identify required columns (Material and Date/Week columns)"
                )

            forecasts_to_insert = []
            current_year = datetime.now().year

            for _, row in df.iterrows():
                material = str(row.get(material_col, "")).strip()
                if not material or material.lower() in ["nan", "none", ""]:
                    continue

                description = str(row.get(desc_col, "")) if desc_col else ""
                unit_price = (
                    float(str(row.get(price_col, 0)).replace(",", "."))
                    if price_col
                    else 0
                )
                unit = str(row.get(unit_col, "EA")) if unit_col else "EA"

                if vendor_name:
                    vendor = vendor_name
                elif vendor_col:
                    vendor = str(row.get(vendor_col, "")).strip()
                else:
                    vendor = None

                for date_col in date_columns:
                    qty = row.get(date_col)
                    if pd.isna(qty) or qty == 0:
                        continue

                    try:
                        qty = int(float(qty))
                    except Exception:
                        continue

                    forecast_date, week_num, month_num = self._parse_date_from_column(
                        date_col, current_year
                    )

                    if forecast_date:
                        forecasts_to_insert.append(
                            {
                                "vendor_name": vendor,
                                "material_code": material,
                                "short_text": description,
                                "forecast_date": forecast_date.strftime("%Y-%m-%d"),
                                "forecast_qty": qty,
                                "unit": unit,
                                "unit_price": unit_price,
                                "total_amount": qty * unit_price,
                                "currency": "EUR",
                                "week_number": week_num,
                                "month_number": month_num,
                                "year_number": forecast_date.year,
                            }
                        )

            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                for forecast in forecasts_to_insert:
                    cursor.execute(
                        """
                        INSERT INTO forecasts (vendor_name, material_code, short_text, forecast_date,
                        forecast_qty, unit, unit_price, total_amount, currency, week_number, month_number, year_number)
                        VALUES (:vendor_name, :material_code, :short_text, :forecast_date,
                        :forecast_qty, :unit, :unit_price, :total_amount, :currency, :week_number, :month_number, :year_number)
                    """,
                        forecast,
                    )
                conn.commit()

            return len(forecasts_to_insert)

        except Exception as e:
            raise ValueError(f"Failed to process forecast file: {str(e)}")

    def upload_requisitions_file(self, file_path):
        """
        Upload requisitions from Excel file.
        ENHANCED: Automatically creates/updates material master data from requisitions
        - Upserts requisition lines found in the file, ensuring their status is 'Open'.
        - Marks any existing open requisition lines in the database that are NOT in the file as 'Closed'.
        - Auto-creates materials in master data with lead time, vendor, and description
        """
        try:
            df = pd.read_excel(file_path, engine=None)
            df.columns = [str(c).strip() for c in df.columns]

            col_map = {
                "pr": [
                    "PR",
                    "PR Number",
                    "Requisition",
                    "Req Number",
                    "Purchase Requisition",
                    "Req. Tracking Number",
                ],
                "item": ["Item", "Line", "Position", "Item Number", "Line Item"],
                "material": [
                    "Material",
                    "Material Code",
                    "Part Number",
                    "Material Number",
                ],
                "desc": [
                    "Description",
                    "Short Text",
                    "Material Description",
                    "Short text",
                ],
                "vendor": [
                    "Vendor",
                    "Supplier",
                    "Vendor Name",
                    "Name of Vendor",
                    "Fixed Vendor",
                ],
                "qty": ["Quantity", "Qty", "Requested Quantity", "Quantity requested"],
                "date": [
                    "Delivery Date",
                    "Required Date",
                    "Del Date",
                    "Planned Deliv. Time",
                    "Delivery date",
                ],
                "price": ["Unit Price", "Price", "Price/Unit", "Net Order Price"],
                "unit": ["Unit", "UoM", "Unit of Measure"],
                "lead_time": [
                    "Planned Deliv. Time",
                    "Lead Time",
                    "Delivery Time",
                    "Planned Delivery Time",
                ],  # NEW
            }

            pr_col = self._find_column(df, col_map["pr"])
            item_col = self._find_column(df, col_map["item"])
            material_col = self._find_column(df, col_map["material"])
            desc_col = self._find_column(df, col_map["desc"])
            vendor_col = self._find_column(df, col_map["vendor"])
            qty_col = self._find_column(df, col_map["qty"])
            date_col = self._find_column(df, col_map["date"])
            price_col = self._find_column(df, col_map["price"])
            unit_col = self._find_column(df, col_map["unit"])
            lead_time_col = self._find_column(df, col_map["lead_time"])  # NEW

            if not pr_col or not material_col:
                raise ValueError(
                    "Could not identify required columns (PR Number and Material)"
                )

            requisitions_to_insert = []
            vendors_to_upsert = {}
            lines_in_excel = set()
            materials_to_upsert = {}  # NEW: Track materials to add/update

            # Track item numbers per PR if item column doesn't exist
            pr_item_counters = {}

            for _, row in df.iterrows():
                pr_num = str(row.get(pr_col, "")).strip()
                if not pr_num or pr_num.lower() in ["nan", "none", ""]:
                    continue

                # Handle item number
                if item_col:
                    item_num_val = row.get(item_col)
                    if pd.isna(item_num_val):
                        continue
                    item_num = (
                        str(int(item_num_val))
                        if isinstance(item_num_val, (int, float))
                        else str(item_num_val).strip()
                    )
                else:
                    if pr_num not in pr_item_counters:
                        pr_item_counters[pr_num] = 10
                    else:
                        pr_item_counters[pr_num] += 10
                    item_num = str(pr_item_counters[pr_num])

                lines_in_excel.add((pr_num, item_num))

                # Extract data
                material_code = str(row.get(material_col, "")).strip()
                if not material_code or material_code.lower() in ["nan", "none", ""]:
                    continue

                description = str(row.get(desc_col, "")) if desc_col else ""
                vendor_display = (
                    str(row.get(vendor_col, "")).strip() if vendor_col else ""
                )
                unit = str(row.get(unit_col, "EA")) if unit_col else "EA"

                # NEW: Extract lead time (in days)
                lead_time_days = 0
                if lead_time_col:
                    lead_time_val = row.get(lead_time_col)
                    if pd.notna(lead_time_val):
                        try:
                            # Handle various formats: "14", "14 days", "2 weeks"
                            lead_time_str = str(lead_time_val).lower().strip()
                            if "week" in lead_time_str:
                                # Extract number and convert weeks to days
                                weeks = int("".join(filter(str.isdigit, lead_time_str)))
                                lead_time_days = weeks * 7
                            else:
                                # Assume it's already in days
                                lead_time_days = int(
                                    "".join(filter(str.isdigit, lead_time_str))
                                )
                        except Exception:
                            lead_time_days = 0

                # Upsert vendor
                vendor_name = None
                if vendor_display:
                    vendor_name = unidecode(vendor_display).strip().lower()
                    if vendor_name not in vendors_to_upsert:
                        vendors_to_upsert[vendor_name] = {
                            "display_name": vendor_display,
                            "vendor_name": vendor_name,
                        }

                # NEW: Track material for master data creation
                if material_code not in materials_to_upsert:
                    materials_to_upsert[material_code] = {
                        "material_code": material_code,
                        "description": description,
                        "unit": unit,
                        "lead_time_days": lead_time_days,
                        "preferred_vendor": vendor_display if vendor_display else None,
                        "safety_stock": 0,  # Default
                        "min_order_qty": 1,  # Default
                        "lot_size_rule": "LOT_FOR_LOT",  # Default
                    }

                # Format date
                def format_date(val):
                    if pd.isna(val):
                        return None
                    try:
                        return pd.to_datetime(val).strftime("%d.%m.%Y")
                    except Exception:
                        return None

                req_date = format_date(row.get(date_col)) if date_col else None

                requisitions_to_insert.append(
                    {
                        "req_number": pr_num,
                        "item": item_num,
                        "vendor_name": vendor_name,
                        "material_code": material_code,
                        "short_text": description,
                        "requested_qty": int(row.get(qty_col, 0)) if qty_col else 0,
                        "unit": unit,
                        "requested_del_date": req_date,
                        "unit_price": (
                            float(str(row.get(price_col, 0)).replace(",", "."))
                            if price_col
                            else 0
                        ),
                        "status": "Open",
                        "pr_status": "Open",
                    }
                )

            # Database operations
            closed_count = 0
            materials_created = 0
            materials_updated = 0

            with self.db.get_connection() as conn:
                cursor = conn.cursor()

                # Step 1: Upsert vendors
                for vendor in vendors_to_upsert.values():
                    cursor.execute(
                        "INSERT INTO vendors (vendor_name, display_name) VALUES (?, ?) ON CONFLICT(vendor_name) DO NOTHING",
                        (vendor["vendor_name"], vendor["display_name"]),
                    )

                # Step 2: NEW - Upsert materials to master data
                for mat_code, mat_data in materials_to_upsert.items():
                    # Check if material exists
                    cursor.execute(
                        "SELECT material_code FROM materials WHERE material_code = ?",
                        (mat_code,),
                    )
                    existing = cursor.fetchone()

                    if existing:
                        # Update existing material (only update if lead time is greater or vendor is missing)
                        cursor.execute(
                            """
                            UPDATE materials 
                            SET description = COALESCE(NULLIF(description, ''), ?),
                                unit = COALESCE(NULLIF(unit, ''), ?),
                                lead_time_days = CASE 
                                    WHEN ? > lead_time_days THEN ? 
                                    ELSE lead_time_days 
                                END,
                                preferred_vendor = COALESCE(NULLIF(preferred_vendor, ''), ?),
                                last_updated = ?
                            WHERE material_code = ?
                        """,
                            (
                                mat_data["description"],
                                mat_data["unit"],
                                mat_data["lead_time_days"],
                                mat_data["lead_time_days"],
                                mat_data["preferred_vendor"],
                                datetime.now().isoformat(),
                                mat_code,
                            ),
                        )
                        materials_updated += 1
                    else:
                        # Create new material
                        cursor.execute(
                            """
                            INSERT INTO materials (
                                material_code, description, unit, lead_time_days,
                                preferred_vendor, safety_stock, min_order_qty, lot_size_rule,
                                created_date, last_updated
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                            (
                                mat_code,
                                mat_data["description"],
                                mat_data["unit"],
                                mat_data["lead_time_days"],
                                mat_data["preferred_vendor"],
                                mat_data["safety_stock"],
                                mat_data["min_order_qty"],
                                mat_data["lot_size_rule"],
                                datetime.now().isoformat(),
                                datetime.now().isoformat(),
                            ),
                        )
                        materials_created += 1

                # Step 3: Get existing open requisition lines
                cursor.execute(
                    "SELECT req_number, item FROM requisitions WHERE status = 'Open'"
                )
                fetched_rows = cursor.fetchall()
                existing_open_lines = set(
                    (row["req_number"], row["item"]) for row in fetched_rows
                )

                # Step 4: Upsert requisition lines
                for req in requisitions_to_insert:
                    cursor.execute(
                        """
                        INSERT INTO requisitions (
                            req_number, item, vendor_name, material_code, short_text,
                            requested_qty, unit, requested_del_date, unit_price, status, pr_status
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ON CONFLICT(req_number, item) DO UPDATE SET
                            vendor_name = excluded.vendor_name,
                            material_code = excluded.material_code,
                            short_text = excluded.short_text,
                            requested_qty = excluded.requested_qty,
                            unit = excluded.unit,
                            requested_del_date = excluded.requested_del_date,
                            unit_price = excluded.unit_price,
                            status = 'Open',
                            pr_status = 'Open'
                    """,
                        (
                            req["req_number"],
                            req["item"],
                            req["vendor_name"],
                            req["material_code"],
                            req["short_text"],
                            req["requested_qty"],
                            req["unit"],
                            req["requested_del_date"],
                            req["unit_price"],
                            req["status"],
                            req["pr_status"],
                        ),
                    )

                # Step 5: Close lines not in Excel
                lines_to_close = existing_open_lines - lines_in_excel
                if lines_to_close:
                    lines_to_close_list = [
                        (req_num, item) for req_num, item in lines_to_close
                    ]
                    update_query = "UPDATE requisitions SET status = 'Closed' WHERE req_number = ? AND item = ?"
                    cursor.executemany(update_query, lines_to_close_list)
                    closed_count = cursor.rowcount
                    print(f"ℹ️ INFO: Marked {closed_count} requisition lines as 'Closed'.")

                conn.commit()

            print(
                f"ℹ️ INFO: Materials Master Data - Created: {materials_created}, Updated: {materials_updated}"
            )
            return (
                len(requisitions_to_insert),
                closed_count,
                materials_created,
                materials_updated,
            )

        except Exception as e:
            raise ValueError(f"Failed to process requisitions file: {str(e)}")

    def clear_all_requisitions(self):
        """Delete all requisitions from the database"""
        try:
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("DELETE FROM requisitions")
                deleted_count = cursor.rowcount
                conn.commit()
            return deleted_count
        except Exception as e:
            raise ValueError(f"Failed to clear requisitions: {str(e)}")

    def link_requisitions_to_forecast(self, vendor_name=None, date_range_days=30):
        """Link requisitions to forecast data"""
        query = """
            SELECT r.*, v.display_name as vendor_display
            FROM requisitions r
            LEFT JOIN vendors v ON r.vendor_name = v.vendor_name
            WHERE r.pr_status = 'Pending'
        """

        params = []
        if vendor_name:
            query += " AND v.display_name = ?"
            params.append(vendor_name)

        reqs = self.db.execute_query(query, tuple(params), fetchall=True)

        linked_count = 0

        for req in reqs:
            if not req.get("requested_del_date"):
                continue

            try:
                req_date = datetime.strptime(req["requested_del_date"], "%d.%m.%Y")

                forecast_query = """
                    SELECT * FROM forecasts
                    WHERE vendor_name = ?
                    AND material_code = ?
                    AND forecast_date BETWEEN ? AND ?
                """

                date_start = (req_date - timedelta(days=date_range_days)).strftime(
                    "%Y-%m-%d"
                )
                date_end = (req_date + timedelta(days=date_range_days)).strftime(
                    "%Y-%m-%d"
                )

                matching_forecasts = self.db.execute_query(
                    forecast_query,
                    (req["vendor_name"], req["material_code"], date_start, date_end),
                    fetchall=True,
                )

                if matching_forecasts:
                    linked_count += 1
                    self.db.execute_query(
                        "UPDATE requisitions SET comments = ? WHERE req_number = ? AND item = ?",
                        (
                            f"Linked to forecast (found {len(matching_forecasts)} matching periods)",
                            req["req_number"],
                            req["item"],
                        ),
                        commit=True,
                    )
            except Exception:
                continue

        return linked_count

    def generate_forecast_template(self, vendor_name, num_weeks=13, start_date=None):
        """Generate a blank forecast template Excel file for a vendor"""
        if not start_date:
            start_date = datetime.now()

        query = """
            SELECT DISTINCT material_code, short_text, unit, unit_price
            FROM (
                SELECT material_code, short_text, unit, unit_price FROM open_orders WHERE vendor_name = ?
                UNION
                SELECT material_code, short_text, unit, unit_price FROM forecasts WHERE vendor_name = ?
            )
            ORDER BY material_code
        """

        vendor_internal = unidecode(vendor_name).strip().lower()
        materials = self.db.execute_query(
            query, (vendor_internal, vendor_internal), fetchall=True
        )

        data = []
        for mat in materials:
            row = {
                "Material": mat["material_code"],
                "Description": mat["short_text"],
                "Unit": mat["unit"],
                "Unit Price": mat["unit_price"],
            }

            for week in range(num_weeks):
                week_date = start_date + timedelta(weeks=week)
                week_label = f"Week_{week_date.strftime('%W_%Y')}"
                row[week_label] = 0

            data.append(row)

        df = pd.DataFrame(data)

        output_buffer = io.BytesIO()
        with pd.ExcelWriter(output_buffer, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Forecast", index=False)

            worksheet = writer.sheets["Forecast"]

            from openpyxl.styles import Font, PatternFill, Alignment

            header_fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )
            header_font = Font(color="FFFFFF", bold=True)

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:  # TODO: Add proper error handling
                        pass  # TODO: Add proper error handling
                adjusted_width = min(max_length + 2, 30)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        output_buffer.seek(0)
        return output_buffer

    def get_forecast_summary(self, vendor_name=None, start_date=None, end_date=None):
        """Get summary of forecast data with filters"""
        query = """
            SELECT 
                v.display_name as vendor,
                f.material_code,
                f.short_text,
                SUM(f.forecast_qty) as total_qty,
                SUM(f.total_amount) as total_value,
                f.currency,
                MIN(f.forecast_date) as earliest_date,
                MAX(f.forecast_date) as latest_date
            FROM forecasts f
            LEFT JOIN vendors v ON f.vendor_name = v.vendor_name
            WHERE 1=1
        """

        params = []

        if vendor_name:
            query += " AND v.display_name = ?"
            params.append(vendor_name)

        if start_date:
            query += " AND f.forecast_date >= ?"
            params.append(start_date)

        if end_date:
            query += " AND f.forecast_date <= ?"
            params.append(end_date)

        query += " GROUP BY v.display_name, f.material_code, f.short_text, f.currency ORDER BY v.display_name, f.material_code"

        return self.db.execute_query(query, tuple(params), fetchall=True)

    def get_open_requisitions(self, vendor_name=None):
        """Get all open requisitions"""
        query = """
            SELECT r.*, v.display_name as vendor_display
            FROM requisitions r
            LEFT JOIN vendors v ON r.vendor_name = v.vendor_name
            WHERE r.status = 'Open'
        """

        params = []
        if vendor_name:
            query += " AND v.display_name = ?"
            params.append(vendor_name)

        query += " ORDER BY r.requested_del_date, r.req_number, r.item"

        return self.db.execute_query(query, tuple(params), fetchall=True)

    def convert_requisition_to_po(self, req_numbers):
        """
        Convert approved requisitions to actual Purchase Orders
        ENHANCED: Uses net_price and price_per_unit from materials master when available
        """
        if not req_numbers:
            return 0

        # Get requisition details with vendor information and material prices
        placeholders = ",".join("?" * len(req_numbers))
        query = f"""
            SELECT r.*,
                   v.vendor_name, v.display_name, v.transport_days,
                   m.net_price, m.price_per_unit, m.currency
            FROM requisitions r
            LEFT JOIN vendors v ON r.vendor_name = v.vendor_name
            LEFT JOIN materials m ON r.material_code = m.material_code
            WHERE r.req_number IN ({placeholders})
            AND r.approval_status = 'APPROVED'
            ORDER BY r.vendor_name, r.req_number, r.item
        """
        requisitions = self.db.execute_query(query, tuple(req_numbers), fetchall=True)

        if not requisitions:
            return 0

        # Group requisitions by vendor to create one PO per vendor
        from collections import defaultdict

        po_groups = defaultdict(list)

        for req in requisitions:
            vendor_key = req["vendor_name"] or "UNKNOWN"
            po_groups[vendor_key].append(req)

        total_converted = 0
        created_pos = []

        # Create a PO for each vendor
        for vendor_name, vendor_items in po_groups.items():
            try:
                # Generate new PO number
                po_number = self._generate_po_number()

                # Get transport days for this vendor
                transport_days = int(vendor_items[0].get("transport_days", 0) or 0)

                # Insert each item into open_orders
                for idx, item in enumerate(vendor_items, start=10):
                    # Calculate ETD from ETA
                    pr_eta_str = item.get("requested_del_date", "")

                    if pr_eta_str and transport_days > 0:
                        try:
                            pr_eta = datetime.strptime(pr_eta_str, "%d.%m.%Y")
                            po_etd = subtract_working_days(pr_eta, transport_days)
                            po_etd_str = po_etd.strftime("%d.%m.%Y")
                        except Exception:
                            po_etd_str = pr_eta_str
                    else:
                        po_etd_str = pr_eta_str

                    # NEW: Use material master prices if available, otherwise use requisition prices
                    unit_price = item.get("net_price")  # From materials table
                    price_per = item.get("price_per_unit")  # From materials table
                    currency = item.get("currency")  # From materials table - NEW

                    if unit_price is None or unit_price == 0:
                        # Fall back to requisition price
                        unit_price = item.get("unit_price", 0.0)

                    if price_per is None or price_per == 0:
                        price_per = 1

                    if not currency:
                        # Fall back to requisition currency
                        currency = item.get("currency", "EUR")

                    # Calculate total amount
                    requested_qty = item.get("requested_qty", 0)
                    try:
                        total_amount = (requested_qty * float(unit_price)) / int(
                            price_per
                        )
                    except Exception:
                        total_amount = float(requested_qty or 0) * float(
                            unit_price or 0
                        )

                    self.db.execute_query(
                        """
                        INSERT INTO open_orders (
                            po, item, vendor_name, material_code, short_text,
                            requested_qty, requested_del_date, unit, 
                            unit_price, price_per_unit, total_amount, currency,
                            status, pdf_status, email_status
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Open', 'Pending', 'Pending')
                    """,
                        (
                            po_number,
                            str(idx),
                            vendor_name,
                            item.get("material_code", ""),
                            item.get("short_text", ""),
                            requested_qty,
                            po_etd_str,
                            item.get("unit", "EA"),
                            float(unit_price or 0),
                            int(price_per or 1),
                            total_amount,
                            currency,  # Uses material master currency or falls back to requisition currency
                        ),
                        commit=True,
                    )

                    # Log pricing source for transparency
                    price_source = (
                        "material master" if item.get("net_price") else "requisition"
                    )
                    print(
                        f"ℹ️ INFO: PO {po_number} Item {idx} - Using price from {price_source}: {unit_price}/{price_per} {currency}"
                    )

                created_pos.append(po_number)
                total_converted += len(vendor_items)

                # Update requisitions to reference the PO number
                req_nums_for_vendor = [r["req_number"] for r in vendor_items]
                req_placeholders = ",".join("?" * len(req_nums_for_vendor))

                self.db.execute_query(
                    f"""
                    UPDATE requisitions 
                    SET pr_status = ?,
                        status = 'Closed'
                    WHERE req_number IN ({req_placeholders})
                """,
                    (f"Converted to PO {po_number}", *req_nums_for_vendor),
                    commit=True,
                )

            except Exception as e:
                print(f"Error creating PO for vendor {vendor_name}: {e}")
                import traceback

                traceback.print_exc()
                continue

        return total_converted

    def _generate_po_number(self):
        """Generate a unique PO number"""
        # Get the highest existing PO number
        result = self.db.execute_query(
            """
            SELECT MAX(CAST(po AS INTEGER)) as max_po 
            FROM open_orders 
            WHERE po GLOB '[0-9]*'
        """,
            fetchone=True,
        )

        last_po = result["max_po"] if result and result["max_po"] else 4500000000
        return str(last_po + 1)

    def _find_column(self, df, possible_names):
        """Find column by checking multiple possible names"""
        for name in possible_names:
            if name in df.columns:
                return name
        return None

    def _parse_date_from_column(self, col_name, default_year):
        """Parse date, week number, and month number from column name"""
        col_lower = col_name.lower()

        if "week" in col_lower or col_lower.startswith("w"):
            parts = (
                col_name.replace("Week_", "")
                .replace("week_", "")
                .replace("W", "")
                .replace("w", "")
                .split("_")
            )
            try:
                week_num = int(parts[0])
                year = int(parts[1]) if len(parts) > 1 else default_year
                date = datetime.strptime(f"{year}-W{week_num:02d}-1", "%Y-W%W-%w")
                return date, week_num, date.month
            except Exception:
                return None, None, None

        month_names = {m.lower(): i for i, m in enumerate(calendar.month_abbr) if m}
        month_names_full = {
            m.lower(): i for i, m in enumerate(calendar.month_name) if m
        }

        for month_name, month_num in {**month_names, **month_names_full}.items():
            if month_name in col_lower:
                parts = col_name.lower().split("_")
                year = default_year
                for part in parts:
                    if part.isdigit() and len(part) == 4:
                        year = int(part)

                date = datetime(year, month_num, 1)
                week_num = date.isocalendar()[1]
                return date, week_num, month_num

        return None, None, None

    def generate_outbound_forecast_from_requisitions(
        self,
        vendor_name=None,
        weeks_ahead=13,
        group_by="week",
        include_open_orders=True,
    ):
        """
        Generate demand forecast to send to vendors based on requisitions and open orders.

        Args:
            vendor_name: Specific vendor or None for all
            weeks_ahead: Number of weeks to forecast
            group_by: 'week' or 'month'
            include_open_orders: Include existing open orders in forecast

        Returns:
            Dictionary with vendor forecasts
        """
        # Get requisitions
        req_query = """
            SELECT r.*, v.display_name as vendor_display
            FROM requisitions r
            LEFT JOIN vendors v ON r.vendor_name = v.vendor_name
            WHERE r.status = 'Open' AND r.requested_del_date IS NOT NULL
        """

        params = []
        if vendor_name:
            req_query += " AND v.display_name = ?"
            params.append(vendor_name)

        requisitions = self.db.execute_query(req_query, tuple(params), fetchall=True)

        # Get open orders if requested
        orders = []
        if include_open_orders:
            order_query = """
                SELECT oo.*, v.display_name as vendor_display
                FROM open_orders oo
                LEFT JOIN vendors v ON oo.vendor_name = v.vendor_name
                WHERE oo.status = 'Open' AND oo.requested_del_date IS NOT NULL
            """

            params = []
            if vendor_name:
                order_query += " AND v.display_name = ?"
                params.append(vendor_name)

            orders = self.db.execute_query(order_query, tuple(params), fetchall=True)

        # Combine data
        all_demand = []

        for req in requisitions:
            if req.get("requested_del_date"):
                try:
                    del_date = datetime.strptime(req["requested_del_date"], "%d.%m.%Y")
                    all_demand.append(
                        {
                            "vendor": req.get("vendor_display", "Unknown"),
                            "material": req["material_code"],
                            "description": req["short_text"],
                            "qty": req["requested_qty"],
                            "date": del_date,
                            "unit": req.get("unit", "EA"),
                            "source": "Requisition",
                            "ref": req["req_number"],
                        }
                    )
                except Exception:  # TODO: Add proper error handling
                    pass  # TODO: Add proper error handling
        # MODIFIED: For open orders, use rescheduling_date (ETA) if available, otherwise requested_del_date
        # MODIFIED: For open orders, use rescheduling_date (ETA) if available, otherwise requested_del_date + transport days
        for order in orders:
            # Get transport days for this vendor
            transport_days = int(order.get("transport_days", 0) or 0)

            # Priority: use rescheduling_date if available (already ETA), otherwise requested_del_date (ETD - needs conversion)
            if order.get("rescheduling_date"):
                # Rescheduling date is already ETA, use it directly
                date_to_use = order.get("rescheduling_date")
            elif order.get("requested_del_date"):
                # Requested date is ETD, need to add transport days to get ETA
                date_to_use = order.get("requested_del_date")
                # Convert to datetime, add transport days, then back to string
                try:
                    etd_date = datetime.strptime(date_to_use, "%d.%m.%Y")
                    if transport_days > 0:
                        eta_date = add_working_days(etd_date, transport_days)
                        date_to_use = eta_date.strftime("%d.%m.%Y")
                except Exception:
                    pass  # Keep original date if conversion fails
            else:
                date_to_use = None

            if date_to_use:
                try:
                    del_date = datetime.strptime(date_to_use, "%d.%m.%Y")
                    all_demand.append(
                        {
                            "vendor": order.get("vendor_display", "Unknown"),
                            "material": order["material_code"],
                            "description": order["short_text"],
                            "qty": order["requested_qty"],
                            "date": del_date,  # This is now ETA for both cases
                            "unit": order.get("unit", "EA"),
                            "source": "Open Order",
                            "ref": order["po"],
                        }
                    )
                except Exception:  # TODO: Add proper error handling
                    pass  # TODO: Add proper error handling
        # Group by vendor, material, and period
        forecast_data = {}
        today = datetime.now()
        end_date = today + timedelta(weeks=weeks_ahead)

        for item in all_demand:
            # Skip items outside forecast window
            if item["date"] > end_date:
                continue

            vendor = item["vendor"]
            material = item["material"]

            if vendor not in forecast_data:
                forecast_data[vendor] = {}

            if material not in forecast_data[vendor]:
                forecast_data[vendor][material] = {
                    "description": item["description"],
                    "unit": item["unit"],
                    "periods": {},
                    "details": [],
                }

            # Determine period (week or month)
            if group_by == "week":
                period_key = f"Week_{item['date'].strftime('%W_%Y')}"
                period_label = f"Week {item['date'].strftime('%W')} ({item['date'].strftime('%Y')})"
            else:  # month
                period_key = f"Month_{item['date'].strftime('%m_%Y')}"
                period_label = f"{item['date'].strftime('%B %Y')}"

            # Aggregate quantities
            if period_key not in forecast_data[vendor][material]["periods"]:
                forecast_data[vendor][material]["periods"][period_key] = {
                    "qty": 0,
                    "label": period_label,
                    "date": item["date"],
                }

            forecast_data[vendor][material]["periods"][period_key]["qty"] += item["qty"]

            # Store details
            forecast_data[vendor][material]["details"].append(
                {
                    "source": item["source"],
                    "ref": item["ref"],
                    "qty": item["qty"],
                    "date": item["date"].strftime("%d.%m.%Y"),
                }
            )

        return forecast_data

    def create_outbound_forecast_excel(self, vendor_name, forecast_data):
        """
        Create Excel file with outbound forecast for a specific vendor.

        Args:
            vendor_name: Vendor to create forecast for
            forecast_data: Output from generate_outbound_forecast_from_requisitions

        Returns:
            BytesIO buffer with Excel file
        """
        if vendor_name not in forecast_data:
            raise ValueError(f"No forecast data found for vendor: {vendor_name}")

        vendor_forecast = forecast_data[vendor_name]

        # Prepare data for Excel
        rows = []

        # Get all unique periods across all materials and sort them chronologically
        all_periods = {}  # Changed from set to dict to maintain date info
        for material_data in vendor_forecast.values():
            for period_key, period_info in material_data["periods"].items():
                if period_key not in all_periods:
                    all_periods[period_key] = {
                        "label": period_info["label"],
                        "date": period_info["date"],
                    }

        # Sort periods by date
        sorted_periods = sorted(all_periods.items(), key=lambda x: x[1]["date"])

        # Build column headers in order
        ordered_columns = ["Material", "Description", "Unit"]
        period_labels = []
        for period_key, period_info in sorted_periods:
            period_labels.append(period_info["label"])
            ordered_columns.append(period_info["label"])

        # Build rows
        for material, data in sorted(vendor_forecast.items()):
            row = {
                "Material": material,
                "Description": data["description"],
                "Unit": data["unit"],
            }

            # Add quantities for each period in order
            for period_key, period_info in sorted_periods:
                period_label = period_info["label"]
                # Check if this material has data for this period
                if period_key in data["periods"]:
                    qty = data["periods"][period_key]["qty"]
                else:
                    qty = 0  # No forecast for this period

                row[period_label] = qty

            rows.append(row)

        # Create DataFrame with ordered columns
        df = pd.DataFrame(rows, columns=ordered_columns)

        # Create Excel with formatting
        output_buffer = io.BytesIO()
        with pd.ExcelWriter(output_buffer, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Forecast", index=False)

            workbook = writer.book
            worksheet = writer.sheets["Forecast"]

            # Format header
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            header_font = Font(color="FFFFFF", bold=True, size=11)

            # Format header row
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

            # Add borders and formatting to data
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for row in worksheet.iter_rows(
                min_row=1,
                max_row=worksheet.max_row,
                min_col=1,
                max_col=worksheet.max_column,
            ):
                for cell in row:
                    cell.border = thin_border
                    if cell.row > 1:  # Data rows
                        # Align text left for first 3 columns, center for date columns
                        if cell.column <= 3:
                            cell.alignment = Alignment(
                                horizontal="left", vertical="center"
                            )
                        else:
                            cell.alignment = Alignment(
                                horizontal="center", vertical="center"
                            )

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:  # TODO: Add proper error handling
                        pass  # TODO: Add proper error handling
                adjusted_width = min(max_length + 3, 40)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Freeze header row and first 3 columns
            worksheet.freeze_panes = "D2"

            # Add summary sheet
            summary_data = []
            total_qty = 0
            for material, data in vendor_forecast.items():
                material_total = sum(p["qty"] for p in data["periods"].values())
                total_qty += material_total
                summary_data.append(
                    {
                        "Material": material,
                        "Description": data["description"],
                        "Total Quantity": material_total,
                        "Unit": data["unit"],
                    }
                )

            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name="Summary", index=False)

            # Format summary sheet
            summary_ws = writer.sheets["Summary"]
            for cell in summary_ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for column in summary_ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:  # TODO: Add proper error handling
                        pass  # TODO: Add proper error handling
                adjusted_width = min(max_length + 3, 40)
                summary_ws.column_dimensions[column_letter].width = adjusted_width

            # ===== NEW: Add Open Orders Sheet =====
            self._add_open_orders_sheet(
                writer, vendor_name, header_fill, header_font, thin_border
            )

        output_buffer.seek(0)
        return output_buffer

    def _add_open_orders_sheet(
        self, writer, vendor_name, header_fill, header_font, thin_border
    ):
        """
        Add Open Orders sheet to the forecast Excel file.

        Args:
            writer: ExcelWriter object
            vendor_name: Name of the vendor to filter orders
            header_fill: Fill style for headers
            header_font: Font style for headers
            thin_border: Border style for cells
        """
        from openpyxl.styles import Alignment
        from openpyxl.utils import get_column_letter

        # Get vendor's internal name for querying
        vendor_internal = unidecode(vendor_name).strip().lower()

        # Query open orders for this vendor
        query = """
            SELECT oo.*, v.display_name as name, v.transport_days, v.transport_days_secondary
            FROM open_orders oo
            JOIN vendors v ON oo.vendor_name = v.vendor_name
            WHERE oo.status = 'Open' AND v.vendor_name = ?
            ORDER BY oo.po, oo.item
        """

        orders = self.db.execute_query(query, (vendor_internal,), fetchall=True)

        if not orders:
            # Create empty sheet with message
            df_empty = pd.DataFrame(
                {"Message": ["No open orders found for this vendor"]}
            )
            df_empty.to_excel(writer, sheet_name="Open Orders", index=False)
            return

        # Convert to DataFrame for processing
        df = pd.DataFrame(orders)

        # Ensure all text fields are properly converted to strings
        df["comments"] = df["comments"].fillna("").astype(str)
        df["exception_message"] = df["exception_message"].fillna("").astype(str)
        df["material_code"] = df["material_code"].fillna("").astype(str)
        df["short_text"] = df["short_text"].fillna("").astype(str)
        df["conf_delivery_date"] = df["conf_delivery_date"].fillna("").astype(str)
        df["rescheduling_date"] = df["rescheduling_date"].fillna("").astype(str)

        # Get transport days for this vendor
        primary_days = int(df["transport_days"].iloc[0] or 0) if len(df) > 0 else 0
        secondary_days = (
            int(df["transport_days_secondary"].iloc[0] or 0) if len(df) > 0 else 0
        )

        # Parse dates
        df["requested_del_date_dt"] = pd.to_datetime(
            df["requested_del_date"], format="%d.%m.%Y", errors="coerce"
        )
        df["rescheduling_date_dt"] = pd.to_datetime(
            df["rescheduling_date"], format="%d.%m.%Y", errors="coerce"
        )

        def calculate_etd_and_notes(row):
            """Calculates ETD date and generates exception notes."""
            reschedule_dt = row["rescheduling_date_dt"]
            current_exception_notes = str(row.get("exception_message") or "")

            if pd.isna(reschedule_dt):
                return pd.NaT, current_exception_notes

            today = datetime.now()

            # 1. Calculate ETD with primary transport days
            final_etd = subtract_working_days(reschedule_dt, primary_days)
            note = f"ETD calculated using primary transport ({primary_days} days)."

            # 2. Check if primary ETD is in the past and if a secondary option exists
            if final_etd < today and secondary_days > 0:
                final_etd = subtract_working_days(reschedule_dt, secondary_days)
                note = f"Used secondary transport ({secondary_days} days)."
            elif primary_days == 0:
                note = "No transport days configured - ETD not calculated."

            # 3. Combine notes
            if current_exception_notes and note and note not in current_exception_notes:
                updated_notes = f"{current_exception_notes}; {note}"
            else:
                updated_notes = current_exception_notes or note

            return final_etd, updated_notes

        # Apply the function to each row
        results = df.apply(calculate_etd_and_notes, axis=1)
        df[["etd_date", "exception_message"]] = pd.DataFrame(
            results.tolist(), index=df.index
        )
        df["etd_date_str"] = df["etd_date"].dt.strftime("%d.%m.%Y").fillna("")

        # Prepare output DataFrame
        output_df = df[
            [
                "po",
                "item",
                "material_code",
                "short_text",
                "requested_qty",
                "total_amount",
                "currency",
                "requested_del_date",
                "conf_delivery_date",
                "etd_date_str",
                "comments",
                "exception_message",
            ]
        ].copy()

        # Rename columns
        output_df.rename(
            columns={
                "po": "PO Number",
                "item": "Item",
                "material_code": "Material",
                "short_text": "Description",
                "requested_qty": "Quantity",
                "total_amount": "Value",
                "currency": "Currency",
                "requested_del_date": "Requested Date",
                "conf_delivery_date": "Confirmed Date",
                "etd_date_str": "Reschedule To (ETD)",
                "comments": "Supplier Comments",
                "exception_message": "Exception Notes",
            },
            inplace=True,
        )

        # Write to Excel
        output_df.to_excel(writer, sheet_name="Open Orders", index=False)

        # Get the worksheet to apply formatting
        worksheet = writer.sheets["Open Orders"]

        # Format header row
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        # Add borders and formatting to all cells
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=worksheet.max_row,
            min_col=1,
            max_col=worksheet.max_column,
        ):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:  # Data rows
                    cell.alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=True
                    )

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:  # TODO: Add proper error handling
                    pass  # TODO: Add proper error handling
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        worksheet.freeze_panes = "A2"

    def create_outbound_forecast_pdf(self, vendor_name, forecast_data):
        """
        Create PDF file with outbound forecast for a specific vendor.

        Args:
            vendor_name: Vendor to create forecast for
            forecast_data: Output from generate_outbound_forecast_from_requisitions

        Returns:
            BytesIO buffer with PDF file
        """
        if vendor_name not in forecast_data:
            raise ValueError(f"No forecast data found for vendor: {vendor_name}")

        vendor_forecast = forecast_data[vendor_name]

        pdf_buffer = io.BytesIO()
        c = canvas.Canvas(pdf_buffer, pagesize=letter)
        width, height = letter

        # Get company config
        config_query = "SELECT value FROM app_config WHERE key = 'company_config'"
        config_result = self.db.execute_query(config_query, fetchone=True)
        config = json.loads(config_result["value"]) if config_result else {}

        # Header
        c.setFont("Helvetica-Bold", 18)
        c.drawString(0.5 * inch, height - 0.5 * inch, "DEMAND FORECAST")

        c.setFont("Helvetica", 11)
        c.drawString(
            0.5 * inch,
            height - 0.8 * inch,
            f"Date: {datetime.now().strftime('%d.%m.%Y')}",
        )
        c.drawString(0.5 * inch, height - 1.0 * inch, f"To: {vendor_name}")
        c.drawString(
            0.5 * inch,
            height - 1.2 * inch,
            f"From: {config.get('my_company_name', 'Your Company')}",
        )

        # Message
        y_pos = height - 1.6 * inch
        c.setFont("Helvetica", 10)
        message_lines = [
            "Dear Supplier,",
            "",
            "Please find below our demand forecast for the coming weeks.",
            "This forecast is provided to help you plan your production capacity.",
            "Actual purchase orders will follow based on confirmed requirements.",
            "",
        ]

        for line in message_lines:
            c.drawString(0.5 * inch, y_pos, line)
            y_pos -= 0.2 * inch

        # Calculate totals
        total_qty_all = 0
        for material_data in vendor_forecast.values():
            total_qty_all += sum(p["qty"] for p in material_data["periods"].values())

        # Summary section
        y_pos -= 0.2 * inch
        c.setFont("Helvetica-Bold", 12)
        c.drawString(
            0.5 * inch, y_pos, f"Total Forecast Quantity: {total_qty_all:,} units"
        )
        c.drawString(
            0.5 * inch, y_pos - 0.25 * inch, f"Materials: {len(vendor_forecast)}"
        )

        y_pos -= 0.7 * inch

        # Material details
        c.setFont("Helvetica-Bold", 11)
        c.drawString(0.5 * inch, y_pos, "Forecast by Material:")
        y_pos -= 0.3 * inch

        c.setFont("Helvetica", 9)

        for material, data in sorted(vendor_forecast.items()):
            # Check if we need a new page
            if y_pos < 2 * inch:
                c.showPage()
                y_pos = height - 1 * inch
                c.setFont("Helvetica", 9)

            # Material header
            c.setFont("Helvetica-Bold", 10)
            c.drawString(0.5 * inch, y_pos, f"Material: {material}")
            y_pos -= 0.2 * inch

            c.setFont("Helvetica", 9)
            c.drawString(0.75 * inch, y_pos, f"Description: {data['description']}")
            y_pos -= 0.2 * inch

            # Period breakdown
            sorted_periods = sorted(data["periods"].items(), key=lambda x: x[1]["date"])

            for period_key, period_info in sorted_periods:
                period_text = (
                    f"   {period_info['label']}: {period_info['qty']:,} {data['unit']}"
                )
                c.drawString(0.75 * inch, y_pos, period_text)
                y_pos -= 0.18 * inch

            # Material total
            material_total = sum(p["qty"] for p in data["periods"].values())
            c.setFont("Helvetica-Bold", 9)
            c.drawString(
                0.75 * inch, y_pos, f"Total: {material_total:,} {data['unit']}"
            )
            y_pos -= 0.35 * inch
            c.setFont("Helvetica", 9)

        # Footer
        if y_pos < 2 * inch:
            c.showPage()
            y_pos = height - 1 * inch

        y_pos -= 0.3 * inch
        c.setFont("Helvetica", 10)
        footer_lines = [
            "",
            "This forecast is for planning purposes only and does not constitute a commitment.",
            "Please confirm your capacity to meet this forecast or provide alternative proposals.",
            f"Contact: {config.get('buyer_email', 'purchasing@company.com')}",
            "",
            "Thank you for your partnership.",
        ]

        for line in footer_lines:
            if y_pos > 0.5 * inch:
                c.drawString(0.5 * inch, y_pos, line)
                y_pos -= 0.2 * inch

        c.save()
        pdf_buffer.seek(0)
        return pdf_buffer


# ==============================================================================
# GUI WINDOWS FOR FORECAST MANAGEMENT
# ==============================================================================


class ForecastManagementWindow(tk.Toplevel):
    """Main window for forecast management"""

    def __init__(self, parent, log_callback, data_manager, forecast_manager):
        super().__init__(parent)
        self.title("Forecast & Requisition Management")
        self.geometry("1200x800")
        self.log = log_callback
        self.dm = data_manager
        self.fm = forecast_manager

        # Create notebook for tabs
        notebook = ttk.Notebook(self)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Tab 1: Upload & Manage
        upload_frame = ttk.Frame(notebook)
        notebook.add(upload_frame, text="Upload & Manage")
        self.setup_upload_tab(upload_frame)

        # Tab 2: Forecast View
        forecast_frame = ttk.Frame(notebook)
        notebook.add(forecast_frame, text="Forecast Summary")
        self.setup_forecast_tab(forecast_frame)

        # Tab 3: Requisitions View
        req_frame = ttk.Frame(notebook)
        notebook.add(req_frame, text="Requisitions")
        self.setup_requisitions_tab(req_frame)

        # Tab 4: Analysis
        analysis_frame = ttk.Frame(notebook)
        notebook.add(analysis_frame, text="Forecast vs Actual")
        self.setup_analysis_tab(analysis_frame)

        # Tab 5: Outbound Forecast
        outbound_frame = ttk.Frame(notebook)
        notebook.add(outbound_frame, text="Outbound Forecast")
        self.setup_outbound_forecast_tab(outbound_frame)

    def setup_upload_tab(self, parent):
        """Setup the upload and management tab"""
        # Upload section
        upload_section = ttk.LabelFrame(parent, text="Upload Files", padding=10)
        upload_section.pack(fill=tk.X, padx=10, pady=5)

        # Forecast upload
        forecast_frame = ttk.Frame(upload_section)
        forecast_frame.pack(fill=tk.X, pady=5)

        ttk.Label(forecast_frame, text="Vendor Forecast:").pack(
            side=tk.LEFT, padx=(0, 5)
        )

        self.forecast_vendor_var = tk.StringVar()
        vendor_combo = ttk.Combobox(
            forecast_frame,
            textvariable=self.forecast_vendor_var,
            state="readonly",
            width=25,
        )
        vendor_combo["values"] = [v["display_name"] for v in self.dm.get_all_vendors()]
        vendor_combo.pack(side=tk.LEFT, padx=5)

        ttk.Button(
            forecast_frame,
            text="Upload Forecast Excel...",
            command=self.upload_forecast,
        ).pack(side=tk.LEFT, padx=5)
        ttk.Button(
            forecast_frame, text="Download Template", command=self.download_template
        ).pack(side=tk.LEFT)

        # Requisitions upload
        req_frame = ttk.Frame(upload_section)
        req_frame.pack(fill=tk.X, pady=5)

        ttk.Label(req_frame, text="Purchase Requisitions:").pack(
            side=tk.LEFT, padx=(0, 5)
        )
        ttk.Button(
            req_frame,
            text="Upload Requisitions Excel...",
            command=self.upload_requisitions,
        ).pack(side=tk.LEFT, padx=5)
        ttk.Button(
            req_frame, text="Clear All Requisitions", command=self.clear_requisitions
        ).pack(side=tk.LEFT, padx=5)

        # Link section
        link_section = ttk.LabelFrame(
            parent, text="Link Requisitions to Forecast", padding=10
        )
        link_section.pack(fill=tk.X, padx=10, pady=5)

        link_frame = ttk.Frame(link_section)
        link_frame.pack(fill=tk.X)

        ttk.Label(link_frame, text="Vendor (optional):").pack(side=tk.LEFT, padx=(0, 5))
        self.link_vendor_var = tk.StringVar()
        link_combo = ttk.Combobox(
            link_frame, textvariable=self.link_vendor_var, state="readonly", width=25
        )
        link_combo["values"] = ["All Vendors"] + [
            v["display_name"] for v in self.dm.get_all_vendors()
        ]
        link_combo.set("All Vendors")
        link_combo.pack(side=tk.LEFT, padx=5)

        ttk.Button(
            link_frame,
            text="Link Requisitions to Forecast",
            command=self.link_requisitions,
        ).pack(side=tk.LEFT, padx=5)

        # Status section
        status_section = ttk.LabelFrame(parent, text="Status", padding=10)
        status_section.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        self.status_text = scrolledtext.ScrolledText(
            status_section, wrap=tk.WORD, height=15
        )
        self.status_text.pack(fill=tk.BOTH, expand=True)

    def setup_forecast_tab(self, parent):
        """Setup the forecast summary tab"""
        # Filter section
        filter_frame = ttk.LabelFrame(parent, text="Filters", padding=10)
        filter_frame.pack(fill=tk.X, padx=10, pady=5)

        # Vendor filter
        vendor_filter_frame = ttk.Frame(filter_frame)
        vendor_filter_frame.pack(fill=tk.X, pady=2)

        ttk.Label(vendor_filter_frame, text="Vendor:").pack(side=tk.LEFT, padx=(0, 5))
        self.forecast_filter_vendor_var = tk.StringVar()
        forecast_vendor_combo = ttk.Combobox(
            vendor_filter_frame,
            textvariable=self.forecast_filter_vendor_var,
            state="readonly",
            width=30,
        )
        forecast_vendor_combo["values"] = ["All Vendors"] + [
            v["display_name"] for v in self.dm.get_all_vendors()
        ]
        forecast_vendor_combo.set("All Vendors")
        forecast_vendor_combo.pack(side=tk.LEFT, padx=5)

        # Date range filter
        date_filter_frame = ttk.Frame(filter_frame)
        date_filter_frame.pack(fill=tk.X, pady=2)

        ttk.Label(date_filter_frame, text="Date Range:").pack(side=tk.LEFT, padx=(0, 5))
        self.forecast_start_date = tk.StringVar()
        self.forecast_end_date = tk.StringVar()

        ttk.Entry(
            date_filter_frame, textvariable=self.forecast_start_date, width=12
        ).pack(side=tk.LEFT, padx=2)
        ttk.Label(date_filter_frame, text="to").pack(side=tk.LEFT, padx=5)
        ttk.Entry(
            date_filter_frame, textvariable=self.forecast_end_date, width=12
        ).pack(side=tk.LEFT, padx=2)
        ttk.Label(date_filter_frame, text="(YYYY-MM-DD)").pack(side=tk.LEFT, padx=5)

        ttk.Button(
            date_filter_frame, text="Load Forecast", command=self.load_forecast_summary
        ).pack(side=tk.LEFT, padx=10)
        ttk.Button(
            date_filter_frame, text="Export to Excel", command=self.export_forecast
        ).pack(side=tk.LEFT)

        # Tree view
        tree_frame = ttk.Frame(parent)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        cols = (
            "Vendor",
            "Material",
            "Description",
            "Total Qty",
            "Total Value",
            "Currency",
            "Date Range",
        )
        self.forecast_tree = ttk.Treeview(tree_frame, columns=cols, show="headings")

        for col in cols:
            self.forecast_tree.heading(col, text=col)
            if col in ["Total Qty", "Total Value"]:
                self.forecast_tree.column(col, width=100, anchor="e")
            elif col == "Currency":
                self.forecast_tree.column(col, width=80)
            elif col == "Material":
                self.forecast_tree.column(col, width=120)

        forecast_scroll = ttk.Scrollbar(
            tree_frame, orient="vertical", command=self.forecast_tree.yview
        )
        self.forecast_tree.configure(yscrollcommand=forecast_scroll.set)
        self.forecast_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        forecast_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # Summary label
        self.forecast_summary_label = ttk.Label(
            parent, text="", font=("Helvetica", 10, "bold")
        )
        self.forecast_summary_label.pack(pady=5)

    def setup_requisitions_tab(self, parent):
        """Setup the requisitions view tab"""
        # Filter section
        filter_frame = ttk.LabelFrame(parent, text="Filters", padding=10)
        filter_frame.pack(fill=tk.X, padx=10, pady=5)

        filter_controls = ttk.Frame(filter_frame)
        filter_controls.pack(fill=tk.X)

        ttk.Label(filter_controls, text="Vendor:").pack(side=tk.LEFT, padx=(0, 5))
        self.req_filter_vendor_var = tk.StringVar()
        req_vendor_combo = ttk.Combobox(
            filter_controls,
            textvariable=self.req_filter_vendor_var,
            state="readonly",
            width=30,
        )
        req_vendor_combo["values"] = ["All Vendors"] + [
            v["display_name"] for v in self.dm.get_all_vendors()
        ]
        req_vendor_combo.set("All Vendors")
        req_vendor_combo.pack(side=tk.LEFT, padx=5)

        ttk.Button(
            filter_controls, text="Load Requisitions", command=self.load_requisitions
        ).pack(side=tk.LEFT, padx=10)
        ttk.Button(
            filter_controls,
            text="Mark as Converted to PO",
            command=self.convert_selected_to_po,
        ).pack(side=tk.LEFT, padx=5)
        ttk.Button(
            filter_controls, text="Export to Excel", command=self.export_requisitions
        ).pack(side=tk.LEFT)

        # Tree view
        tree_frame = ttk.Frame(parent)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        cols = (
            "PR Number",
            "Item",
            "Material",
            "Description",
            "Vendor",
            "Qty",
            "Delivery Date",
            "Unit Price",
            "Total Value",
            "Status",
            "PR Status",
        )
        self.req_tree = ttk.Treeview(tree_frame, columns=cols, show="headings")

        for col in cols:
            self.req_tree.heading(col, text=col)
            if col in ["PR Number", "Item"]:
                self.req_tree.column(col, width=100)
            elif col in ["Qty", "Unit Price", "Total Value"]:
                self.req_tree.column(col, width=90, anchor="e")
            elif col in ["Status", "PR Status"]:
                self.req_tree.column(col, width=100)

        req_scroll = ttk.Scrollbar(
            tree_frame, orient="vertical", command=self.req_tree.yview
        )
        self.req_tree.configure(yscrollcommand=req_scroll.set)
        self.req_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        req_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # Summary label
        self.req_summary_label = ttk.Label(
            parent, text="", font=("Helvetica", 10, "bold")
        )
        self.req_summary_label.pack(pady=5)

    def setup_analysis_tab(self, parent):
        """Setup the forecast vs actual analysis tab"""
        info_frame = ttk.LabelFrame(
            parent, text="Forecast Accuracy Analysis", padding=10
        )
        info_frame.pack(fill=tk.X, padx=10, pady=5)

        info_text = """This tab will show forecast vs actual comparison.
        
Features:
- Compare forecasted quantities vs actual orders placed
- Calculate variance and accuracy percentages
- Identify materials with largest forecast errors
- Export accuracy reports for vendor performance review"""

        ttk.Label(info_frame, text=info_text, justify=tk.LEFT).pack(anchor="w")

        ttk.Button(
            info_frame,
            text="Generate Accuracy Report",
            command=self.generate_accuracy_report,
        ).pack(pady=10)

        # Placeholder for analysis results
        analysis_frame = ttk.Frame(parent)
        analysis_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        self.analysis_text = scrolledtext.ScrolledText(analysis_frame, wrap=tk.WORD)
        self.analysis_text.pack(fill=tk.BOTH, expand=True)

    # Replace the entire setup_outbound_forecast_tab and add all missing methods to the ForecastManagementWindow class

    # Replace the entire setup_outbound_forecast_tab and add all missing methods to the ForecastManagementWindow class

    def setup_outbound_forecast_tab(self, parent):
        """Setup the outbound forecast generation tab"""
        # Info section
        info_frame = ttk.LabelFrame(
            parent, text="Generate Demand Forecast for Vendors", padding=10
        )
        info_frame.pack(fill=tk.X, padx=10, pady=5)

        info_text = "This tool generates YOUR demand forecast to send TO vendors.\nBased on your requisitions and open orders, the system creates a forecast showing\nvendors what quantities you'll need in upcoming weeks/months."

        ttk.Label(info_frame, text=info_text, justify=tk.LEFT).pack(anchor="w")

        # Configuration section
        config_frame = ttk.LabelFrame(parent, text="Forecast Configuration", padding=10)
        config_frame.pack(fill=tk.X, padx=10, pady=5)

        # Vendor selection with search and checkboxes
        vendor_frame = ttk.LabelFrame(config_frame, text="Select Vendors", padding=5)
        vendor_frame.pack(fill=tk.X, pady=5)

        # Search box
        search_frame = ttk.Frame(vendor_frame)
        search_frame.pack(fill=tk.X, pady=(0, 5))

        ttk.Label(search_frame, text="Search:").pack(side=tk.LEFT, padx=(0, 5))
        self.vendor_search_var = tk.StringVar()
        self.vendor_search_var.trace("w", lambda *args: self.filter_vendor_list())
        search_entry = ttk.Entry(
            search_frame, textvariable=self.vendor_search_var, width=25
        )
        search_entry.pack(side=tk.LEFT, padx=5)

        # Select/Deselect all buttons
        btn_frame = ttk.Frame(search_frame)
        btn_frame.pack(side=tk.LEFT, padx=(10, 0))
        ttk.Button(
            btn_frame, text="Select All", command=self.select_all_vendors, width=10
        ).pack(side=tk.LEFT, padx=2)
        ttk.Button(
            btn_frame, text="Deselect All", command=self.deselect_all_vendors, width=10
        ).pack(side=tk.LEFT, padx=2)

        # Scrollable vendor list with checkboxes - NARROWER
        list_container = ttk.Frame(vendor_frame)
        list_container.pack(fill=tk.BOTH, expand=True)

        # Canvas and scrollbar for custom checkbox list
        self.vendor_canvas = tk.Canvas(
            list_container, height=120, width=400, bg="white"
        )
        scrollbar = ttk.Scrollbar(
            list_container, orient="vertical", command=self.vendor_canvas.yview
        )
        self.vendor_checkbox_frame = ttk.Frame(self.vendor_canvas)

        self.vendor_checkbox_frame.bind(
            "<Configure>",
            lambda e: self.vendor_canvas.configure(
                scrollregion=self.vendor_canvas.bbox("all")
            ),
        )

        canvas_window = self.vendor_canvas.create_window(
            (0, 0), window=self.vendor_checkbox_frame, anchor="nw"
        )
        self.vendor_canvas.configure(yscrollcommand=scrollbar.set)

        # Bind mouse wheel scrolling
        self.vendor_canvas.bind("<Enter>", self._bind_mousewheel)
        self.vendor_canvas.bind("<Leave>", self._unbind_mousewheel)

        self.vendor_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Store vendor checkboxes
        self.vendor_checkboxes = {}
        self.vendor_check_vars = {}

        # Populate vendor list
        self.populate_vendor_checkboxes()

        # Time horizon
        horizon_frame = ttk.Frame(config_frame)
        horizon_frame.pack(fill=tk.X, pady=5)

        ttk.Label(horizon_frame, text="Forecast Horizon:").pack(
            side=tk.LEFT, padx=(0, 5)
        )
        self.forecast_weeks_var = tk.StringVar(value="13")
        ttk.Entry(horizon_frame, textvariable=self.forecast_weeks_var, width=10).pack(
            side=tk.LEFT, padx=5
        )
        ttk.Label(horizon_frame, text="weeks").pack(side=tk.LEFT, padx=5)

        # Grouping
        group_frame = ttk.Frame(config_frame)
        group_frame.pack(fill=tk.X, pady=5)

        ttk.Label(group_frame, text="Group By:").pack(side=tk.LEFT, padx=(0, 5))
        self.outbound_group_var = tk.StringVar(value="week")
        ttk.Radiobutton(
            group_frame, text="Week", variable=self.outbound_group_var, value="week"
        ).pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(
            group_frame, text="Month", variable=self.outbound_group_var, value="month"
        ).pack(side=tk.LEFT, padx=5)

        # Include open orders
        self.include_orders_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            config_frame,
            text="Include existing open orders in forecast",
            variable=self.include_orders_var,
        ).pack(anchor="w", pady=5)

        # Actions
        action_frame = ttk.Frame(config_frame)
        action_frame.pack(fill=tk.X, pady=10)

        ttk.Button(
            action_frame,
            text="Generate Preview",
            command=self.preview_outbound_forecast,
        ).pack(side=tk.LEFT, padx=5)
        ttk.Button(
            action_frame,
            text="Export to Excel",
            command=self.export_outbound_forecast_excel,
        ).pack(side=tk.LEFT, padx=5)
        ttk.Button(
            action_frame,
            text="Export to PDF",
            command=self.export_outbound_forecast_pdf,
        ).pack(side=tk.LEFT, padx=5)
        ttk.Button(
            action_frame,
            text="Generate & Email to Vendors",
            command=self.email_outbound_forecasts,
        ).pack(side=tk.LEFT, padx=5)

        # Preview section
        preview_frame = ttk.LabelFrame(parent, text="Forecast Preview", padding=10)
        preview_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        self.outbound_preview_text = scrolledtext.ScrolledText(
            preview_frame, wrap=tk.WORD
        )
        self.outbound_preview_text.pack(fill=tk.BOTH, expand=True)

    def _bind_mousewheel(self, event):
        """Bind mouse wheel to canvas scrolling"""
        self.vendor_canvas.bind_all("<MouseWheel>", self._on_mousewheel)
        self.vendor_canvas.bind_all(
            "<Button-4>", self._on_mousewheel
        )  # Linux scroll up
        self.vendor_canvas.bind_all(
            "<Button-5>", self._on_mousewheel
        )  # Linux scroll down

    def _unbind_mousewheel(self, event):
        """Unbind mouse wheel from canvas"""
        self.vendor_canvas.unbind_all("<MouseWheel>")
        self.vendor_canvas.unbind_all("<Button-4>")
        self.vendor_canvas.unbind_all("<Button-5>")

    def _on_mousewheel(self, event):
        """Handle mouse wheel scrolling"""
        if event.num == 5 or event.delta < 0:
            # Scroll down
            self.vendor_canvas.yview_scroll(1, "units")
        elif event.num == 4 or event.delta > 0:
            # Scroll up
            self.vendor_canvas.yview_scroll(-1, "units")

    def populate_vendor_checkboxes(self):
        """Populate the vendor checkbox list"""
        # Clear existing checkboxes
        for widget in self.vendor_checkbox_frame.winfo_children():
            widget.destroy()

        self.vendor_checkboxes.clear()
        self.vendor_check_vars.clear()

        # Get all vendors
        vendors = self.dm.get_all_vendors()

        # Create checkbox for each vendor
        for idx, vendor in enumerate(vendors):
            vendor_name = vendor["display_name"]
            var = tk.BooleanVar(value=False)
            self.vendor_check_vars[vendor_name] = var

            cb = ttk.Checkbutton(
                self.vendor_checkbox_frame, text=vendor_name, variable=var
            )
            cb.grid(row=idx, column=0, sticky="w", padx=5, pady=2)
            self.vendor_checkboxes[vendor_name] = cb

    def filter_vendor_list(self):
        """Filter vendor list based on search term"""
        search_term = self.vendor_search_var.get().lower()

        for vendor_name, checkbox in self.vendor_checkboxes.items():
            if search_term in vendor_name.lower():
                checkbox.grid()
            else:
                checkbox.grid_remove()

    def select_all_vendors(self):
        """Select all visible vendors"""
        search_term = self.vendor_search_var.get().lower()

        for vendor_name, var in self.vendor_check_vars.items():
            # Only select if visible (matches search)
            if search_term in vendor_name.lower():
                var.set(True)

    def deselect_all_vendors(self):
        """Deselect all vendors"""
        for var in self.vendor_check_vars.values():
            var.set(False)

    def get_selected_vendors(self):
        """Get list of selected vendor names"""
        selected = []
        for vendor_name, var in self.vendor_check_vars.items():
            if var.get():
                selected.append(vendor_name)
        return selected

    # Action methods for Upload tab
    def upload_forecast(self):
        """Upload vendor forecast file"""
        vendor_name = self.forecast_vendor_var.get()
        if not vendor_name:
            messagebox.showwarning(
                "No Vendor", "Please select a vendor first.", parent=self
            )
            return

        file_path = filedialog.askopenfilename(
            title="Select Forecast Excel File",
            filetypes=[("Excel Files", "*.xlsx;*.xls"), ("All Files", "*.*")],
        )

        if not file_path:
            return

        try:
            count = self.fm.upload_forecast_file(file_path, vendor_name)
            message = (
                f"Successfully imported {count} forecast entries for {vendor_name}"
            )
            self.status_text.insert(
                tk.END, f"{datetime.now().strftime('%H:%M:%S')} - {message}\n"
            )
            self.status_text.see(tk.END)
            self.log(f"✓ SUCCESS: {message}")
            messagebox.showinfo("✅ Success", message, parent=self)
        except Exception as e:
            error_msg = f"Failed to upload forecast: {str(e)}"
            self.status_text.insert(
                tk.END, f"{datetime.now().strftime('%H:%M:%S')} - ERROR: {error_msg}\n"
            )
            self.status_text.see(tk.END)
            self.log(f"✗ ERROR: {error_msg}")
            messagebox.showerror("Upload Error", error_msg, parent=self)

    def upload_requisitions(self):
        """Upload requisitions file - ENHANCED with material master auto-population"""
        file_path = filedialog.askopenfilename(
            title="Select Requisitions Excel File",
            filetypes=[("Excel Files", "*.xlsx;*.xls"), ("All Files", "*.*")],
        )

        if not file_path:
            return

        try:
            result = self.fm.upload_requisitions_file(file_path)

            # Handle tuple return (processed_count, closed_count, materials_created, materials_updated)
            if isinstance(result, tuple) and len(result) == 4:
                processed_count, closed_count, materials_created, materials_updated = (
                    result
                )
                message = f" Requisitions Upload Complete!\n\n"
                message += f" Requisitions: {processed_count} lines imported"

                if closed_count > 0:
                    message += f"\n Marked {closed_count} missing lines as 'Closed'"

                message += f"\n\n Materials Master Data:"
                message += f"\n   Created: {materials_created} new materials"
                message += f"\n   Updated: {materials_updated} existing materials"

                if materials_created > 0 or materials_updated > 0:
                    message += f"\n\n Material master data has been automatically populated with:"
                    message += f"\n   Material codes & descriptions"
                    message += f"\n   Lead times (from 'Planned Deliv. Time')"
                    message += f"\n   Preferred vendors"
                    message += f"\n\n You can now run MRP immediately!"

            elif isinstance(result, tuple) and len(result) == 2:
                # Old format compatibility
                processed_count, closed_count = result
                message = f"Successfully imported {processed_count} requisition lines."
                if closed_count > 0:
                    message += f"\nMarked {closed_count} missing lines as 'Closed'."
            else:
                processed_count = result
                message = f"Successfully imported {processed_count} requisition lines."

            self.status_text.insert(
                tk.END, f"{datetime.now().strftime('%H:%M:%S')} - {message}\n"
            )
            self.status_text.see(tk.END)
            self.log(f"✓ SUCCESS: Requisitions uploaded")
            messagebox.showinfo("Upload Complete", message, parent=self)

        except Exception as e:
            error_msg = f"Failed to upload requisitions: {str(e)}"
            self.status_text.insert(
                tk.END, f"{datetime.now().strftime('%H:%M:%S')} - ERROR: {error_msg}\n"
            )
            self.status_text.see(tk.END)
            self.log(f"✗ ERROR: {error_msg}")
            messagebox.showerror("Upload Error", error_msg, parent=self)

    def clear_requisitions(self):
        """Clear all requisitions from database"""
        if not messagebox.askyesno(
            "Confirm Clear",
            "Are you sure you want to delete ALL requisitions?\n\nThis action cannot be undone.",
            parent=self,
        ):
            return

        try:
            deleted_count = self.fm.clear_all_requisitions()
            message = f"Successfully deleted {deleted_count} requisition(s)."
            self.status_text.insert(
                tk.END, f"{datetime.now().strftime('%H:%M:%S')} - {message}\n"
            )
            self.status_text.see(tk.END)
            self.log(f"✓ SUCCESS: {message}")
            messagebox.showinfo("✅ Success", message, parent=self)

            # Refresh the requisitions view if it's loaded
            if hasattr(self, "req_tree"):
                self.load_requisitions()
        except Exception as e:
            error_msg = f"Failed to clear requisitions: {str(e)}"
            self.status_text.insert(
                tk.END, f"{datetime.now().strftime('%H:%M:%S')} - ERROR: {error_msg}\n"
            )
            self.status_text.see(tk.END)
            self.log(f"✗ ERROR: {error_msg}")
            messagebox.showerror("❌ Error", error_msg, parent=self)

    def download_template(self):
        """Download forecast template"""
        vendor_name = self.forecast_vendor_var.get()
        if not vendor_name:
            messagebox.showwarning(
                "No Vendor", "Please select a vendor first.", parent=self
            )
            return

        try:
            template_buffer = self.fm.generate_forecast_template(vendor_name)

            save_path = filedialog.asksaveasfilename(
                title="Save Forecast Template",
                defaultextension=".xlsx",
                initialfile=f"Forecast_Template_{vendor_name}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
            )

            if save_path:
                with open(save_path, "wb") as f:
                    f.write(template_buffer.read())

                message = f"Template saved to {save_path}"
                self.status_text.insert(
                    tk.END, f"{datetime.now().strftime('%H:%M:%S')} - {message}\n"
                )
                self.status_text.see(tk.END)
                self.log(f"✓ SUCCESS: {message}")
                messagebox.showinfo("✅ Success", message, parent=self)
        except Exception as e:
            error_msg = f"Failed to generate template: {str(e)}"
            self.status_text.insert(
                tk.END, f"{datetime.now().strftime('%H:%M:%S')} - ERROR: {error_msg}\n"
            )
            self.status_text.see(tk.END)
            self.log(f"✗ ERROR: {error_msg}")
            messagebox.showerror("❌ Error", error_msg, parent=self)

    def link_requisitions(self):
        """Link requisitions to forecast data"""
        vendor = self.link_vendor_var.get()
        vendor_name = None if vendor == "All Vendors" else vendor

        try:
            count = self.fm.link_requisitions_to_forecast(vendor_name)
            message = f"Linked {count} requisitions to forecast data"
            self.status_text.insert(
                tk.END, f"{datetime.now().strftime('%H:%M:%S')} - {message}\n"
            )
            self.status_text.see(tk.END)
            self.log(f"✓ SUCCESS: {message}")
            messagebox.showinfo("✅ Success", message, parent=self)
        except Exception as e:
            error_msg = f"Failed to link requisitions: {str(e)}"
            self.status_text.insert(
                tk.END, f"{datetime.now().strftime('%H:%M:%S')} - ERROR: {error_msg}\n"
            )
            self.status_text.see(tk.END)
            self.log(f"✗ ERROR: {error_msg}")
            messagebox.showerror("❌ Error", error_msg, parent=self)

    def load_forecast_summary(self):
        """Load and display forecast summary"""
        for item in self.forecast_tree.get_children():
            self.forecast_tree.delete(item)

        vendor = self.forecast_filter_vendor_var.get()
        vendor_name = None if vendor == "All Vendors" else vendor

        start_date = self.forecast_start_date.get().strip() or None
        end_date = self.forecast_end_date.get().strip() or None

        try:
            forecasts = self.fm.get_forecast_summary(vendor_name, start_date, end_date)

            total_value = 0

            for fc in forecasts:
                values = (
                    fc["vendor_display"],
                    fc["material_code"],
                    fc["short_text"],
                    f"{fc['total_qty']:,}",
                    f"{fc['total_value']:.2f}",
                    fc["currency"],
                    f"{fc['date_range']}",
                )
                self.forecast_tree.insert("", "end", values=values)
                total_value += fc["total_value"]

            currency = forecasts[0]["currency"] if forecasts else "EUR"
            summary = f"Total: {len(forecasts)} forecast items | {total_value:,.2f} {currency}"
            self.forecast_summary_label.config(text=summary)

            self.log(f"ℹ️ INFO: Loaded {len(forecasts)} forecast items")
        except Exception as e:
            self.log(f"✗ ERROR: Failed to load forecast: {e}")
            messagebox.showerror("❌ Error", f"Failed to load forecast:\n{str(e)}", parent=self
            )

    def load_requisitions(self):
        """Load and display requisitions"""
        for item in self.req_tree.get_children():
            self.req_tree.delete(item)

        vendor = self.req_filter_vendor_var.get()
        vendor_name = None if vendor == "All Vendors" else vendor

        try:
            requisitions = self.fm.get_open_requisitions(vendor_name)

            total_value = 0

            for req in requisitions:
                values = (
                    req["req_number"],
                    req["item"],
                    req["material_code"],
                    req["short_text"],
                    req.get("vendor_display", "N/A"),
                    f"{req['requested_qty']:,}",
                    req.get("requested_del_date", "N/A"),
                    f"{req['unit_price']:.2f}",
                    f"{req['total_amount']:.2f}",
                    req["status"],
                    req["pr_status"],
                )
                self.req_tree.insert("", "end", values=values)
                total_value += req["total_amount"]

            currency = requisitions[0]["currency"] if requisitions else "EUR"
            summary = f"Total: {len(requisitions)} requisition lines | {total_value:,.2f} {currency}"
            self.req_summary_label.config(text=summary)

            self.log(f"ℹ️ INFO: Loaded {len(requisitions)} requisition items")
        except Exception as e:
            self.log(f"✗ ERROR: Failed to load requisitions: {e}")
            messagebox.showerror("❌ Error", f"Failed to load requisitions:\n{str(e)}", parent=self
            )

    def convert_selected_to_po(self):
        """Mark selected requisitions as converted to PO"""
        selected = self.req_tree.selection()
        if not selected:
            messagebox.showwarning(
                "No Selection", "Please select requisitions to convert.", parent=self
            )
            return

        req_numbers = list(
            set([self.req_tree.item(item, "values")[0] for item in selected])
        )

        if not messagebox.askyesno("❓ Confirm",
            f"Mark {len(req_numbers)} requisition(s) as converted to PO?",
            parent=self,
        ):
            return

        try:
            count = self.fm.convert_requisition_to_po(req_numbers)
            self.log(f"✓ SUCCESS: Converted {count} requisitions")
            messagebox.showinfo("✅ Success", f"Converted {count} requisitions to PO status", parent=self
            )
            self.load_requisitions()
        except Exception as e:
            self.log(f"✗ ERROR: Failed to convert requisitions: {e}")
            messagebox.showerror("❌ Error", f"Failed to convert:\n{str(e)}", parent=self)

    def export_forecast(self):
        """Export forecast data to Excel"""
        vendor = self.forecast_filter_vendor_var.get()
        vendor_name = None if vendor == "All Vendors" else vendor

        start_date = self.forecast_start_date.get().strip() or None
        end_date = self.forecast_end_date.get().strip() or None

        try:
            forecasts = self.fm.get_forecast_summary(vendor_name, start_date, end_date)

            if not forecasts:
                messagebox.showinfo(
                    "No Data", "No forecast data to export.", parent=self
                )
                return

            df = pd.DataFrame(forecasts)

            save_path = filedialog.asksaveasfilename(
                title="Export Forecast",
                defaultextension=".xlsx",
                initialfile=f"Forecast_Export_{datetime.now().strftime('%Y%m%d')}.xlsx",
                filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
            )

            if save_path:
                with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
                    df.to_excel(writer, sheet_name="Forecast", index=False)

                self.log(f"✓ SUCCESS: Exported forecast to {save_path}")
                messagebox.showinfo("✅ Success", f"Exported to {save_path}", parent=self)
        except Exception as e:
            self.log(f"✗ ERROR: Failed to export: {e}")
            messagebox.showerror("❌ Error", f"Failed to export:\n{str(e)}", parent=self)

    def export_requisitions(self):
        """Export requisitions to Excel"""
        vendor = self.req_filter_vendor_var.get()
        vendor_name = None if vendor == "All Vendors" else vendor

        try:
            requisitions = self.fm.get_open_requisitions(vendor_name)

            if not requisitions:
                messagebox.showinfo(
                    "No Data", "No requisitions to export.", parent=self
                )
                return

            df = pd.DataFrame(requisitions)

            save_path = filedialog.asksaveasfilename(
                title="Export Requisitions",
                defaultextension=".xlsx",
                initialfile=f"Requisitions_Export_{datetime.now().strftime('%Y%m%d')}.xlsx",
                filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
            )

            if save_path:
                with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
                    df.to_excel(writer, sheet_name="Requisitions", index=False)

                self.log(f"✓ SUCCESS: Exported requisitions to {save_path}")
                messagebox.showinfo("✅ Success", f"Exported to {save_path}", parent=self)
        except Exception as e:
            self.log(f"✗ ERROR: Failed to export: {e}")
            messagebox.showerror("❌ Error", f"Failed to export:\n{str(e)}", parent=self)

    def generate_accuracy_report(self):
        """Generate forecast accuracy analysis report"""
        self.analysis_text.delete("1.0", tk.END)
        self.analysis_text.insert(tk.END, "Generating forecast accuracy report...\n\n")

        # This would compare forecast quantities against actual orders
        # For now, show a placeholder
        report = """FORECAST ACCURACY ANALYSIS

==========================

This feature will compare:
- Forecasted quantities by period
- Actual orders received in those periods
- Variance (actual - forecast)
- Accuracy percentage

Materials will be ranked by:
- Largest absolute variance
- Largest percentage error
- Most consistent accuracy

Report will help identify:
- Which materials need better forecasting
- Which vendors provide accurate forecasts
- Seasonal patterns and trends
"""

        self.analysis_text.insert(tk.END, report)
        self.log("ℹ️ INFO: Accuracy report generated")

    def preview_outbound_forecast(self):
        """Generate and display forecast preview"""
        self.outbound_preview_text.delete("1.0", tk.END)

        selected_vendors = self.get_selected_vendors()

        if not selected_vendors:
            self.outbound_preview_text.insert(
                tk.END, "Please select at least one vendor."
            )
            return

        try:
            weeks = int(self.forecast_weeks_var.get())
        except Exception:
            weeks = 13

        group_by = self.outbound_group_var.get()
        include_orders = self.include_orders_var.get()

        try:
            # Generate forecast for all vendors, then filter
            forecast_data = self.fm.generate_outbound_forecast_from_requisitions(
                None, weeks, group_by, include_orders
            )

            # Filter to only selected vendors
            forecast_data = {
                v: data for v, data in forecast_data.items() if v in selected_vendors
            }

            if not forecast_data:
                self.outbound_preview_text.insert(
                    tk.END, "No forecast data available for the selected vendors."
                )
                return

            # Display preview
            preview_text = f"OUTBOUND DEMAND FORECAST PREVIEW\n"
            preview_text += f"={'=' * 60}\n\n"
            preview_text += f"Generated: {datetime.now().strftime('%d.%m.%Y %H:%M')}\n"
            preview_text += f"Forecast Horizon: {weeks} weeks\n"
            preview_text += f"Grouping: {group_by.capitalize()}\n"
            preview_text += (
                f"Include Open Orders: {'Yes' if include_orders else 'No'}\n"
            )
            preview_text += f"Selected Vendors: {len(selected_vendors)}\n\n"

            for vendor, materials in forecast_data.items():
                preview_text += f"\n{'=' * 60}\n"
                preview_text += f"VENDOR: {vendor}\n"
                preview_text += f"{'=' * 60}\n\n"

                vendor_total = 0

                for material, data in sorted(materials.items()):
                    preview_text += f"Material: {material}\n"
                    preview_text += f"Description: {data['description']}\n"
                    preview_text += f"Unit: {data['unit']}\n"
                    preview_text += f"\nForecast by Period:\n"

                    sorted_periods = sorted(
                        data["periods"].items(), key=lambda x: x[1]["date"]
                    )

                    material_total = 0
                    for period_key, period_info in sorted_periods:
                        qty = period_info["qty"]
                        material_total += qty
                        preview_text += (
                            f"  {period_info['label']}: {qty:,} {data['unit']}\n"
                        )

                    preview_text += (
                        f"\nMaterial Total: {material_total:,} {data['unit']}\n"
                    )
                    preview_text += f"{'-' * 40}\n\n"

                    vendor_total += material_total

                preview_text += f"\nVENDOR TOTAL: {vendor_total:,} units\n\n"

            self.outbound_preview_text.insert(tk.END, preview_text)
            self.log("ℹ️ INFO: Outbound forecast preview generated")

        except Exception as e:
            error_msg = f"Failed to generate forecast preview: {str(e)}"
            self.outbound_preview_text.insert(tk.END, error_msg)
            self.log(f"✗ ERROR: {error_msg}")

    def export_outbound_forecast_excel(self):
        """Export outbound forecast to Excel"""
        selected_vendors = self.get_selected_vendors()

        if not selected_vendors:
            messagebox.showinfo(
                "No Selection", "Please select at least one vendor.", parent=self
            )
            return

        try:
            weeks = int(self.forecast_weeks_var.get())
        except Exception:
            weeks = 13

        group_by = self.outbound_group_var.get()
        include_orders = self.include_orders_var.get()

        try:
            forecast_data = self.fm.generate_outbound_forecast_from_requisitions(
                None, weeks, group_by, include_orders
            )

            # Filter to selected vendors
            forecast_data = {
                v: data for v, data in forecast_data.items() if v in selected_vendors
            }

            if not forecast_data:
                messagebox.showinfo(
                    "No Data",
                    "No forecast data available for the selected vendors.",
                    parent=self,
                )
                return

            # If single vendor, export single file
            if len(selected_vendors) == 1:
                vendor = selected_vendors[0]
                excel_buffer = self.fm.create_outbound_forecast_excel(
                    vendor, forecast_data
                )

                safe_vendor = re.sub(r'[\\/*?:"<>|]', "_", vendor)
                save_path = filedialog.asksaveasfilename(
                    title="Save Outbound Forecast",
                    defaultextension=".xlsx",
                    initialfile=f"Demand_Forecast_{safe_vendor}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
                )

                if save_path:
                    with open(save_path, "wb") as f:
                        f.write(excel_buffer.read())

                    self.log(f"✓ SUCCESS: Exported outbound forecast to {save_path}")
                    messagebox.showinfo("✅ Success", f"Forecast exported to:\n{save_path}", parent=self
                    )
            else:
                # Multiple vendors - create zip file
                save_path = filedialog.asksaveasfilename(
                    title="Save Outbound Forecasts (ZIP)",
                    defaultextension=".zip",
                    initialfile=f"Demand_Forecasts_{datetime.now().strftime('%Y%m%d')}.zip",
                    filetypes=[("ZIP Files", "*.zip"), ("All Files", "*.*")],
                )

                if save_path:
                    import zipfile

                    with zipfile.ZipFile(save_path, "w") as zf:
                        for vendor in selected_vendors:
                            if vendor in forecast_data:
                                excel_buffer = self.fm.create_outbound_forecast_excel(
                                    vendor, {vendor: forecast_data[vendor]}
                                )
                                safe_vendor = re.sub(r'[\\/*?:"<>|]', "_", vendor)
                                zf.writestr(
                                    f"{safe_vendor}_forecast.xlsx", excel_buffer.read()
                                )

                    self.log(
                        f"✓ SUCCESS: Exported {len(selected_vendors)} forecasts to {save_path}"
                    )
                    messagebox.showinfo("✅ Success",
                        f"Exported {len(selected_vendors)} forecasts to:\n{save_path}",
                        parent=self,
                    )

        except Exception as e:
            error_msg = f"Failed to export forecast: {str(e)}"
            self.log(f"✗ ERROR: {error_msg}")
            messagebox.showerror("❌ Error", error_msg, parent=self)

    def export_outbound_forecast_pdf(self):
        """Export outbound forecast to PDF"""
        selected_vendors = self.get_selected_vendors()

        if not selected_vendors:
            messagebox.showinfo(
                "No Selection", "Please select at least one vendor.", parent=self
            )
            return

        try:
            weeks = int(self.forecast_weeks_var.get())
        except Exception:
            weeks = 13

        group_by = self.outbound_group_var.get()
        include_orders = self.include_orders_var.get()

        try:
            forecast_data = self.fm.generate_outbound_forecast_from_requisitions(
                None, weeks, group_by, include_orders
            )

            # Filter to selected vendors
            forecast_data = {
                v: data for v, data in forecast_data.items() if v in selected_vendors
            }

            if not forecast_data:
                messagebox.showinfo(
                    "No Data",
                    "No forecast data available for the selected vendors.",
                    parent=self,
                )
                return

            # If single vendor, export single file
            if len(selected_vendors) == 1:
                vendor = selected_vendors[0]
                pdf_buffer = self.fm.create_outbound_forecast_pdf(
                    vendor, {vendor: forecast_data[vendor]}
                )

                safe_vendor = re.sub(r'[\\/*?:"<>|]', "_", vendor)
                save_path = filedialog.asksaveasfilename(
                    title="Save Outbound Forecast",
                    defaultextension=".pdf",
                    initialfile=f"Demand_Forecast_{safe_vendor}_{datetime.now().strftime('%Y%m%d')}.pdf",
                    filetypes=[("PDF Files", "*.pdf"), ("All Files", "*.*")],
                )

                if save_path:
                    with open(save_path, "wb") as f:
                        f.write(pdf_buffer.read())

                    self.log(f"✓ SUCCESS: Exported outbound forecast PDF to {save_path}")
                    messagebox.showinfo("✅ Success", f"Forecast exported to:\n{save_path}", parent=self
                    )
            else:
                # Multiple vendors - create zip file
                save_path = filedialog.asksaveasfilename(
                    title="Save Outbound Forecasts (ZIP)",
                    defaultextension=".zip",
                    initialfile=f"Demand_Forecasts_{datetime.now().strftime('%Y%m%d')}.zip",
                    filetypes=[("ZIP Files", "*.zip"), ("All Files", "*.*")],
                )

                if save_path:
                    import zipfile

                    with zipfile.ZipFile(save_path, "w") as zf:
                        for vendor in selected_vendors:
                            if vendor in forecast_data:
                                pdf_buffer = self.fm.create_outbound_forecast_pdf(
                                    vendor, {vendor: forecast_data[vendor]}
                                )
                                safe_vendor = re.sub(r'[\\/*?:"<>|]', "_", vendor)
                                zf.writestr(
                                    f"{safe_vendor}_forecast.pdf", pdf_buffer.read()
                                )

                    self.log(
                        f"✓ SUCCESS: Exported {len(selected_vendors)} forecasts to {save_path}"
                    )
                    messagebox.showinfo("✅ Success",
                        f"Exported {len(selected_vendors)} forecasts to:\n{save_path}",
                        parent=self,
                    )

        except Exception as e:
            error_msg = f"Failed to export forecast PDF: {str(e)}"
            self.log(f"✗ ERROR: {error_msg}")
            messagebox.showerror("❌ Error", error_msg, parent=self)

    def email_outbound_forecasts(self):
        """Generate and email outbound forecasts to vendors using templates"""
        selected_vendors = self.get_selected_vendors()

        if not selected_vendors:
            messagebox.showinfo(
                "No Selection", "Please select at least one vendor.", parent=self
            )
            return

        try:
            weeks = int(self.forecast_weeks_var.get())
        except Exception:
            weeks = 13

        group_by = self.outbound_group_var.get()
        include_orders = self.include_orders_var.get()

        try:
            forecast_data = self.fm.generate_outbound_forecast_from_requisitions(
                None, weeks, group_by, include_orders
            )

            # Filter to selected vendors
            forecast_data = {
                v: data for v, data in forecast_data.items() if v in selected_vendors
            }

            if not forecast_data:
                messagebox.showinfo(
                    "No Data",
                    "No forecast data available for the selected vendors.",
                    parent=self,
                )
                return

            # Ask for format
            format_choice = messagebox.askquestion(
                "File Format",
                "Send forecasts as Excel files?\n\n"
                "Yes = Excel format\n"
                "No = PDF format",
                parent=self,
            )

            use_excel = format_choice == "yes"

            # Ask for send method
            method = messagebox.askquestion(
                "Email Method",
                "Send via Outlook?\n\nYes = Outlook\nNo = SMTP",
                parent=self,
            )

            send_method = "Outlook" if method == "yes" else "SMTP"

            if send_method == "Outlook" and not OUTLOOK_AVAILABLE:
                messagebox.showerror("❌ Error",
                    "Outlook is not available. Please use SMTP method.",
                    parent=self,
                )
                return

            # Generate and send for each vendor
            sender = EmailSender(self.log, self.dm)
            sent_count = 0
            failed_count = 0
            failed_details = []

            # Get vendor email mapping
            vendor_emails = {}
            for v in self.dm.get_all_vendors():
                if v.get("emails"):
                    vendor_emails[v["display_name"]] = v["emails"]

            for vendor_name in selected_vendors:
                if vendor_name not in forecast_data:
                    continue

                if vendor_name not in vendor_emails:
                    failed_count += 1
                    failed_details.append(f"{vendor_name}: No email address")
                    continue

                try:
                    # Generate file
                    if use_excel:
                        file_buffer = self.fm.create_outbound_forecast_excel(
                            vendor_name, {vendor_name: forecast_data[vendor_name]}
                        )
                        file_ext = ".xlsx"
                    else:
                        file_buffer = self.fm.create_outbound_forecast_pdf(
                            vendor_name, {vendor_name: forecast_data[vendor_name]}
                        )
                        file_ext = ".pdf"

                    # Save to temp file
                    safe_vendor = re.sub(r'[\\/*?:"<>|]', "_", vendor_name)
                    temp_filename = f"Demand_Forecast_{safe_vendor}_{datetime.now().strftime('%Y%m%d')}{file_ext}"
                    temp_path = os.path.join(APP_DATA_FOLDER, temp_filename)

                    with open(temp_path, "wb") as f:
                        f.write(file_buffer.read())

                    # Generate email content using template
                    subject, body = sender.generate_forecast_email_content(vendor_name)
                    to_emails = vendor_emails[vendor_name]

                    # Send email
                    success = False
                    if send_method == "Outlook":
                        success, message = sender._send_outlook(
                            to_emails, subject, body, temp_path
                        )
                    else:
                        success, message = sender._send_smtp(
                            to_emails, subject, body, temp_path
                        )

                    if success:
                        sent_count += 1
                        self.log(f"✓ SUCCESS: Sent forecast to {vendor_name}")
                    else:
                        failed_count += 1
                        failed_details.append(f"{vendor_name}: {message}")

                    # Clean up temp file
                    try:
                        os.remove(temp_path)
                    except Exception:  # TODO: Add proper error handling
                        pass  # TODO: Add proper error handling
                except Exception as e:
                    failed_count += 1
                    failed_details.append(f"{vendor_name}: {str(e)}")
                    self.log(f"✗ ERROR: Failed to send forecast to {vendor_name}: {e}")

            # Show summary
            summary = f"Sent {sent_count} forecast email(s), Failed {failed_count}"
            if failed_details:
                summary += "\n\nFailures:\n" + "\n".join(failed_details)

            messagebox.showinfo("Email Results", summary, parent=self)
            self.log(
                f"ℹ️ INFO: Forecast email batch complete - {sent_count} sent, {failed_count} failed"
            )

        except Exception as e:
            messagebox.showerror("❌ Error", f"Failed to send forecasts:\n{str(e)}", parent=self
            )
            self.log(f"✗ ERROR: Failed to send forecasts: {e}")

    def email_single_vendor_forecast(self, vendor_name):
        """Email forecast to a single vendor"""
        try:
            weeks = int(self.forecast_weeks_var.get())
        except Exception:
            weeks = 13

        group_by = self.outbound_group_var.get()
        include_orders = self.include_orders_var.get()

        try:
            forecast_data = self.fm.generate_outbound_forecast_from_requisitions(
                vendor_name, weeks, group_by, include_orders
            )

            if not forecast_data or vendor_name not in forecast_data:
                messagebox.showinfo(
                    "No Data",
                    f"No forecast data available for {vendor_name}.",
                    parent=self,
                )
                return

            # Get vendor email
            vendor_data = next(
                (
                    v
                    for v in self.dm.get_all_vendors()
                    if v["display_name"] == vendor_name
                ),
                None,
            )
            if not vendor_data or not vendor_data.get("emails"):
                messagebox.showerror("❌ Error", f"No email address found for {vendor_name}.", parent=self
                )
                return

            # Ask for format
            format_choice = messagebox.askquestion(
                "File Format",
                "Send forecast as Excel file?\n\n"
                "Yes = Excel format\n"
                "No = PDF format",
                parent=self,
            )

            use_excel = format_choice == "yes"

            # Ask for send method
            method = messagebox.askquestion(
                "Email Method",
                "Send via Outlook?\n\nYes = Outlook\nNo = SMTP",
                parent=self,
            )

            send_method = "Outlook" if method == "yes" else "SMTP"

            if send_method == "Outlook" and not OUTLOOK_AVAILABLE:
                messagebox.showerror("❌ Error",
                    "Outlook is not available. Please use SMTP method.",
                    parent=self,
                )
                return

            # Generate file
            if use_excel:
                file_buffer = self.fm.create_outbound_forecast_excel(
                    vendor_name, forecast_data
                )
                file_ext = ".xlsx"
            else:
                file_buffer = self.fm.create_outbound_forecast_pdf(
                    vendor_name, forecast_data
                )
                file_ext = ".pdf"

            # Save to temp file
            safe_vendor = re.sub(r'[\\/*?:"<>|]', "_", vendor_name)
            temp_filename = f"Demand_Forecast_{safe_vendor}_{datetime.now().strftime('%Y%m%d')}{file_ext}"
            temp_path = os.path.join(APP_DATA_FOLDER, temp_filename)

            with open(temp_path, "wb") as f:
                f.write(file_buffer.read())

            # Generate email content using template
            sender = EmailSender(self.log, self.dm)
            subject, body = sender.generate_forecast_email_content(vendor_name)
            to_emails = vendor_data["emails"]

            # Send email
            success = False
            if send_method == "Outlook":
                success, message = sender._send_outlook(
                    to_emails, subject, body, temp_path
                )
            else:
                success, message = sender._send_smtp(
                    to_emails, subject, body, temp_path
                )

            if success:
                messagebox.showinfo("✅ Success",
                    f"Forecast sent successfully to {vendor_name}!",
                    parent=self,
                )
                self.log(f"✓ SUCCESS: Sent forecast to {vendor_name}")
            else:
                messagebox.showerror(
                    "Failed", f"Failed to send forecast: {message}", parent=self
                )

            # Clean up temp file
            try:
                os.remove(temp_path)
            except Exception:  # TODO: Add proper error handling
                pass  # TODO: Add proper error handling
        except Exception as e:
            messagebox.showerror("❌ Error", f"Failed to send forecast:\n{str(e)}", parent=self
            )
            self.log(f"✗ ERROR: Failed to send forecast: {e}")


# ==============================================================================
# MRP (Material Requirements Planning) Engine
# ==============================================================================


class MRPEngine:
    """Core MRP calculation engine for forecast-driven requisition generation"""

    def __init__(self, database_manager, log_callback=None):
        self.db = database_manager
        self.log = log_callback or print

    def run_mrp(self, horizon_weeks=13, materials=None, create_requisitions=True):
        """
        Run MRP calculation for specified materials and time horizon

        Args:
            horizon_weeks: Planning horizon in weeks
            materials: List of material codes, or None for all
            create_requisitions: Auto-create PRs for net requirements

        Returns:
            run_id: MRP run identifier for tracking
        """
        self.log("ℹ️ INFO: Starting MRP run...")

        # Create MRP run record
        run_params = {
            "horizon_weeks": horizon_weeks,
            "materials": materials or "ALL",
            "create_requisitions": create_requisitions,
        }

        run_query = """
            INSERT INTO mrp_runs (run_date, horizon_weeks, status, parameters)
            VALUES (?, ?, ?, ?)
        """

        run_date = datetime.now().isoformat()
        cursor = self.db.execute_query(
            run_query,
            (run_date, horizon_weeks, "RUNNING", json.dumps(run_params)),
            commit=True,
        )
        run_id = cursor.lastrowid

        try:
            # Get materials to plan
            if materials:
                placeholders = ",".join("?" * len(materials))
                materials_query = (
                    f"SELECT * FROM materials WHERE material_code IN ({placeholders})"
                )
                material_list = self.db.execute_query(
                    materials_query, tuple(materials), fetchall=True
                )
            else:
                material_list = self.db.execute_query(
                    "SELECT * FROM materials", fetchall=True
                )

            self.log(
                f"ℹ️ INFO: Planning {len(material_list)} materials over {horizon_weeks} weeks"
            )

            # Calculate planning periods
            start_date = datetime.now().date()
            periods = self._generate_periods(start_date, horizon_weeks)

            requisitions_created = 0

            # Run MRP for each material
            for material in material_list:
                material_code = material["material_code"]

                try:
                    # Run MRP calculation for this material
                    mrp_result = self._calculate_material_mrp(
                        material_code, periods, run_id
                    )

                    # Create requisitions if needed
                    if create_requisitions and mrp_result["requisitions"]:
                        for req in mrp_result["requisitions"]:
                            self._create_requisition_from_mrp(req, run_id)
                            requisitions_created += 1

                except Exception as e:
                    self.log(f"✗ ERROR: MRP calculation failed for {material_code}: {e}")
                    continue

            # Mark run as completed
            self.db.execute_query(
                "UPDATE mrp_runs SET status = ? WHERE run_id = ?",
                ("COMPLETED", run_id),
                commit=True,
            )

            self.log(
                f"✓ SUCCESS: MRP run {run_id} completed. Created {requisitions_created} requisitions."
            )
            return run_id

        except Exception as e:
            self.log(f"✗ ERROR: MRP run failed: {e}")
            self.db.execute_query(
                "UPDATE mrp_runs SET status = ? WHERE run_id = ?",
                ("FAILED", run_id),
                commit=True,
            )
            raise

    def _generate_periods(self, start_date, num_weeks):
        """Generate weekly planning periods"""
        periods = []
        current_date = start_date

        for week in range(num_weeks):
            # Start of week (Monday)
            week_start = current_date + timedelta(weeks=week)
            week_start = week_start - timedelta(days=week_start.weekday())

            periods.append(
                {
                    "week_num": week + 1,
                    "date": week_start,
                    "date_str": week_start.isoformat(),
                }
            )

        return periods

    def _calculate_material_mrp(self, material_code, periods, run_id):
        """
        Calculate MRP for a single material across all periods
        Time-phased MRP calculation with period-by-period logic
        """
        # Get material master data
        material = self.db.execute_query(
            "SELECT * FROM materials WHERE material_code = ?",
            (material_code,),
            fetchone=True,
        )

        if not material:
            raise ValueError(f"Material {material_code} not found")

        # Get current inventory
        inventory = self.db.execute_query(
            "SELECT * FROM inventory WHERE material_code = ?",
            (material_code,),
            fetchone=True,
        )

        on_hand = inventory["available_qty"] if inventory else 0
        lead_time_days = material.get("lead_time_days", 0)
        safety_stock = material.get("safety_stock", 0)

        # Initialize MRP table
        mrp_table = []
        requisitions = []

        for period in periods:
            period_date = period["date"]

            # 1. Get Gross Requirements (Forecast + Safety Stock)
            gross_req = self._get_gross_requirements(material_code, period_date)

            # 2. Get Scheduled Receipts (Open POs due this period)
            scheduled_receipts = self._get_scheduled_receipts(
                material_code, period_date
            )

            # 3. Calculate Available Inventory at start of period
            on_hand_start = on_hand

            # 4. Calculate projected on-hand at end of period
            on_hand_end = on_hand_start + scheduled_receipts - gross_req

            # 5. Check if we need to order (consider safety stock)
            net_requirement = 0
            planned_order_qty = 0
            planned_order_date = None

            if on_hand_end < safety_stock:
                # Calculate net requirement
                net_requirement = (
                    gross_req + safety_stock - (on_hand_start + scheduled_receipts)
                )

                if net_requirement > 0:
                    # Apply lot sizing rules
                    planned_order_qty = self._apply_lot_sizing(
                        material, net_requirement
                    )

                    # Calculate order date (consider lead time)
                    planned_order_date = period_date - timedelta(days=lead_time_days)

                    # Adjust on-hand end with planned order
                    on_hand_end += planned_order_qty

                    # Create requisition data
                    requisitions.append(
                        {
                            "material_code": material_code,
                            "quantity": planned_order_qty,
                            "required_date": period_date.isoformat(),
                            "order_date": planned_order_date.isoformat(),
                            "reason": f"MRP Net Requirement: {net_requirement:.2f}",
                        }
                    )

            # Store period calculation
            mrp_record = {
                "run_id": run_id,
                "material_code": material_code,
                "period_date": period_date.isoformat(),
                "gross_requirement": gross_req,
                "scheduled_receipts": scheduled_receipts,
                "on_hand_start": on_hand_start,
                "on_hand_end": on_hand_end,
                "net_requirement": net_requirement,
                "planned_order_qty": planned_order_qty,
                "planned_order_date": (
                    planned_order_date.isoformat() if planned_order_date else None
                ),
                "vendor_name": material.get("preferred_vendor"),
            }

            mrp_table.append(mrp_record)

            # Save to database
            self._save_mrp_calculation(mrp_record)

            # Update on_hand for next period
            on_hand = on_hand_end

        return {
            "material_code": material_code,
            "mrp_table": mrp_table,
            "requisitions": requisitions,
        }

    def _get_gross_requirements(self, material_code, period_date):
        """
        Get total demand for a material in a specific period
        Sources: Forecast + Sales Orders + Safety Stock needs
        """
        # Get forecast for this period
        forecast_query = """
            SELECT SUM(forecast_qty) as total_forecast
            FROM forecast_demand
            WHERE material_code = ?
            AND forecast_date = ?
        """

        forecast_result = self.db.execute_query(
            forecast_query, (material_code, period_date.isoformat()), fetchone=True
        )

        forecast_qty = (
            forecast_result["total_forecast"]
            if forecast_result and forecast_result["total_forecast"]
            else 0
        )

        # TODO: Add sales orders demand when that module exists
        # sales_demand = self._get_sales_orders_demand(material_code, period_date)

        return forecast_qty

    def _get_scheduled_receipts(self, material_code, period_date):
        """
        Get expected supply arriving in this period
        Includes BOTH:
        - Open POs (confirmed orders with vendors)
        - Existing PRs (pending requisitions not yet converted to POs)
        """

        # 1. Get Open POs arriving in this period
        po_query = """
            SELECT SUM(requested_qty) as incoming_qty
            FROM open_orders
            WHERE material_code = ?
            AND status = 'Open'
            AND (requested_del_date BETWEEN ? AND ?)
        """

        # 2. Get Existing PRs (requisitions) arriving in this period
        pr_query = """
            SELECT SUM(requested_qty) as incoming_qty
            FROM requisitions
            WHERE material_code = ?
            AND status = 'Open'
            AND pr_status != 'Converted'
            AND (requested_del_date BETWEEN ? AND ?)
        """

        # Period is one week
        period_start = period_date.isoformat()
        period_end = (period_date + timedelta(days=6)).isoformat()

        # Get PO receipts
        po_result = self.db.execute_query(
            po_query, (material_code, period_start, period_end), fetchone=True
        )
        po_qty = (
            po_result["incoming_qty"] if po_result and po_result["incoming_qty"] else 0
        )

        # Get PR receipts (existing requisitions)
        pr_result = self.db.execute_query(
            pr_query, (material_code, period_start, period_end), fetchone=True
        )
        pr_qty = (
            pr_result["incoming_qty"] if pr_result and pr_result["incoming_qty"] else 0
        )

        # Total scheduled receipts = POs + existing PRs
        total_receipts = po_qty + pr_qty

        # Log for transparency
        if pr_qty > 0:
            self.log(
                f"ℹ️ INFO: Material {material_code}, Period {period_date.isoformat()}: "
                f"PO Receipts={po_qty}, Existing PR Receipts={pr_qty}, Total={total_receipts}"
            )

        return total_receipts

    def _apply_lot_sizing(self, material, net_requirement):
        """
        Apply lot sizing rules to determine order quantity
        """
        lot_size_rule = material.get("lot_size_rule", "LOT_FOR_LOT")
        min_order_qty = material.get("min_order_qty", 1)

        if lot_size_rule == "LOT_FOR_LOT":
            # Order exactly what's needed
            order_qty = net_requirement

        elif lot_size_rule == "FIXED_LOT":
            # Order in fixed lot sizes
            fixed_lot = material.get("fixed_lot_size", 100)
            # Round up to nearest lot size
            order_qty = math.ceil(net_requirement / fixed_lot) * fixed_lot

        elif lot_size_rule == "EOQ":
            # Economic Order Quantity (simplified)
            # EOQ = sqrt((2 * D * S) / H)
            # For now, use a simple multiple of net requirement
            order_qty = math.ceil(net_requirement / 100) * 100

        else:
            order_qty = net_requirement

        # Ensure minimum order quantity
        return max(order_qty, min_order_qty)

    def _save_mrp_calculation(self, mrp_record):
        """Save MRP calculation to database"""
        query = """
            INSERT INTO mrp_calculations (
                run_id, material_code, period_date, gross_requirement,
                scheduled_receipts, on_hand_start, on_hand_end,
                net_requirement, planned_order_qty, planned_order_date, vendor_name
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """

        self.db.execute_query(
            query,
            (
                mrp_record["run_id"],
                mrp_record["material_code"],
                mrp_record["period_date"],
                mrp_record["gross_requirement"],
                mrp_record["scheduled_receipts"],
                mrp_record["on_hand_start"],
                mrp_record["on_hand_end"],
                mrp_record["net_requirement"],
                mrp_record["planned_order_qty"],
                mrp_record["planned_order_date"],
                mrp_record["vendor_name"],
            ),
            commit=True,
        )

    def _create_requisition_from_mrp(self, req_data, run_id):
        """Create purchase requisition from MRP result"""
        # Get material details
        material = self.db.execute_query(
            "SELECT * FROM materials WHERE material_code = ?",
            (req_data["material_code"],),
            fetchone=True,
        )

        # Get vendor
        vendor = self.db.execute_query(
            "SELECT vendor_name FROM vendors WHERE display_name = ?",
            (material.get("preferred_vendor", ""),),
            fetchone=True,
        )

        vendor_name = vendor["vendor_name"] if vendor else None

        # Generate requisition number
        req_number = self._generate_req_number()

        query = """
            INSERT INTO requisitions (
                req_number, item, vendor_name, material_code, short_text,
                requested_qty, unit, requested_del_date, status, pr_status,
                source, mrp_run_id, priority, approval_status, notes
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """

        self.db.execute_query(
            query,
            (
                req_number,
                "10",  # Line item
                vendor_name,
                req_data["material_code"],
                material.get("description", ""),
                req_data["quantity"],
                material.get("unit", "EA"),
                req_data["required_date"],
                "Open",
                "Open",
                "MRP",
                run_id,
                "NORMAL",
                "PENDING",
                req_data["reason"],
            ),
            commit=True,
        )

        self.log(
            f"ℹ️ INFO: Created requisition {req_number} for {req_data['material_code']}"
        )

    def _generate_req_number(self):
        """Generate unique requisition number"""
        # Get last req number
        result = self.db.execute_query(
            "SELECT MAX(CAST(req_number AS INTEGER)) as last_num FROM requisitions WHERE req_number GLOB '[0-9]*'",
            fetchone=True,
        )

        last_num = result["last_num"] if result and result["last_num"] else 4500000000
        return str(last_num + 1)

    def get_mrp_exceptions(self, run_id):
        """
        Get MRP exceptions (materials with issues)
        - Negative projected inventory
        - Past due planned orders
        - No preferred vendor
        """
        query = """
            SELECT m.material_code, m.description, mc.period_date,
                   mc.on_hand_end, mc.net_requirement, mc.vendor_name,
                   'STOCKOUT' as exception_type
            FROM mrp_calculations mc
            JOIN materials m ON mc.material_code = m.material_code
            WHERE mc.run_id = ?
            AND mc.on_hand_end < 0
            
            UNION ALL
            
            SELECT m.material_code, m.description, mc.period_date,
                   mc.on_hand_end, mc.net_requirement, mc.vendor_name,
                   'NO_VENDOR' as exception_type
            FROM mrp_calculations mc
            JOIN materials m ON mc.material_code = m.material_code
            WHERE mc.run_id = ?
            AND mc.net_requirement > 0
            AND (mc.vendor_name IS NULL OR mc.vendor_name = '')
            
            ORDER BY period_date, material_code
        """

        return self.db.execute_query(query, (run_id, run_id), fetchall=True)


# ==============================================================================
# MRP Planning Workbench Window
# ==============================================================================


class MRPPlanningWindow(tk.Toplevel):
    """Main window for MRP planning and requisition management"""

    def __init__(self, parent, log_callback, data_manager):
        super().__init__(parent)
        self.title("MRP Planning Workbench")
        self.geometry("1400x900")
        self.log = log_callback
        self.dm = data_manager
        self.fm = ForecastDataManager(self.dm.db)
        self.mrp_engine = MRPEngine(self.dm.db, self.log)

        # Create main notebook with tabs
        notebook = ttk.Notebook(self)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Tab 1: Run MRP
        run_frame = ttk.Frame(notebook)
        notebook.add(run_frame, text=" Run MRP")
        self.setup_run_mrp_tab(run_frame)

        # Tab 2: View Results
        results_frame = ttk.Frame(notebook)
        notebook.add(results_frame, text=" MRP Results")
        self.setup_results_tab(results_frame)

        # Tab 3: Manage Requisitions
        req_frame = ttk.Frame(notebook)
        notebook.add(req_frame, text=" Requisitions")
        self.setup_requisitions_tab(req_frame)

        # Tab 4: Master Data
        master_frame = ttk.Frame(notebook)
        notebook.add(master_frame, text=" Master Data")
        self.setup_master_data_tab(master_frame)

    def setup_run_mrp_tab(self, parent):
        """Setup MRP run configuration and execution"""
        # Configuration Frame
        config_frame = ttk.LabelFrame(parent, text="MRP Run Configuration", padding=10)
        config_frame.pack(fill=tk.X, padx=10, pady=10)

        # Planning Horizon
        horizon_frame = ttk.Frame(config_frame)
        horizon_frame.pack(fill=tk.X, pady=5)
        ttk.Label(horizon_frame, text="Planning Horizon:").pack(side=tk.LEFT, padx=5)
        self.horizon_var = tk.StringVar(value="13")
        ttk.Entry(horizon_frame, textvariable=self.horizon_var, width=10).pack(
            side=tk.LEFT
        )
        ttk.Label(horizon_frame, text="weeks").pack(side=tk.LEFT, padx=5)

        # Material Selection
        material_frame = ttk.Frame(config_frame)
        material_frame.pack(fill=tk.X, pady=5)
        ttk.Label(material_frame, text="Materials:").pack(side=tk.LEFT, padx=5)
        self.material_scope_var = tk.StringVar(value="all")
        ttk.Radiobutton(
            material_frame,
            text="All Materials",
            variable=self.material_scope_var,
            value="all",
        ).pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(
            material_frame,
            text="Selected Materials",
            variable=self.material_scope_var,
            value="selected",
        ).pack(side=tk.LEFT)

        # Options
        options_frame = ttk.LabelFrame(config_frame, text="Options", padding=10)
        options_frame.pack(fill=tk.X, pady=10)

        self.create_pr_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            options_frame,
            text="Auto-create Purchase Requisitions",
            variable=self.create_pr_var,
        ).pack(anchor="w", pady=2)

        self.consider_safety_stock_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            options_frame,
            text="Consider Safety Stock",
            variable=self.consider_safety_stock_var,
        ).pack(anchor="w", pady=2)

        self.include_forecast_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            options_frame,
            text="Include Forecast Demand",
            variable=self.include_forecast_var,
        ).pack(anchor="w", pady=2)

        # Run Button
        run_btn_frame = ttk.Frame(config_frame)
        run_btn_frame.pack(fill=tk.X, pady=10)

        ttk.Button(
            run_btn_frame,
            text=" Run MRP Calculation",
            command=self.run_mrp,
            style="Accent.TButton",
        ).pack(side=tk.LEFT, padx=5)
        ttk.Button(
            run_btn_frame, text="View Last Run", command=self.view_last_run
        ).pack(side=tk.LEFT, padx=5)

        # Status Frame
        status_frame = ttk.LabelFrame(parent, text="MRP Run Status", padding=10)
        status_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.mrp_status_text = scrolledtext.ScrolledText(
            status_frame, wrap=tk.WORD, height=15
        )
        self.mrp_status_text.pack(fill=tk.BOTH, expand=True)

        # Recent Runs
        recent_frame = ttk.LabelFrame(parent, text="Recent MRP Runs", padding=10)
        recent_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        cols = ("Run ID", "Date", "Horizon", "Status", "PRs Created")
        self.runs_tree = ttk.Treeview(
            recent_frame, columns=cols, show="headings", height=6
        )
        for col in cols:
            self.runs_tree.heading(col, text=col)
            self.runs_tree.column(col, width=120)

        self.runs_tree.pack(fill=tk.BOTH, expand=True)
        self.runs_tree.bind("<Double-1>", self.on_run_double_click)

        self.load_recent_runs()

    def setup_results_tab(self, parent):
        """Setup MRP results visualization"""
        # Top controls
        control_frame = ttk.Frame(parent, padding=10)
        control_frame.pack(fill=tk.X)

        ttk.Label(control_frame, text="Select MRP Run:").pack(side=tk.LEFT, padx=5)
        self.run_selector_var = tk.StringVar()
        self.run_selector = ttk.Combobox(
            control_frame,
            textvariable=self.run_selector_var,
            width=30,
            state="readonly",
        )
        self.run_selector.pack(side=tk.LEFT, padx=5)
        self.run_selector.bind("<<ComboboxSelected>>", self.load_mrp_results)

        ttk.Button(control_frame, text="Refresh", command=self.refresh_run_list).pack(
            side=tk.LEFT, padx=5
        )
        ttk.Button(
            control_frame, text="Export to Excel", command=self.export_results
        ).pack(side=tk.LEFT, padx=5)

        # Create PanedWindow for split view
        paned = ttk.PanedWindow(parent, orient=tk.VERTICAL)
        paned.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Top: Material Summary
        summary_frame = ttk.LabelFrame(paned, text="Material Summary", padding=5)

        cols = (
            "Material",
            "Description",
            "Total Demand",
            "On Hand",
            "Net Req",
            "PRs Generated",
        )
        self.summary_tree = ttk.Treeview(
            summary_frame, columns=cols, show="headings", height=10
        )
        for col in cols:
            self.summary_tree.heading(col, text=col)

        self.summary_tree.column("Material", width=100)
        self.summary_tree.column("Description", width=200)
        self.summary_tree.column("Total Demand", width=100)
        self.summary_tree.column("On Hand", width=80)
        self.summary_tree.column("Net Req", width=80)
        self.summary_tree.column("PRs Generated", width=100)

        summary_scroll = ttk.Scrollbar(
            summary_frame, orient=tk.VERTICAL, command=self.summary_tree.yview
        )
        self.summary_tree.configure(yscrollcommand=summary_scroll.set)

        self.summary_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        summary_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        self.summary_tree.bind("<Double-1>", self.on_material_double_click)

        paned.add(summary_frame, weight=1)

        # Bottom: Period Detail
        detail_frame = ttk.LabelFrame(paned, text="Period-by-Period Detail", padding=5)

        cols = (
            "Period",
            "Date",
            "Gross Req",
            "Scheduled",
            "On Hand Start",
            "On Hand End",
            "Net Req",
            "Planned Order",
        )
        self.detail_tree = ttk.Treeview(
            detail_frame, columns=cols, show="headings", height=10
        )
        for col in cols:
            self.detail_tree.heading(col, text=col)
            self.detail_tree.column(col, width=100)

        detail_scroll = ttk.Scrollbar(
            detail_frame, orient=tk.VERTICAL, command=self.detail_tree.yview
        )
        self.detail_tree.configure(yscrollcommand=detail_scroll.set)

        self.detail_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        detail_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        paned.add(detail_frame, weight=1)

        self.refresh_run_list()

    def setup_requisitions_tab(self, parent):
        """Setup requisition management interface"""
        # Filter controls
        filter_frame = ttk.Frame(parent, padding=10)
        filter_frame.pack(fill=tk.X)

        ttk.Label(filter_frame, text="Filter by Status:").pack(side=tk.LEFT, padx=5)
        self.req_status_filter = tk.StringVar(value="PENDING")
        for status in ["ALL", "PENDING", "APPROVED", "REJECTED", "CONVERTED"]:
            ttk.Radiobutton(
                filter_frame,
                text=status,
                variable=self.req_status_filter,
                value=status,
                command=self.filter_requisitions,
            ).pack(side=tk.LEFT, padx=2)

        # Requisitions tree
        tree_frame = ttk.Frame(parent, padding=10)
        tree_frame.pack(fill=tk.BOTH, expand=True)

        cols = (
            "Req#",
            "Material",
            "Description",
            "Qty",
            "Required Date",
            "Vendor",
            "Source",
            "Status",
            "Priority",
        )
        self.req_tree = ttk.Treeview(tree_frame, columns=cols, show="headings")
        for col in cols:
            self.req_tree.heading(col, text=col)

        self.req_tree.column("Req#", width=100)
        self.req_tree.column("Material", width=100)
        self.req_tree.column("Description", width=200)
        self.req_tree.column("Qty", width=80)
        self.req_tree.column("Required Date", width=100)
        self.req_tree.column("Vendor", width=150)
        self.req_tree.column("Source", width=80)
        self.req_tree.column("Status", width=80)
        self.req_tree.column("Priority", width=80)

        req_scroll = ttk.Scrollbar(
            tree_frame, orient=tk.VERTICAL, command=self.req_tree.yview
        )
        self.req_tree.configure(yscrollcommand=req_scroll.set)

        self.req_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        req_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # Action buttons
        action_frame = ttk.Frame(parent, padding=10)
        action_frame.pack(fill=tk.X)

        ttk.Button(
            action_frame, text=" Approve Selected", command=self.approve_requisitions
        ).pack(side=tk.LEFT, padx=5)
        ttk.Button(
            action_frame, text=" Reject Selected", command=self.reject_requisitions
        ).pack(side=tk.LEFT, padx=5)
        ttk.Button(
            action_frame, text=" Convert to PO", command=self.convert_to_po
        ).pack(side=tk.LEFT, padx=5)
        ttk.Button(
            action_frame, text=" View Details", command=self.view_req_details
        ).pack(side=tk.LEFT, padx=5)

        self.load_requisitions()

    def setup_master_data_tab(self, parent):
        """Setup material master data management with pricing"""
        # Create PanedWindow
        paned = ttk.PanedWindow(parent, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Left: Material list
        list_frame = ttk.LabelFrame(paned, text="Materials", padding=10)
        paned.add(list_frame, weight=1)

        # Search
        search_frame = ttk.Frame(list_frame)
        search_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(search_frame, text="Search:").pack(side=tk.LEFT, padx=(0, 5))
        self.material_search_var = tk.StringVar()
        self.material_search_var.trace("w", self.filter_materials)
        ttk.Entry(search_frame, textvariable=self.material_search_var).pack(
            side=tk.LEFT, fill=tk.X, expand=True
        )

        # Material listbox
        self.material_listbox = Listbox(list_frame, exportselection=False)
        self.material_listbox.pack(fill=tk.BOTH, expand=True)
        self.material_listbox.bind("<<ListboxSelect>>", self.on_material_select)

        # Right: Material details form
        form_frame = ttk.LabelFrame(paned, text="Material Details", padding=10)
        paned.add(form_frame, weight=2)

        # Create form fields
        self.material_fields = {}
        fields = [
            ("material_code", "Material Code*"),
            ("description", "Description"),
            ("unit", "Unit of Measure"),
            ("net_price", "Net Price"),
            ("price_per_unit", "Price Per Unit (e.g., 100)"),
            ("currency", "Currency"),
            ("lead_time_days", "Lead Time (days)"),
            ("safety_stock", "Safety Stock"),
            ("min_order_qty", "Min Order Qty"),
            ("lot_size_rule", "Lot Size Rule"),
            ("fixed_lot_size", "Fixed Lot Size"),
            ("preferred_vendor", "Preferred Vendor"),
            ("abc_class", "ABC Class"),
        ]

        for idx, (field_name, label_text) in enumerate(fields):
            ttk.Label(form_frame, text=f"{label_text}:").grid(
                row=idx, column=0, sticky="w", padx=5, pady=5
            )
            self.material_fields[field_name] = tk.StringVar()

            if field_name == "lot_size_rule":
                combo = ttk.Combobox(
                    form_frame,
                    textvariable=self.material_fields[field_name],
                    values=["LOT_FOR_LOT", "FIXED_LOT", "EOQ"],
                )
                combo.grid(row=idx, column=1, sticky="ew", padx=5, pady=5)
            elif field_name == "abc_class":
                combo = ttk.Combobox(
                    form_frame,
                    textvariable=self.material_fields[field_name],
                    values=["A", "B", "C"],
                )
                combo.grid(row=idx, column=1, sticky="ew", padx=5, pady=5)
            else:
                entry = ttk.Entry(
                    form_frame, textvariable=self.material_fields[field_name]
                )
                entry.grid(row=idx, column=1, sticky="ew", padx=5, pady=5)

        form_frame.columnconfigure(1, weight=1)

        # NEW: Add pricing info label
        pricing_info = ttk.Label(
            form_frame,
            text=" Tip: Prices are automatically updated when uploading Open Order Book",
            font=("Helvetica", 8, "italic"),
            foreground="gray",
        )
        pricing_info.grid(
            row=len(fields), column=0, columnspan=2, sticky="w", padx=5, pady=(10, 5)
        )

        # Buttons
        button_frame = ttk.Frame(form_frame)
        button_frame.grid(row=len(fields) + 1, column=0, columnspan=2, pady=20)

        ttk.Button(button_frame, text="Save Material", command=self.save_material).pack(
            side=tk.LEFT, padx=5
        )
        ttk.Button(
            button_frame, text="New Material", command=self.clear_material_form
        ).pack(side=tk.LEFT, padx=5)
        ttk.Button(
            button_frame, text="Delete Material", command=self.delete_material
        ).pack(side=tk.LEFT, padx=5)

        self.load_materials()

    # === MRP Run Functions ===

    def run_mrp(self):
        """Execute MRP calculation"""
        try:
            horizon = int(self.horizon_var.get())
            create_prs = self.create_pr_var.get()

            self.mrp_status_text.insert(tk.END, f"\n{'='*60}\n")
            self.mrp_status_text.insert(
                tk.END, f"Starting MRP Run at {datetime.now().strftime('%H:%M:%S')}\n"
            )
            self.mrp_status_text.insert(tk.END, f"{'='*60}\n")

            def run_thread():
                try:
                    run_id = self.mrp_engine.run_mrp(
                        horizon_weeks=horizon,
                        materials=None,
                        create_requisitions=create_prs,
                    )

                    self.after(
                        0,
                        lambda run_id=run_id: self.mrp_status_text.insert(
                            tk.END, f"\n MRP Run {run_id} completed successfully!\n"
                        ),
                    )
                    self.after(0, self.load_recent_runs)
                    self.after(
                        0,
                        lambda run_id=run_id: messagebox.showinfo(
                            "MRP Complete",
                            f"MRP run {run_id} completed successfully!",
                            parent=self,
                        ),
                    )

                except Exception as e:
                    # FIXED: Capture exception message in lambda using default argument
                    error_msg = str(e)
                    self.after(
                        0,
                        lambda error_msg=error_msg: self.mrp_status_text.insert(
                            tk.END, f"\n ERROR: {error_msg}\n"
                        ),
                    )
                    self.after(
                        0,
                        lambda error_msg=error_msg: messagebox.showerror(
                            "MRP Error", f"MRP run failed:\n{error_msg}", parent=self
                        ),
                    )

            threading.Thread(target=run_thread, daemon=True).start()

        except ValueError:
            messagebox.showerror(
                "Invalid Input",
                "Please enter a valid number for planning horizon",
                parent=self,
            )

    def load_recent_runs(self):
        """Load recent MRP runs"""
        for item in self.runs_tree.get_children():
            self.runs_tree.delete(item)

        query = """
            SELECT run_id, run_date, horizon_weeks, status,
                   (SELECT COUNT(*) FROM requisitions WHERE mrp_run_id = run_id) as pr_count
            FROM mrp_runs
            ORDER BY run_date DESC
            LIMIT 10
        """

        runs = self.dm.db.execute_query(query, fetchall=True)

        for run in runs:
            run_date = datetime.fromisoformat(run["run_date"]).strftime(
                "%Y-%m-%d %H:%M"
            )
            self.runs_tree.insert(
                "",
                "end",
                values=(
                    run["run_id"],
                    run_date,
                    f"{run['horizon_weeks']} weeks",
                    run["status"],
                    run["pr_count"],
                ),
            )

    def refresh_run_list(self):
        """Refresh MRP run dropdown"""
        query = "SELECT run_id, run_date FROM mrp_runs ORDER BY run_date DESC LIMIT 20"
        runs = self.dm.db.execute_query(query, fetchall=True)

        run_options = [
            f"Run {r['run_id']} - {datetime.fromisoformat(r['run_date']).strftime('%Y-%m-%d %H:%M')}"
            for r in runs
        ]

        self.run_selector["values"] = run_options
        if run_options:
            self.run_selector.current(0)
            self.load_mrp_results(None)

    def load_mrp_results(self, event):
        """Load results for selected MRP run"""
        selection = self.run_selector_var.get()
        if not selection:
            return

        # Extract run_id from selection
        run_id = int(selection.split()[1])

        # Clear trees
        for item in self.summary_tree.get_children():
            self.summary_tree.delete(item)

        # Load summary
        query = """
            SELECT mc.material_code, m.description,
                   SUM(mc.gross_requirement) as total_demand,
                   MAX(mc.on_hand_start) as on_hand,
                   SUM(mc.net_requirement) as net_req,
                   COUNT(DISTINCT r.req_number) as pr_count
            FROM mrp_calculations mc
            LEFT JOIN materials m ON mc.material_code = m.material_code
            LEFT JOIN requisitions r ON r.material_code = mc.material_code 
                AND r.mrp_run_id = mc.run_id
            WHERE mc.run_id = ?
            GROUP BY mc.material_code, m.description
        """

        results = self.dm.db.execute_query(query, (run_id,), fetchall=True)

        for row in results:
            self.summary_tree.insert(
                "",
                "end",
                values=(
                    row["material_code"],
                    row["description"],
                    f"{row['total_demand']:.2f}",
                    f"{row['on_hand']:.2f}",
                    f"{row['net_req']:.2f}",
                    row["pr_count"],
                ),
            )

    def on_material_double_click(self, event):
        """Show period detail for selected material"""
        selection = self.summary_tree.selection()
        if not selection:
            return

        material_code = self.summary_tree.item(selection[0], "values")[0]
        run_selection = self.run_selector_var.get()
        run_id = int(run_selection.split()[1])

        # Clear detail tree
        for item in self.detail_tree.get_children():
            self.detail_tree.delete(item)

        # Load period details
        query = """
            SELECT period_date, gross_requirement, scheduled_receipts,
                   on_hand_start, on_hand_end, net_requirement, planned_order_qty
            FROM mrp_calculations
            WHERE run_id = ? AND material_code = ?
            ORDER BY period_date
        """

        periods = self.dm.db.execute_query(
            query, (run_id, material_code), fetchall=True
        )

        for i, period in enumerate(periods, 1):
            date_obj = datetime.fromisoformat(period["period_date"])
            self.detail_tree.insert(
                "",
                "end",
                values=(
                    f"Week {i}",
                    date_obj.strftime("%Y-%m-%d"),
                    f"{period['gross_requirement']:.2f}",
                    f"{period['scheduled_receipts']:.2f}",
                    f"{period['on_hand_start']:.2f}",
                    f"{period['on_hand_end']:.2f}",
                    f"{period['net_requirement']:.2f}",
                    (
                        f"{period['planned_order_qty']:.2f}"
                        if period["planned_order_qty"] > 0
                        else ""
                    ),
                ),
            )

    # === Requisition Management ===

    def load_requisitions(self):
        """Load requisitions based on filter"""
        for item in self.req_tree.get_children():
            self.req_tree.delete(item)

        status_filter = self.req_status_filter.get()

        query = """
            SELECT r.req_number, r.material_code, r.short_text, r.requested_qty,
                   r.requested_del_date, v.display_name, r.source, 
                   r.approval_status, r.priority
            FROM requisitions r
            LEFT JOIN vendors v ON r.vendor_name = v.vendor_name
            WHERE r.status = 'Open'
        """

        if status_filter != "ALL":
            query += f" AND r.approval_status = '{status_filter}'"

        query += " ORDER BY r.requested_del_date, r.req_number"

        reqs = self.dm.db.execute_query(query, fetchall=True)

        for req in reqs:
            self.req_tree.insert(
                "",
                "end",
                values=(
                    req["req_number"],
                    req["material_code"],
                    req["short_text"],
                    req["requested_qty"],
                    req["requested_del_date"],
                    req["display_name"] or "No Vendor",
                    req["source"],
                    req["approval_status"],
                    req["priority"],
                ),
            )

    def filter_requisitions(self):
        """Filter requisitions by status"""
        self.load_requisitions()

    def approve_requisitions(self):
        """Approve selected requisitions"""
        selected = self.req_tree.selection()
        if not selected:
            messagebox.showwarning(
                "No Selection", "Please select requisitions to approve", parent=self
            )
            return

        req_numbers = [self.req_tree.item(item, "values")[0] for item in selected]

        if messagebox.askyesno(
            "Confirm Approval",
            f"Approve {len(req_numbers)} requisition(s)?",
            parent=self,
        ):
            placeholders = ",".join("?" * len(req_numbers))
            query = f"""
                UPDATE requisitions 
                SET approval_status = 'APPROVED',
                    approved_date = ?,
                    approved_by = 'User'
                WHERE req_number IN ({placeholders})
            """

            self.dm.db.execute_query(
                query, (datetime.now().isoformat(), *req_numbers), commit=True
            )

            self.log(f"ℹ️ INFO: Approved {len(req_numbers)} requisitions")
            self.load_requisitions()

    def reject_requisitions(self):
        """Reject selected requisitions"""
        selected = self.req_tree.selection()
        if not selected:
            messagebox.showwarning(
                "No Selection", "Please select requisitions to reject", parent=self
            )
            return

        req_numbers = [self.req_tree.item(item, "values")[0] for item in selected]

        if messagebox.askyesno(
            "Confirm Rejection",
            f"Reject {len(req_numbers)} requisition(s)?",
            parent=self,
        ):
            placeholders = ",".join("?" * len(req_numbers))
            query = f"UPDATE requisitions SET approval_status = 'REJECTED' WHERE req_number IN ({placeholders})"

            self.dm.db.execute_query(query, tuple(req_numbers), commit=True)

            self.log(f"ℹ️ INFO: Rejected {len(req_numbers)} requisitions")
            self.load_requisitions()

    def convert_to_po(self):
        """Convert approved requisitions to PO"""
        selected = self.req_tree.selection()
        if not selected:
            messagebox.showwarning(
                "No Selection", "Please select requisitions to convert", parent=self
            )
            return

        # Check if all selected are approved
        req_numbers = []
        for item in selected:
            values = self.req_tree.item(item, "values")
            if values[7] != "APPROVED":
                messagebox.showerror("❌ Error",
                    "Only APPROVED requisitions can be converted to PO",
                    parent=self,
                )
                return
            req_numbers.append(values[0])

        if messagebox.askyesno(
            "Convert to PO",
            f"Convert {len(req_numbers)} requisition(s) to Purchase Order?",
            parent=self,
        ):
            # TODO: Implement actual PO creation
            # For now, just mark as converted
            self.fm.convert_requisition_to_po(req_numbers)

            self.log(f"ℹ️ INFO: Converted {len(req_numbers)} requisitions to PO")
            messagebox.showinfo("✅ Success",
                f"Converted {len(req_numbers)} requisition(s) to PO",
                parent=self,
            )
            self.load_requisitions()

    # === Material Master Functions ===

    def load_materials(self):
        """Load material list"""
        self.material_listbox.delete(0, tk.END)
        materials = self.dm.db.execute_query(
            "SELECT material_code, description FROM materials ORDER BY material_code",
            fetchall=True,
        )

        for mat in materials:
            self.material_listbox.insert(
                tk.END, f"{mat['material_code']} - {mat['description']}"
            )

    def filter_materials(self, *args):
        """Filter material list"""
        search_term = self.material_search_var.get().lower()
        self.material_listbox.delete(0, tk.END)

        query = """
            SELECT material_code, description FROM materials 
            WHERE LOWER(material_code) LIKE ? OR LOWER(description) LIKE ?
            ORDER BY material_code
        """

        materials = self.dm.db.execute_query(
            query, (f"%{search_term}%", f"%{search_term}%"), fetchall=True
        )

        for mat in materials:
            self.material_listbox.insert(
                tk.END, f"{mat['material_code']} - {mat['description']}"
            )

    def on_material_select(self, event):
        """Load selected material details with pricing"""
        selection = self.material_listbox.curselection()
        if not selection:
            return

        material_text = self.material_listbox.get(selection[0])
        material_code = material_text.split(" - ")[0]

        material = self.dm.db.execute_query(
            "SELECT * FROM materials WHERE material_code = ?",
            (material_code,),
            fetchone=True,
        )

        if material:
            for field, var in self.material_fields.items():
                value = material.get(field, "")
                # Format numeric/price fields
                if field in [
                    "net_price",
                    "safety_stock",
                    "min_order_qty",
                    "fixed_lot_size",
                ]:
                    try:
                        value = f"{float(value or 0):.2f}"
                    except Exception:
                        value = str(value or "")
                elif field in ["price_per_unit", "lead_time_days"]:
                    try:
                        value = str(int(value)) if value not in (None, "") else ""
                    except Exception:
                        value = str(value or "")
                else:
                    value = value or ""
                var.set(value)

    def save_material(self):
        """Save material master data with pricing and currency"""
        material_code = self.material_fields["material_code"].get().strip()
        if not material_code:
            messagebox.showerror("❌ Error", "Material code is required", parent=self)
            return

        # Check if material exists
        existing = self.dm.db.execute_query(
            "SELECT 1 FROM materials WHERE material_code = ?",
            (material_code,),
            fetchone=True,
        )

        values = {field: var.get() for field, var in self.material_fields.items()}

        if existing:
            # Update
            query = """
                UPDATE materials SET description=?, unit=?, net_price=?, price_per_unit=?, 
                currency=?, lead_time_days=?, safety_stock=?, min_order_qty=?, lot_size_rule=?, 
                fixed_lot_size=?, preferred_vendor=?, abc_class=?, last_updated=?
                WHERE material_code=?
            """

            self.dm.db.execute_query(
                query,
                (
                    values["description"],
                    values["unit"],
                    float(values["net_price"] or 0),
                    int(values["price_per_unit"] or 1),
                    values["currency"] or "EUR",  # NEW
                    int(values["lead_time_days"] or 0),
                    float(values["safety_stock"] or 0),
                    float(values["min_order_qty"] or 1),
                    values["lot_size_rule"],
                    float(values["fixed_lot_size"] or 0),
                    values["preferred_vendor"],
                    values["abc_class"],
                    datetime.now().isoformat(),
                    material_code,
                ),
                commit=True,
            )

            self.log(f"ℹ️ INFO: Updated material {material_code}")
            messagebox.showinfo("✅ Success", f"Material {material_code} updated", parent=self
            )
        else:
            # Insert
            query = """
                INSERT INTO materials (material_code, description, unit, net_price, price_per_unit,
                currency, lead_time_days, safety_stock, min_order_qty, lot_size_rule, fixed_lot_size,
                preferred_vendor, abc_class, created_date)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """

            self.dm.db.execute_query(
                query,
                (
                    material_code,
                    values["description"],
                    values["unit"],
                    float(values["net_price"] or 0),
                    int(values["price_per_unit"] or 1),
                    values["currency"] or "EUR",  # NEW
                    int(values["lead_time_days"] or 0),
                    float(values["safety_stock"] or 0),
                    float(values["min_order_qty"] or 1),
                    values["lot_size_rule"],
                    float(values["fixed_lot_size"] or 0),
                    values["preferred_vendor"],
                    values["abc_class"],
                    datetime.now().isoformat(),
                ),
                commit=True,
            )

            self.log(f"ℹ️ INFO: Created material {material_code}")
            messagebox.showinfo("✅ Success", f"Material {material_code} created", parent=self
            )

        self.load_materials()

    def clear_material_form(self):
        """Clear form for new material"""
        for var in self.material_fields.values():
            var.set("")

    def delete_material(self):
        """Delete material"""
        material_code = self.material_fields["material_code"].get().strip()
        if not material_code:
            messagebox.showwarning(
                "No Selection", "Please select a material to delete", parent=self
            )
            return

        if not messagebox.askyesno(
            "Confirm Delete",
            f"Delete material {material_code}? This cannot be undone.",
            parent=self,
        ):
            return

        try:
            self.dm.db.execute_query(
                "DELETE FROM materials WHERE material_code = ?",
                (material_code,),
                commit=True,
            )

            self.log(f"ℹ️ INFO: Deleted material {material_code}")
            self.clear_material_form()
            self.load_materials()
            messagebox.showinfo("✅ Success", f"Material {material_code} deleted", parent=self
            )

        except Exception as e:
            messagebox.showerror("❌ Error", f"Cannot delete material:\n{str(e)}", parent=self
            )

    def view_last_run(self):
        """View results from last MRP run"""
        last_run = self.dm.db.execute_query(
            "SELECT run_id FROM mrp_runs ORDER BY run_date DESC LIMIT 1", fetchone=True
        )

        if last_run:
            # Switch to results tab and load
            self.after(100, lambda: self.refresh_run_list())
        else:
            messagebox.showinfo("No Runs", "No MRP runs found", parent=self)

    def on_run_double_click(self, event):
        """View detailed results for selected run"""
        selection = self.runs_tree.selection()
        if not selection:
            return

        run_id = self.runs_tree.item(selection[0], "values")[0]

        # Show exceptions
        exceptions = self.mrp_engine.get_mrp_exceptions(run_id)

        if exceptions:
            msg = f"MRP Run {run_id} - Exceptions Found:\n\n"
            for exc in exceptions[:10]:  # Show first 10
                msg += f" {exc['material_code']}: {exc['exception_type']}\n"

            if len(exceptions) > 10:
                msg += f"\n... and {len(exceptions) - 10} more"

            messagebox.showwarning("MRP Exceptions", msg, parent=self)
        else:
            messagebox.showinfo(
                "No Exceptions",
                f"MRP Run {run_id} completed with no exceptions!",
                parent=self,
            )

    def view_req_details(self):
        """Show detailed view of selected requisition"""
        selection = self.req_tree.selection()
        if not selection:
            messagebox.showwarning(
                "No Selection", "Please select a requisition", parent=self
            )
            return

        req_number = self.req_tree.item(selection[0], "values")[0]

        # Get full requisition details
        req = self.dm.db.execute_query(
            "SELECT * FROM requisitions WHERE req_number = ?",
            (req_number,),
            fetchone=True,
        )

        if not req:
            return

        # Create detail window
        detail_win = tk.Toplevel(self)
        detail_win.title(f"Requisition Details - {req_number}")
        detail_win.geometry("600x500")

        # Header
        header_frame = ttk.LabelFrame(
            detail_win, text="Requisition Information", padding=10
        )
        header_frame.pack(fill=tk.X, padx=10, pady=10)

        info_text = f"""
Requisition Number: {req['req_number']}
Material: {req['material_code']}
Description: {req['short_text']}
Quantity: {req['requested_qty']} {req.get('unit', 'EA')}
Required Date: {req['requested_del_date']}

Source: {req.get('source', 'MANUAL')}
Priority: {req.get('priority', 'NORMAL')}
Status: {req.get('approval_status', 'PENDING')}

Vendor: {req.get('vendor_name', 'Not assigned')}
"""

        if req.get("mrp_run_id"):
            info_text += f"MRP Run ID: {req['mrp_run_id']}\n"

        if req.get("notes"):
            info_text += f"\nNotes:\n{req['notes']}"

        ttk.Label(header_frame, text=info_text, justify=tk.LEFT).pack(anchor="w")

        # Action buttons
        btn_frame = ttk.Frame(detail_win, padding=10)
        btn_frame.pack(fill=tk.X, padx=10, pady=10)

        ttk.Button(btn_frame, text="Close", command=detail_win.destroy).pack(
            side=tk.RIGHT
        )

    def export_results(self):
        """Export MRP results to Excel"""
        selection = self.run_selector_var.get()
        if not selection:
            messagebox.showwarning(
                "No Selection", "Please select an MRP run to export", parent=self
            )
            return

        run_id = int(selection.split()[1])

        file_path = filedialog.asksaveasfilename(
            title="Save MRP Results",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
            parent=self,
        )

        if not file_path:
            return

        try:
            # Get MRP results
            query = """
                SELECT mc.material_code, m.description, mc.period_date,
                       mc.gross_requirement, mc.scheduled_receipts, mc.on_hand_start,
                       mc.on_hand_end, mc.net_requirement, mc.planned_order_qty,
                       mc.planned_order_date, mc.vendor_name
                FROM mrp_calculations mc
                LEFT JOIN materials m ON mc.material_code = m.material_code
                WHERE mc.run_id = ?
                ORDER BY mc.material_code, mc.period_date
            """

            results = self.dm.db.execute_query(query, (run_id,), fetchall=True)

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = f"MRP Run {run_id}"

            # Headers
            headers = [
                "Material",
                "Description",
                "Period Date",
                "Gross Req",
                "Scheduled",
                "On Hand Start",
                "On Hand End",
                "Net Req",
                "Planned Order",
                "Order Date",
                "Vendor",
            ]
            ws.append(headers)

            # Data
            for row in results:
                ws.append(
                    [
                        row["material_code"],
                        row["description"],
                        row["period_date"],
                        row["gross_requirement"],
                        row["scheduled_receipts"],
                        row["on_hand_start"],
                        row["on_hand_end"],
                        row["net_requirement"],
                        row["planned_order_qty"],
                        row["planned_order_date"],
                        row["vendor_name"],
                    ]
                )

            # Format as table
            tab = Table(displayName="MRPResults", ref=f"A1:K{len(results) + 1}")
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            tab.tableStyleInfo = style
            ws.add_table(tab)

            # Save
            wb.save(file_path)

            self.log(f"✓ SUCCESS: Exported MRP results to {file_path}")
            messagebox.showinfo("✅ Success", f"MRP results exported to:\n{file_path}", parent=self
            )

        except Exception as e:
            self.log(f"✗ ERROR: Failed to export MRP results: {e}")
            messagebox.showerror("❌ Error", f"Failed to export:\n{str(e)}", parent=self)


class OpenOrderBookWindow(tk.Toplevel):
    def __init__(self, parent, log_callback, data_manager):
        super().__init__(parent)
        self.title("Open Order Book - Advanced View")
        self.geometry("1800x800")
        self.log = log_callback
        self.dm = data_manager
        self.all_orders = []
        self.filtered_orders = []

        # Create main container with paned window
        main_paned = ttk.PanedWindow(self, orient=tk.HORIZONTAL)
        main_paned.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Left panel for filters - INCREASED WIDTH
        filter_frame = ttk.LabelFrame(main_paned, text="Filters & Search", padding=10)
        main_paned.add(filter_frame, weight=0)

        # Right panel for data display - expandable
        data_frame = ttk.Frame(main_paned)
        main_paned.add(data_frame, weight=1)

        self.setup_filters(filter_frame)
        self.setup_data_display(data_frame)

        # Set minimum width for filter panel AFTER adding to paned window
        self.after(100, lambda: main_paned.sashpos(0, 300))

        # Load initial data
        self.refresh_data()

    def setup_filters(self, parent):
        """Setup the filter controls"""
        # REMOVED the canvas/scrollbar wrapper - direct frame instead
        filter_content = ttk.Frame(parent)
        filter_content.pack(fill=tk.BOTH, expand=True)

        # Text Search
        search_frame = ttk.LabelFrame(filter_content, text="Text Search", padding=5)
        search_frame.pack(fill=tk.X, pady=2)

        self.search_var = tk.StringVar()
        self.search_var.trace("w", self.apply_filters)
        ttk.Label(search_frame, text="Search in all fields:").pack(anchor="w")
        ttk.Entry(search_frame, textvariable=self.search_var).pack(fill=tk.X, pady=2)

        # PO Number Filter
        po_frame = ttk.LabelFrame(filter_content, text="PO Number", padding=5)
        po_frame.pack(fill=tk.X, pady=2)

        self.po_filter_var = tk.StringVar()
        self.po_filter_var.trace("w", self.apply_filters)
        ttk.Entry(po_frame, textvariable=self.po_filter_var).pack(fill=tk.X)

        # Supplier Filter
        supplier_frame = ttk.LabelFrame(filter_content, text="Suppliers", padding=5)
        supplier_frame.pack(fill=tk.X, pady=2)

        supplier_btn_frame = ttk.Frame(supplier_frame)
        supplier_btn_frame.pack(fill=tk.X, pady=2)
        ttk.Button(
            supplier_btn_frame, text="All", command=self.select_all_suppliers, width=8
        ).pack(side=tk.LEFT, padx=2)
        ttk.Button(
            supplier_btn_frame,
            text="None",
            command=self.deselect_all_suppliers,
            width=8,
        ).pack(side=tk.LEFT)

        # Listbox with scrollbar
        list_container = ttk.Frame(supplier_frame)
        list_container.pack(fill=tk.BOTH, expand=True)

        self.supplier_listbox = Listbox(
            list_container, selectmode=tk.MULTIPLE, height=6, exportselection=False
        )
        supplier_scroll = ttk.Scrollbar(
            list_container, orient="vertical", command=self.supplier_listbox.yview
        )
        self.supplier_listbox.configure(yscrollcommand=supplier_scroll.set)
        self.supplier_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        supplier_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.supplier_listbox.bind("<<ListboxSelect>>", lambda e: self.apply_filters())

        # Date Filters
        date_frame = ttk.LabelFrame(filter_content, text="Date Filters", padding=5)
        date_frame.pack(fill=tk.X, pady=2)

        ttk.Label(date_frame, text="Requested Date Range:").pack(anchor="w")
        ttk.Label(
            date_frame, text="(YYYY-MM-DD)", font=("Helvetica", 8), foreground="gray"
        ).pack(anchor="w")

        req_date_frame = ttk.Frame(date_frame)
        req_date_frame.pack(fill=tk.X, pady=2)

        self.req_date_start = tk.StringVar()
        self.req_date_end = tk.StringVar()
        self.req_date_start.trace("w", self.apply_filters)
        self.req_date_end.trace("w", self.apply_filters)

        ttk.Entry(req_date_frame, textvariable=self.req_date_start, width=12).pack(
            side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 2)
        )
        ttk.Label(req_date_frame, text="to").pack(side=tk.LEFT, padx=2)
        ttk.Entry(req_date_frame, textvariable=self.req_date_end, width=12).pack(
            side=tk.LEFT, fill=tk.X, expand=True, padx=(2, 0)
        )

        # Value Filters
        value_frame = ttk.LabelFrame(filter_content, text="Value Range", padding=5)
        value_frame.pack(fill=tk.X, pady=2)

        ttk.Label(value_frame, text="Total Amount Range:").pack(anchor="w")
        val_range_frame = ttk.Frame(value_frame)
        val_range_frame.pack(fill=tk.X, pady=2)

        self.value_min = tk.StringVar()
        self.value_max = tk.StringVar()
        self.value_min.trace("w", self.apply_filters)
        self.value_max.trace("w", self.apply_filters)

        ttk.Entry(val_range_frame, textvariable=self.value_min, width=12).pack(
            side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 2)
        )
        ttk.Label(val_range_frame, text="to").pack(side=tk.LEFT, padx=2)
        ttk.Entry(val_range_frame, textvariable=self.value_max, width=12).pack(
            side=tk.LEFT, fill=tk.X, expand=True, padx=(2, 0)
        )

        # Status Filters
        status_frame = ttk.LabelFrame(filter_content, text="Order Status", padding=5)
        status_frame.pack(fill=tk.X, pady=2)

        self.status_vars = {}
        status_options = [
            ("Has Confirmation Date", "has_confirmation"),
            ("Has Reschedule Date", "has_reschedule"),
            ("No Confirmation", "no_confirmation"),
            ("No Reschedule", "no_reschedule"),
        ]

        for text, key in status_options:
            self.status_vars[key] = tk.BooleanVar()
            self.status_vars[key].trace("w", self.apply_filters)
            ttk.Checkbutton(
                status_frame, text=text, variable=self.status_vars[key]
            ).pack(anchor="w", pady=1)

        # Action Buttons - SIMPLIFIED
        action_frame = ttk.Frame(filter_content)
        action_frame.pack(fill=tk.X, pady=10)

        ttk.Button(
            action_frame,
            text="❌ Mark Selected as Closed",
            command=self.mark_selected_closed,
        ).pack(fill=tk.X, pady=2)
        ttk.Button(
            action_frame, text="🗑️ Clear All Filters", command=self.clear_filters
        ).pack(fill=tk.X, pady=2)
        ttk.Button(
            action_frame, text="📊 Export Filtered Data", command=self.export_filtered_data
        ).pack(fill=tk.X, pady=2)
        ttk.Button(action_frame, text="🔄 Refresh Data", command=self.refresh_data).pack(
            fill=tk.X, pady=2
        )

    def setup_data_display(self, parent):
        """Setup the data display area"""
        # Level 2: Method body (8 spaces from class start)

        # Toolbar at top
        top_toolbar = ttk.Frame(parent)
        top_toolbar.pack(fill=tk.X, padx=5, pady=5)

        ttk.Button(
            top_toolbar,
            text="📤 Upload Order Book (Excel)",
            command=self.upload_order_book,
        ).pack(side=tk.LEFT, padx=5)

        ttk.Button(top_toolbar, text="🔄 Refresh Data", command=self.refresh_data).pack(
            side=tk.LEFT, padx=5
        )

        ttk.Button(
            top_toolbar,
            text=" Arriving This Week",
            command=self.show_arriving_orders_this_week
        ).pack(side=tk.LEFT, padx=5)

        # ADD THIS NEW SECTION - Level 2: Add checkbox for auto-close control (8 spaces)
        # Create checkbox variable for controlling auto-close behavior
        self.auto_close_var = tk.BooleanVar(
            value=True
        )  # Default to True (current behavior)

        # Level 2: Add checkbox to toolbar (8 spaces)
        ttk.Checkbutton(
            top_toolbar,
            text="Auto-close missing lines",
            variable=self.auto_close_var,
        ).pack(side=tk.LEFT, padx=15)

        # Level 2: Add help icon/label (8 spaces)
        help_label = ttk.Label(top_toolbar, text="", cursor="hand2", foreground="blue")
        help_label.pack(side=tk.LEFT)

        # Level 2: Bind help tooltip (8 spaces)
        def show_help(event):
            # Level 3: Function body (12 spaces from class start, 4 from def)
            messagebox.showinfo(
                "Auto-Close Help",
                "When enabled (): Lines missing from the uploaded file will be marked as 'Closed'\n"
                " Use for FULL order book replacements\n\n"
                "When disabled ( ): Existing lines remain open, only new/updated lines are added\n"
                " Use for INCREMENTAL additions",
                parent=self,
            )

        # Level 2: Bind click event (8 spaces)
        help_label.bind("<Button-1>", show_help)

        # Info bar (existing code continues...)
        info_frame = ttk.Frame(parent)
        info_frame.pack(fill=tk.X, padx=5, pady=2)

        self.info_label = ttk.Label(info_frame, text="", font=("Helvetica", 10))
        self.info_label.pack(side=tk.LEFT)

        # Data table
        table_frame = ttk.LabelFrame(
            parent, text="Order Data (Right-click to copy)", padding=5
        )
        table_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=2)

        # Create Treeview with columns
        columns = (
            "PO",
            "Item",
            "Supplier",
            "Material",
            "Description",
            "Qty",
            "Unit Price",
            "Per",
            "Total Amount",
            "Currency",
            "Req. Date",
            "Conf. Date",
            "Reschedule Date",
            "Comments",
            "Status",
        )

        self.tree = ttk.Treeview(
            table_frame, columns=columns, show="headings", height=20
        )

        # Configure column headings and widths
        column_widths = {
            "PO": 100,
            "Item": 60,
            "Supplier": 150,
            "Material": 120,
            "Description": 200,
            "Qty": 70,
            "Unit Price": 80,
            "Per": 50,
            "Total Amount": 90,
            "Currency": 60,
            "Req. Date": 90,
            "Conf. Date": 90,
            "Reschedule Date": 110,
            "Comments": 150,
            "Status": 100,
        }
        for col in columns:
            self.tree.heading(
                col, text=col, command=lambda c=col: self.sort_by_column(c)
            )
            self.tree.column(col, width=column_widths.get(col, 100), minwidth=50)

        # Scrollbars
        tree_scroll_v = ttk.Scrollbar(
            table_frame, orient="vertical", command=self.tree.yview
        )
        tree_scroll_h = ttk.Scrollbar(
            table_frame, orient="horizontal", command=self.tree.xview
        )
        self.tree.configure(
            yscrollcommand=tree_scroll_v.set, xscrollcommand=tree_scroll_h.set
        )

        self.tree.bind("<Double-1>", self.on_po_double_click)

        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tree_scroll_v.pack(side=tk.RIGHT, fill=tk.Y)
        tree_scroll_h.pack(side=tk.BOTTOM, fill=tk.X)

        # Create context menu for copying
        self.create_context_menu()

        # Summary frame
        summary_frame = ttk.LabelFrame(parent, text="Summary", padding=5)
        summary_frame.pack(fill=tk.X, padx=5, pady=2)

        self.summary_label = ttk.Label(summary_frame, text="", font=("Helvetica", 9))
        self.summary_label.pack()

    def upload_order_book(self):
        """Upload order book Excel file with price updates"""
        # Level 2: Method body (8 spaces from class start)
        file_path = filedialog.askopenfilename(
            title="Select Order Book Excel File",
            filetypes=[("Excel Files", "*.xlsx;*.xls"), ("All Files", "*.*")],
        )

        if not file_path:
            return

        try:
            # Level 3: Get the auto-close setting from checkbox (8 spaces)
            auto_close_missing = self.auto_close_var.get()

            # Level 3: Upload and get results - PASS THE PARAMETER (8 spaces)
            result = self.dm.upload_order_book(
                file_path, auto_close_missing=auto_close_missing
            )

            # Level 3: Handle tuple return (8 spaces)
            if isinstance(result, tuple) and len(result) == 4:
                # Level 4: Inside if block (12 spaces)
                processed_count, closed_count, updated_prices, created_materials = (
                    result
                )

                message = f" Order Book Upload Complete!\n\n"
                message += f" Orders: {processed_count} lines imported"

                # Level 4: Conditional message for closed lines (12 spaces)
                if auto_close_missing:
                    # Level 5: Inside nested if (16 spaces)
                    if closed_count > 0:
                        # Level 6: Display closure info (20 spaces)
                        message += (
                            f"\n Marked {closed_count} missing lines as 'Closed'"
                        )
                    else:
                        # Level 6: No closures (20 spaces)
                        message += f"\n No lines needed to be closed"
                else:
                    # Level 5: Auto-close was disabled (16 spaces)
                    message += f"\n Auto-close disabled - existing lines remain open"

                message += f"\n\n Material Prices Updated:"
                message += f"\n   Updated: {updated_prices} materials"
                message += f"\n   Created: {created_materials} new materials"

                # Level 4: Additional info (12 spaces)
                if updated_prices > 0 or created_materials > 0:
                    # Level 5: Inside if (16 spaces)
                    message += f"\n\n Material master prices have been automatically updated!"
                    message += (
                        f"\n   These prices will be used when converting PRs to POs."
                    )

                messagebox.showinfo("Upload Complete", message, parent=self)
                self.log(f"✓ SUCCESS: {message.replace(chr(10), ' ')}")
            else:
                # Level 4: Handle old return format (12 spaces)
                processed_count = result
                message = f"Successfully uploaded {processed_count} order lines"
                messagebox.showinfo("Upload Complete", message, parent=self)
                self.log(f"✓ SUCCESS: Uploaded {processed_count} order lines")

            # Level 3: Refresh the display (8 spaces)
            self.refresh_data()

        except Exception as e:
            # Level 3: Error handling (8 spaces)
            error_msg = f"Failed to upload order book: {str(e)}"
            self.log(f"✗ ERROR: {error_msg}")
            messagebox.showerror("Upload Failed", error_msg, parent=self)

    def create_context_menu(self):
        """Create right-click context menu for copying data"""
        self.context_menu = tk.Menu(self, tearoff=0)
        self.context_menu.add_command(
            label="Copy Cell", command=self.copy_selected_cell
        )
        self.context_menu.add_command(label="Copy Row", command=self.copy_selected_row)
        self.context_menu.add_separator()
        self.context_menu.add_command(
            label="Copy PO Number", command=lambda: self.copy_column_value("PO")
        )
        self.context_menu.add_command(
            label="Copy Material", command=lambda: self.copy_column_value("Material")
        )

        self.tree.bind("<Button-3>", self.show_context_menu)  # Right-click

    def show_context_menu(self, event):
        """Show context menu on right-click"""
        try:
            self.context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.context_menu.grab_release()

    def copy_selected_cell(self):
        """Copy the value of the clicked cell"""
        selection = self.tree.selection()
        if not selection:
            return

        # Get the column that was clicked
        column = self.tree.identify_column(
            self.tree.winfo_pointerx() - self.tree.winfo_rootx()
        )
        col_index = int(column.replace("#", "")) - 1

        item = selection[0]
        values = self.tree.item(item, "values")

        if col_index < len(values):
            value = str(values[col_index])
            self.clipboard_clear()
            self.clipboard_append(value)
            self.log(f"ℹ️ INFO: Copied '{value}' to clipboard")

    def copy_selected_row(self):
        """Copy all values in the selected row as tab-separated"""
        selection = self.tree.selection()
        if not selection:
            return

        item = selection[0]
        values = self.tree.item(item, "values")
        row_text = "\t".join(str(v) for v in values)

        self.clipboard_clear()
        self.clipboard_append(row_text)
        self.log(f"ℹ️ INFO: Copied row data to clipboard")

    def copy_column_value(self, column_name):
        """Copy a specific column value from selected row"""
        selection = self.tree.selection()
        if not selection:
            return

        columns = (
            "PO",
            "Item",
            "Supplier",
            "Material",
            "Description",
            "Qty",
            "Unit Price",
            "Per",
            "Total Amount",
            "Currency",
            "Req. Date",
            "Conf. Date",
            "Reschedule Date",
            "Status",
        )

        try:
            col_index = columns.index(column_name)
            item = selection[0]
            values = self.tree.item(item, "values")

            if col_index < len(values):
                value = str(values[col_index])
                self.clipboard_clear()
                self.clipboard_append(value)
                self.log(f"ℹ️ INFO: Copied {column_name} '{value}' to clipboard")
        except (ValueError, IndexError) as e:
            self.log(f"✗ ERROR: Could not copy {column_name}: {e}")

    def mark_selected_closed(self):
        """Marks the selected order lines in the treeview as closed."""
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showwarning(
                "No Selection",
                "Please select one or more order lines to mark as closed.",
                parent=self,
            )
            return

        lines_to_close = []
        for item_id in selected_items:
            values = self.tree.item(item_id, "values")
            # Column indices: 0=PO, 1=Item
            po_number = values[0]
            item_number = values[1]
            if po_number and item_number:
                lines_to_close.append((po_number, item_number))

        if not lines_to_close:
            messagebox.showerror("❌ Error",
                "Could not identify the selected lines. Please try again.",
                parent=self,
            )
            return

        if messagebox.askyesno("❓ Confirm",
            f"⚠️ Manual Order Closure"
            f"Are you sure you want to close {len(lines_to_close)} order line(s)?"

            f"📌 Manually closed orders will NOT reopen during batch uploads"

            f"⚙️ This ensures your manual closures are preserved.",
            parent=self,
        ):
            try:
                closed_count = self.dm.close_order_lines(lines_to_close, closed_by_user=True)
                self.log(f"✓ Marked {closed_count} order line(s) as manually closed.")
                messagebox.showinfo("✅ Success", 
                    f"✅ {closed_count} line(s) have been manually closed."

                    f"These orders will remain closed even if they appear in future uploads.",
                    parent=self
                )
                # Refresh the data to remove closed orders from the view
                self.refresh_data()
            except Exception as e:
                self.log(f"✗ ERROR: Failed to close order lines: {e}")
                messagebox.showerror(
                    "Database Error",
                    f"Failed to close order lines:\n{str(e)}",
                    parent=self,
                )

    def on_po_double_click(self, event):
        """Handler for double-clicking a PO line to open the preview window."""
        selected_item = self.tree.focus()  # .focus() gets the ID of the selected item
        if not selected_item:
            return

        # Extract PO number from the selected row (it's the first value)
        po_number = self.tree.item(selected_item, "values")[0]
        if po_number:
            # Open the new preview window
            POPreviewWindow(self, self.log, self.dm, po_number)
        else:
            self.log("⚠ WARNING: Could not identify PO number from the selected line.")

    def refresh_data(self):
        """Load fresh data from database"""
        try:
            self.all_orders = self.dm.get_all_open_orders()
            self.load_suppliers()
            self.apply_filters()
            self.log("ℹ️ INFO: Order book data refreshed")
        except Exception as e:
            messagebox.showerror("❌ Error", f"Failed to load order data:\n{str(e)}", parent=self
            )
            self.log(f"✗ ERROR: Failed to refresh order data: {e}")

    def load_suppliers(self):
        """Load supplier list for filter"""
        suppliers = set()
        for order in self.all_orders:
            if order.get("name"):
                suppliers.add(order["name"])

        self.supplier_listbox.delete(0, tk.END)
        for supplier in sorted(suppliers):
            self.supplier_listbox.insert(tk.END, supplier)

    def apply_filters(self, *args):
        """Apply all active filters to the data"""
        filtered = self.all_orders.copy()

        # Text search filter
        search_term = self.search_var.get().lower().strip()
        if search_term:
            filtered = [
                order for order in filtered if self.matches_search(order, search_term)
            ]

        # PO number filter
        po_filter = self.po_filter_var.get().strip()
        if po_filter:
            filtered = [
                order
                for order in filtered
                if po_filter.lower() in str(order.get("po", "")).lower()
            ]

        # Supplier filter
        selected_suppliers = [
            self.supplier_listbox.get(i) for i in self.supplier_listbox.curselection()
        ]
        if selected_suppliers:
            filtered = [
                order for order in filtered if order.get("name") in selected_suppliers
            ]

        # Date filters
        filtered = self.apply_date_filters(filtered)

        # Value filters
        filtered = self.apply_value_filters(filtered)

        # Status filters
        filtered = self.apply_status_filters(filtered)

        self.filtered_orders = filtered
        self.update_display()

    def matches_search(self, order, search_term):
        """Check if order matches search term in any field"""
        searchable_fields = [
            "po",
            "item",
            "name",
            "material_code",
            "short_text",
            "comments",
        ]
        for field in searchable_fields:
            if search_term in str(order.get(field, "")).lower():
                return True
        return False

    def apply_date_filters(self, orders):
        """Apply date range filters"""
        try:
            start_date = self.req_date_start.get().strip()
            end_date = self.req_date_end.get().strip()

            if start_date or end_date:
                filtered = []
                for order in orders:
                    req_date_str = order.get("requested_del_date", "")
                    if req_date_str:
                        try:
                            order_date = datetime.strptime(req_date_str, "%d.%m.%Y")

                            if start_date:
                                filter_start = datetime.strptime(start_date, "%Y-%m-%d")
                                if order_date < filter_start:
                                    continue

                            if end_date:
                                filter_end = datetime.strptime(end_date, "%Y-%m-%d")
                                if order_date > filter_end:
                                    continue

                            filtered.append(order)
                        except ValueError:
                            filtered.append(order)
                    else:
                        filtered.append(order)
                return filtered
        except ValueError:
            pass

        return orders

    def apply_value_filters(self, orders):
        """Apply value range filters"""
        try:
            min_val = self.value_min.get().strip()
            max_val = self.value_max.get().strip()

            if min_val or max_val:
                filtered = []
                for order in orders:
                    try:
                        order_value = float(order.get("total_amount", 0))

                        if min_val and order_value < float(min_val):
                            continue
                        if max_val and order_value > float(max_val):
                            continue

                        filtered.append(order)
                    except (ValueError, TypeError):
                        filtered.append(order)
                return filtered
        except ValueError:
            pass

        return orders

    def apply_status_filters(self, orders):
        """Apply status-based filters with proper AND logic"""
        # If no status filters are active, return all orders
        if not any(var.get() for var in self.status_vars.values()):
            return orders

        filtered = []
        for order in orders:
            # Safe string checking with None handling
            conf_date = order.get("conf_delivery_date") or ""
            reschedule_date = order.get("rescheduling_date") or ""

            has_conf = bool(str(conf_date).strip())
            has_reschedule = bool(str(reschedule_date).strip())

            # Check each active filter - order must pass ALL active filters
            passes_all_filters = True

            # If "Has Confirmation" is checked, order MUST have confirmation
            if self.status_vars["has_confirmation"].get():
                if not has_conf:
                    passes_all_filters = False

            # If "Has Reschedule" is checked, order MUST have reschedule date
            if self.status_vars["has_reschedule"].get():
                if not has_reschedule:
                    passes_all_filters = False

            # If "No Confirmation" is checked, order MUST NOT have confirmation
            if self.status_vars["no_confirmation"].get():
                if has_conf:
                    passes_all_filters = False

            # If "No Reschedule" is checked, order MUST NOT have reschedule date
            if self.status_vars["no_reschedule"].get():
                if has_reschedule:
                    passes_all_filters = False

            if passes_all_filters:
                filtered.append(order)

        return filtered

    def update_display(self):
        """Update the treeview with filtered data"""
        for item in self.tree.get_children():
            self.tree.delete(item)

        total_value = 0
        for order in self.filtered_orders:
            status = self.get_order_status(order)

            unit_price = (
                f"{float(order.get('unit_price', 0)):.2f}"
                if order.get("unit_price")
                else ""
            )
            price_per = order.get("price_per_unit", 1)
            total_amount = float(order.get("total_amount", 0))
            total_value += total_amount

            values = (
                order.get("po", ""),
                order.get("item", ""),
                order.get("name", ""),
                order.get("material_code", ""),
                order.get("short_text", ""),
                order.get("requested_qty", ""),
                unit_price,
                str(price_per) if price_per and price_per > 1 else "1",
                f"{total_amount:.2f}" if total_amount else "",
                order.get("currency", ""),
                order.get("requested_del_date", ""),
                order.get("conf_delivery_date", ""),
                order.get("rescheduling_date", ""),
                order.get("comments", ""),
                status,
            )

            item = self.tree.insert("", "end", values=values)

            if status == "Confirmed":
                self.tree.item(item, tags=("confirmed",))
            elif status == "Rescheduled":
                self.tree.item(item, tags=("rescheduled",))
            elif status == "Pending":
                self.tree.item(item, tags=("pending",))

        self.tree.tag_configure("confirmed", background="#d4edda")
        self.tree.tag_configure("rescheduled", background="#fff3cd")
        self.tree.tag_configure("pending", background="#f8d7da")

        self.info_label.config(
            text=f"Showing {len(self.filtered_orders)} of {len(self.all_orders)} orders"
        )

        currency = (
            self.filtered_orders[0].get("currency", "EUR")
            if self.filtered_orders
            else "EUR"
        )
        self.summary_label.config(
            text=(
                f"Total Value: {total_value:.2f} {currency} | "
                f"Average: {total_value/len(self.filtered_orders):.2f} {currency}"
                if self.filtered_orders
                else "No data"
            )
        )

    def get_order_status(self, order):
        """Determine order status based on dates"""
        conf_date = order.get("conf_delivery_date") or ""
        reschedule_date = order.get("rescheduling_date") or ""

        has_conf = bool(str(conf_date).strip())
        has_reschedule = bool(str(reschedule_date).strip())

        if has_conf and has_reschedule:
            return "Rescheduled"
        elif has_conf:
            return "Confirmed"
        elif has_reschedule:
            return "Rescheduled"
        else:
            return "Pending"

    def sort_by_column(self, col):
        """Sort the treeview by column"""
        try:
            data = [
                (self.tree.set(child, col), child)
                for child in self.tree.get_children("")
            ]

            try:
                data.sort(key=lambda x: float(x[0]) if x[0] else 0)
            except ValueError:
                data.sort(key=lambda x: x[0].lower())

            for index, (val, child) in enumerate(data):
                self.tree.move(child, "", index)

        except Exception as e:
            self.log(f"⚠ WARNING: Sort failed for column {col}: {e}")

    def clear_filters(self):
        """Clear all active filters"""
        self.search_var.set("")
        self.po_filter_var.set("")
        self.req_date_start.set("")
        self.req_date_end.set("")
        self.value_min.set("")
        self.value_max.set("")

        for var in self.status_vars.values():
            var.set(False)

        self.supplier_listbox.selection_clear(0, tk.END)
        self.apply_filters()

    def select_all_suppliers(self):
        """Select all suppliers"""
        self.supplier_listbox.select_set(0, tk.END)
        self.apply_filters()

    def deselect_all_suppliers(self):
        """Deselect all suppliers"""
        self.supplier_listbox.selection_clear(0, tk.END)
        self.apply_filters()

    def export_filtered_data(self):
        """Export filtered data to Excel"""
        if not self.filtered_orders:
            messagebox.showwarning(
                "No Data",
                "No data to export. Apply filters to get results.",
                parent=self,
            )
            return

        file_path = filedialog.asksaveasfilename(
            title="Export Filtered Orders",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
            parent=self,
        )

        if not file_path:
            return

        try:
            export_data = []
            for order in self.filtered_orders:
                export_data.append(
                    {
                        "PO Number": order.get("po", ""),
                        "Item": order.get("item", ""),
                        "Supplier": order.get("name", ""),
                        "Material Code": order.get("material_code", ""),
                        "Description": order.get("short_text", ""),
                        "Quantity": order.get("requested_qty", ""),
                        "Unit Price": order.get("unit_price", ""),
                        "Per": order.get("price_per_unit", 1),
                        "Total Amount": order.get("total_amount", ""),
                        "Currency": order.get("currency", ""),
                        "Requested Date": order.get("requested_del_date", ""),
                        "Confirmed Date": order.get("conf_delivery_date", ""),
                        "Reschedule Date": order.get("rescheduling_date", ""),
                        "Comments": order.get("comments", ""),
                        "Status": self.get_order_status(order),
                    }
                )

            df = pd.DataFrame(export_data)

            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name="Filtered_Orders", index=False)

                worksheet = writer.sheets["Filtered_Orders"]
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except Exception:  # TODO: Add proper error handling
                            pass  # TODO: Add proper error handling
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            messagebox.showinfo(
                "Export Complete",
                f"Exported {len(export_data)} filtered orders to:\n{file_path}",
                parent=self,
            )
            self.log(f"✓ SUCCESS: Exported {len(export_data)} filtered orders to Excel")

        except Exception as e:
            messagebox.showerror(
                "Export Error", f"Failed to export data:\n{str(e)}", parent=self
            )
            self.log(f"✗ ERROR: Export failed: {e}")

    def show_arriving_orders_this_week(self):
        """Show orders arriving in current week (Monday-Sunday)"""
        try:
            # Get current week range
            today = datetime.now()
            # Find Monday of current week
            days_since_monday = today.weekday()  # Monday=0, Sunday=6
            week_start = today - timedelta(days=days_since_monday)
            week_start = week_start.replace(hour=0, minute=0, second=0, microsecond=0)
            week_end = week_start + timedelta(days=6, hours=23, minutes=59, seconds=59)
            
            # Query all open orders
            query = """
                SELECT oo.*, v.transport_days
                FROM open_orders oo
                LEFT JOIN vendors v ON oo.vendor_name = v.vendor_name
                WHERE oo.status = 'Open'
            """
            
            orders = self.dm.db.execute_query(query, (), fetchall=True)
            
            if not orders:
                messagebox.showinfo(
                    "No Data",
                    "No open orders found.",
                    parent=self
                )
                return
            
            # Calculate arrival dates and filter
            arriving_orders = []
            
            for order in orders:
                try:
                    transport_days = int(order.get('transport_days', 0))
                    
                    # Priority: Use confirmation date if available, else requested date
                    if order.get('conf_delivery_date'):
                        base_date_str = order['conf_delivery_date']
                    elif order.get('requested_del_date'):
                        base_date_str = order['requested_del_date']
                    else:
                        continue  # Skip orders without dates
                    
                    # Parse base date
                    base_date = datetime.strptime(base_date_str, "%d.%m.%Y")
                    
                    # Calculate arrival date (base_date + transport days)
                    arrival_date = base_date
                    if transport_days > 0:
                        # Add working days
                        days_added = 0
                        while days_added < transport_days:
                            arrival_date += timedelta(days=1)
                            # Skip weekends
                            if arrival_date.weekday() < 5:  # Monday=0, Friday=4
                                days_added += 1
                    
                    # Check if arrival is in current week
                    if week_start <= arrival_date <= week_end:
                        order_dict = dict(order)
                        order_dict['calculated_arrival'] = arrival_date.strftime("%d.%m.%Y")
                        order_dict['base_date_used'] = 'Confirmation' if order.get('conf_delivery_date') else 'Requested'
                        arriving_orders.append(order_dict)
                        
                except Exception as e:
                    self.log(f"✗ ERROR: Failed to process order {order.get('po')}: {e}")
                    continue
            
            if not arriving_orders:
                messagebox.showinfo(
                    "No Arrivals",
                    f"No orders arriving this week\n({week_start.strftime('%d.%m.%Y')} - {week_end.strftime('%d.%m.%Y')})",
                    parent=self
                )
                return
            
            # Open window with results
            ArrivingOrdersWindow(
                self,
                arriving_orders,
                week_start.strftime("%d.%m.%Y"),
                week_end.strftime("%d.%m.%Y"),
                self.log
            )
            
        except Exception as e:
            messagebox.showerror("❌ Error",
                f"Failed to calculate arriving orders:\n{str(e)}",
                parent=self
            )
            self.log(f"✗ ERROR: Arriving orders calculation failed: {e}")


# ==============================================================================
# ARRIVING ORDERS WINDOW
# ==============================================================================


class ArrivingOrdersWindow(tk.Toplevel):
    """Display orders arriving in the current week"""
    def __init__(self, parent, orders, week_start, week_end, log_callback):
        super().__init__(parent)
        self.title(f"Orders Arriving This Week ({week_start} - {week_end})")
        self.geometry("1200x600")
        self.orders = orders
        self.log = log_callback
        
        # Title
        title_frame = ttk.Frame(self, padding=10)
        title_frame.pack(fill=tk.X)
        
        ttk.Label(
            title_frame,
            text=f" Orders Arriving: {week_start} - {week_end}",
            font=("Helvetica", 14, "bold")
        ).pack(side=tk.LEFT)
        
        ttk.Label(
            title_frame,
            text=f"Total: {len(orders)} orders",
            font=("Helvetica", 12)
        ).pack(side=tk.LEFT, padx=20)
        
        # Buttons
        button_frame = ttk.Frame(self, padding=10)
        button_frame.pack(fill=tk.X)
        
        ttk.Button(
            button_frame,
            text=" Export to Excel",
            command=self.export_to_excel
        ).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            button_frame,
            text=" Refresh",
            command=lambda: [self.destroy(), parent.show_arriving_orders_this_week()]
        ).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            button_frame,
            text=" Close",
            command=self.destroy
        ).pack(side=tk.RIGHT, padx=5)
        
        # Tree view
        tree_frame = ttk.Frame(self, padding=10)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        columns = (
            "PO",
            "Vendor",
            "Item",
            "Quantity",
            "Unit",
            "Conf. Date",
            "Req. Date",
            "Transport Days",
            "Arrival Date",
            "Date Used",
            "Status"
        )
        
        self.tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=20)
        
        # Configure columns
        column_widths = {
            "PO": 100,
            "Vendor": 150,
            "Item": 200,
            "Quantity": 80,
            "Unit": 60,
            "Conf. Date": 100,
            "Req. Date": 100,
            "Transport Days": 100,
            "Arrival Date": 100,
            "Date Used": 100,
            "Status": 80
        }
        
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=column_widths.get(col, 100))
        
        # Scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        # Populate data
        self.populate_tree()
        
    def populate_tree(self):
        """Populate tree with order data"""
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        for order in self.orders:
            values = (
                order.get('po', ''),
                order.get('vendor_name', ''),
                order.get('item', ''),
                order.get('quantity', ''),
                order.get('unit', ''),
                order.get('conf_delivery_date', 'N/A'),
                order.get('requested_del_date', 'N/A'),
                order.get('transport_days', '0'),
                order.get('calculated_arrival', ''),
                order.get('base_date_used', ''),
                order.get('status', '')
            )
            self.tree.insert("", "end", values=values)
    
    def export_to_excel(self):
        """Export arriving orders to Excel"""
        try:
            from tkinter import filedialog
            
            # Ask for save location
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=f"arriving_orders_{datetime.now().strftime('%Y%m%d')}.xlsx",
                parent=self
            )
            
            if not file_path:
                return
            
            # Create DataFrame
            df = pd.DataFrame(self.orders)
            
            # Select and order columns
            columns_to_export = [
                'po', 'vendor_name', 'item', 'quantity', 'unit',
                'conf_delivery_date', 'requested_del_date', 'transport_days',
                'calculated_arrival', 'base_date_used', 'status'
            ]
            
            df_export = df[[col for col in columns_to_export if col in df.columns]]
            
            # Rename columns for clarity
            df_export.columns = [
                'PO Number', 'Vendor', 'Item', 'Quantity', 'Unit',
                'Confirmation Date', 'Requested Date', 'Transport Days',
                'Calculated Arrival', 'Date Used For Calculation', 'Status'
            ]
            
            # Export to Excel
            df_export.to_excel(file_path, index=False, sheet_name="Arriving Orders")
            
            messagebox.showinfo("✅ Success",
                f"Exported {len(self.orders)} orders to:\n{file_path}",
                parent=self
            )
            self.log(f"ℹ️ INFO: Exported arriving orders to {file_path}")
            
        except Exception as e:
            messagebox.showerror(
                "Export Error",
                f"Failed to export:\n{str(e)}",
                parent=self
            )
            self.log(f"✗ ERROR: Export failed: {e}")


# ==============================================================================
# 3. EMAIL SENDER CLASS
# ==============================================================================


class EmailSender:
    """Handles all email sending logic via SMTP or Outlook."""

    def __init__(self, log_callback, data_manager):
        self.log = log_callback
        self.dm = data_manager
        self.smtp_config = self.dm.get_config("smtp_settings")
        self.email_templates = self.dm.get_config("email_templates", {})
        self.company_config = self.dm.get_config("company_config", {})

    def find_po_pdf(self, po_number):
        for folder in [ORDERS_FOLDER, ORDERS_SENT_FOLDER]:
            pdf_path = os.path.join(folder, f"PO_{po_number}.pdf")
            if os.path.exists(pdf_path):
                return pdf_path

            try:
                for filename in os.listdir(folder):
                    if filename.lower().endswith(".pdf") and po_number in filename:
                        return os.path.join(folder, filename)
            except Exception:  # TODO: Add proper error handling
                pass  # TODO: Add proper error handling
        return None

    def generate_po_email_content(self, po_data):
        """Generate email content for PO emails"""
        po_number = po_data["po"]
        supplier_name = po_data["supplier"]
        api_key = po_data.get("api_key")
        has_portal_access = bool(api_key and api_key.strip())

        # Get templates
        po_template = self.email_templates.get("po_email", {})

        # Prepare replacements
        replacements = {
            "po_number": po_number,
            "supplier_name": supplier_name,
            "company_name": self.company_config.get("my_company_name", "Your Company"),
            "buyer_name": self.company_config.get("buyer_name", "Purchasing Team"),
            "buyer_email": self.company_config.get(
                "buyer_email", "purchasing@company.com"
            ),
            "portal_url": self.email_templates.get("portal_base_url", ""),
            "api_key": api_key or "",
        }

        # Generate portal link section
        portal_link_section = ""
        if self.email_templates.get("include_portal_link", True) and has_portal_access:
            extra_template = po_template.get("extra", "")
            portal_link_section = extra_template.format(**replacements)

        replacements["portal_link_section"] = portal_link_section

        # Generate subject and body
        subject = po_template.get("subject", "Purchase Order {po_number}").format(
            **replacements
        )
        body = po_template.get("body", "Please find attached PO {po_number}").format(
            **replacements
        )

        return subject, body

    def generate_email_content(self, po_data, is_fallback=False):
        """
        Wrapper method for backward compatibility.
        Calls generate_po_email_content internally.
        """
        return self.generate_po_email_content(po_data)

    def generate_reschedule_email_content(self, supplier_name):
        """Generate email content for reschedule emails"""
        template = self.email_templates.get("reschedule_email", {})

        replacements = {
            "supplier_name": supplier_name,
            "company_name": self.company_config.get("my_company_name", "Your Company"),
            "buyer_name": self.company_config.get("buyer_name", "Purchasing Team"),
            "buyer_email": self.company_config.get(
                "buyer_email", "purchasing@company.com"
            ),
        }

        subject = template.get("subject", "Delivery Schedule Update").format(
            **replacements
        )
        body = template.get(
            "body", "Please find attached the updated delivery schedule."
        ).format(**replacements)

        return subject, body

    def generate_reminder_email_content(self, supplier_name):
        """Generate email content for reminder emails"""
        template = self.email_templates.get("reminder_email", {})

        replacements = {
            "supplier_name": supplier_name,
            "company_name": self.company_config.get("my_company_name", "Your Company"),
            "buyer_name": self.company_config.get("buyer_name", "Purchasing Team"),
            "buyer_email": self.company_config.get(
                "buyer_email", "purchasing@company.com"
            ),
        }

        subject = template.get(
            "subject", "Reminder: Please Confirm Purchase Orders"
        ).format(**replacements)
        body = template.get(
            "body", "This is a reminder to confirm your outstanding purchase orders."
        ).format(**replacements)

        return subject, body

    def generate_forecast_email_content(self, supplier_name):
        """Generate email content for forecast emails"""
        template = self.email_templates.get("forecast_email", {})

        replacements = {
            "supplier_name": supplier_name,
            "company_name": self.company_config.get("my_company_name", "Your Company"),
            "buyer_name": self.company_config.get("buyer_name", "Purchasing Team"),
            "buyer_email": self.company_config.get(
                "buyer_email", "purchasing@company.com"
            ),
        }

        subject = template.get("subject", "Demand Forecast").format(**replacements)
        body = template.get("body", "Please find attached our demand forecast.").format(
            **replacements
        )

        return subject, body

    def _get_outlook_signature(self):
        """Retrieve the default Outlook signature"""
        try:
            import win32com.client

            outlook = win32com.client.Dispatch("outlook.application")

            appdata = os.environ.get("APPDATA")
            sig_path = os.path.join(appdata, "Microsoft", "Signatures")

            if os.path.exists(sig_path):
                sig_files = [f for f in os.listdir(sig_path) if f.endswith(".htm")]
                if sig_files:
                    with open(
                        os.path.join(sig_path, sig_files[0]), "r", encoding="utf-8"
                    ) as f:
                        return f.read()
            return ""
        except Exception as e:
            self.log(f"⚠ WARNING: Could not load Outlook signature: {e}")
            return ""

    def _send_smtp(self, to_emails, subject, body, attachment_path=None, use_custom_signature=False):
        """Generic SMTP send method"""
        if not self.smtp_config:
            return False, "SMTP configuration is missing."

        if not to_emails:
            return False, "No email address provided."

        # Get Outlook signature by default, custom signature only if requested
        outlook_signature = self._get_outlook_signature()
        custom_signature = self.dm.get_signature() if use_custom_signature else ""
        
        # Combine signatures
        combined_signature = ""
        if outlook_signature:
            combined_signature = outlook_signature
        if custom_signature:
            combined_signature = (combined_signature + "<br>" + custom_signature) if combined_signature else custom_signature
        
        msg = MIMEMultipart("alternative")
        msg["From"] = (
            f"{self.smtp_config['from_name']} <{self.smtp_config['from_email']}>"
        )
        msg["To"] = to_emails
        msg["Subject"] = subject
        
        # Plain text version
        plain_body = body
        if combined_signature:
            plain_body += "\n\n---\n[Email Signature]"
        msg.attach(MIMEText(plain_body, "plain"))
        
        # HTML version with signature
        if combined_signature:
            html_body = f"""
            <html>
                <body>
                    <pre style="font-family: Arial, sans-serif;">{body}</pre>
                    <br>
                    {combined_signature}
                </body>
            </html>
            """
            msg.attach(MIMEText(html_body, "html"))
        else:
            msg.attach(MIMEText(body, "plain"))

        if attachment_path and os.path.exists(attachment_path):
            with open(attachment_path, "rb") as f:
                part = MIMEApplication(f.read(), Name=os.path.basename(attachment_path))
            part["Content-Disposition"] = (
                f'attachment; filename="{os.path.basename(attachment_path)}"'
            )
            msg.attach(part)

        server = smtplib.SMTP(
            self.smtp_config["smtp_server"], self.smtp_config["smtp_port"]
        )
        if self.smtp_config.get("use_tls", True):
            server.starttls()
        server.login(
            self.smtp_config["smtp_username"], self.smtp_config["smtp_password"]
        )
        server.send_message(msg)
        server.quit()
        return True, "Email sent via SMTP"

    def _send_outlook(self, to_emails, subject, body, attachment_path=None, use_custom_signature=False):
        """Generic Outlook send method"""
        if not to_emails:
            return False, "No email address provided."

        # Initialize COM for this thread
        pythoncom.CoInitialize()

        try:
            outlook = win32com.client.Dispatch("outlook.application")
            mail = outlook.CreateItem(0)
            mail.To = to_emails
            mail.Subject = subject

            # Get Outlook default signature and custom signature only if requested
            outlook_signature = self._get_outlook_signature()
            custom_signature = self.dm.get_signature() if use_custom_signature else ""
            
            # Combine signatures
            combined_signature = ""
            if outlook_signature:
                combined_signature = outlook_signature
            if custom_signature:
                combined_signature = (combined_signature + "<br>" + custom_signature) if combined_signature else custom_signature
            
            if combined_signature:
                mail.HTMLBody = body.replace("\n", "<br>") + "<br><br>" + combined_signature
            else:
                mail.Body = body

            if attachment_path and os.path.exists(attachment_path):
                mail.Attachments.Add(attachment_path)

            mail.Send()
            return True, "Email sent via Outlook"
        finally:
            # Uninitialize COM when done
            pythoncom.CoUninitialize()

    def send_all_pending_emails(self, preferred_method="Outlook", pos_to_send=None, use_custom_signature=False):
        """Send PO emails"""
        if preferred_method == "Outlook" and not OUTLOOK_AVAILABLE:
            return (
                0,
                0,
                "Outlook method selected, but pywin32 library is not installed.",
            )

        if pos_to_send:
            pending_pos = pos_to_send
        else:
            pending_pos = self.dm.get_pending_pos_with_portal_info()

        if not pending_pos:
            return 0, 0, "No pending POs found to send."

        sent_count, failed_count, failed_details = 0, 0, []

        for po_data in pending_pos:
            po_number = po_data["po"]
            is_fallback = po_data.get("is_fallback", False)

            pdf_path = self.find_po_pdf(po_number)
            if not pdf_path:
                failed_count += 1
                failed_details.append(f"PO {po_number}: PDF not found")
                continue

            to_emails = po_data.get("emails")
            subject, body = self.generate_po_email_content(po_data)

            success, message = False, ""
            try:
                if preferred_method == "Outlook":
                    success, message = self._send_outlook(
                        to_emails, subject, body, pdf_path, use_custom_signature
                    )
                else:
                    success, message = self._send_smtp(
                        to_emails, subject, body, pdf_path, use_custom_signature
                    )
            except Exception as e:
                self.log(f"✗ ERROR: Failed to send PO {po_number}: {e}")
                message = str(e)

            if success:
                sent_count += 1

                if not is_fallback:
                    self.dm.mark_email_sent(po_number)

                sent_pdf_path = os.path.join(
                    ORDERS_SENT_FOLDER, os.path.basename(pdf_path)
                )
                if os.path.exists(pdf_path):
                    os.rename(pdf_path, sent_pdf_path)

                fallback_tag = " (fallback)" if is_fallback else ""
                self.log(f"✓ SUCCESS: PO {po_number} sent{fallback_tag}. {message}")
            else:
                failed_count += 1
                failed_details.append(f"PO {po_number}: {message}")

        summary = f"Sent {sent_count}, Failed {failed_count}"
        if failed_details:
            summary += "\n\nFailures:\n" + "\n".join(failed_details)
        return sent_count, failed_count, summary

    def _send_smtp_multiple_attachments(self, to_emails, subject, body, attachment_paths, use_custom_signature=False):
        """Send email via SMTP with multiple attachments"""
        if not self.smtp_config:
            return False, "SMTP configuration is missing."

        if not to_emails:
            return False, "No email address provided."

        # Get Outlook signature by default, custom signature only if requested
        outlook_signature = self._get_outlook_signature()
        custom_signature = self.dm.get_signature() if use_custom_signature else ""
        
        # Combine signatures
        combined_signature = ""
        if outlook_signature:
            combined_signature = outlook_signature
        if custom_signature:
            combined_signature = (combined_signature + "<br>" + custom_signature) if combined_signature else custom_signature
        
        msg = MIMEMultipart("alternative")
        msg["From"] = f"{self.smtp_config['from_name']} <{self.smtp_config['from_email']}>"
        msg["To"] = to_emails
        msg["Subject"] = subject
        
        # Plain text version
        plain_body = body
        if combined_signature:
            plain_body += "\n\n---\n[Email Signature]"
        msg.attach(MIMEText(plain_body, "plain"))
        
        # HTML version with signature
        if combined_signature:
            html_body = f"""
            <html>
                <body>
                    <pre style="font-family: Arial, sans-serif;">{body}</pre>
                    <br>
                    {combined_signature}
                </body>
            </html>
            """
            msg.attach(MIMEText(html_body, "html"))

        # Add all attachments
        for attachment_path in attachment_paths:
            if attachment_path and os.path.exists(attachment_path):
                with open(attachment_path, "rb") as f:
                    part = MIMEApplication(f.read(), Name=os.path.basename(attachment_path))
                part["Content-Disposition"] = (
                    f'attachment; filename="{os.path.basename(attachment_path)}"'
                )
                msg.attach(part)

        try:
            server = smtplib.SMTP(
                self.smtp_config["smtp_server"], self.smtp_config["smtp_port"]
            )
            if self.smtp_config.get("use_tls", True):
                server.starttls()
            server.login(
                self.smtp_config["smtp_username"], self.smtp_config["smtp_password"]
            )
            server.send_message(msg)
            server.quit()
            return True, "Email sent via SMTP"
        except Exception as e:
            return False, str(e)


class POManagementWindow(tk.Toplevel):
    def __init__(self, parent, log_callback, data_manager, send_method_getter):
        super().__init__(parent)
        self.title("Purchase Order Management")
        self.geometry("900x700")
        self.log = log_callback
        self.dm = data_manager
        self.get_send_method = send_method_getter

        # Create notebook for tabs
        notebook = ttk.Notebook(self)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Tab 1: Create PO
        create_frame = ttk.Frame(notebook)
        notebook.add(create_frame, text="Create New PO")
        self.setup_create_tab(create_frame)

        # Tab 2: Generate PDFs
        generate_frame = ttk.Frame(notebook)
        notebook.add(generate_frame, text="Generate PDFs")
        self.setup_generate_tab(generate_frame)

        # Tab 3: Send Emails
        send_frame = ttk.Frame(notebook)
        notebook.add(send_frame, text="Send Emails")
        self.setup_send_tab(send_frame)

        # Tab 4: Reminders
        reminder_frame = ttk.Frame(notebook)
        notebook.add(reminder_frame, text="Send Reminders")
        self.setup_reminder_tab(reminder_frame)

    def setup_create_tab(self, parent):
        """Setup PO creation interface"""
        info_frame = ttk.LabelFrame(parent, text="Create Purchase Order", padding=10)
        info_frame.pack(fill=tk.X, padx=10, pady=10)

        ttk.Label(
            info_frame, text="Create new purchase orders manually or load from database"
        ).pack(anchor="w")

        ttk.Button(
            info_frame,
            text="Open PO Creator",
            command=lambda: POCreatorWindow(self, self.log, self.dm),
        ).pack(pady=10)

    def setup_generate_tab(self, parent):
        """Setup PDF generation interface"""
        info_frame = ttk.LabelFrame(
            parent, text="Generate PDFs from Pending Orders", padding=10
        )
        info_frame.pack(fill=tk.X, padx=10, pady=10)

        # Status display
        self.pending_count_label = ttk.Label(info_frame, text="Loading...")
        self.pending_count_label.pack(pady=5)

        ttk.Button(
            info_frame, text="Refresh Count", command=self.update_pending_count
        ).pack(pady=5)

        ttk.Button(
            info_frame,
            text="Generate All PDFs",
            command=self.generate_pending_pdfs,
            style="Accent.TButton",
        ).pack(pady=10)

        # Status log
        log_frame = ttk.LabelFrame(parent, text="Generation Log", padding=10)
        log_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.gen_log = scrolledtext.ScrolledText(log_frame, wrap=tk.WORD, height=15)
        self.gen_log.pack(fill=tk.BOTH, expand=True)

        self.update_pending_count()

    def setup_send_tab(self, parent):
        """Setup email sending interface"""
        # Method selection
        method_frame = ttk.LabelFrame(parent, text="Email Method", padding=10)
        method_frame.pack(fill=tk.X, padx=10, pady=10)

        self.send_method_var = tk.StringVar()
        send_methods = ["Outlook (Primary)", "SMTP (Secondary)"]
        if not OUTLOOK_AVAILABLE:
            send_methods = ["SMTP (Outlook not available)"]

        for method in send_methods:
            ttk.Radiobutton(
                method_frame, text=method, variable=self.send_method_var, value=method
            ).pack(anchor="w")

        self.send_method_var.set(send_methods[0])

        # Action buttons
        action_frame = ttk.LabelFrame(parent, text="Send Actions", padding=10)
        action_frame.pack(fill=tk.X, padx=10, pady=10)

        ttk.Button(
            action_frame,
            text="Preview & Send Emails...",
            command=self.open_enhanced_email_preview,
        ).pack(fill=tk.X, pady=5)

        ttk.Button(
            action_frame, text="Quick Send All Pending", command=self.send_pending_pos
        ).pack(fill=tk.X, pady=5)

    def setup_reminder_tab(self, parent):
        """Setup reminder email interface"""
        info_frame = ttk.LabelFrame(
            parent, text="Send Reminders for Unconfirmed Orders", padding=10
        )
        info_frame.pack(fill=tk.X, padx=10, pady=10)

        ttk.Label(
            info_frame,
            text="Send reminder emails to suppliers who haven't confirmed their orders",
        ).pack(anchor="w")

        ttk.Button(
            info_frame, text="Open Reminder Tool", command=self.open_reminder_window
        ).pack(pady=10)

    def update_pending_count(self):
        """Update count of pending PDFs"""
        try:
            pending = self.dm.get_pending_po_data()
            count = len(set(order["po"] for order in pending))
            self.pending_count_label.config(
                text=f"Pending POs: {count} purchase orders need PDFs generated"
            )
        except Exception as e:
            self.pending_count_label.config(text=f"Error loading count: {e}")

    def generate_pending_pdfs(self):
        """Generate PDFs for all pending lines"""
        self.gen_log.insert(
            tk.END,
            f"[{datetime.now().strftime('%H:%M:%S')}] Starting PDF generation...\n",
        )

        def process_pdfs():
            try:
                pending_data = self.dm.get_pending_po_data()
                if not pending_data:
                    self.after(
                        0,
                        lambda: messagebox.showinfo(
                            "No Pending POs",
                            "No pending purchase orders found to generate PDFs for.",
                            parent=self,
                        ),
                    )
                    self.gen_log.insert(tk.END, "No pending POs found.\n")
                    return

                df = pd.DataFrame(pending_data)
                po_groups = df.groupby("po")
                generated_count = 0
                processed_pos = []

                for po_number, lines_df in po_groups:
                    pdf_buffer = io.BytesIO()
                    if self.dm._generate_single_po_pdf(pdf_buffer, po_number, lines_df):
                        pdf_path = os.path.join(ORDERS_FOLDER, f"PO_{po_number}.pdf")
                        with open(pdf_path, "wb") as f:
                            f.write(pdf_buffer.getvalue())
                        generated_count += 1
                        processed_pos.append(po_number)
                        self.gen_log.insert(tk.END, f" Generated PO {po_number}\n")

                self.dm.mark_pos_as_created(processed_pos)

                self.after(
                    0,
                    lambda: messagebox.showinfo("✅ Success",
                        f"Successfully generated {generated_count} PO PDF(s).",
                        parent=self,
                    ),
                )
                self.gen_log.insert(
                    tk.END, f"\nSuccess: Generated {generated_count} PDFs.\n"
                )
                self.update_pending_count()
            except Exception as e:
                self.gen_log.insert(tk.END, f"\nERROR: {e}\n")
                self.after(
                    0,
                    lambda: messagebox.showerror("❌ Error", f"Failed to generate PDFs: {e}", parent=self
                    ),
                )

        threading.Thread(target=process_pdfs, daemon=True).start()

    def open_enhanced_email_preview(self):
        EnhancedEmailPreviewWindow(self, self.log, self.dm, self.get_send_method)

    def send_pending_pos(self):
        method = (
            "Outlook"
            if "Outlook" in self.send_method_var.get() and OUTLOOK_AVAILABLE
            else "SMTP"
        )
        if not messagebox.askyesno(
            "Confirm Quick Send",
            f"Send all pending PO emails using '{method}'?",
            parent=self,
        ):
            return

        def send_thread():
            email_sender = EmailSender(self.log, self.dm)
            sent, failed, summary = email_sender.send_all_pending_emails(
                preferred_method=method
            )
            self.after(
                0,
                lambda: messagebox.showinfo(
                    "Email Sending Complete", summary, parent=self
                ),
            )

        threading.Thread(target=send_thread, daemon=True).start()

    def open_reminder_window(self):
        EmailReminderWindow(self, self.log, self.dm, self.get_send_method)


class SettingsWindow(tk.Toplevel):
    def __init__(self, parent, log_callback, data_manager):
        super().__init__(parent)
        self.title("System Settings")
        self.geometry("900x700")
        self.log = log_callback
        self.dm = data_manager

        # Create notebook for tabs
        notebook = ttk.Notebook(self)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Tab 1: Email Templates
        email_frame = ttk.Frame(notebook)
        notebook.add(email_frame, text="Email Templates")
        self.setup_email_templates_tab(email_frame)

        # Tab 2: SMTP Settings
        smtp_frame = ttk.Frame(notebook)
        notebook.add(smtp_frame, text="SMTP")
        self.setup_smtp_tab(smtp_frame)

        # Tab 3: Company Info
        company_frame = ttk.Frame(notebook)
        notebook.add(company_frame, text="Company Info")
        self.setup_company_tab(company_frame)

        # Tab 4: Email Signature
        signature_frame = ttk.Frame(notebook)
        notebook.add(signature_frame, text="Email Signature")
        self.setup_signature_tab(signature_frame)

    def setup_email_templates_tab(self, parent):
        """Setup email templates for all functions"""
        # Create canvas for scrolling
        canvas = tk.Canvas(parent)
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # Portal Settings (for PO emails)
        portal_frame = ttk.LabelFrame(
            scrollable_frame, text="Supplier Portal Settings", padding=10
        )
        portal_frame.pack(fill=tk.X, padx=10, pady=5)
        portal_frame.columnconfigure(1, weight=1)

        ttk.Label(portal_frame, text="Portal Base URL:").grid(
            row=0, column=0, sticky="w", pady=2
        )
        self.portal_url_var = tk.StringVar()
        ttk.Entry(portal_frame, textvariable=self.portal_url_var).grid(
            row=0, column=1, sticky="ew", pady=2
        )

        self.include_portal_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            portal_frame,
            text="Include portal link in PO emails",
            variable=self.include_portal_var,
        ).grid(row=1, column=1, sticky="w", pady=2)

        # Template selector
        template_selector_frame = ttk.LabelFrame(
            scrollable_frame, text="Select Email Template to Edit", padding=10
        )
        template_selector_frame.pack(fill=tk.X, padx=10, pady=5)

        self.template_type_var = tk.StringVar(value="po_email")

        templates = [
            ("po_email", "Purchase Order Email"),
            ("reschedule_email", "Reschedule Email"),
            ("reminder_email", "Confirmation Reminder Email"),
            ("forecast_email", "Outbound Forecast Email"),
        ]

        for idx, (value, text) in enumerate(templates):
            ttk.Radiobutton(
                template_selector_frame,
                text=text,
                variable=self.template_type_var,
                value=value,
                command=self.load_selected_template,
            ).grid(row=idx // 2, column=idx % 2, sticky="w", padx=10, pady=2)

        # Email Template Editor
        editor_frame = ttk.LabelFrame(
            scrollable_frame, text="Email Template Editor", padding=10
        )
        editor_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        editor_frame.columnconfigure(1, weight=1)

        ttk.Label(editor_frame, text="Subject:").grid(
            row=0, column=0, sticky="w", pady=5
        )
        self.subject_var = tk.StringVar()
        ttk.Entry(editor_frame, textvariable=self.subject_var).grid(
            row=0, column=1, sticky="ew", pady=5
        )

        ttk.Label(editor_frame, text="Body:").grid(row=1, column=0, sticky="nw", pady=5)
        self.body_text = scrolledtext.ScrolledText(editor_frame, height=8, wrap=tk.WORD)
        self.body_text.grid(row=1, column=1, sticky="nsew", pady=5)

        ttk.Label(editor_frame, text="Portal/Extra Section:").grid(
            row=2, column=0, sticky="nw", pady=5
        )
        self.extra_text = scrolledtext.ScrolledText(
            editor_frame, height=4, wrap=tk.WORD
        )
        self.extra_text.grid(row=2, column=1, sticky="nsew", pady=5)

        editor_frame.rowconfigure(1, weight=2)
        editor_frame.rowconfigure(2, weight=1)

        # Available placeholders info
        placeholders_frame = ttk.LabelFrame(
            scrollable_frame, text="Available Placeholders", padding=10
        )
        placeholders_frame.pack(fill=tk.X, padx=10, pady=5)

        self.placeholders_label = ttk.Label(
            placeholders_frame, text="", font=("Helvetica", 8), foreground="gray"
        )
        self.placeholders_label.pack(anchor="w")

        # Save button
        ttk.Button(
            scrollable_frame,
            text="Save Email Templates",
            command=self.save_email_templates,
        ).pack(pady=10, padx=10, anchor="e")

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        self.load_email_templates()
        self.load_selected_template()

    def load_email_templates(self):
        """Load all email templates from config"""
        config = self.dm.get_config("email_templates", {})

        # Portal settings
        self.portal_url_var.set(
            config.get("portal_base_url", "http://example.com/supplier_portal.html")
        )
        self.include_portal_var.set(config.get("include_portal_link", True))

        # Store all templates
        self.templates = {
            "po_email": config.get(
                "po_email",
                {
                    "subject": "Purchase Order {po_number} from {company_name}",
                    "body": "Dear {supplier_name},\n\nPlease find attached Purchase Order {po_number}.\n\n{portal_link_section}\n\nBest regards,\n{buyer_name}\n{buyer_email}",
                    "extra": "Access our Supplier Portal: {portal_url}?key={api_key}",
                },
            ),
            "reschedule_email": config.get(
                "reschedule_email",
                {
                    "subject": "Delivery Schedule Update - {supplier_name}",
                    "body": "Dear {supplier_name},\n\nPlease find attached the updated delivery schedule for your open purchase orders.\n\nThe attached file contains line items that require schedule updates. Please review and confirm the updated delivery dates.\n\nIf you have any questions, please contact us at {buyer_email}.\n\nBest regards,\n{buyer_name}",
                    "extra": "",
                },
            ),
            "reminder_email": config.get(
                "reminder_email",
                {
                    "subject": "Reminder: Please Confirm Purchase Order(s) - {supplier_name}",
                    "body": "Dear {supplier_name},\n\nThis is a friendly reminder that we are still awaiting delivery confirmation for the purchase order(s) listed in the attached document(s).\n\nPlease review and confirm the delivery dates at your earliest convenience.\n\nIf you have any questions, please contact us at {buyer_email}.\n\nBest regards,\n{buyer_name}",
                    "extra": "",
                },
            ),
            "forecast_email": config.get(
                "forecast_email",
                {
                    "subject": "Demand Forecast - {supplier_name}",
                    "body": "Dear {supplier_name},\n\nPlease find attached our demand forecast for the coming weeks.\n\nThis forecast is provided to help you plan your production capacity. Actual purchase orders will follow based on confirmed requirements.\n\nPlease confirm your capacity to meet this forecast or provide alternative proposals.\n\nIf you have any questions, please contact us at {buyer_email}.\n\nBest regards,\n{buyer_name}",
                    "extra": "",
                },
            ),
        }

    def load_selected_template(self):
        """Load the selected template into the editor"""
        template_type = self.template_type_var.get()
        template = self.templates.get(template_type, {})

        self.subject_var.set(template.get("subject", ""))

        self.body_text.delete("1.0", tk.END)
        self.body_text.insert("1.0", template.get("body", ""))

        self.extra_text.delete("1.0", tk.END)
        self.extra_text.insert("1.0", template.get("extra", ""))

        # Update placeholders info
        placeholders = {
            "po_email": "{po_number}, {supplier_name}, {company_name}, {buyer_name}, {buyer_email}, {portal_url}, {api_key}, {portal_link_section}",
            "reschedule_email": "{supplier_name}, {company_name}, {buyer_name}, {buyer_email}",
            "reminder_email": "{supplier_name}, {company_name}, {buyer_name}, {buyer_email}",
            "forecast_email": "{supplier_name}, {company_name}, {buyer_name}, {buyer_email}",
        }

        self.placeholders_label.config(
            text=f"Available: {placeholders.get(template_type, '')}"
        )

    def save_email_templates(self):
        """Save all email templates"""
        # Save current template being edited
        template_type = self.template_type_var.get()
        self.templates[template_type] = {
            "subject": self.subject_var.get(),
            "body": self.body_text.get("1.0", "end-1c"),
            "extra": self.extra_text.get("1.0", "end-1c"),
        }

        config = {
            "portal_base_url": self.portal_url_var.get(),
            "include_portal_link": self.include_portal_var.get(),
            "po_email": self.templates["po_email"],
            "reschedule_email": self.templates["reschedule_email"],
            "reminder_email": self.templates["reminder_email"],
            "forecast_email": self.templates["forecast_email"],
        }

        self.dm.save_config("email_templates", config)
        messagebox.showinfo("✅ Success", "Email templates saved successfully.", parent=self
        )
        self.log("ℹ️ INFO: Email templates saved")

    def setup_smtp_tab(self, parent):
        """SMTP configuration (unchanged)"""
        main_frame = ttk.Frame(parent, padding=15)
        main_frame.pack(fill=tk.BOTH, expand=True)
        main_frame.columnconfigure(1, weight=1)

        self.smtp_fields = {}
        field_labels = [
            ("smtp_server", "SMTP Server:"),
            ("smtp_port", "Port:"),
            ("smtp_username", "Username:"),
            ("smtp_password", "Password:"),
            ("from_name", "From Name:"),
            ("from_email", "From Email:"),
        ]

        for i, (key, label) in enumerate(field_labels):
            ttk.Label(main_frame, text=label).grid(row=i, column=0, sticky="w", pady=2)
            self.smtp_fields[key] = tk.StringVar()
            entry = ttk.Entry(main_frame, textvariable=self.smtp_fields[key])
            if "password" in key:
                entry.config(show="*")
            entry.grid(row=i, column=1, sticky="ew", pady=2)

        self.smtp_fields["use_tls"] = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            main_frame, text="Use TLS", variable=self.smtp_fields["use_tls"]
        ).grid(row=len(field_labels), column=1, sticky="w", pady=5)

        button_frame = ttk.Frame(main_frame)
        button_frame.grid(
            row=len(field_labels) + 1, column=0, columnspan=2, pady=10, sticky="e"
        )
        ttk.Button(
            button_frame, text="Test Connection", command=self.test_smtp_connection
        ).pack(side=tk.LEFT, padx=5)
        ttk.Button(
            button_frame, text="Save SMTP Settings", command=self.save_smtp_config
        ).pack(side=tk.LEFT)

        self.load_smtp_config()

    def setup_company_tab(self, parent):
        """Company information with logo upload"""
        main_frame = ttk.Frame(parent, padding=15)
        main_frame.pack(fill=tk.BOTH, expand=True)
        main_frame.columnconfigure(1, weight=1)

        # Logo Upload Section
        logo_frame = ttk.LabelFrame(main_frame, text="Company Logo", padding=10)
        logo_frame.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 10))

        self.logo_path_var = tk.StringVar()

        logo_info_frame = ttk.Frame(logo_frame)
        logo_info_frame.pack(fill=tk.X)

        ttk.Label(logo_info_frame, text="Logo File:").pack(side=tk.LEFT, padx=(0, 5))
        ttk.Entry(
            logo_info_frame, textvariable=self.logo_path_var, state="readonly"
        ).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        ttk.Button(logo_info_frame, text="Browse...", command=self.browse_logo).pack(
            side=tk.LEFT, padx=(0, 5)
        )
        ttk.Button(logo_info_frame, text="Clear Logo", command=self.clear_logo).pack(
            side=tk.LEFT
        )

        ttk.Label(
            logo_frame,
            text="Recommended: PNG or JPG, max width 2 inches, transparent background preferred",
            font=("Helvetica", 8),
            foreground="gray",
        ).pack(anchor="w", pady=(5, 0))

        # Company fields
        self.company_fields = {}
        field_labels = [
            ("my_company_name", "Company Name:"),
            ("my_company_address", "Company Address:"),
            ("buyer_name", "Buyer Name:"),
            ("buyer_email", "Buyer Email:"),
            ("ship_to_name", "Ship-To Name:"),
            ("ship_to_address", "Ship-To Address:"),
            ("terms_and_conditions", "Terms and Conditions:"),
        ]

        row = 1
        for key, label in field_labels:
            ttk.Label(main_frame, text=label).grid(
                row=row, column=0, sticky="nw", pady=2
            )
            if "address" in key or "terms_and_conditions" in key:
                height = 5 if "terms_and_conditions" in key else 3
                widget = tk.Text(main_frame, height=height)
                widget.widget_type = "text"  # Marker for Text widget
            else:
                widget = ttk.Entry(main_frame)
                widget.widget_type = "entry"  # Marker for Entry widget

            # Store widget in dictionary
            self.company_fields[key] = widget
            # Grid the widget
            widget.grid(row=row, column=1, sticky="ew", pady=2)
            row += 1

        ttk.Button(
            main_frame,
            text="Save Company Configuration",
            command=self.save_company_config,
        ).grid(row=row, column=0, columnspan=2, pady=10, sticky="e")

        self.load_company_config()

    def browse_logo(self):
        """Browse for logo file"""
        filename = filedialog.askopenfilename(
            title="Select Company Logo",
            filetypes=[
                ("Image Files", "*.png *.jpg *.jpeg"),
                ("PNG Files", "*.png"),
                ("JPEG Files", "*.jpg *.jpeg"),
                ("All Files", "*.*"),
            ],
        )
        if filename:
            # Copy logo to app data folder
            import shutil

            logo_folder = os.path.join(APP_DATA_FOLDER, "logos")
            os.makedirs(logo_folder, exist_ok=True)

            # Get file extension
            _, ext = os.path.splitext(filename)
            logo_filename = f"company_logo{ext}"
            logo_dest = os.path.join(logo_folder, logo_filename)

            try:
                shutil.copy2(filename, logo_dest)
                self.logo_path_var.set(logo_dest)
                messagebox.showinfo("✅ Success", "Logo uploaded successfully!", parent=self
                )
            except Exception as e:
                messagebox.showerror("❌ Error", f"Failed to copy logo: {e}", parent=self)

    def clear_logo(self):
        """Clear the logo"""
        if messagebox.askyesno("❓ Confirm", "Remove company logo from PO template?", parent=self
        ):
            self.logo_path_var.set("")

    def save_company_config(self):
        """Save company configuration including logo"""
        config = {}
        for key, widget in self.company_fields.items():
            # Use widget marker to determine type
            if hasattr(widget, "widget_type") and widget.widget_type == "text":
                # Text widget requires indices
                config[key] = widget.get("1.0", "end-1c").strip()
            else:
                # Entry widget uses get() without arguments
                config[key] = widget.get().strip()

        # Add logo path
        config["company_logo_path"] = self.logo_path_var.get()

        self.dm.save_config("company_config", config)
        messagebox.showinfo("✅ Success", "Company configuration saved.", parent=self)
        self.log("ℹ️ INFO: Company configuration saved")

    def load_company_config(self):
        """Load company configuration including logo"""
        config = self.dm.get_config("company_config", {})
        for key, widget in self.company_fields.items():
            value = config.get(key, "")

            # Try-except approach - most reliable
            try:
                # Try Text widget method first
                widget.delete("1.0", tk.END)
                widget.insert("1.0", value)
            except Exception as e:
                # If Text method fails, use Entry method
                try:
                    widget.delete(0, tk.END)
                    widget.insert(0, value)
                except Exception as e2:
                    print(f"ERROR loading {key}: {e2}")

        # Load logo path
        logo_path = config.get("company_logo_path", "")
        self.logo_path_var.set(logo_path)

    def load_smtp_config(self):
        config = self.dm.get_config("smtp_settings", {})
        for key, var in self.smtp_fields.items():
            if isinstance(var, tk.BooleanVar):
                var.set(config.get(key, True))
            else:
                var.set(config.get(key, ""))

    def save_smtp_config(self):
        config = {key: var.get() for key, var in self.smtp_fields.items()}
        try:
            config["smtp_port"] = int(config["smtp_port"])
        except ValueError:
            messagebox.showerror("❌ Error", "Port must be a number.", parent=self)
            return

        self.dm.save_config("smtp_settings", config)
        messagebox.showinfo("✅ Success", "SMTP settings saved.", parent=self)

    def test_smtp_connection(self):
        config = {key: var.get() for key, var in self.smtp_fields.items()}
        try:
            port = int(config["smtp_port"])
            server = smtplib.SMTP(config["smtp_server"], port, timeout=10)
            if config["use_tls"]:
                server.starttls()
            server.login(config["smtp_username"], config["smtp_password"])
            server.quit()
            messagebox.showinfo("✅ Success", "SMTP connection successful!", parent=self)
        except Exception as e:
            messagebox.showerror(
                "Connection Failed", f"Could not connect:\n{e}", parent=self
            )

    def setup_signature_tab(self, parent):
        """Setup email signature editor"""
        main_frame = ttk.Frame(parent, padding=15)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Title
        title_label = ttk.Label(
            main_frame,
            text="Email Signature Configuration",
            font=("Helvetica", 14, "bold")
        )
        title_label.pack(pady=(0, 15))
        
        # Description
        desc_label = ttk.Label(
            main_frame,
            text="Create a professional email signature that will be automatically added to all outgoing emails.",
            wraplength=800
        )
        desc_label.pack(pady=(0, 15))
        
        # Editor Frame
        editor_frame = ttk.LabelFrame(main_frame, text="Signature Editor", padding=10)
        editor_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        # Buttons
        button_frame = ttk.Frame(editor_frame)
        button_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Button(
            button_frame,
            text="🖼️ Open Visual Editor",
            command=self.open_signature_editor,
            width=25
        ).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            button_frame,
            text="💾 Save Signature",
            command=self.save_signature,
            width=20
        ).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            button_frame,
            text="📧 Test Email",
            command=self.test_signature_email,
            width=20
        ).pack(side=tk.LEFT, padx=5)
        
        # Add checkbox for using custom signature
        self.use_custom_signature_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            button_frame,
            text="Use this custom signature (instead of Outlook default)",
            variable=self.use_custom_signature_var
        ).pack(side=tk.LEFT, padx=15)
        
        # HTML Editor
        html_label = ttk.Label(editor_frame, text="HTML Signature Code:")
        html_label.pack(anchor="w", pady=(10, 5))
        
        self.signature_html_text = scrolledtext.ScrolledText(
            editor_frame, height=15, wrap=tk.WORD, font=("Courier", 10)
        )
        self.signature_html_text.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # Preview
        preview_label = ttk.Label(
            editor_frame, 
            text="HTML Code Preview (Note: Images/logos won't display here - use 'Test Email' button to see rendered signature):",
            wraplength=700
        )
        preview_label.pack(anchor="w", pady=(10, 5))
        
        self.signature_preview = scrolledtext.ScrolledText(
            editor_frame, height=8, wrap=tk.WORD, state='disabled'
        )
        self.signature_preview.pack(fill=tk.BOTH, expand=True)
        
        # Load existing signature
        self.load_signature()
        
    def open_signature_editor(self):
        """Open improved signature editor with live preview"""
        try:
            # Pass self.update_signature_display as the callback
            ImprovedSignatureEditor(self, self.dm.db, self.log, on_save_callback=self.update_signature_display)
        except Exception as e:
            self.log(f"✗ ERROR: Failed to open signature editor: {e}")
            messagebox.showerror("❌ Error", f"Failed to open signature editor:\n{e}", parent=self)
        
        
    def update_signature_display(self, html_content):
        """Callback function to update the HTML text box and preview"""
        self.signature_html_text.delete("1.0", tk.END)
        self.signature_html_text.insert("1.0", html_content)

        # Update preview as well
        self.signature_preview.config(state='normal')
        self.signature_preview.delete("1.0", tk.END)
        self.signature_preview.insert("1.0", html_content)
        self.signature_preview.config(state='disabled')
        self.log("ℹ️ INFO: Signature display updated from visual editor.")

    def save_signature(self):
        """Save signature to database"""
        try:
            html_content = self.signature_html_text.get("1.0", tk.END).strip()
            
            # Save HTML content to database
            self.dm.db.execute_query(
                """INSERT OR REPLACE INTO email_signatures (id, html_content, last_updated)
                   VALUES (1, ?, CURRENT_TIMESTAMP)""",
                (html_content,)
            )
            
            # Save the preference whether to use custom signature
            self.dm.save_config("use_custom_signature", self.use_custom_signature_var.get())
            
            # Update preview
            self.signature_preview.config(state='normal')
            self.signature_preview.delete("1.0", tk.END)
            self.signature_preview.insert("1.0", html_content)
            self.signature_preview.config(state='disabled')
            
            messagebox.showinfo("✅ Success", "Email signature saved successfully!", parent=self)
            self.log("ℹ️ INFO: Email signature saved")
            
        except Exception as e:
            messagebox.showerror("❌ Error", f"Failed to save signature:\n{e}", parent=self)
            self.log(f"✗ ERROR: Failed to save signature: {e}")
            
    def load_signature(self):
        """Load signature from database"""
        try:
            # Ensure table exists
            self.dm.db.execute_query(
                """CREATE TABLE IF NOT EXISTS email_signatures (
                    id INTEGER PRIMARY KEY,
                    html_content TEXT,
                    last_updated TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )""",
                ()
            )
            
            # Load signature
            result = self.dm.db.execute_query(
                "SELECT html_content FROM email_signatures WHERE id = 1",
                (),
                fetchone=True
            )
            
            if result and isinstance(result, dict) and result.get('html_content'):
                self.signature_html_text.delete("1.0", tk.END)
                self.signature_html_text.insert("1.0", result['html_content'])
                
                # Update preview
                self.signature_preview.config(state='normal')
                self.signature_preview.delete("1.0", tk.END)
                self.signature_preview.insert("1.0", result['html_content'])
                self.signature_preview.config(state='disabled')
                
        except Exception as e:
            self.log(f"ℹ️ INFO: No existing signature found or error loading: {e}")
        
        # Load the use_custom_signature preference
        try:
            use_custom = self.dm.get_config("use_custom_signature", False)
            self.use_custom_signature_var.set(use_custom)
        except Exception as e:
            self.log(f"ℹ️ INFO: Could not load use_custom_signature preference: {e}")
            
    def test_signature_email(self):
        """Send a test email with signature using EmailSender"""
        try:
            # Ask for test email address
            test_email = simpledialog.askstring(
                "Test Email",
                "Enter email address to send test:",
                parent=self
            )
            if not test_email: 
                return

            # Ask if they want to test with custom signature only or with Outlook + custom
            include_both = messagebox.askyesno(
                "Signature Test",
                "Include Outlook signature?\n\nYes = Show Outlook + Custom signature (like normal emails)\nNo = Show Custom signature only (for testing custom part)",
                parent=self
            )

            # Determine which signature to use based on the test type
            use_custom_for_test = not include_both  # If *not* including both, use *only* custom

            # Send using EmailSender
            sender = EmailSender(self.log, self.dm)
            subject = "Email Signature Test"
            body = "This is a test email to preview your signature configuration.\n\n"

            # Decide method (prefer SMTP for testing to avoid Outlook prompts if possible)
            method = "SMTP" if self.dm.get_config("smtp_settings") else "Outlook"

            if method == "Outlook" and not OUTLOOK_AVAILABLE:
                messagebox.showerror("❌ Error", "Outlook not available and SMTP not configured.", parent=self)
                return

            success = False
            message = ""

            if method == "Outlook":
                success, message = sender._send_outlook(
                    test_email, subject, body, None, use_custom_for_test
                )
            else:
                success, message = sender._send_smtp(
                    test_email, subject, body, None, use_custom_for_test
                )

            if success:
                sig_type = "Outlook + Custom" if include_both else "Custom Only"
                messagebox.showinfo("✅ Success", f"Test email sent to {test_email}!\n\nShowing: {sig_type}", parent=self)
                self.log(f"ℹ️ INFO: Test signature email sent ({sig_type})")
            else:
                messagebox.showerror("❌ Error", f"Failed to send test email:\n{message}", parent=self)
                self.log(f"✗ ERROR: Failed to send test email: {message}")

        except Exception as e:
            messagebox.showerror("❌ Error", f"Failed to send test email:\n{e}", parent=self)
            self.log(f"✗ ERROR: Failed to send test email: {e}")
            
class SignatureEditorWindow(tk.Toplevel):
    """Visual editor for creating email signatures with editable text on canvas"""
    def __init__(self, parent, text_widget_callback):
        super().__init__(parent)
        self.title("Email Signature Visual Editor")
        self.geometry("1000x750")
        self.text_widget = text_widget_callback
        self.elements = []
        self.selected_element = None
        self.drag_data = {"x": 0, "y": 0, "item": None}
        
        # Main container
        main_frame = ttk.Frame(self, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Toolbar
        toolbar = ttk.Frame(main_frame)
        toolbar.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Button(
            toolbar, text=" Add Text Box", command=self.add_editable_text
        ).pack(side=tk.LEFT, padx=2)
        
        ttk.Button(
            toolbar, text=" Add Image", command=self.add_image_element
        ).pack(side=tk.LEFT, padx=2)
        
        ttk.Button(
            toolbar, text=" Quick Contact", command=self.add_contact_template
        ).pack(side=tk.LEFT, padx=2)
        
        ttk.Label(toolbar, text=" | ").pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            toolbar, text=" Edit Selected", command=self.edit_selected_text
        ).pack(side=tk.LEFT, padx=2)
        
        ttk.Button(
            toolbar, text=" Delete Selected", command=self.delete_selected
        ).pack(side=tk.LEFT, padx=2)
        
        ttk.Label(toolbar, text=" | ").pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            toolbar, text=" Generate HTML", command=self.generate_html
        ).pack(side=tk.LEFT, padx=2)
        
        ttk.Button(
            toolbar, text=" Clear All", command=self.clear_all
        ).pack(side=tk.LEFT, padx=2)
        
        # Formatting toolbar
        format_toolbar = ttk.Frame(main_frame)
        format_toolbar.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(format_toolbar, text="Format Selected:").pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            format_toolbar, text="Bold (B)", command=lambda: self.apply_format("bold")
        ).pack(side=tk.LEFT, padx=2)
        
        ttk.Button(
            format_toolbar, text="Italic (I)", command=lambda: self.apply_format("italic")
        ).pack(side=tk.LEFT, padx=2)
        
        ttk.Label(format_toolbar, text="Size:").pack(side=tk.LEFT, padx=5)
        
        self.size_var = tk.StringVar(value="12")
        size_combo = ttk.Combobox(
            format_toolbar, textvariable=self.size_var, width=5,
            values=["8", "10", "11", "12", "14", "16", "18", "20", "24"]
        )
        size_combo.pack(side=tk.LEFT, padx=2)
        size_combo.bind("<<ComboboxSelected>>", lambda e: self.apply_size())
        
        # Canvas for design
        canvas_frame = ttk.LabelFrame(main_frame, text="Design Area (Click elements to select, drag to move)", padding=10)
        canvas_frame.pack(fill=tk.BOTH, expand=True)
        
        self.canvas = tk.Canvas(canvas_frame, bg="white", height=450)
        self.canvas.pack(fill=tk.BOTH, expand=True)
        
        # Bind canvas events
        self.canvas.bind("<Button-1>", self.on_canvas_click)
        self.canvas.bind("<B1-Motion>", self.on_canvas_drag)
        self.canvas.bind("<ButtonRelease-1>", self.on_canvas_release)
        self.canvas.bind("<Double-Button-1>", self.on_canvas_double_click)
        
        # Instructions
        instructions = ttk.Label(
            main_frame,
            text=" Double-click text to edit | Drag to move | Click to select | Generate HTML when done",
            foreground="blue",
            font=("Arial", 9)
        )
        instructions.pack(pady=10)
        
    def add_editable_text(self):
        """Add an editable text box to canvas"""
        # Create text input dialog
        text = simpledialog.askstring(
            "Add Text",
            "Enter text (you can edit it later by double-clicking):",
            parent=self
        )
        
        if text:
            x = 50
            y = 50 + len(self.elements) * 40
            
            text_id = self.canvas.create_text(
                x, y, text=text, anchor="w",
                font=("Arial", 12),
                fill="black",
                tags="draggable"
            )
            
            # Create selection box (hidden initially)
            bbox = self.canvas.bbox(text_id)
            rect_id = self.canvas.create_rectangle(
                bbox[0]-2, bbox[1]-2, bbox[2]+2, bbox[3]+2,
                outline="blue",
                dash=(2, 2),
                width=2,
                tags="selection",
                state="hidden"
            )
            
            element = {
                "type": "text",
                "content": text,
                "id": text_id,
                "rect_id": rect_id,
                "x": x,
                "y": y,
                "size": 12,
                "style": "normal",
                "color": "black"
            }
            self.elements.append(element)
            
    def add_image_element(self):
        """Add image element"""
        from tkinter import filedialog
        import base64
        from PIL import Image, ImageTk
        
        file_path = filedialog.askopenfilename(
            title="Select Image (Company Logo)",
            filetypes=[("Image files", "*.png *.jpg *.jpeg *.gif"), ("All files", "*.*")],
            parent=self
        )
        
        if file_path:
            try:
                # Load and display image on canvas
                pil_image = Image.open(file_path)
                
                # Resize if too large
                max_width, max_height = 300, 200
                pil_image.thumbnail((max_width, max_height), Image.Resampling.LANCZOS)
                
                # Convert to PhotoImage for canvas
                photo = ImageTk.PhotoImage(pil_image)
                
                # Convert to base64 for HTML
                with open(file_path, "rb") as f:
                    img_data = base64.b64encode(f.read()).decode()
                
                # Determine image type
                ext = file_path.lower().split('.')[-1]
                mime_type = f"image/{ext if ext in ['png', 'jpg', 'jpeg', 'gif'] else 'png'}"
                
                x = 50
                y = 50 + len(self.elements) * 120
                
                # Create image on canvas
                img_id = self.canvas.create_image(
                    x, y, image=photo, anchor="nw", tags="draggable"
                )
                
                # Create selection box (hidden initially)
                bbox = self.canvas.bbox(img_id)
                rect_id = self.canvas.create_rectangle(
                    bbox[0]-2, bbox[1]-2, bbox[2]+2, bbox[3]+2,
                    outline="blue",
                    dash=(2, 2),
                    width=2,
                    tags="selection",
                    state="hidden"
                )
                
                element = {
                    "type": "image",
                    "id": img_id,
                    "rect_id": rect_id,
                    "photo": photo,  # Keep reference
                    "data": img_data,
                    "mime": mime_type,
                    "x": x,
                    "y": y,
                    "width": pil_image.width,
                    "height": pil_image.height
                }
                self.elements.append(element)
                
            except Exception as e:
                messagebox.showerror("❌ Error", f"Failed to load image:\n{e}", parent=self)
                
    def add_contact_template(self):
        """Add contact information template"""
        name = simpledialog.askstring("Name", "Enter your name:", parent=self)
        if not name:
            return
            
        title = simpledialog.askstring("Title", "Enter your title (optional):", parent=self)
        phone = simpledialog.askstring("Phone", "Enter phone number (optional):", parent=self)
        email = simpledialog.askstring("Email", "Enter email address (optional):", parent=self)
        
        # Create contact block with proper spacing
        y_start = 50 + len(self.elements) * 30
        
        contact_items = []
        if name:
            contact_items.append((name, 14, "bold"))
        if title:
            contact_items.append((title, 11, "italic"))
        if phone:
            contact_items.append((f" {phone}", 10, "normal"))
        if email:
            contact_items.append((f" {email}", 10, "normal"))
        
        for text, size, style in contact_items:
            font_tuple = ("Arial", size, style) if style != "normal" else ("Arial", size)
            text_id = self.canvas.create_text(
                50, y_start, text=text, anchor="w",
                font=font_tuple,
                tags="draggable"
            )
            
            bbox = self.canvas.bbox(text_id)
            rect_id = self.canvas.create_rectangle(
                bbox[0]-2, bbox[1]-2, bbox[2]+2, bbox[3]+2,
                outline="blue",
                dash=(2, 2),
                width=2,
                tags="selection",
                state="hidden"
            )
            
            self.elements.append({
                "type": "text",
                "content": text,
                "id": text_id,
                "rect_id": rect_id,
                "x": 50,
                "y": y_start,
                "size": size,
                "style": style,
                "color": "black"
            })
            y_start += size + 10
    
    def on_canvas_click(self, event):
        """Handle canvas click to select elements"""
        # Find clicked item
        item = self.canvas.find_closest(event.x, event.y)[0]
        
        # Hide all selection boxes
        for elem in self.elements:
            self.canvas.itemconfig(elem.get("rect_id"), state="hidden")
        
        # Find and select the element
        self.selected_element = None
        for elem in self.elements:
            if elem["id"] == item:
                self.selected_element = elem
                # Show selection box
                self.canvas.itemconfig(elem["rect_id"], state="normal")
                # Update size dropdown
                if elem["type"] == "text":
                    self.size_var.set(str(elem.get("size", 12)))
                break
        
        # Store drag data
        self.drag_data["x"] = event.x
        self.drag_data["y"] = event.y
        self.drag_data["item"] = item
    
    def on_canvas_drag(self, event):
        """Handle dragging elements"""
        if self.drag_data["item"]:
            # Calculate delta
            dx = event.x - self.drag_data["x"]
            dy = event.y - self.drag_data["y"]
            
            # Move the item
            self.canvas.move(self.drag_data["item"], dx, dy)
            
            # Update stored position
            if self.selected_element:
                self.selected_element["x"] += dx
                self.selected_element["y"] += dy
                
                # Move selection box too
                self.canvas.move(self.selected_element["rect_id"], dx, dy)
            
            # Update drag data
            self.drag_data["x"] = event.x
            self.drag_data["y"] = event.y
    
    def on_canvas_release(self, event):
        """Handle mouse release"""
        self.drag_data["item"] = None
    
    def on_canvas_double_click(self, event):
        """Handle double-click to edit text"""
        if self.selected_element and self.selected_element["type"] == "text":
            self.edit_selected_text()
    
    def edit_selected_text(self):
        """Edit the selected text element"""
        if not self.selected_element or self.selected_element["type"] != "text":
            messagebox.showinfo("No Selection", "Please select a text element first.", parent=self)
            return
        
        current_text = self.selected_element["content"]
        new_text = simpledialog.askstring(
            "Edit Text",
            "Enter new text:",
            initialvalue=current_text,
            parent=self
        )
        
        if new_text:
            # Update canvas text
            self.canvas.itemconfig(self.selected_element["id"], text=new_text)
            self.selected_element["content"] = new_text
            
            # Update selection box
            bbox = self.canvas.bbox(self.selected_element["id"])
            self.canvas.coords(
                self.selected_element["rect_id"],
                bbox[0]-2, bbox[1]-2, bbox[2]+2, bbox[3]+2
            )
    
    def delete_selected(self):
        """Delete the selected element"""
        if not self.selected_element:
            messagebox.showinfo("No Selection", "Please select an element first.", parent=self)
            return
        
        if messagebox.askyesno("Confirm Delete", "Delete selected element?", parent=self):
            # Delete from canvas
            self.canvas.delete(self.selected_element["id"])
            self.canvas.delete(self.selected_element["rect_id"])
            
            # Remove from elements list
            self.elements.remove(self.selected_element)
            self.selected_element = None
    
    def apply_format(self, style):
        """Apply formatting to selected text"""
        if not self.selected_element or self.selected_element["type"] != "text":
            messagebox.showinfo("No Selection", "Please select a text element first.", parent=self)
            return
        
        self.selected_element["style"] = style
        size = self.selected_element.get("size", 12)
        
        if style == "bold":
            font_tuple = ("Arial", size, "bold")
        elif style == "italic":
            font_tuple = ("Arial", size, "italic")
        else:
            font_tuple = ("Arial", size)
        
        self.canvas.itemconfig(self.selected_element["id"], font=font_tuple)
    
    def apply_size(self):
        """Apply size to selected text"""
        if not self.selected_element or self.selected_element["type"] != "text":
            return
        
        try:
            size = int(self.size_var.get())
            self.selected_element["size"] = size
            
            style = self.selected_element.get("style", "normal")
            if style == "bold":
                font_tuple = ("Arial", size, "bold")
            elif style == "italic":
                font_tuple = ("Arial", size, "italic")
            else:
                font_tuple = ("Arial", size)
            
            self.canvas.itemconfig(self.selected_element["id"], font=font_tuple)
            
            # Update selection box
            bbox = self.canvas.bbox(self.selected_element["id"])
            self.canvas.coords(
                self.selected_element["rect_id"],
                bbox[0]-2, bbox[1]-2, bbox[2]+2, bbox[3]+2
            )
        except ValueError:
            pass
            
    def generate_html(self):
        """Generate HTML from canvas elements"""
        if not self.elements:
            messagebox.showwarning("No Elements", "Please add some elements first.", parent=self)
            return
            
        html_parts = ['<div style="font-family: Arial, sans-serif; padding: 10px;">']
        
        # Sort elements by Y position for proper HTML order
        sorted_elements = sorted(self.elements, key=lambda e: e["y"])
        
        for elem in sorted_elements:
            if elem["type"] == "text":
                style_attrs = []
                size = elem.get("size", 12)
                style_attrs.append(f"font-size: {size}px")
                
                if elem.get("style") == "bold":
                    style_attrs.append("font-weight: bold")
                elif elem.get("style") == "italic":
                    style_attrs.append("font-style: italic")
                
                color = elem.get("color", "black")
                if color != "black":
                    style_attrs.append(f"color: {color}")
                    
                style = "; ".join(style_attrs)
                html_parts.append(f'<p style="{style}; margin: 5px 0;">{elem["content"]}</p>')
                
            elif elem["type"] == "image":
                width = elem.get("width", 100)
                height = elem.get("height", 50)
                html_parts.append(
                    f'<img src="data:{elem["mime"]};base64,{elem["data"]}" '
                    f'width="{width}" height="{height}" style="display: block; margin: 10px 0;" />'
                )
                
        html_parts.append('</div>')
        html_content = "\n".join(html_parts)
        
        # Update parent text widget
        self.text_widget.delete("1.0", tk.END)
        self.text_widget.insert("1.0", html_content)
        
        messagebox.showinfo("✅ Success",
            "HTML generated! Close this window and save your signature.",
            parent=self
        )
        
    def clear_all(self):
        """Clear all elements"""
        if messagebox.askyesno("❓ Confirm", "Clear all elements from canvas?", parent=self):
            self.canvas.delete("all")
            self.elements = []
            self.selected_element = None

        messagebox.showinfo("✅ Success", "HTML generated! Close this window and save your signature.", parent=self)
        
    def clear_all(self):
        """Clear all elements"""
        if messagebox.askyesno("❓ Confirm", "Clear all elements?", parent=self):
            self.canvas.delete("all")
            self.elements = []


# ==============================================================================
# 4. GUI APPLICATION (Tkinter)
# ==============================================================================


class RemoteOperationsApp:
    def __init__(self, root, data_manager):
        self.root = root
        self.dm = data_manager
        self.root.title("Standalone Operations Tool v3.1 (SQLite)")

        # FURTHER REDUCED window size to fit content
        self.root.geometry("850x800")  # Reduced from 800x650

        # Set minimum window size
        self.root.minsize(700, 550)

        # Main scrollable frame - REMOVED, not needed for this compact layout
        main_frame = ttk.Frame(root)
        main_frame.pack(fill="both", expand=True)

        # Header with quick stats
        self.create_header(main_frame)

        # Main dashboard with large action cards
        self.create_dashboard_cards(main_frame)

        # Log Section - REDUCED height
        log_frame = ttk.LabelFrame(main_frame, text="Activity Log", padding=5)
        log_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        self.log_output = scrolledtext.ScrolledText(
            log_frame, wrap=tk.WORD, height=5
        )  # Reduced from 8
        self.log_output.pack(fill=tk.BOTH, expand=True)

    def create_header(self, parent):
        """Create header with company name and quick stats"""
        header_frame = ttk.Frame(parent, padding=5)  # Reduced padding from 10
        header_frame.pack(fill=tk.X, padx=10, pady=(5, 0))  # Reduced padding

        title_label = ttk.Label(
            header_frame,
            text="Operations Management System",
            font=("Helvetica", 14, "bold"),
        )  # Reduced from 16
        title_label.pack(side=tk.LEFT)

        stats_frame = ttk.Frame(header_frame)
        stats_frame.pack(side=tk.RIGHT)

        self.stats_label = ttk.Label(
            stats_frame, text="Loading...", font=("Helvetica", 8)
        )  # Reduced from 9
        self.stats_label.pack()

        self.update_header_stats()

    def update_header_stats(self):
        """Update header with current counts"""
        try:
            open_orders = len(self.dm.get_all_open_orders())
            pending_pdfs = len(self.dm.get_pending_po_data())
            vendors = len(self.dm.get_all_vendors())

            stats_text = f"Open Orders: {open_orders} | Pending PDFs: {pending_pdfs} | Vendors: {vendors}"
            self.stats_label.config(text=stats_text)
        except Exception as e:
            self.stats_label.config(text=f"Error loading stats")

    def create_dashboard_cards(self, parent):
        """Create main dashboard with COMPACT action cards"""
        dashboard_frame = ttk.Frame(parent, padding=5)
        dashboard_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=2)

        # Configure grid to be more compact - UPDATED to 4 rows
        for i in range(4):
            dashboard_frame.rowconfigure(i, weight=1)
        dashboard_frame.columnconfigure(0, weight=1)
        dashboard_frame.columnconfigure(1, weight=1)

        # Card 1: Purchase Orders
        self.create_card(
            dashboard_frame,
            title="Purchase Orders",
            description="Create, manage, and send purchase orders",
            icon="📋",
            command=self.open_po_management,
            row=0,
            column=0,
        )

        # Card 2: Order Book
        self.create_card(
            dashboard_frame,
            title="Order Book",
            description="View and manage open orders",
            icon="📚",
            command=self.open_order_book_window,
            row=0,
            column=1,
        )

        # Card 3: Suppliers
        self.create_card(
            dashboard_frame,
            title="Supplier Management",
            description="Manage vendor master data",
            icon="🏢",
            command=self.open_vendor_manager,
            row=1,
            column=0,
        )

        # Card 4: MRP Planning
        self.create_card(
            dashboard_frame,
            title="MRP Planning",
            description="Material Requirements Planning & Requisitions",
            icon="📏",
            command=self.open_mrp_planning,
            row=1,
            column=1,
        )

        # Card 5: Forecasting
        self.create_card(
            dashboard_frame,
            title="Forecasting & Planning",
            description="Manage forecasts and requisitions",
            icon="📆",
            command=self.open_forecast_management,
            row=2,
            column=0,
        )

        # Card 6: Reschedule
        self.create_card(
            dashboard_frame,
            title="Reschedule Management",
            description="Generate and send reschedule files",
            icon="📦",
            command=self.open_reschedule_config,
            row=2,
            column=1,
        )

        # Card 7: Settings
        self.create_card(
            dashboard_frame,
            title="System Settings",
            description="Configure email, SMTP, and company info",
            icon="📏",
            command=self.open_settings_window,
            row=3,
            column=0,
        )

        # Card 8: Email Scanner (add this to your create_dashboard_cards method)
        self.create_card(
            dashboard_frame,
            title="Email Scanner",
            description="Scan emails for order confirmations",
            icon="📧",
            command=self.open_email_scanner,
            row=3,
            column=1,
        )

    def create_card(self, parent, title, description, icon, command, row, column):
        """Create a dashboard card button - VERY COMPACT VERSION"""
        # Outer frame with border - MINIMAL padding
        card_frame = ttk.LabelFrame(parent, text="", relief="raised", borderwidth=2)
        card_frame.grid(
            row=row, column=column, padx=5, pady=4, sticky="nsew"
        )  # Further reduced

        # Inner frame for content - MINIMAL padding
        inner_frame = ttk.Frame(card_frame)
        inner_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=8)  # Further reduced

        # Icon - SMALLER
        icon_label = ttk.Label(
            inner_frame, text=icon, font=("Helvetica", 24)
        )  # Reduced from 28
        icon_label.pack(pady=(0, 4))  # Reduced spacing

        # Title - SMALLER
        title_label = ttk.Label(
            inner_frame, text=title, font=("Helvetica", 10, "bold")
        )  # Reduced from 11
        title_label.pack(pady=2)  # Reduced spacing

        # Description - SMALLER
        desc_label = ttk.Label(
            inner_frame, text=description, font=("Helvetica", 8)
        )  # Keep at 8
        desc_label.pack(pady=2)  # Reduced spacing

        # Button - SMALLER
        btn = ttk.Button(inner_frame, text="Open", command=command)
        btn.pack(pady=(4, 0))  # Reduced from 8

    def log(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        formatted_message = f"[{timestamp}] {message}"
        self.root.after(
            0, lambda: self.log_output.insert(tk.END, formatted_message + "\n")
        )
        self.root.after(0, lambda: self.log_output.see(tk.END))

    def get_selected_send_method(self):
        return "Outlook" if OUTLOOK_AVAILABLE else "SMTP"

    # Window openers remain the same
    def open_po_management(self):
        try:
            POManagementWindow(
                self.root, self.log, self.dm, self.get_selected_send_method
            )
        except Exception as e:
            self.log(f"✗ ERROR: Failed to open PO Management: {e}")
            import traceback

            traceback.print_exc()

    def open_settings_window(self):
        try:
            SettingsWindow(self.root, self.log, self.dm)
        except Exception as e:
            self.log(f"✗ ERROR: Failed to open Settings: {e}")
            import traceback

            traceback.print_exc()

    def open_vendor_manager(self):
        try:
            VendorManagerWindow(self.root, self.log, self.dm)
        except Exception as e:
            self.log(f"✗ ERROR: Failed to open Vendor Manager: {e}")
            import traceback

            traceback.print_exc()

    def open_order_book_window(self):
        try:
            OpenOrderBookWindow(self.root, self.log, self.dm)
        except Exception as e:
            self.log(f"✗ ERROR: Failed to open Order Book: {e}")
            import traceback

            traceback.print_exc()

    def open_reschedule_config(self):
        try:
            RescheduleConfigWindow(self.root, self.log, self.dm)
        except Exception as e:
            self.log(f"✗ ERROR: Failed to open Reschedule Config: {e}")
            import traceback

            traceback.print_exc()

    def open_forecast_management(self):
        try:
            if not hasattr(self, "forecast_manager"):
                self.forecast_manager = ForecastDataManager(self.dm.db)
            ForecastManagementWindow(
                self.root, self.log, self.dm, self.forecast_manager
            )
        except Exception as e:
            self.log(f"✗ ERROR: Failed to open Forecast Management: {e}")
            import traceback

            traceback.print_exc()

    def open_mrp_planning(self):
        try:
            MRPPlanningWindow(self.root, self.log, self.dm)
        except Exception as e:
            self.log(f"✗ ERROR: Failed to open MRP Planning: {e}")
            import traceback

            traceback.print_exc()

    def open_email_scanner(self):
        """Open email confirmation scanner window"""
        try:
            EmailConfirmationScannerWindow(self.root, self.log, self.dm)
        except Exception as e:
            self.log(f"✗ ERROR: Failed to open Email Scanner: {e}")
            import traceback

            traceback.print_exc()


# ==============================================================================
# 5. GUI CHILD WINDOWS
# ==============================================================================


class VendorManagerWindow(tk.Toplevel):
    def __init__(self, parent, log_callback, data_manager):
        super().__init__(parent)
        self.title("Vendor Master Data Management")
        self.geometry("1200x750")
        self.log = log_callback
        self.dm = data_manager
        self.vendors_data = []
        self.current_vendor_name = None

        paned_window = ttk.PanedWindow(self, orient=tk.HORIZONTAL)
        paned_window.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        list_frame = ttk.LabelFrame(paned_window, text="Vendor List", padding=10)
        paned_window.add(list_frame, weight=1)

        # Search and batch upload section
        top_controls_frame = ttk.Frame(list_frame)
        top_controls_frame.pack(fill=tk.X, pady=(0, 10))

        # Search functionality
        search_frame = ttk.Frame(top_controls_frame)
        search_frame.pack(fill=tk.X, pady=(0, 5))
        self.search_var = tk.StringVar()
        self.search_var.trace("w", self.filter_vendors)
        ttk.Label(search_frame, text="Search:").pack(side=tk.LEFT, padx=(0, 5))
        ttk.Entry(search_frame, textvariable=self.search_var).pack(
            side=tk.LEFT, fill=tk.X, expand=True
        )

        # Batch upload section
        batch_frame = ttk.Frame(top_controls_frame)
        batch_frame.pack(fill=tk.X, pady=(5, 0))
        ttk.Button(
            batch_frame,
            text="Batch Upload from Excel...",
            command=self.batch_upload_vendors,
        ).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(
            batch_frame, text="Export to Excel...", command=self.export_vendors_to_excel
        ).pack(side=tk.LEFT)

        self.vendor_listbox = Listbox(list_frame, exportselection=False)
        self.vendor_listbox.pack(fill=tk.BOTH, expand=True)
        self.vendor_listbox.bind("<<ListboxSelect>>", self.on_vendor_select)

        form_frame = ttk.LabelFrame(paned_window, text="Vendor Details", padding=10)
        paned_window.add(form_frame, weight=3)
        form_frame.columnconfigure(1, weight=1)

        self.fields = {}
        ## --- MODIFIED --- ##
        # Added "Secondary Transport Days" to the form fields
        field_labels = [
            "Display Name",
            "Emails",
            "Contact Person",
            "Transport Days",
            "Secondary Transport Days",
            "Delivery Terms",
            "Payment Terms",
        ]
        row_num = 0
        for i, label in enumerate(field_labels):
            ttk.Label(form_frame, text=f"{label}:").grid(
                row=i, column=0, sticky="w", padx=5, pady=5
            )
            self.fields[label] = tk.StringVar()
            ttk.Entry(form_frame, textvariable=self.fields[label]).grid(
                row=i, column=1, sticky="ew", padx=5, pady=5
            )
            row_num = i + 1

        ttk.Label(form_frame, text="Address:").grid(
            row=row_num, column=0, sticky="nw", padx=5, pady=5
        )
        self.address_text = tk.Text(form_frame, height=4)
        self.address_text.grid(row=row_num, column=1, sticky="ew", padx=5, pady=5)

        ttk.Label(form_frame, text="API Key:").grid(
            row=row_num + 1, column=0, sticky="nw", padx=5, pady=5
        )
        api_key_frame = ttk.Frame(form_frame)
        api_key_frame.grid(row=row_num + 1, column=1, sticky="ew")
        self.api_key_var = tk.StringVar()
        ttk.Entry(api_key_frame, textvariable=self.api_key_var, state="readonly").pack(
            side=tk.LEFT, fill=tk.X, expand=True
        )
        ttk.Button(
            api_key_frame, text="Generate New", command=self.generate_api_key
        ).pack(side=tk.LEFT)

        button_frame = ttk.Frame(self)
        button_frame.pack(fill=tk.X, padx=10, pady=10)
        ttk.Button(button_frame, text="➕ Add New", command=self.prepare_for_new).pack(
            side=tk.LEFT
        )
        ttk.Button(button_frame, text="🗑️ Delete", command=self.delete_vendor).pack(
            side=tk.LEFT, padx=5
        )
        ttk.Button(button_frame, text="💾 Save Changes", command=self.save_vendor).pack(
            side=tk.RIGHT
        )

        self.refresh_vendor_list()

    def batch_upload_vendors(self):
        """Open file dialog and process Excel file for batch vendor upload"""
        file_path = filedialog.askopenfilename(
            title="Select Vendor Excel File",
            filetypes=[("Excel Files", "*.xlsx;*.xls"), ("All Files", "*.*")],
        )

        if not file_path:
            return

        try:
            # Read Excel file
            df = pd.read_excel(file_path, engine=None)
            df.columns = [str(c).strip() for c in df.columns]

            # Look for vendor and email columns (flexible column names)
            vendor_col = None
            email_col = None
            ## --- ADDED --- ##
            # Look for secondary transport days column
            sec_transport_col = None

            # Possible column names for vendor
            vendor_possibilities = [
                "Vendor",
                "vendor",
                "Supplier",
                "supplier",
                "Name",
                "name",
                "Company",
                "company",
            ]
            for col in vendor_possibilities:
                if col in df.columns:
                    vendor_col = col
                    break

            # Possible column names for email
            email_possibilities = [
                "e-mail",
                "email",
                "Email",
                "E-mail",
                "emails",
                "Emails",
                "Email Address",
                "email_address",
            ]
            for col in email_possibilities:
                if col in df.columns:
                    email_col = col
                    break

            ## --- ADDED --- ##
            # Possible column names for secondary transport days
            sec_transport_possibilities = [
                "Secondary Transport Days",
                "secondary_transport_days",
                "Fast Transport",
                "transport_days_secondary",
            ]
            for col in sec_transport_possibilities:
                if col in df.columns:
                    sec_transport_col = col
                    break

            if not vendor_col or not email_col:
                available_cols = ", ".join(df.columns.tolist())
                messagebox.showerror(
                    "Column Not Found",
                    f"Could not find vendor and email columns.\n\nAvailable columns: {available_cols}\n\nExpected columns like: 'Vendor', 'Supplier', 'e-mail', 'Email'",
                    parent=self,
                )
                return

            # Show preview and confirmation dialog
            preview_result = self.show_upload_preview(
                df, vendor_col, email_col, sec_transport_col
            )
            if not preview_result:
                return

            # Process the upload
            self.process_vendor_upload(df, vendor_col, email_col, sec_transport_col)

        except Exception as e:
            messagebox.showerror(
                "Upload Error", f"Failed to process Excel file:\n{str(e)}", parent=self
            )
            self.log(f"✗ ERROR: Batch upload failed: {e}")

    def show_upload_preview(self, df, vendor_col, email_col, sec_transport_col):
        """Show preview dialog for batch upload"""
        preview_window = tk.Toplevel(self)
        preview_window.title("Upload Preview")
        preview_window.geometry("800x600")
        preview_window.transient(self)
        preview_window.grab_set()

        # Info frame
        info_frame = ttk.Frame(preview_window, padding=10)
        info_frame.pack(fill=tk.X)

        ttk.Label(
            info_frame,
            text=f"Found {len(df)} vendors to import",
            font=("Helvetica", 12, "bold"),
        ).pack()
        info_text = f"Vendor Column: '{vendor_col}' | Email Column: '{email_col}'"
        if sec_transport_col:
            info_text += f" | Secondary Transport: '{sec_transport_col}'"
        ttk.Label(info_frame, text=info_text).pack()

        # Preview data
        preview_frame = ttk.LabelFrame(
            preview_window, text="Preview (First 20 rows)", padding=10
        )
        preview_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # Create treeview for preview
        ## --- MODIFIED --- ##
        preview_cols = ("vendor", "email", "sec_transport")
        preview_tree = ttk.Treeview(
            preview_frame, columns=preview_cols, show="headings", height=15
        )
        preview_tree.heading("vendor", text="Vendor Name")
        preview_tree.heading("email", text="Email Address(es)")
        preview_tree.heading("sec_transport", text="Sec. Transport Days")
        preview_tree.column("vendor", width=300)
        preview_tree.column("email", width=350)
        preview_tree.column("sec_transport", width=120)

        # Add scrollbar
        preview_scroll = ttk.Scrollbar(
            preview_frame, orient="vertical", command=preview_tree.yview
        )
        preview_tree.configure(yscrollcommand=preview_scroll.set)
        preview_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        preview_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # Populate preview (first 20 rows)
        for i, (_, row) in enumerate(df.head(20).iterrows()):
            vendor_name = (
                str(row[vendor_col]).strip() if pd.notna(row[vendor_col]) else ""
            )
            email_addr = str(row[email_col]).strip() if pd.notna(row[email_col]) else ""
            sec_transport = (
                str(row[sec_transport_col]).strip()
                if sec_transport_col and pd.notna(row[sec_transport_col])
                else ""
            )
            if vendor_name:  # Only show rows with vendor names
                preview_tree.insert(
                    "", "end", values=(vendor_name, email_addr, sec_transport)
                )

        # Options frame
        options_frame = ttk.LabelFrame(
            preview_window, text="Import Options", padding=10
        )
        options_frame.pack(fill=tk.X, padx=10, pady=5)

        update_existing_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            options_frame,
            text="Update existing vendors (if unchecked, existing vendors will be skipped)",
            variable=update_existing_var,
        ).pack(anchor="w")

        # Button frame
        button_frame = ttk.Frame(preview_window, padding=10)
        button_frame.pack(fill=tk.X)

        result = {"confirmed": False, "update_existing": True}

        def confirm_upload():
            result["confirmed"] = True
            result["update_existing"] = update_existing_var.get()
            preview_window.destroy()

        def cancel_upload():
            result["confirmed"] = False
            preview_window.destroy()

        ttk.Button(button_frame, text="Cancel", command=cancel_upload).pack(
            side=tk.RIGHT, padx=(5, 0)
        )
        ttk.Button(button_frame, text="Import Vendors", command=confirm_upload).pack(
            side=tk.RIGHT
        )

        # Wait for user decision
        preview_window.wait_window()
        return result if result["confirmed"] else None

    def process_vendor_upload(self, df, vendor_col, email_col, sec_transport_col):
        """Process the actual vendor upload"""
        update_existing = True  # Default behavior

        created_count = 0
        updated_count = 0
        skipped_count = 0
        error_count = 0

        # Get existing vendors for comparison
        existing_vendors = {
            v["display_name"].lower(): v for v in self.dm.get_all_vendors()
        }

        for _, row in df.iterrows():
            try:
                vendor_name = (
                    str(row[vendor_col]).strip() if pd.notna(row[vendor_col]) else ""
                )
                email_addr = (
                    str(row[email_col]).strip() if pd.notna(row[email_col]) else ""
                )

                if not vendor_name:
                    continue  # Skip empty vendor names

                # Clean up email addresses (remove extra spaces around semicolons)
                if email_addr:
                    email_parts = [
                        email.strip()
                        for email in email_addr.split(";")
                        if email.strip()
                    ]
                    email_addr = "; ".join(email_parts)

                ## --- ADDED --- ##
                # Safely get secondary transport days value
                sec_transport_val = 0
                if sec_transport_col and pd.notna(row.get(sec_transport_col)):
                    try:
                        sec_transport_val = int(row[sec_transport_col])
                    except (ValueError, TypeError):
                        sec_transport_val = 0

                vendor_data = {
                    "display_name": vendor_name,
                    "emails": email_addr,
                    "contact_person": "",
                    "address": "",
                    "transport_days": 0,  # Default, can be extended to support more columns
                    "transport_days_secondary": sec_transport_val,
                    "delivery_terms": "",
                    "payment_terms": "",
                }

                # Check if vendor already exists
                if vendor_name.lower() in existing_vendors:
                    if update_existing:
                        # Update existing vendor
                        original_name = existing_vendors[vendor_name.lower()][
                            "display_name"
                        ]
                        self.dm.update_vendor(original_name, vendor_data)
                        updated_count += 1
                        self.log(f"ℹ️ INFO: Updated vendor '{vendor_name}'")
                    else:
                        skipped_count += 1
                        self.log(f"ℹ️ INFO: Skipped existing vendor '{vendor_name}'")
                else:
                    # Create new vendor
                    self.dm.create_vendor(vendor_data)
                    created_count += 1
                    self.log(f"ℹ️ INFO: Created new vendor '{vendor_name}'")

            except Exception as e:
                error_count += 1
                self.log(f"✗ ERROR: Failed to process vendor '{vendor_name}': {e}")

        # Show results
        result_message = f"Batch Upload Complete:\n\n"
        result_message += f"Created: {created_count} vendors\n"
        result_message += f"Updated: {updated_count} vendors\n"
        result_message += f"Skipped: {skipped_count} vendors\n"
        result_message += f"Errors: {error_count} vendors"

        messagebox.showinfo("Upload Results", result_message, parent=self)
        self.log(
            f"✓ SUCCESS: Batch upload completed - Created: {created_count}, Updated: {updated_count}, Skipped: {skipped_count}, Errors: {error_count}"
        )

        # Refresh the vendor list
        self.refresh_vendor_list()

    def export_vendors_to_excel(self):
        """Export current vendors to Excel file"""
        file_path = filedialog.asksaveasfilename(
            title="Save Vendors to Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
        )

        if not file_path:
            return

        try:
            vendors = self.dm.get_all_vendors()

            # Prepare data for export
            export_data = []
            for vendor in vendors:
                ## --- MODIFIED --- ##
                # Added "Secondary Transport Days" to export
                export_data.append(
                    {
                        "Vendor": vendor["display_name"],
                        "e-mail": vendor.get("emails", ""),
                        "Contact Person": vendor.get("contact_person", ""),
                        "Address": vendor.get("address", ""),
                        "Transport Days": vendor.get("transport_days", 0),
                        "Secondary Transport Days": vendor.get(
                            "transport_days_secondary", 0
                        ),
                        "Delivery Terms": vendor.get("delivery_terms", ""),
                        "Payment Terms": vendor.get("payment_terms", ""),
                        "API Key": vendor.get("api_key", ""),
                    }
                )

            # Create Excel file
            df = pd.DataFrame(export_data)

            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name="Vendors", index=False)

                # Auto-adjust column widths
                worksheet = writer.sheets["Vendors"]
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except Exception:  # TODO: Add proper error handling
                            pass  # TODO: Add proper error handling
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            messagebox.showinfo(
                "Export Complete",
                f"Exported {len(vendors)} vendors to:\n{file_path}",
                parent=self,
            )
            self.log(f"✓ SUCCESS: Exported {len(vendors)} vendors to Excel file")

        except Exception as e:
            messagebox.showerror(
                "Export Error", f"Failed to export vendors:\n{str(e)}", parent=self
            )
            self.log(f"✗ ERROR: Export failed: {e}")

    def refresh_vendor_list(self):
        self.vendors_data = self.dm.get_all_vendors()
        self.filter_vendors()
        self.clear_form()

    def filter_vendors(self, *args):
        search_term = self.search_var.get().lower()
        self.vendor_listbox.delete(0, tk.END)
        for vendor in self.vendors_data:
            if search_term in vendor["display_name"].lower():
                self.vendor_listbox.insert(tk.END, vendor["display_name"])

    def on_vendor_select(self, event=None):
        selected = self.vendor_listbox.curselection()
        if not selected:
            return
        name = self.vendor_listbox.get(selected[0])
        self.current_vendor_name = name

        vendor = next((v for v in self.vendors_data if v["display_name"] == name), None)
        if vendor:
            # Populate form fields
            self.fields["Display Name"].set(vendor.get("display_name", ""))
            self.fields["Emails"].set(vendor.get("emails", ""))
            self.fields["Contact Person"].set(vendor.get("contact_person", ""))
            self.fields["Transport Days"].set(vendor.get("transport_days", 0))
            self.fields["Secondary Transport Days"].set(
                vendor.get("transport_days_secondary", 0)
            )
            self.fields["Delivery Terms"].set(vendor.get("delivery_terms", ""))
            self.fields["Payment Terms"].set(vendor.get("payment_terms", ""))

            # Fix for address_text.insert error
            self.address_text.delete("1.0", tk.END)

            # Get address and ensure it's a string, not None
            address = vendor.get("address", "")
            if address is None:
                address = ""

            # Insert address (now guaranteed to be a string)
            self.address_text.insert("1.0", str(address))

            # Set API key
            self.api_key_var.set(vendor.get("api_key", "N/A") or "N/A")

    def save_vendor(self):
        data = {
            "display_name": self.fields["Display Name"].get().strip(),
            "emails": self.fields["Emails"].get().strip(),
            "contact_person": self.fields["Contact Person"].get().strip(),
            "address": self.address_text.get("1.0", "end-1c").strip(),
            "delivery_terms": self.fields["Delivery Terms"].get().strip(),
            "payment_terms": self.fields["Payment Terms"].get().strip(),
        }
        ## --- MODIFIED --- ##
        # Save both primary and secondary transport days, ensuring they are numbers
        try:
            data["transport_days"] = int(self.fields["Transport Days"].get() or 0)
            data["transport_days_secondary"] = int(
                self.fields["Secondary Transport Days"].get() or 0
            )
        except ValueError:
            messagebox.showerror("❌ Error", "Transport Days fields must be numbers.", parent=self
            )
            return

        try:
            if (
                self.current_vendor_name
                and self.current_vendor_name != "New Vendor Name"
            ):
                self.dm.update_vendor(self.current_vendor_name, data)
                self.log(f"ℹ️ INFO: Updated vendor '{data['display_name']}'")
            else:
                new_vendor = self.dm.create_vendor(data)
                self.log(f"ℹ️ INFO: Created new vendor '{data['display_name']}'")
                messagebox.showinfo(
                    "API Key",
                    f"Vendor created.\n\nAPI Key: {new_vendor['api_key']}",
                    parent=self,
                )
            self.refresh_vendor_list()
        except Exception as e:
            messagebox.showerror(
                "Database Error", f"Could not save vendor: {e}", parent=self
            )

    def delete_vendor(self):
        if not self.current_vendor_name:
            messagebox.showwarning(
                "No Selection", "Please select a vendor to delete.", parent=self
            )
            return
        if not messagebox.askyesno("❓ Confirm",
            f"Delete '{self.current_vendor_name}'? This cannot be undone.",
            parent=self,
        ):
            return

        try:
            self.dm.delete_vendor(self.current_vendor_name)
            self.log(f"ℹ️ INFO: Deleted vendor '{self.current_vendor_name}'")
            self.refresh_vendor_list()
        except Exception as e:
            messagebox.showerror("❌ Error", f"Could not delete vendor: {e}", parent=self)

    def generate_api_key(self):
        if not self.current_vendor_name:
            messagebox.showwarning(
                "No Selection", "Please select a vendor first.", parent=self
            )
            return
        if not messagebox.askyesno("❓ Confirm",
            f"Generate a new API key for '{self.current_vendor_name}'? The old key will be invalidated.",
            parent=self,
        ):
            return
        new_key = self.dm.generate_new_api_key(self.current_vendor_name)
        self.api_key_var.set(new_key)
        messagebox.showinfo(
            "New API Key",
            f"New key for '{self.current_vendor_name}':\n\n{new_key}",
            parent=self,
        )
        self.refresh_vendor_list()

    def clear_form(self):
        for var in self.fields.values():
            var.set("")
        self.address_text.delete("1.0", tk.END)
        self.api_key_var.set("")
        self.current_vendor_name = None
        self.vendor_listbox.selection_clear(0, tk.END)

    def prepare_for_new(self):
        self.clear_form()
        self.fields["Display Name"].set("New Vendor Name")
        self.fields["Transport Days"].set("0")
        ## --- ADDED --- ##
        self.fields["Secondary Transport Days"].set("0")
        self.current_vendor_name = "New Vendor Name"


class SMTPConfigWindow(tk.Toplevel):
    def __init__(self, parent, log_callback, data_manager):
        super().__init__(parent)
        self.title("SMTP Configuration")
        self.geometry("500x350")
        self.log = log_callback
        self.dm = data_manager

        main_frame = ttk.Frame(self, padding=15)
        main_frame.pack(fill=tk.BOTH, expand=True)
        main_frame.columnconfigure(1, weight=1)

        self.fields = {}
        field_labels = [
            ("smtp_server", "SMTP Server:"),
            ("smtp_port", "Port:"),
            ("smtp_username", "Username:"),
            ("smtp_password", "Password:"),
            ("from_name", "From Name:"),
            ("from_email", "From Email:"),
        ]

        for i, (key, label) in enumerate(field_labels):
            ttk.Label(main_frame, text=label).grid(row=i, column=0, sticky="w", pady=2)
            self.fields[key] = tk.StringVar()
            entry = ttk.Entry(main_frame, textvariable=self.fields[key])
            if "password" in key:
                entry.config(show="*")
            entry.grid(row=i, column=1, sticky="ew", pady=2)

        self.fields["use_tls"] = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            main_frame, text="Use TLS", variable=self.fields["use_tls"]
        ).grid(row=len(field_labels), column=1, sticky="w", pady=5)

        button_frame = ttk.Frame(main_frame)
        button_frame.grid(
            row=len(field_labels) + 1, column=0, columnspan=2, pady=10, sticky="e"
        )
        ttk.Button(
            button_frame, text="Test Connection", command=self.test_connection
        ).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Save", command=self.save_config).pack(
            side=tk.LEFT
        )

        self.load_config()

    def load_config(self):
        config = self.dm.get_config("smtp_settings", {})
        for key, var in self.fields.items():
            # Handle boolean separately
            if isinstance(var, tk.BooleanVar):
                var.set(config.get(key, True))
            else:
                var.set(config.get(key, ""))

    def save_config(self):
        config = {key: var.get() for key, var in self.fields.items()}
        try:
            config["smtp_port"] = int(config["smtp_port"])
        except ValueError:
            messagebox.showerror("❌ Error", "Port must be a number.", parent=self)
            return

        self.dm.save_config("smtp_settings", config)
        self.log("ℹ️ INFO: SMTP settings saved.")
        messagebox.showinfo("✅ Success", "SMTP settings saved.", parent=self)
        self.destroy()

    def test_connection(self):
        config = {key: var.get() for key, var in self.fields.items()}
        try:
            port = int(config["smtp_port"])
            server = smtplib.SMTP(config["smtp_server"], port, timeout=10)
            if config["use_tls"]:
                server.starttls()
            server.login(config["smtp_username"], config["smtp_password"])
            server.quit()
            messagebox.showinfo("✅ Success", "SMTP connection successful!", parent=self)
        except Exception as e:
            messagebox.showerror(
                "Connection Failed",
                f"Could not connect to SMTP server:\n{e}",
                parent=self,
            )


class EmailConfigWindow(tk.Toplevel):
    def __init__(self, parent, log_callback, data_manager):
        super().__init__(parent)
        self.title("Email & Portal Configuration")
        self.geometry("700x600")
        self.log = log_callback
        self.dm = data_manager

        main_frame = ttk.Frame(self, padding=15)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Portal Settings
        portal_frame = ttk.LabelFrame(main_frame, text="Supplier Portal", padding=10)
        portal_frame.pack(fill=tk.X, pady=5)
        portal_frame.columnconfigure(1, weight=1)
        ttk.Label(portal_frame, text="Portal Base URL:").grid(
            row=0, column=0, sticky="w"
        )
        self.portal_url_var = tk.StringVar()
        ttk.Entry(portal_frame, textvariable=self.portal_url_var).grid(
            row=0, column=1, sticky="ew"
        )
        self.include_portal_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            portal_frame,
            text="Include portal link in emails",
            variable=self.include_portal_var,
        ).grid(row=1, column=1, sticky="w")

        # Email Templates
        template_frame = ttk.LabelFrame(main_frame, text="Email Templates", padding=10)
        template_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        template_frame.columnconfigure(1, weight=1)

        ttk.Label(template_frame, text="Subject:").grid(row=0, column=0, sticky="w")
        self.subject_var = tk.StringVar()
        ttk.Entry(template_frame, textvariable=self.subject_var).grid(
            row=0, column=1, sticky="ew"
        )
        ttk.Label(template_frame, text="Body:").grid(row=1, column=0, sticky="nw")
        self.body_text = tk.Text(template_frame, height=8)
        self.body_text.grid(row=1, column=1, sticky="nsew")
        ttk.Label(template_frame, text="Portal Link Section:").grid(
            row=2, column=0, sticky="nw"
        )
        self.portal_text = tk.Text(template_frame, height=4)
        self.portal_text.grid(row=2, column=1, sticky="nsew")
        template_frame.rowconfigure(1, weight=1)
        template_frame.rowconfigure(2, weight=1)

        ttk.Button(self, text="Save", command=self.save_config).pack(
            pady=10, padx=10, anchor="e"
        )
        self.load_config()

    def load_config(self):
        config = self.dm.get_config("email_config", {})
        self.portal_url_var.set(
            config.get("portal_base_url", "http://example.com/supplier_portal.html")
        )
        self.include_portal_var.set(config.get("include_portal_link", True))
        self.subject_var.set(
            config.get(
                "email_subject_template", "Purchase Order {po_number} - {supplier_name}"
            )
        )
        self.body_text.insert(
            "1.0",
            config.get(
                "email_body_template",
                "Dear {supplier_name},\n\nPlease find attached PO {po_number}.\n\n{portal_link_section}\n\nRegards",
            ),
        )
        self.portal_text.insert(
            "1.0",
            config.get(
                "portal_link_template", "Access our portal: {portal_url}?key={api_key}"
            ),
        )

    def save_config(self):
        config = {
            "portal_base_url": self.portal_url_var.get(),
            "include_portal_link": self.include_portal_var.get(),
            "email_subject_template": self.subject_var.get(),
            "email_body_template": self.body_text.get("1.0", "end-1c"),
            "portal_link_template": self.portal_text.get("1.0", "end-1c"),
        }
        self.dm.save_config("email_config", config)
        messagebox.showinfo("✅ Success", "Email configuration saved.", parent=self)
        self.destroy()


class CompanyConfigWindow(tk.Toplevel):
    def __init__(self, parent, log_callback, data_manager):
        super().__init__(parent)
        self.title("Company (PO Template) Configuration")
        self.geometry("600x650")
        self.log = log_callback
        self.dm = data_manager

        main_frame = ttk.Frame(self, padding=15)
        main_frame.pack(fill=tk.BOTH, expand=True)
        main_frame.columnconfigure(1, weight=1)

        self.fields = {}
        field_labels = [
            ("my_company_name", "Company Name:"),
            ("my_company_address", "Company Address:"),
            ("buyer_name", "Buyer Name:"),
            ("buyer_email", "Buyer Email:"),
            ("ship_to_name", "Ship-To Name:"),
            ("ship_to_address", "Ship-To Address:"),
            ("terms_and_conditions", "Terms and Conditions:"),
        ]

        row = 0
        for key, label in field_labels:
            ttk.Label(main_frame, text=label).grid(
                row=row, column=0, sticky="nw", pady=2
            )
            if "address" in key or "terms_and_conditions" in key:
                height = 5 if "terms_and_conditions" in key else 3
                self.fields[key] = tk.Text(main_frame, height=height)
            else:
                self.fields[key] = ttk.Entry(main_frame)
            self.fields[key].grid(row=row, column=1, sticky="ew", pady=2)
            row += 1

        ttk.Button(self, text="Save", command=self.save_config).pack(
            pady=10, padx=10, anchor="e"
        )
        self.load_config()

    def load_config(self):
        config = self.dm.get_config("company_config", {})
        for key, widget in self.fields.items():
            # Clear existing content before inserting
            if isinstance(widget, tk.Text):
                widget.delete("1.0", tk.END)
                widget.insert("1.0", config.get(key, ""))
            else:
                widget.delete(0, tk.END)
                widget.insert(0, config.get(key, ""))

    def save_config(self):
        config = {}
        for key, widget in self.fields.items():
            if isinstance(widget, tk.Text):
                config[key] = widget.get("1.0", "end-1c").strip()
            else:
                config[key] = widget.get().strip()
        self.dm.save_config("company_config", config)
        messagebox.showinfo("✅ Success", "Company configuration saved.", parent=self)
        self.destroy()


class POCreatorWindow(tk.Toplevel):
    def __init__(self, parent, log_callback, data_manager):
        super().__init__(parent)
        self.title("Create New Purchase Order")
        self.geometry("950x650")
        self.log = log_callback
        self.dm = data_manager

        # Header
        header_frame = ttk.LabelFrame(self, text="Header Info", padding=10)
        header_frame.pack(fill=tk.X, padx=10, pady=5)
        header_frame.columnconfigure(1, weight=1)
        header_frame.columnconfigure(3, weight=1)

        ttk.Label(header_frame, text="PO Number:").grid(
            row=0, column=0, sticky="w", padx=5
        )
        self.po_var = tk.StringVar()
        po_entry = ttk.Entry(header_frame, textvariable=self.po_var)
        po_entry.grid(row=0, column=1, sticky="ew", padx=(0, 10))
        po_entry.bind("<Return>", self.load_po_from_database)  # Load on Enter key
        po_entry.bind("<FocusOut>", self.load_po_from_database)  # Load on focus lost

        ttk.Button(
            header_frame, text="Load PO", command=self.load_po_from_database
        ).grid(row=0, column=2, padx=5)

        ttk.Label(header_frame, text="Supplier:").grid(
            row=1, column=0, sticky="w", padx=5
        )
        self.supplier_var = tk.StringVar()
        self.supplier_combo = ttk.Combobox(
            header_frame, textvariable=self.supplier_var, state="readonly"
        )
        self.supplier_combo["values"] = [
            v["display_name"] for v in self.dm.get_all_vendors()
        ]
        self.supplier_combo.grid(row=1, column=1, sticky="ew", columnspan=3)

        # Lines
        lines_frame = ttk.LabelFrame(self, text="Line Items", padding=10)
        lines_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        cols = (
            "Item",
            "Material",
            "Description",
            "Qty",
            "Unit",
            "Price",
            "Per",
            "Req. Date",
        )
        self.lines_tree = ttk.Treeview(
            lines_frame, columns=cols, show="headings", height=10
        )
        for col in cols:
            self.lines_tree.heading(col, text=col)
            if col in ["Item", "Qty", "Unit", "Per"]:
                self.lines_tree.column(col, width=60)
            elif col == "Price":
                self.lines_tree.column(col, width=80)
            elif col == "Req. Date":
                self.lines_tree.column(col, width=90)

        tree_scroll = ttk.Scrollbar(
            lines_frame, orient="vertical", command=self.lines_tree.yview
        )
        self.lines_tree.configure(yscrollcommand=tree_scroll.set)
        self.lines_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # Entry fields for manual adding (optional)
        entry_frame = ttk.LabelFrame(
            lines_frame, text="Add/Edit Line Item Manually", padding=5
        )
        entry_frame.pack(fill=tk.X, pady=(10, 0))

        self.line_vars = {col: tk.StringVar() for col in cols}
        for i, col in enumerate(cols):
            ttk.Label(entry_frame, text=col, font=("Helvetica", 8)).grid(
                row=0, column=i, padx=2
            )
            width = (
                8
                if col in ["Item", "Qty", "Unit", "Per"]
                else (10 if col in ["Price", "Req. Date"] else 15)
            )
            ttk.Entry(entry_frame, textvariable=self.line_vars[col], width=width).grid(
                row=1, column=i, padx=2, sticky="ew"
            )
            entry_frame.columnconfigure(i, weight=1)

        btn_frame = ttk.Frame(entry_frame)
        btn_frame.grid(row=2, column=0, columnspan=len(cols), pady=5)
        ttk.Button(btn_frame, text="Add Line", command=self.add_line).pack(side=tk.LEFT)
        ttk.Button(btn_frame, text="Remove Selected", command=self.remove_line).pack(
            side=tk.LEFT, padx=5
        )

        ttk.Button(self, text="Generate and Save PO", command=self.submit_po).pack(
            pady=10
        )

    def load_po_from_database(self, event=None):
        """Load PO lines from the open orders database"""
        po_number = self.po_var.get().strip()
        if not po_number:
            return

        try:
            # Query database for this PO
            query = """
                SELECT oo.*, v.display_name as supplier_name
                FROM open_orders oo
                LEFT JOIN vendors v ON oo.vendor_name = v.vendor_name
                WHERE oo.po = ? AND oo.status = 'Open'
                ORDER BY oo.item
            """
            results = self.dm.db.execute_query(query, (po_number,), fetchall=True)

            if not results:
                messagebox.showinfo(
                    "PO Not Found",
                    f"No open orders found for PO: {po_number}",
                    parent=self,
                )
                return

            # Clear existing lines
            for item in self.lines_tree.get_children():
                self.lines_tree.delete(item)

            # Set supplier from first line
            if results[0]["supplier_name"]:
                self.supplier_var.set(results[0]["supplier_name"])

            # Add all lines to tree
            for order in results:
                values = (
                    order.get("item", ""),
                    order.get("material_code", ""),
                    order.get("short_text", ""),
                    order.get("requested_qty", ""),
                    order.get("unit", "EA"),
                    f"{float(order.get('unit_price', 0)):.2f}",
                    str(order.get("price_per_unit", 1)),
                    order.get("requested_del_date", ""),
                )
                self.lines_tree.insert("", "end", values=values)

            self.log(f"ℹ️ INFO: Loaded {len(results)} line(s) for PO {po_number}")
            messagebox.showinfo(
                "PO Loaded",
                f"Loaded {len(results)} line item(s) for PO {po_number}",
                parent=self,
            )

        except Exception as e:
            messagebox.showerror("❌ Error", f"Failed to load PO: {str(e)}", parent=self)
            self.log(f"✗ ERROR: Failed to load PO {po_number}: {e}")

    def add_line(self):
        """Manually add a line item"""
        values = [v.get().strip() for v in self.line_vars.values()]
        if not all([values[0], values[1], values[3]]):  # Item, Material, Qty required
            messagebox.showwarning(
                "Incomplete", "Item, Material, and Quantity are required.", parent=self
            )
            return

        self.lines_tree.insert("", "end", values=values)
        for var in self.line_vars.values():
            var.set("")

    def remove_line(self):
        """Remove selected line"""
        selected = self.lines_tree.selection()
        if selected:
            self.lines_tree.delete(selected)

    def submit_po(self):
        """Generate and save the PO PDF"""
        po_number = self.po_var.get().strip()
        vendor = self.supplier_var.get().strip()

        if not po_number or not vendor:
            messagebox.showerror("❌ Error", "PO Number and Supplier are required.", parent=self
            )
            return

        # Get all lines from tree
        lines = []
        for item_id in self.lines_tree.get_children():
            vals = self.lines_tree.item(item_id, "values")
            try:
                qty = int(vals[3])
                price = float(vals[5])
                # Handle "Per" field - it's at index 6, date is now at index 7
                price_per_str = (
                    str(vals[6]).strip() if len(vals) > 6 and vals[6] else "1"
                )
                try:
                    price_per = int(price_per_str) if price_per_str else 1
                except (ValueError, TypeError):
                    price_per = 1

                lines.append(
                    {
                        "item": vals[0],
                        "material_code": vals[1],
                        "short_text": vals[2],
                        "requested_qty": qty,
                        "unit": vals[4] if vals[4] else "EA",
                        "unit_price": price,
                        "price_per_unit": price_per,
                        "total_amount": qty * price,
                        "requested_del_date": vals[7] if len(vals) > 7 else "",
                    }
                )
            except (ValueError, IndexError) as e:
                messagebox.showerror(
                    "Data Error",
                    f"Invalid data in line: {vals}\nError: {e}",
                    parent=self,
                )
                return

        if not lines:
            messagebox.showerror("❌ Error", "At least one line item is required.", parent=self
            )
            return

        po_data = {
            "po_number": po_number,
            "vendor_display_name": vendor,
            "lines": lines,
        }

        try:
            pdf_buffer = self.dm.create_po_from_data(po_data)
            save_path = filedialog.asksaveasfilename(
                defaultextension=".pdf", initialfile=f"PO_{po_number}.pdf", parent=self
            )
            if save_path:
                with open(save_path, "wb") as f:
                    f.write(pdf_buffer.read())

                # Also save to OrdersToSend folder
                local_path = os.path.join(ORDERS_FOLDER, f"PO_{po_number}.pdf")
                with open(local_path, "wb") as f:
                    pdf_buffer.seek(0)
                    f.write(pdf_buffer.read())

                messagebox.showinfo("✅ Success",
                    f"PO PDF saved to {save_path} and is ready to be sent.",
                    parent=self,
                )
                self.log(f"✓ SUCCESS: Created PO {po_number} and saved PDF.")
                self.destroy()
        except Exception as e:
            self.log(f"✗ ERROR: Failed to create PO: {e}")
            messagebox.showerror("❌ Error", f"Failed to create PO: {e}", parent=self)


class RescheduleConfigWindow(tk.Toplevel):
    def __init__(self, parent, log_callback, data_manager):
        super().__init__(parent)
        self.title("Advanced Reschedule File Generation")
        self.geometry("700x850")  # Adjusted for proper fit
        self.log = log_callback
        self.dm = data_manager

        # Create scrollable main frame
        main_canvas = tk.Canvas(self)
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=main_canvas.yview)
        main_frame = ttk.Frame(main_canvas)
        main_frame.bind(
            "<Configure>",
            lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all")),
        )
        main_canvas.create_window((0, 0), window=main_frame, anchor="nw")
        main_canvas.configure(yscrollcommand=scrollbar.set)
        main_canvas.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        scrollbar.pack(side="right", fill="y")

        # Output Format Selection
        format_frame = ttk.LabelFrame(main_frame, text="Output Format", padding=10)
        format_frame.pack(fill=tk.X, pady=5)

        self.output_format_var = tk.StringVar(value="excel")
        ttk.Radiobutton(
            format_frame,
            text="Excel Format (.xlsx) - Editable spreadsheet",
            variable=self.output_format_var,
            value="excel",
        ).pack(anchor="w")
        ttk.Radiobutton(
            format_frame,
            text="PDF Format (.pdf) - Professional document with date changes",
            variable=self.output_format_var,
            value="pdf",
        ).pack(anchor="w")
        # MANUAL RESCHEDULE DATE OVERRIDE
        override_frame = ttk.LabelFrame(
            main_frame, text="Manual Reschedule Date Override (Optional)", padding=10
        )
        override_frame.pack(fill=tk.X, pady=5)

        self.use_manual_override = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            override_frame,
            text="Apply manual reschedule date to all order lines",
            variable=self.use_manual_override,
            command=self.toggle_manual_override,
        ).pack(anchor="w", pady=(0, 5))

        date_entry_frame = ttk.Frame(override_frame)
        date_entry_frame.pack(fill=tk.X, pady=5)

        ttk.Label(date_entry_frame, text="Reschedule Date (DD.MM.YYYY):").pack(
            side=tk.LEFT, padx=(0, 5)
        )
        self.manual_reschedule_date = tk.StringVar()
        self.manual_date_entry = ttk.Entry(
            date_entry_frame, textvariable=self.manual_reschedule_date, width=15
        )
        self.manual_date_entry.pack(side=tk.LEFT, padx=5)
        self.manual_date_entry.config(state="disabled")

        # Calendar button (optional - for easier date selection)
        ttk.Button(
            date_entry_frame, text="Today", command=self.set_today_date, width=8
        ).pack(side=tk.LEFT, padx=2)

        help_text = ttk.Label(
            override_frame,
            text="This date will be used as the ETD (Expected Time of Departure) in reschedule files.\nThe open order book will NOT be modified.",
            font=("Helvetica", 8),
            foreground="gray",
        )
        help_text.pack(anchor="w", pady=(5, 0))

        # Vendor Selection
        vendor_frame = ttk.LabelFrame(main_frame, text="Select Suppliers", padding=10)
        vendor_frame.pack(fill=tk.X, pady=5)

        # Search functionality
        search_frame = ttk.Frame(vendor_frame)
        search_frame.pack(fill=tk.X, pady=(0, 5))
        ttk.Label(search_frame, text="Search Supplier:").pack(side=tk.LEFT, padx=(0, 5))
        self.supplier_search_var = tk.StringVar()
        self.supplier_search_var.trace("w", self.filter_suppliers)
        ttk.Entry(search_frame, textvariable=self.supplier_search_var).pack(
            side=tk.LEFT, fill=tk.X, expand=True
        )

        # Select All/None buttons
        vendor_btn_frame = ttk.Frame(vendor_frame)
        vendor_btn_frame.pack(fill=tk.X, pady=(0, 5))
        ttk.Button(
            vendor_btn_frame, text="Select All", command=self.select_all_vendors
        ).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(
            vendor_btn_frame, text="Deselect All", command=self.deselect_all_vendors
        ).pack(side=tk.LEFT)

        self.vendor_listbox = Listbox(
            vendor_frame, selectmode=tk.MULTIPLE, exportselection=False, height=6
        )
        vendor_scroll = ttk.Scrollbar(
            vendor_frame, orient="vertical", command=self.vendor_listbox.yview
        )
        self.vendor_listbox.configure(yscrollcommand=vendor_scroll.set)
        self.vendor_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vendor_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # Reschedule Out Filters
        out_frame = ttk.LabelFrame(
            main_frame,
            text="Filters for 'Reschedule Out' Lines (Later than requested/confirmed)",
            padding=10,
        )
        out_frame.pack(fill=tk.X, pady=5)

        out_date_frame = ttk.Frame(out_frame)
        out_date_frame.pack(fill=tk.X, pady=2)
        out_date_frame.columnconfigure(1, weight=1)
        out_date_frame.columnconfigure(3, weight=1)

        self.out_vars = {
            "start_date": tk.StringVar(),
            "end_date": tk.StringVar(),
            "min_val": tk.StringVar(),
            "max_val": tk.StringVar(),
        }

        ttk.Label(out_date_frame, text="Req. Date Range (YYYY-MM-DD):").grid(
            row=0, column=0, sticky="w", padx=(0, 5)
        )
        ttk.Entry(
            out_date_frame, textvariable=self.out_vars["start_date"], width=12
        ).grid(row=0, column=1, sticky="ew", padx=(0, 5))
        ttk.Label(out_date_frame, text="to").grid(row=0, column=2, padx=5)
        ttk.Entry(
            out_date_frame, textvariable=self.out_vars["end_date"], width=12
        ).grid(row=0, column=3, sticky="ew")

        out_val_frame = ttk.Frame(out_frame)
        out_val_frame.pack(fill=tk.X, pady=2)
        out_val_frame.columnconfigure(1, weight=1)
        out_val_frame.columnconfigure(3, weight=1)

        ttk.Label(out_val_frame, text="Value Range:").grid(
            row=0, column=0, sticky="w", padx=(0, 5)
        )
        ttk.Entry(out_val_frame, textvariable=self.out_vars["min_val"], width=12).grid(
            row=0, column=1, sticky="ew", padx=(0, 5)
        )
        ttk.Label(out_val_frame, text="to").grid(row=0, column=2, padx=5)
        ttk.Entry(out_val_frame, textvariable=self.out_vars["max_val"], width=12).grid(
            row=0, column=3, sticky="ew"
        )

        # Reschedule In Filters
        in_frame = ttk.LabelFrame(
            main_frame,
            text="Filters for 'Reschedule In' Lines (Earlier than requested/confirmed)",
            padding=10,
        )
        in_frame.pack(fill=tk.X, pady=5)

        in_date_frame = ttk.Frame(in_frame)
        in_date_frame.pack(fill=tk.X, pady=2)
        in_date_frame.columnconfigure(1, weight=1)
        in_date_frame.columnconfigure(3, weight=1)

        self.in_vars = {"start_date": tk.StringVar(), "end_date": tk.StringVar()}

        ttk.Label(in_date_frame, text="Date Range (YYYY-MM-DD):").grid(
            row=0, column=0, sticky="w", padx=(0, 5)
        )
        ttk.Entry(
            in_date_frame, textvariable=self.in_vars["start_date"], width=12
        ).grid(row=0, column=1, sticky="ew", padx=(0, 5))
        ttk.Label(in_date_frame, text="to").grid(row=0, column=2, padx=5)
        ttk.Entry(in_date_frame, textvariable=self.in_vars["end_date"], width=12).grid(
            row=0, column=3, sticky="ew"
        )

        # MODIFIED: Open Orders Option (changed from "Unconfirmed Orders")
        open_orders_frame = ttk.LabelFrame(main_frame, text="Open Orders", padding=10)
        open_orders_frame.pack(fill=tk.X, pady=5)

        self.include_all_open_orders = tk.BooleanVar()
        ttk.Checkbutton(
            open_orders_frame,
            text="Include All Open Orders (regardless of confirmation status)",
            variable=self.include_all_open_orders,
        ).pack(anchor="w")

        open_orders_help = ttk.Label(
            open_orders_frame,
            text="When checked: All open orders will be included. Orders filtered out by date/value criteria will have their reschedule dates cleared.",
            font=("Helvetica", 8),
            foreground="gray",
        )
        open_orders_help.pack(anchor="w", pady=(5, 0))

        # Action Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=15)

        generate_btn = ttk.Button(
            button_frame, text="Generate Files", command=self.generate_files
        )
        generate_btn.pack(side=tk.RIGHT)
        # Style the generate button
        style = ttk.Style(self)
        try:
            style.configure("Accent.TButton", font=("Helvetica", 10, "bold"))
            generate_btn.configure(style="Accent.TButton")
        except Exception:  # TODO: Add proper error handling
            pass  # TODO: Add proper error handling
        # Add custom signature checkbox
        self.use_custom_sig_var = tk.BooleanVar(value=self.dm.get_config('use_custom_signature', False))
        ttk.Checkbutton(
            button_frame,
            text="Use Custom Signature",
            variable=self.use_custom_sig_var
        ).pack(side=tk.RIGHT, padx=5)

        ttk.Button(
            button_frame, text="Send Emails", command=self.send_reschedule_emails_only
        ).pack(side=tk.RIGHT, padx=(5, 0))

        ttk.Button(button_frame, text="Cancel", command=self.destroy).pack(
            side=tk.RIGHT, padx=(10, 0)
        )

        # Help text
        help_frame = ttk.LabelFrame(main_frame, text="Help", padding=10)
        help_frame.pack(fill=tk.X, pady=5)

        help_text = """Date Format: Use YYYY-MM-DD format (e.g., 2024-01-15)
Suppliers: Leave none selected to include all suppliers
Value Range: Enter numeric values only (e.g., 1000.50)
Reschedule Out: Orders where new date is LATER than confirmed (or requested if no confirmation)
Reschedule In: Orders where new date is EARLIER than confirmed (or requested if no confirmation)
Open Orders: When checked, includes all open orders regardless of confirmation status
Output Format: Excel for editable files, PDF for professional presentation with date changes
Note: Reschedule comparison prioritizes confirmed date over requested date when available
      PDF format automatically calculates ETD dates using transport days"""

        ttk.Label(help_frame, text=help_text, font=("Helvetica", 8)).pack(anchor="w")

        self.load_vendors()

    def load_vendors(self):
        """Load vendors into the listbox"""
        self.all_vendors = [
            vendor["display_name"] for vendor in self.dm.get_all_vendors()
        ]
        self.filter_suppliers()

    def filter_suppliers(self, *args):
        """Filter supplier listbox based on search term"""
        search_term = self.supplier_search_var.get().lower()
        self.vendor_listbox.delete(0, tk.END)

        for vendor in self.all_vendors:
            if search_term in vendor.lower():
                self.vendor_listbox.insert(tk.END, vendor)

    def select_all_vendors(self):
        """Select all visible vendors in listbox"""
        self.vendor_listbox.select_set(0, tk.END)

    def deselect_all_vendors(self):
        """Deselect all vendors"""
        self.vendor_listbox.select_clear(0, tk.END)

    def toggle_manual_override(self):
        """Enable/disable manual date entry based on checkbox"""
        if self.use_manual_override.get():
            self.manual_date_entry.config(state="normal")
        else:
            self.manual_date_entry.config(state="disabled")

    def set_today_date(self):
        """Set the manual reschedule date to today"""
        if self.use_manual_override.get():
            today = datetime.now().strftime("%d.%m.%Y")
            self.manual_reschedule_date.set(today)

    def generate_files(self):
        """Generate reschedule files with advanced filtering"""
        # Gather filter data
        selected_indices = self.vendor_listbox.curselection()
        selected_vendors = [self.vendor_listbox.get(i) for i in selected_indices]

        # Helper to get value or None if empty
        def get_val(var):
            val = var.get().strip()
            return val if val else None

        # Validate manual reschedule date if enabled
        manual_date = None
        if self.use_manual_override.get():
            date_str = get_val(self.manual_reschedule_date)
            if not date_str:
                messagebox.showerror(
                    "Date Required",
                    "Please enter a reschedule date or disable the manual override.",
                    parent=self,
                )
                return

            try:
                manual_date = datetime.strptime(date_str, "%d.%m.%Y")
            except ValueError:
                messagebox.showerror(
                    "Date Format Error",
                    "Manual reschedule date must be in DD.MM.YYYY format (e.g., 15.01.2025).",
                    parent=self,
                )
                return

        filters = {
            "suppliers": selected_vendors if selected_vendors else None,
            # Date filters
            "reschedule_out_req_date_start": get_val(self.out_vars["start_date"]),
            "reschedule_out_req_date_end": get_val(self.out_vars["end_date"]),
            "reschedule_in_req_date_start": get_val(self.in_vars["start_date"]),
            "reschedule_in_req_date_end": get_val(self.in_vars["end_date"]),
            # Include all open orders option
            "include_all_open_orders": self.include_all_open_orders.get(),
            # Output format
            "use_pdf_format": (self.output_format_var.get() == "pdf"),
            # MANUAL OVERRIDE
            "manual_reschedule_date": manual_date,
        }

        # Value filters
        try:
            min_val = get_val(self.out_vars["min_val"])
            max_val = get_val(self.out_vars["max_val"])
            filters["reschedule_out_value_min"] = float(min_val) if min_val else None
            filters["reschedule_out_value_max"] = float(max_val) if max_val else None
        except ValueError:
            messagebox.showerror(
                "Input Error", "Value range must contain valid numbers.", parent=self
            )
            return

        # Validate date formats for other date filters
        date_fields = [
            (filters["reschedule_out_req_date_start"], "Reschedule Out Start Date"),
            (filters["reschedule_out_req_date_end"], "Reschedule Out End Date"),
            (filters["reschedule_in_req_date_start"], "Reschedule In Start Date"),
            (filters["reschedule_in_req_date_end"], "Reschedule In End Date"),
        ]

        for date_val, field_name in date_fields:
            if date_val:
                try:
                    datetime.strptime(date_val, "%Y-%m-%d")
                except ValueError:
                    messagebox.showerror(
                        "Date Format Error",
                        f"{field_name} must be in YYYY-MM-DD format.",
                        parent=self,
                    )
                    return

        # Generate files
        def generate_files_thread():
            try:
                files_created = self.dm.generate_reschedule_files(filters)
                self.after(0, lambda: self.show_generation_result(files_created))
            except Exception as e:
                self.after(0, lambda: self.show_generation_error(str(e)))

        # Show loading message
        format_type = "PDF" if filters["use_pdf_format"] else "Excel"
        self.log(
            f"ℹ️ INFO: Generating {format_type} reschedule files with advanced filters..."
        )
        threading.Thread(target=generate_files_thread, daemon=True).start()

    def show_generation_result(self, files_created):
        """Show the result of file generation"""
        if files_created > 0:
            format_type = "PDF" if self.output_format_var.get() == "pdf" else "Excel"
            message = f"Successfully generated {files_created} {format_type} reschedule file(s) in the 'reschedule_output' folder."
            messagebox.showinfo("✅ Success", message, parent=self)
            self.log(
                f"✓ SUCCESS: Generated {files_created} {format_type} reschedule files with advanced filters."
            )
        else:
            message = (
                "No orders found matching the specified criteria to generate files for."
            )
            messagebox.showinfo("No Data", message, parent=self)
            self.log("ℹ️ INFO: No reschedule data found for the given filters.")

    def show_generation_error(self, error_message):
        """Show error message for file generation failure"""
        self.log(f"✗ ERROR: Failed to generate reschedule files: {error_message}")
        messagebox.showerror("❌ Error", f"Failed to generate files:\n{error_message}", parent=self
        )

    def send_reschedule_emails_only(self):
        """Send already-generated reschedule files via email to suppliers"""
        try:
            # Check if there are files to send
            reschedule_folder = RESCHEDULE_OUTPUT_FOLDER

            if not os.path.exists(reschedule_folder):
                messagebox.showwarning(
                    "No Files",
                    "No reschedule files found. Please generate files first.",
                    parent=self,
                )
                return

            # Check for recent files (generated in last 24 hours)
            recent_files = []
            cutoff_time = datetime.now() - timedelta(hours=24)

            for filename in os.listdir(reschedule_folder):
                if (
                    filename.endswith(".xlsx")
                    or filename.endswith(".pdf")
                    or filename.endswith(".zip")
                ):
                    file_path = os.path.join(reschedule_folder, filename)
                    file_time = datetime.fromtimestamp(os.path.getmtime(file_path))
                    if file_time > cutoff_time:
                        recent_files.append(filename)

            if not recent_files:
                messagebox.showwarning(
                    "No Recent Files",
                    "No recent reschedule files found (last 24 hours).\n\nPlease generate files first.",
                    parent=self,
                )
                return

            # Show which files will be sent
            file_list = "\n".join(recent_files[:10])
            if len(recent_files) > 10:
                file_list += f"\n... and {len(recent_files) - 10} more"

            if not messagebox.askyesno(
                "Confirm Send",
                f"Send the following {len(recent_files)} reschedule file(s)?\n\n{file_list}",
                parent=self,
            ):
                return

            # Ask for send method
            method = messagebox.askquestion(
                "Email Method",
                "Send via Outlook?\n\nYes = Outlook\nNo = SMTP",
                parent=self,
            )

            send_method = "Outlook" if method == "yes" else "SMTP"

            # Check if Outlook is available
            if send_method == "Outlook":
                try:
                    import win32com.client

                    OUTLOOK_AVAILABLE = True
                except ImportError:
                    OUTLOOK_AVAILABLE = False
                    messagebox.showerror("❌ Error",
                        "Outlook is not available. Please use SMTP method.",
                        parent=self,
                    )
                    return

            # Get vendor email mapping
            email_map = {}
            for vendor in self.dm.get_all_vendors():
                if vendor.get("emails"):
                    email_map[vendor["display_name"]] = vendor["emails"]

            sent_count = 0
            failed_count = 0
            failed_details = []

            config = self.dm.get_config("company_config", {})
            buyer_email = config.get("buyer_email", "purchasing@company.com")
            
            # Get checkbox value for custom signature
            use_custom_sig = self.use_custom_sig_var.get()

            # Check if files are ZIP or individual files
            zip_files = [f for f in recent_files if f.endswith(".zip")]
            individual_files = [
                f for f in recent_files if f.endswith(".xlsx") or f.endswith(".pdf")
            ]

            if zip_files:
                # ZIP file contains multiple suppliers - need to extract and send individually
                for zip_file in zip_files:
                    zip_path = os.path.join(reschedule_folder, zip_file)

                    with zipfile.ZipFile(zip_path, "r") as zf:
                        for file_name in zf.namelist():
                            # Extract supplier name from filename
                            supplier_name = (
                                file_name.replace("_reschedule.xlsx", "")
                                .replace("_reschedule.pdf", "")
                                .replace("_", " ")
                            )

                            if supplier_name not in email_map:
                                failed_count += 1
                                failed_details.append(
                                    f"{supplier_name}: No email address found"
                                )
                                continue

                            # Extract file to temp location
                            temp_path = os.path.join(reschedule_folder, file_name)
                            with open(temp_path, "wb") as f:
                                f.write(zf.read(file_name))

                            # Send email
                            try:
                                if self.send_single_reschedule_email(
                                    supplier_name,
                                    temp_path,
                                    email_map[supplier_name],
                                    buyer_email,
                                    send_method,
                                    use_custom_sig,
                                ):
                                    sent_count += 1
                                    self.log(
                                        f"✓ SUCCESS: Sent reschedule email to {supplier_name}"
                                    )
                                else:
                                    failed_count += 1
                                    failed_details.append(
                                        f"{supplier_name}: Failed to send"
                                    )
                            except Exception as e:
                                failed_count += 1
                                failed_details.append(f"{supplier_name}: {str(e)}")
                                self.log(
                                    f"✗ ERROR: Failed to send to {supplier_name}: {e}"
                                )
                            finally:
                                # Clean up temp file
                                if os.path.exists(temp_path):
                                    os.remove(temp_path)

            if individual_files:
                # Individual files (Excel or PDF)
                for file_name in individual_files:
                    # Extract supplier name from filename
                    supplier_name = (
                        file_name.replace("_reschedule.xlsx", "")
                        .replace("_reschedule.pdf", "")
                        .replace("_", " ")
                    )

                    if supplier_name not in email_map:
                        failed_count += 1
                        failed_details.append(
                            f"{supplier_name}: No email address found"
                        )
                        continue

                    file_path = os.path.join(reschedule_folder, file_name)

                    try:
                        if self.send_single_reschedule_email(
                            supplier_name,
                            file_path,
                            email_map[supplier_name],
                            buyer_email,
                            send_method,
                            use_custom_sig,
                        ):
                            sent_count += 1
                            self.log(
                                f"✓ SUCCESS: Sent reschedule email to {supplier_name}"
                            )
                        else:
                            failed_count += 1
                            failed_details.append(f"{supplier_name}: Failed to send")
                    except Exception as e:
                        failed_count += 1
                        failed_details.append(f"{supplier_name}: {str(e)}")
                        self.log(f"✗ ERROR: Failed to send to {supplier_name}: {e}")

            # Show summary
            summary = f"Sent {sent_count} reschedule email(s), Failed {failed_count}"
            if failed_details:
                summary += "\n\nFailures:\n" + "\n".join(failed_details)

            messagebox.showinfo("Email Results", summary, parent=self)
            self.log(f"ℹ️ INFO: Reschedule email batch complete: {summary}")

        except Exception as e:
            self.log(f"✗ ERROR: Failed to send reschedule emails: {e}")
            messagebox.showerror("❌ Error", f"Failed to send emails:\n{str(e)}", parent=self
            )

    def send_single_reschedule_email(
        self, supplier_name, file_path, to_emails, buyer_email, method, use_custom_signature=False
    ):
        """Send a single reschedule email using template"""
        sender = EmailSender(self.log, self.dm)
        subject, body = sender.generate_reschedule_email_content(supplier_name)

        try:
            if method == "Outlook":
                success, message = sender._send_outlook(
                    to_emails, subject, body, file_path, use_custom_signature
                )
                return success
            else:
                success, message = sender._send_smtp(
                    to_emails, subject, body, file_path, use_custom_signature
                )
                return success
        except Exception as e:
            self.log(f"✗ ERROR: Failed to send email: {e}")
            return False


class EnhancedEmailPreviewWindow(tk.Toplevel):
    def __init__(self, parent, log_callback, data_manager, send_method_getter):
        super().__init__(parent)
        self.title("Email Preview & Sender")
        self.geometry("1000x700")
        self.log = log_callback
        self.dm = data_manager
        self.get_send_method = send_method_getter

        paned_window = ttk.PanedWindow(self, orient=tk.HORIZONTAL)
        paned_window.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Left: PO List
        left_frame = ttk.LabelFrame(
            paned_window, text="POs Ready to Send (with PDFs)", padding=5
        )
        cols = ("PO", "Supplier", "Email", "Source")
        self.po_tree = ttk.Treeview(left_frame, columns=cols, show="headings")
        for col in cols:
            self.po_tree.heading(col, text=col)
        self.po_tree.column("PO", width=100)
        self.po_tree.column("Supplier", width=150)
        self.po_tree.column("Email", width=180)
        self.po_tree.column("Source", width=70)
        self.po_tree.pack(fill=tk.BOTH, expand=True)
        self.po_tree.bind("<<TreeviewSelect>>", self.on_po_select)
        paned_window.add(left_frame, weight=2)

        # Right: Preview
        right_frame = ttk.LabelFrame(paned_window, text="Email Preview", padding=5)
        self.email_to_label = ttk.Label(right_frame, text="To:")
        self.email_to_label.pack(anchor="w", pady=2)
        self.email_subject_label = ttk.Label(right_frame, text="Subject:")
        self.email_subject_label.pack(anchor="w", pady=2)
        self.pdf_status_label = ttk.Label(right_frame, text="PDF:")
        self.pdf_status_label.pack(anchor="w", pady=2)
        self.email_body_text = tk.Text(right_frame, wrap=tk.WORD)
        self.email_body_text.pack(fill=tk.BOTH, expand=True, pady=5)
        paned_window.add(right_frame, weight=3)

        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill=tk.X, padx=10, pady=5)
        
        # Add custom signature checkbox on the left
        self.use_custom_sig_var = tk.BooleanVar(value=self.dm.get_config('use_custom_signature', False))
        ttk.Checkbutton(
            btn_frame, 
            text="Use Custom Signature (default: Outlook signature)",
            variable=self.use_custom_sig_var
        ).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(btn_frame, text="Send All", command=self.send_all).pack(
            side=tk.RIGHT, padx=5
        )
        ttk.Button(btn_frame, text="Send Selected", command=self.send_selected).pack(
            side=tk.RIGHT
        )

        self.refresh_data()

    def refresh_data(self):
        for i in self.po_tree.get_children():
            self.po_tree.delete(i)

        # Get pending POs from database
        db_pending_pos = self.dm.get_pending_pos_with_portal_info()
        sender = EmailSender(self.log, self.dm)

        self.pending_pos = []
        db_po_numbers = set()

        # Add database POs that have PDFs
        for po in db_pending_pos:
            pdf_path = sender.find_po_pdf(po["po"])
            if pdf_path:
                self.pending_pos.append(po)
                db_po_numbers.add(po["po"])
                self.po_tree.insert(
                    "",
                    "end",
                    values=(po["po"], po["supplier"], po["emails"], "Database"),
                )

        # Scan OrdersToSend folder for PDFs not in database (fallback)
        try:
            email_map = load_email_mapping(self.dm.db, self.log)

            # Get display names for vendor extraction
            all_vendors = self.dm.get_all_vendors()
            known_vendors = [v["display_name"] for v in all_vendors]

            for filename in os.listdir(ORDERS_FOLDER):
                if not filename.lower().endswith(".pdf"):
                    continue

                pdf_path = os.path.join(ORDERS_FOLDER, filename)

                # Try to extract PO number from filename
                po_match = re.search(r"PO_(\d+)\.pdf", filename, re.IGNORECASE)
                po_num = po_match.group(1) if po_match else None

                # Check if this PDF is already matched to a database PO
                # by checking if any database PO's PDF path matches this file
                already_in_db = False
                for db_po in db_pending_pos:
                    db_pdf_path = sender.find_po_pdf(db_po["po"])
                    if db_pdf_path and os.path.basename(db_pdf_path) == filename:
                        already_in_db = True
                        break

                if already_in_db:
                    continue

                # Skip if PO number extracted and it's in database
                if po_num and po_num in db_po_numbers:
                    continue

                # PDF not in database - use fallback extraction
                self.log(
                    f"ℹ️ INFO: Found PDF {filename} not in database, using fallback extraction..."
                )

                # Extract supplier name (returns display name)
                supplier_name = extract_supplier_name(pdf_path, known_vendors, self.log)

                if not supplier_name:
                    self.log(
                        f"⚠ WARNING: Could not extract supplier name from {filename}. Skipping."
                    )
                    continue

                emails = email_map.get(supplier_name)

                if not emails:
                    self.log(
                        f"⚠ WARNING: No email found for supplier '{supplier_name}' from {filename}. Skipping."
                    )
                    continue

                # Add to pending list as a fallback order (no portal access)
                fallback_po = {
                    "po": filename.replace(
                        ".pdf", ""
                    ),  # Use full filename as PO identifier
                    "supplier": supplier_name,
                    "emails": "; ".join(emails),
                    "api_key": None,
                    "is_fallback": True,
                }
                self.pending_pos.append(fallback_po)
                self.po_tree.insert(
                    "",
                    "end",
                    values=(fallback_po["po"], supplier_name, emails[0], "Fallback"),
                )
                self.log(
                    f"✓ SUCCESS: Added {filename} via fallback (supplier: {supplier_name})"
                )

        except Exception as e:
            self.log(f"✗ ERROR: Fallback scan failed: {e}")
            import traceback

            traceback.print_exc()

    def clear_preview(self):
        self.email_to_label.config(text="To:")
        self.email_subject_label.config(text="Subject:")
        self.pdf_status_label.config(text="PDF:")
        self.email_body_text.delete("1.0", tk.END)

    def on_po_select(self, event):
        selected = self.po_tree.selection()
        if not selected:
            return
        po_num = self.po_tree.item(selected[0], "values")[0]
        po_data = next((p for p in self.pending_pos if p["po"] == po_num), None)
        if po_data:
            sender = EmailSender(self.log, self.dm)
            is_fallback = po_data.get("is_fallback", False)
            subject, body = sender.generate_email_content(po_data, is_fallback)
            self.email_to_label.config(text=f"To: {po_data['emails']}")
            self.email_subject_label.config(text=f"Subject: {subject}")
            self.email_body_text.delete("1.0", tk.END)
            self.email_body_text.insert("1.0", body)
            pdf_path = sender.find_po_pdf(po_num)
            self.pdf_status_label.config(
                text=f"PDF: {'Found' if pdf_path else 'Missing'}",
                foreground="green" if pdf_path else "red",
            )

    def _send_emails_threaded(self, pos_to_send):
        method = self.get_send_method()
        use_custom_sig = self.use_custom_sig_var.get()
        
        if not messagebox.askyesno("❓ Confirm", f"Send {len(pos_to_send)} email(s) via {method}?\n{'Using custom signature' if use_custom_sig else 'Using Outlook signature'}", parent=self
        ):
            return

        def send_thread():
            sender = EmailSender(self.log, self.dm)
            sent, failed, summary = sender.send_all_pending_emails(
                preferred_method=method, pos_to_send=pos_to_send, use_custom_signature=use_custom_sig
            )
            self.after(0, lambda: messagebox.showinfo("Result", summary, parent=self))
            self.after(0, self.refresh_data)

        threading.Thread(target=send_thread, daemon=True).start()

    def send_all(self):
        if not self.pending_pos:
            messagebox.showinfo("Info", "No pending emails to send.", parent=self)
            return
        self._send_emails_threaded(self.pending_pos)

    def send_selected(self):
        selected = self.po_tree.selection()
        if not selected:
            messagebox.showwarning(
                "No Selection", "Please select one or more POs to send.", parent=self
            )
            return

        pos_to_send = []
        for item in selected:
            po_num = self.po_tree.item(item, "values")[0]
            po_data = next((p for p in self.pending_pos if p["po"] == po_num), None)
            if po_data:
                pos_to_send.append(po_data)

        if pos_to_send:
            self._send_emails_threaded(pos_to_send)


# Complete replacement for POPreviewWindow class with PDF generation integrated


class POPreviewWindow(tk.Toplevel):
    def __init__(self, parent, log_callback, data_manager, po_number):
        super().__init__(parent)
        self.title(f"PO Preview & Confirmation - {po_number}")
        self.geometry("900x700")  # Increased height for PDF status
        self.log = log_callback
        self.dm = data_manager
        self.po_number = po_number
        self.changes_to_save = {}  # Key: item_id, Value: new_conf_date
        self.pdf_path = None  # Track generated PDF

        # Main layout
        main_frame = ttk.Frame(self, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Header Info Frame
        header_frame = ttk.LabelFrame(
            main_frame, text="PO Header Information", padding=10
        )
        header_frame.pack(fill=tk.X, pady=(0, 10))
        header_frame.columnconfigure(1, weight=1)
        header_frame.columnconfigure(3, weight=1)

        self.vendor_label = ttk.Label(header_frame, text="Supplier:")
        self.vendor_label.grid(row=0, column=0, columnspan=2, sticky="w", padx=5)
        self.terms_label = ttk.Label(header_frame, text="Terms:")
        self.terms_label.grid(row=1, column=0, columnspan=2, sticky="w", padx=5)
        self.transport_label = ttk.Label(header_frame, text="Transport Days:")
        self.transport_label.grid(row=0, column=2, sticky="w", padx=5)

        # PDF Status Frame (NEW)
        pdf_status_frame = ttk.LabelFrame(main_frame, text="PDF Status", padding=10)
        pdf_status_frame.pack(fill=tk.X, pady=(0, 10))

        self.pdf_status_label = ttk.Label(
            pdf_status_frame, text="PDF: Not Generated", foreground="gray"
        )
        self.pdf_status_label.pack(side=tk.LEFT, padx=5)

        ttk.Button(
            pdf_status_frame, text="📄 Generate PDF", command=self.generate_pdf
        ).pack(side=tk.LEFT, padx=5)
        ttk.Button(
            pdf_status_frame, text="Open PDF", command=self.open_pdf, state=tk.DISABLED
        ).pack(side=tk.LEFT, padx=5)

        self.open_pdf_button = pdf_status_frame.winfo_children()[
            -1
        ]  # Reference for enabling later

        # Lines Frame
        lines_frame = ttk.LabelFrame(
            main_frame,
            text="Line Items (Double-click to edit Confirmation Date or Comments)",
            padding=10,
        )
        lines_frame.pack(fill=tk.BOTH, expand=True)

        cols = (
            "Item",
            "Material",
            "Description",
            "Qty",
            "Req. Date",
            "Conf. Date",
            "Reschedule Date",
            "Comments",
            "Status",
        )
        self.lines_tree = ttk.Treeview(
            lines_frame, columns=cols, show="headings", height=10
        )

        for col in cols:
            self.lines_tree.heading(col, text=col)

        self.lines_tree.column("Item", width=50)
        self.lines_tree.column("Material", width=100)
        self.lines_tree.column("Description", width=200)
        self.lines_tree.column("Qty", width=60)
        self.lines_tree.column("Req. Date", width=80)
        self.lines_tree.column("Conf. Date", width=80)
        self.lines_tree.column("Reschedule Date", width=100)
        self.lines_tree.column("Comments", width=150)
        self.lines_tree.column("Status", width=60)

        self.lines_tree.pack(fill=tk.BOTH, expand=True)
        self.lines_tree.bind("<Double-1>", self.edit_conf_date)

        # Action buttons
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(10, 0))

        ttk.Button(btn_frame, text="💾 Save Changes", command=self.save_changes).pack(
            side=tk.LEFT, padx=5
        )
        
        # Add custom signature checkbox
        self.use_custom_sig_var = tk.BooleanVar(value=self.dm.get_config('use_custom_signature', False))
        ttk.Checkbutton(
            btn_frame,
            text="Use Custom Signature",
            variable=self.use_custom_sig_var
        ).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            btn_frame, text="Generate & Send Email", command=self.generate_and_send
        ).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Close", command=self.destroy).pack(
            side=tk.RIGHT, padx=5
        )

        # Load data
        self.load_po_data()
        self.check_pdf_exists()

    def load_po_data(self):
        """Load PO data from database"""
        query = """
            SELECT oo.item, oo.material_code, oo.short_text, oo.requested_qty, 
                   oo.requested_del_date, oo.conf_delivery_date, oo.rescheduling_date, oo.comments, oo.status,
                   v.display_name, v.delivery_terms, v.payment_terms, v.transport_days
            FROM open_orders oo
            JOIN vendors v ON oo.vendor_name = v.vendor_name
            WHERE oo.po = ? AND oo.status = 'Open'
            ORDER BY oo.item
        """
        results = self.dm.db.execute_query(query, (self.po_number,), fetchall=True)

        if not results:
            messagebox.showwarning(
                "No Data",
                f"No open order lines found for PO {self.po_number}",
                parent=self,
            )
            self.destroy()
            return

        # Update header info from first line
        first_line = results[0]
        self.vendor_label.config(
            text=f"Supplier: {first_line.get('display_name', 'N/A')}"
        )
        self.terms_label.config(
            text=f"Payment Terms: {first_line.get('payment_terms', 'N/A')} | Delivery Terms: {first_line.get('delivery_terms', 'N/A')}"
        )
        self.transport_label.config(
            text=f"Transport Days: {first_line.get('transport_days', 0)}"
        )

        # Populate lines
        for line in results:
            # Format dates
            req_date = line.get("requested_del_date", "")
            conf_date = line.get("conf_delivery_date", "")
            reschedule_date = line.get("rescheduling_date", "")

            if req_date:
                req_date = (
                    datetime.fromisoformat(req_date).strftime("%d.%m.%Y")
                    if "T" in req_date
                    else req_date
                )
            if conf_date:
                conf_date = (
                    datetime.fromisoformat(conf_date).strftime("%d.%m.%Y")
                    if "T" in conf_date
                    else conf_date
                )
            if reschedule_date:
                reschedule_date = (
                    datetime.fromisoformat(reschedule_date).strftime("%d.%m.%Y")
                    if "T" in reschedule_date
                    else reschedule_date
                )

            values = (
                line.get("item", ""),
                line.get("material_code", ""),
                line.get("short_text", ""),
                line.get("requested_qty", ""),
                req_date,
                conf_date,
                reschedule_date,
                line.get("comments", ""),
                line.get("status", ""),
            )
            self.lines_tree.insert("", "end", values=values)

    def edit_conf_date(self, event):
        """Edit confirmation date and comments for selected line"""
        selected_iid = self.lines_tree.focus()
        if not selected_iid:
            return

        # Create a dialog for date and comments entry
        dialog = tk.Toplevel(self)
        dialog.title("Edit Line Item")
        dialog.transient(self)
        dialog.grab_set()
        dialog.geometry("450x250")

        item_values = self.lines_tree.item(selected_iid, "values")
        item_num = item_values[0]
        current_date = item_values[5]
        current_comments = item_values[7]

        # Main frame
        main_frame = ttk.Frame(dialog, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Header
        ttk.Label(
            main_frame, 
            text=f"Edit Item {item_num}", 
            font=("Helvetica", 10, "bold")
        ).pack(pady=(0, 10))

        # Confirmation Date section
        date_frame = ttk.LabelFrame(main_frame, text="Confirmation Date", padding=10)
        date_frame.pack(fill=tk.X, pady=5)

        ttk.Label(date_frame, text="Format: DD.MM.YYYY").pack(anchor="w")
        date_var = tk.StringVar(value=current_date)
        date_entry = ttk.Entry(date_frame, textvariable=date_var, width=40)
        date_entry.pack(fill=tk.X, pady=5)

        # Comments section
        comments_frame = ttk.LabelFrame(main_frame, text="Comments", padding=10)
        comments_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        comments_text = tk.Text(comments_frame, height=5, width=40, wrap=tk.WORD)
        comments_text.insert("1.0", current_comments)
        comments_text.pack(fill=tk.BOTH, expand=True)

        # Buttons
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=10)

        def on_ok():
            new_date = date_var.get().strip()
            new_comments = comments_text.get("1.0", tk.END).strip()
            
            # Basic validation (format DD.MM.YYYY)
            if new_date and not re.match(r"^\d{2}\.\d{2}\.\d{4}$", new_date):
                messagebox.showwarning(
                    "Invalid Format",
                    "Please use DD.MM.YYYY format for date or leave blank.",
                    parent=dialog,
                )
                return

            # Update the treeview visually
            current_values = list(self.lines_tree.item(selected_iid, "values"))
            current_values[5] = new_date
            current_values[7] = new_comments
            self.lines_tree.item(selected_iid, values=tuple(current_values))

            # Store the changes to be saved later
            # Store as tuple: (new_date, new_comments)
            self.changes_to_save[item_num] = (new_date, new_comments)

            # Add a visual tag to show the line has been edited
            self.lines_tree.item(selected_iid, tags=("edited",))
            self.lines_tree.tag_configure("edited", background="lightyellow")

            dialog.destroy()

        ttk.Button(btn_frame, text="OK", command=on_ok).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Cancel", command=dialog.destroy).pack(
            side=tk.LEFT, padx=5
        )

        # Focus on date entry
        date_entry.focus()

    def save_changes(self):
        """Save confirmation date and comments changes to database"""
        if not self.changes_to_save:
            messagebox.showinfo(
                "No Changes", "No changes were made.", parent=self
            )
            return

        conf_date_updates = []
        comment_updates = []
        
        for item_num, changes in self.changes_to_save.items():
            # Changes is now a tuple: (new_date, new_comments)
            if isinstance(changes, tuple):
                conf_date, comments = changes
                # Update confirmation date
                if conf_date:
                    conf_date_updates.append((conf_date, self.po_number, item_num))
                # Update comments
                comment_updates.append((comments, self.po_number, item_num))
            else:
                # Backward compatibility: if only date was stored
                conf_date_updates.append((changes, self.po_number, item_num))

        try:
            updated_count = 0
            
            # Update confirmation dates
            if conf_date_updates:
                updated_count += self.dm.update_confirmation_dates(conf_date_updates)
            
            # Update comments
            if comment_updates:
                for comments, po, item in comment_updates:
                    self.dm.db.execute_query(
                        "UPDATE open_orders SET comments = ? WHERE po = ? AND item = ?",
                        (comments, po, item),
                        commit=True
                    )
                    updated_count += 1
            
            self.log(
                f"ℹ️ INFO: Updated {updated_count} lines on PO {self.po_number}."
            )
            messagebox.showinfo("✅ Success",
                f"Successfully updated {updated_count} line(s).\n\n"
                "TIP: You may want to regenerate the PDF with updated information.",
                parent=self,
            )

            # Clear the changes after saving
            self.changes_to_save.clear()

            # Remove yellow highlighting
            for item in self.lines_tree.get_children():
                self.lines_tree.item(item, tags=())

            # Reload data to show saved changes
            for item in self.lines_tree.get_children():
                self.lines_tree.delete(item)
            self.load_po_data()

        except Exception as e:
            self.log(f"✗ ERROR: Failed to save changes: {e}")
            messagebox.showerror("❌ Error", f"Failed to save changes:\n{str(e)}", parent=self
            )

    def check_pdf_exists(self):
        """Check if PDF already exists for this PO"""
        sender = EmailSender(self.log, self.dm)
        pdf_path = sender.find_po_pdf(self.po_number)

        if pdf_path:
            self.pdf_path = pdf_path
            self.pdf_status_label.config(
                text=f"PDF: Found ({os.path.basename(pdf_path)})", foreground="green"
            )
            self.open_pdf_button.config(state=tk.NORMAL)
        else:
            self.pdf_status_label.config(text="PDF: Not Generated", foreground="gray")
            self.open_pdf_button.config(state=tk.DISABLED)

    def generate_pdf(self):
        """Generate PDF for this PO"""
        if self.changes_to_save:
            if not messagebox.askyesno(
                "Unsaved Changes",
                "You have unsaved confirmation date changes.\n\n"
                "Do you want to save them before generating the PDF?",
                parent=self,
            ):
                return
            self.save_changes()

        self.log(f"ℹ️ INFO: Generating PDF for PO {self.po_number}...")

        try:
            # Get PO data from database
            query = """
                SELECT oo.po, oo.item, oo.material_code, oo.short_text, oo.requested_qty, 
                       oo.requested_del_date, oo.conf_delivery_date, oo.rescheduling_date,
                       oo.unit, oo.unit_price, oo.total_amount, oo.currency,
                       v.display_name AS "Name", v.address as "Vendor Address", v.transport_days
                FROM open_orders oo
                JOIN vendors v ON oo.vendor_name = v.vendor_name
                WHERE oo.po = ? AND oo.status = 'Open'
                ORDER BY oo.item
            """
            results = self.dm.db.execute_query(query, (self.po_number,), fetchall=True)

            if not results:
                messagebox.showerror("❌ Error", "No data found for PDF generation", parent=self
                )
                return

            # Create DataFrame
            lines_df = pd.DataFrame(results)

            # Generate PDF
            pdf_buffer = io.BytesIO()
            if self.dm._generate_single_po_pdf(pdf_buffer, self.po_number, lines_df):
                # Save PDF
                pdf_path = os.path.join(ORDERS_FOLDER, f"PO_{self.po_number}.pdf")
                with open(pdf_path, "wb") as f:
                    f.write(pdf_buffer.getvalue())

                self.pdf_path = pdf_path
                self.pdf_status_label.config(
                    text=f"PDF: Generated ({os.path.basename(pdf_path)})",
                    foreground="green",
                )
                self.open_pdf_button.config(state=tk.NORMAL)

                # Update database status
                self.dm.mark_pos_as_created([self.po_number])

                self.log(f"✓ SUCCESS: Generated PDF for PO {self.po_number}")
                messagebox.showinfo("✅ Success", f"PDF generated successfully!\n\n{pdf_path}", parent=self
                )
            else:
                messagebox.showerror("❌ Error", "Failed to generate PDF", parent=self)

        except Exception as e:
            self.log(f"✗ ERROR: Failed to generate PDF: {e}")
            messagebox.showerror("❌ Error", f"Failed to generate PDF:\n{str(e)}", parent=self
            )
            import traceback

            traceback.print_exc()

    def open_pdf(self):
        """Open the generated PDF in default viewer"""
        if not self.pdf_path or not os.path.exists(self.pdf_path):
            messagebox.showwarning(
                "PDF Not Found",
                "PDF file not found. Please generate it first.",
                parent=self,
            )
            return

        try:
            import subprocess
            import platform

            if platform.system() == "Windows":
                os.startfile(self.pdf_path)
            elif platform.system() == "Darwin":  # macOS
                subprocess.run(["open", self.pdf_path])
            else:  # Linux
                subprocess.run(["xdg-open", self.pdf_path])

            self.log(f"ℹ️ INFO: Opened PDF: {self.pdf_path}")

        except Exception as e:
            self.log(f"✗ ERROR: Failed to open PDF: {e}")
            messagebox.showerror("❌ Error", f"Failed to open PDF:\n{str(e)}", parent=self)

    def generate_and_send(self):
        """Generate PDF (if needed) and send email"""
        # Check if PDF exists
        if not self.pdf_path or not os.path.exists(self.pdf_path):
            if not messagebox.askyesno(
                "Generate PDF", "PDF not found. Generate it now?", parent=self
            ):
                return
            self.generate_pdf()

            # Check again if generation was successful
            if not self.pdf_path or not os.path.exists(self.pdf_path):
                return

        # Check for unsaved changes
        if self.changes_to_save:
            messagebox.showwarning(
                "Unsaved Changes",
                "Please save your changes and regenerate the PDF before sending.",
                parent=self,
            )
            return

        # Get PO data for email
        query = """
            SELECT v.display_name, v.emails, v.api_key
            FROM open_orders oo
            JOIN vendors v ON oo.vendor_name = v.vendor_name
            WHERE oo.po = ?
            LIMIT 1
        """
        result = self.dm.db.execute_query(query, (self.po_number,), fetchone=True)

        if not result or not result.get("emails"):
            messagebox.showerror("❌ Error",
                f"No email found for supplier. Please update vendor contact info.",
                parent=self,
            )
            return

        # Create PO data for email sender
        po_data = {
            "po": self.po_number,
            "supplier": result["display_name"],
            "emails": result["emails"],
            "api_key": result.get("api_key"),
        }

        # Send email
        try:
            sender = EmailSender(self.log, self.dm)
            method = "Outlook" if OUTLOOK_AVAILABLE else "SMTP"
            use_custom_sig = self.use_custom_sig_var.get()

            if messagebox.askyesno(
                "Confirm Send",
                f"Send PO {self.po_number} to {result['display_name']} via {method}?\n{'Using custom signature' if use_custom_sig else 'Using Outlook signature'}",
                parent=self,
            ):
                subject, body = sender.generate_po_email_content(po_data)

                if method == "Outlook":
                    success, message = sender._send_outlook(
                        result["emails"], subject, body, self.pdf_path, use_custom_sig
                    )
                else:
                    success, message = sender._send_smtp(
                        result["emails"], subject, body, self.pdf_path, use_custom_sig
                    )

                if success:
                    # Mark as sent and move PDF
                    self.dm.mark_email_sent(self.po_number)
                    sent_pdf_path = os.path.join(
                        ORDERS_SENT_FOLDER, os.path.basename(self.pdf_path)
                    )
                    os.rename(self.pdf_path, sent_pdf_path)
                    self.pdf_path = sent_pdf_path

                    self.log(
                        f"✓ SUCCESS: Sent PO {self.po_number} to {result['display_name']}"
                    )
                    messagebox.showinfo("✅ Success",
                        f"Email sent successfully to {result['display_name']}!",
                        parent=self,
                    )
                    self.destroy()
                else:
                    messagebox.showerror("❌ Error", f"Failed to send email:\n{message}", parent=self
                    )

        except Exception as e:
            self.log(f"✗ ERROR: Failed to send email: {e}")
            messagebox.showerror("❌ Error", f"Failed to send email:\n{str(e)}", parent=self
            )
            import traceback

            traceback.print_exc()


class EmailReminderWindow(tk.Toplevel):
    def __init__(self, parent, log_callback, data_manager, send_method_getter):
        super().__init__(parent)
        self.title("Send Reminder Emails for Unconfirmed Orders")
        self.geometry("800x950")
        self.log = log_callback
        self.dm = data_manager
        self.get_send_method = send_method_getter

        # Create scrollable main frame
        main_canvas = tk.Canvas(self)
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=main_canvas.yview)
        main_frame = ttk.Frame(main_canvas)
        main_frame.bind(
            "<Configure>",
            lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all")),
        )
        main_canvas.create_window((0, 0), window=main_frame, anchor="nw")
        main_canvas.configure(yscrollcommand=scrollbar.set)
        main_canvas.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        scrollbar.pack(side="right", fill="y")

        # Filter Section
        filter_frame = ttk.LabelFrame(main_frame, text="Advanced Filters", padding=10)
        filter_frame.pack(fill=tk.X, pady=5)

        # Search Fields Frame
        search_frame = ttk.Frame(filter_frame)
        search_frame.pack(fill=tk.X, pady=5)
        search_frame.columnconfigure(1, weight=1)
        search_frame.columnconfigure(3, weight=1)

        # Supplier Search
        ttk.Label(search_frame, text="Search Supplier:").grid(
            row=0, column=0, sticky="w", padx=(0, 5)
        )
        self.supplier_search_var = tk.StringVar()
        self.supplier_search_var.trace("w", self.filter_suppliers)
        ttk.Entry(search_frame, textvariable=self.supplier_search_var).grid(
            row=0, column=1, sticky="ew", padx=(0, 10)
        )

        # PO Search
        ttk.Label(search_frame, text="Search PO:").grid(
            row=0, column=2, sticky="w", padx=(10, 5)
        )
        self.po_search_var = tk.StringVar()
        ttk.Entry(search_frame, textvariable=self.po_search_var).grid(
            row=0, column=3, sticky="ew"
        )

        # Vendor Selection
        vendor_select_frame = ttk.Frame(filter_frame)
        vendor_select_frame.pack(fill=tk.X, pady=5)

        ttk.Label(vendor_select_frame, text="Suppliers:").pack(anchor="w")
        vendor_btn_frame = ttk.Frame(vendor_select_frame)
        vendor_btn_frame.pack(fill=tk.X, pady=2)
        ttk.Button(
            vendor_btn_frame, text="Select All", command=self.select_all_vendors
        ).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(
            vendor_btn_frame, text="Deselect All", command=self.deselect_all_vendors
        ).pack(side=tk.LEFT)

        self.vendor_listbox = Listbox(
            vendor_select_frame, selectmode=tk.MULTIPLE, exportselection=False, height=6
        )
        vendor_scroll = ttk.Scrollbar(
            vendor_select_frame, orient="vertical", command=self.vendor_listbox.yview
        )
        self.vendor_listbox.configure(yscrollcommand=vendor_scroll.set)
        self.vendor_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vendor_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # Date Filters
        date_frame = ttk.LabelFrame(
            filter_frame, text="Requested Delivery Date Range", padding=5
        )
        date_frame.pack(fill=tk.X, pady=5)
        date_frame.columnconfigure(1, weight=1)
        date_frame.columnconfigure(3, weight=1)

        self.date_vars = {"start_date": tk.StringVar(), "end_date": tk.StringVar()}

        ttk.Label(date_frame, text="From (YYYY-MM-DD):").grid(
            row=0, column=0, sticky="w", padx=(0, 5)
        )
        ttk.Entry(date_frame, textvariable=self.date_vars["start_date"], width=12).grid(
            row=0, column=1, sticky="ew", padx=(0, 10)
        )
        ttk.Label(date_frame, text="To (YYYY-MM-DD):").grid(
            row=0, column=2, sticky="w", padx=(10, 5)
        )
        ttk.Entry(date_frame, textvariable=self.date_vars["end_date"], width=12).grid(
            row=0, column=3, sticky="ew"
        )

        # Value Filters
        value_frame = ttk.LabelFrame(filter_frame, text="Order Value Range", padding=5)
        value_frame.pack(fill=tk.X, pady=5)
        value_frame.columnconfigure(1, weight=1)
        value_frame.columnconfigure(3, weight=1)

        self.value_vars = {"min_val": tk.StringVar(), "max_val": tk.StringVar()}

        ttk.Label(value_frame, text="Min Value:").grid(
            row=0, column=0, sticky="w", padx=(0, 5)
        )
        ttk.Entry(value_frame, textvariable=self.value_vars["min_val"], width=12).grid(
            row=0, column=1, sticky="ew", padx=(0, 10)
        )
        ttk.Label(value_frame, text="Max Value:").grid(
            row=0, column=2, sticky="w", padx=(10, 5)
        )
        ttk.Entry(value_frame, textvariable=self.value_vars["max_val"], width=12).grid(
            row=0, column=3, sticky="ew"
        )

        # Age Filter
        age_frame = ttk.LabelFrame(
            filter_frame, text="Order Age (Days without confirmation)", padding=5
        )
        age_frame.pack(fill=tk.X, pady=5)

        self.min_age_var = tk.StringVar(value="3")
        ttk.Label(age_frame, text="Minimum days since PO creation:").pack(
            side=tk.LEFT, padx=(0, 5)
        )
        ttk.Entry(age_frame, textvariable=self.min_age_var, width=10).pack(side=tk.LEFT)

        # Preview Section
        preview_frame = ttk.LabelFrame(
            main_frame, text="Preview - Unconfirmed Orders", padding=10
        )
        preview_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        preview_btn_frame = ttk.Frame(preview_frame)
        preview_btn_frame.pack(fill=tk.X, pady=(0, 5))
        ttk.Button(
            preview_btn_frame, text="Load Preview", command=self.load_preview
        ).pack(side=tk.LEFT)
        self.preview_count_label = ttk.Label(preview_btn_frame, text="")
        self.preview_count_label.pack(side=tk.LEFT, padx=10)

        # ETD Calculation Option
        etd_frame = ttk.LabelFrame(filter_frame, text="ETD Calculation", padding=5)
        etd_frame.pack(fill=tk.X, pady=5)

        self.include_etd_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            etd_frame,
            text="Calculate and display ETD date (subtracts transport days from rescheduled delivery date)",
            variable=self.include_etd_var,
        ).pack(anchor="w")

        etd_help = ttk.Label(
            etd_frame,
            text="When checked: ETD = Reschedule Date - Transport Days (working days only)",
            font=("Helvetica", 8),
            foreground="gray",
        )
        etd_help.pack(anchor="w", pady=(5, 0))

        # Preview Tree
        cols = (
            "PO",
            "Vendor",
            "Lines",
            "Total Value",
            "Currency",
            "Oldest Req. Date",
            "Days Old",
        )
        self.preview_tree = ttk.Treeview(
            preview_frame, columns=cols, show="headings", height=12
        )
        for col in cols:
            self.preview_tree.heading(col, text=col)
        self.preview_tree.column("PO", width=100)
        self.preview_tree.column("Vendor", width=180)
        self.preview_tree.column("Lines", width=60)
        self.preview_tree.column("Total Value", width=100)
        self.preview_tree.column("Currency", width=70)
        self.preview_tree.column("Oldest Req. Date", width=120)
        self.preview_tree.column("Days Old", width=80)

        preview_scroll = ttk.Scrollbar(
            preview_frame, orient="vertical", command=self.preview_tree.yview
        )
        self.preview_tree.configure(yscrollcommand=preview_scroll.set)
        self.preview_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        preview_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # Action Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=10)

        ttk.Button(button_frame, text="Cancel", command=self.destroy).pack(side=tk.LEFT)
        
        # Add custom signature checkbox
        self.use_custom_sig_var = tk.BooleanVar(value=self.dm.get_config('use_custom_signature', False))
        ttk.Checkbutton(
            button_frame,
            text="Use Custom Signature",
            variable=self.use_custom_sig_var
        ).pack(side=tk.LEFT, padx=10)
        
        ttk.Button(
            button_frame, text="Generate PDFs Only", command=self.generate_pdfs_only
        ).pack(side=tk.RIGHT)
        ttk.Button(
            button_frame,
            text="Generate PDFs & Send Reminders",
            command=self.generate_and_send,
        ).pack(side=tk.RIGHT, padx=(5, 0))

        self.load_vendors()

    def load_vendors(self):
        """Load vendors into the listbox"""
        self.all_vendors = [
            vendor["display_name"] for vendor in self.dm.get_all_vendors()
        ]
        self.filter_suppliers()

    def filter_suppliers(self, *args):
        """Filter supplier listbox based on search term"""
        search_term = self.supplier_search_var.get().lower()
        self.vendor_listbox.delete(0, tk.END)

        for vendor in self.all_vendors:
            if search_term in vendor.lower():
                self.vendor_listbox.insert(tk.END, vendor)

    def select_all_vendors(self):
        self.vendor_listbox.select_set(0, tk.END)

    def deselect_all_vendors(self):
        self.vendor_listbox.select_clear(0, tk.END)

    def load_preview(self):
        """Load and display unconfirmed orders based on filters"""
        for item in self.preview_tree.get_children():
            self.preview_tree.delete(item)

        try:
            filters = self.get_filters()
            unconfirmed_orders = self.get_unconfirmed_orders(filters)

            if not unconfirmed_orders:
                self.preview_count_label.config(text="No unconfirmed orders found")
                messagebox.showinfo(
                    "No Data",
                    "No unconfirmed orders match the specified criteria.",
                    parent=self,
                )
                return

            # Group by PO
            po_groups = {}
            for order in unconfirmed_orders:
                po_num = order["po"]
                if po_num not in po_groups:
                    po_groups[po_num] = {
                        "vendor": order["name"],
                        "lines": [],
                        "total_value": 0,
                        "currency": order.get("currency", "EUR"),
                        "oldest_date": None,
                        "days_old": 0,
                    }

                po_groups[po_num]["lines"].append(order)
                po_groups[po_num]["total_value"] += float(order.get("total_amount", 0))

                # Track oldest requested date
                req_date_str = order.get("requested_del_date")
                if req_date_str:
                    try:
                        req_date = datetime.strptime(req_date_str, "%d.%m.%Y")
                        if (
                            po_groups[po_num]["oldest_date"] is None
                            or req_date < po_groups[po_num]["oldest_date"]
                        ):
                            po_groups[po_num]["oldest_date"] = req_date
                            days_old = (datetime.now() - req_date).days
                            po_groups[po_num]["days_old"] = days_old
                    except Exception:  # TODO: Add proper error handling
                        pass  # TODO: Add proper error handling
            # Display in tree
            for po_num, data in po_groups.items():
                oldest_date_str = (
                    data["oldest_date"].strftime("%d.%m.%Y")
                    if data["oldest_date"]
                    else "N/A"
                )
                values = (
                    po_num,
                    data["vendor"],
                    len(data["lines"]),
                    f"{data['total_value']:.2f}",
                    data["currency"],
                    oldest_date_str,
                    data["days_old"],
                )
                self.preview_tree.insert("", "end", values=values)

            self.preview_count_label.config(
                text=f"Found {len(po_groups)} PO(s) with {len(unconfirmed_orders)} unconfirmed line(s)"
            )
            self.unconfirmed_data = po_groups  # Store for PDF generation

        except Exception as e:
            self.log(f"✗ ERROR: Failed to load preview: {e}")
            messagebox.showerror("❌ Error", f"Failed to load preview:\n{str(e)}", parent=self
            )

    def get_filters(self):
        """Collect all filter values"""
        selected_indices = self.vendor_listbox.curselection()
        selected_vendors = [self.vendor_listbox.get(i) for i in selected_indices]

        def get_val(var):
            val = var.get().strip()
            return val if val else None

        filters = {
            "suppliers": selected_vendors if selected_vendors else None,
            "po_search": get_val(self.po_search_var),
            "start_date": get_val(self.date_vars["start_date"]),
            "end_date": get_val(self.date_vars["end_date"]),
            "min_age_days": (
                int(self.min_age_var.get()) if self.min_age_var.get().strip() else 0
            ),
        }

        # Value filters
        try:
            min_val = get_val(self.value_vars["min_val"])
            max_val = get_val(self.value_vars["max_val"])
            filters["min_value"] = float(min_val) if min_val else None
            filters["max_value"] = float(max_val) if max_val else None
        except ValueError:
            messagebox.showerror(
                "Input Error", "Value range must contain valid numbers.", parent=self
            )
            raise

        return filters

    def get_unconfirmed_orders(self, filters):
        """Query database for unconfirmed orders with filters"""
        query = """
            SELECT oo.*, v.display_name as name, v.emails
            FROM open_orders oo
            JOIN vendors v ON oo.vendor_name = v.vendor_name
            WHERE oo.status = 'Open' 
            AND (oo.conf_delivery_date IS NULL OR oo.conf_delivery_date = '')
        """

        params = []

        # Supplier filter
        if filters.get("suppliers"):
            placeholders = ",".join("?" * len(filters["suppliers"]))
            query += f" AND v.display_name IN ({placeholders})"
            params.extend(filters["suppliers"])

        # PO search filter
        if filters.get("po_search"):
            query += " AND oo.po LIKE ?"
            params.append(f"%{filters['po_search']}%")

        orders = self.dm.db.execute_query(query, tuple(params), fetchall=True)

        if not orders:
            return []

        # Convert to DataFrame for easier filtering
        df = pd.DataFrame(orders)
        df["requested_del_date_dt"] = pd.to_datetime(
            df["requested_del_date"], format="%d.%m.%Y", errors="coerce"
        )
        df["total_amount"] = pd.to_numeric(df["total_amount"], errors="coerce").fillna(
            0
        )

        # Date filters - ONLY apply to orders that HAVE dates
        # Orders without dates should still appear (they need confirmation!)
        if filters.get("start_date"):
            start_date = pd.to_datetime(filters["start_date"])
            # Keep orders with no date OR dates >= start_date
            df = df[(df["requested_del_date_dt"].isna()) | (df["requested_del_date_dt"] >= start_date)]

        if filters.get("end_date"):
            end_date = pd.to_datetime(filters["end_date"])
            # Keep orders with no date OR dates <= end_date
            df = df[(df["requested_del_date_dt"].isna()) | (df["requested_del_date_dt"] <= end_date)]

        # Value filters
        if filters.get("min_value") is not None:
            df = df[df["total_amount"] >= filters["min_value"]]

        if filters.get("max_value") is not None:
            df = df[df["total_amount"] <= filters["max_value"]]

        # Age filter - ONLY apply to orders that HAVE dates
        # Orders without dates should still appear (they need confirmation urgently!)
        if filters.get("min_age_days", 0) > 0:
            today = datetime.now()
            # Keep orders with no date OR dates older than min_age_days
            df = df[
                df["requested_del_date_dt"].apply(
                    lambda x: (
                        pd.isna(x) or  # Keep orders without dates
                        (today - x).days >= filters["min_age_days"]  # Keep old enough orders
                    )
                )
            ]

        return df.to_dict("records")

    def generate_pdfs_only(self):
        """Generate reminder PDFs without sending emails"""
        if not hasattr(self, "unconfirmed_data") or not self.unconfirmed_data:
            messagebox.showwarning("No Data", "Please load preview first.", parent=self)
            return

        self.generate_reminder_pdfs()

    def generate_and_send(self):
        """Generate PDFs and send reminder emails"""
        if not hasattr(self, "unconfirmed_data") or not self.unconfirmed_data:
            messagebox.showwarning("No Data", "Please load preview first.", parent=self)
            return

        if not messagebox.askyesno("❓ Confirm",
            f"Generate PDFs and send reminder emails to {len(self.unconfirmed_data)} supplier(s)?",
            parent=self,
        ):
            return

        # Generate PDFs first
        pdf_paths = self.generate_reminder_pdfs()

        if not pdf_paths:
            messagebox.showerror("❌ Error", "Failed to generate PDFs. Emails not sent.", parent=self
            )
            return

        # Send emails
        self.send_reminder_emails(pdf_paths)

    def generate_and_send(self):
        """Generate PDFs and send reminder emails"""
        if not hasattr(self, "unconfirmed_data") or not self.unconfirmed_data:
            messagebox.showwarning("No Data", "Please load preview first.", parent=self)
            return

        # Ask user which format to use
        format_choice = messagebox.askquestion(
            "PDF Format",
            "Generate separate PDFs per PO?\n\n"
            "Yes = One PDF per PO\n"
            "No = One summary PDF per supplier",
            parent=self,
        )

        use_separate = format_choice == "yes"

        if not messagebox.askyesno("❓ Confirm",
            f"Generate PDFs and send reminder emails to {len(set(data['vendor'] for data in self.unconfirmed_data.values()))} supplier(s)?",
            parent=self,
        ):
            return

        # Generate PDFs based on choice
        if use_separate:
            pdf_paths = self.generate_reminder_pdfs()
        else:
            pdf_paths = self.generate_summary_pdfs()

        if not pdf_paths:
            messagebox.showerror("❌ Error", "Failed to generate PDFs. Emails not sent.", parent=self
            )
            return

        # Send emails
        self.send_reminder_emails(pdf_paths)

    def generate_pdfs_only(self):
        """Generate reminder PDFs without sending emails"""
        if not hasattr(self, "unconfirmed_data") or not self.unconfirmed_data:
            messagebox.showwarning("No Data", "Please load preview first.", parent=self)
            return

        # Ask user which format to use
        format_choice = messagebox.askquestion(
            "PDF Format",
            "Generate separate PDFs per PO?\n\n"
            "Yes = One PDF per PO\n"
            "No = One summary PDF per supplier",
            parent=self,
        )

        if format_choice == "yes":
            self.generate_reminder_pdfs()
        else:
            self.generate_summary_pdfs()

    def generate_summary_pdfs(self):
        """Generate one summary PDF per vendor with all their unconfirmed orders"""
        try:
            pdf_paths = {}
            output_folder = os.path.join(APP_DATA_FOLDER, "ReminderPDFs")
            os.makedirs(output_folder, exist_ok=True)

            # Group by vendor
            vendor_orders = {}
            for po_num, data in self.unconfirmed_data.items():
                vendor_name = data["vendor"]
                if vendor_name not in vendor_orders:
                    vendor_orders[vendor_name] = []
                vendor_orders[vendor_name].append((po_num, data))

            # Create one PDF per vendor
            for vendor_name, orders in vendor_orders.items():
                safe_vendor = re.sub(r'[\\/*?:"<>|]', "_", vendor_name)
                pdf_filename = f"Reminder_Summary_{safe_vendor}_{datetime.now().strftime('%Y%m%d')}.pdf"
                pdf_path = os.path.join(output_folder, pdf_filename)

                if self.create_summary_pdf(pdf_path, vendor_name, orders):
                    pdf_paths[vendor_name] = [pdf_path]
                    self.log(
                        f"✓ SUCCESS: Generated summary reminder PDF for {vendor_name}"
                    )

            messagebox.showinfo("✅ Success",
                f"Generated {len(pdf_paths)} summary PDF(s) in:\n{output_folder}",
                parent=self,
            )
            return pdf_paths

        except Exception as e:
            self.log(f"✗ ERROR: Failed to generate summary PDFs: {e}")
            messagebox.showerror("❌ Error", f"Failed to generate PDFs:\n{str(e)}", parent=self
            )
            return {}

    def create_summary_pdf(self, pdf_path, vendor_name, orders):
        """Create a summary PDF with all unconfirmed orders for a vendor"""
        try:
            c = canvas.Canvas(pdf_path, pagesize=letter)
            width, height = letter
            config = self.dm.get_config("company_config", {})

            # Get transport days for ETD calculation
            include_etd = self.include_etd_var.get()
            transport_days = 0

            if include_etd:
                vendor_data = next(
                    (
                        v
                        for v in self.dm.get_all_vendors()
                        if v["display_name"] == vendor_name
                    ),
                    None,
                )
                if vendor_data:
                    transport_days = int(vendor_data.get("transport_days", 0))

            # Header
            c.setFont("Helvetica-Bold", 18)
            c.drawString(0.5 * inch, height - 0.5 * inch, "SUPPLIER REMINDER SUMMARY")

            c.setFont("Helvetica", 11)
            c.drawString(
                0.5 * inch,
                height - 0.8 * inch,
                f"Date: {datetime.now().strftime('%d.%m.%Y')}",
            )
            c.drawString(0.5 * inch, height - 1.0 * inch, f"Supplier: {vendor_name}")

            if include_etd and transport_days > 0:
                c.setFont("Helvetica", 10)
                c.drawString(
                    0.5 * inch,
                    height - 1.2 * inch,
                    f"Transport Days: {transport_days} working days",
                )

            # Message
            y_pos = height - 1.4 * inch if not include_etd else height - 1.6 * inch
            c.setFont("Helvetica", 10)
            message_lines = [
                "Dear Supplier,",
                "",
                "This is a reminder that we have not yet received confirmation for the following purchase orders.",
                "Please review and provide delivery confirmation for all line items at your earliest convenience.",
                "",
            ]

            for line in message_lines:
                c.drawString(0.5 * inch, y_pos, line)
                y_pos -= 0.2 * inch

            # List all POs and their lines
            y_pos -= 0.2 * inch

            for po_num, data in sorted(orders, key=lambda x: x[0]):
                lines = data["lines"]
                currency = data["currency"]

                # Check if we need a new page
                if y_pos < 2 * inch:
                    c.showPage()
                    y_pos = height - 1 * inch

                # PO Header
                c.setFont("Helvetica-Bold", 12)
                c.drawString(0.5 * inch, y_pos, f"PO {po_num}")
                y_pos -= 0.25 * inch

                # Line items
                c.setFont("Helvetica", 9)
                po_total = 0

                for line in lines:
                    # Check page space
                    if y_pos < 1 * inch:
                        c.showPage()
                        y_pos = height - 1 * inch
                        c.setFont("Helvetica", 9)

                    # Determine which date to display
                    display_date = ""
                    date_label = "Req. Date"

                    if include_etd:
                        # Check if reschedule date exists
                        reschedule_date_str = line.get("rescheduling_date")

                        if reschedule_date_str and reschedule_date_str.strip():
                            # Use reschedule date to calculate ETD
                            try:
                                reschedule_date = datetime.strptime(
                                    reschedule_date_str, "%d.%m.%Y"
                                )
                                if transport_days > 0:
                                    etd_date = subtract_working_days(
                                        reschedule_date, transport_days
                                    )
                                else:
                                    etd_date = reschedule_date
                                display_date = etd_date.strftime("%d.%m.%Y")
                                date_label = "ETD"
                            except Exception:
                                display_date = reschedule_date_str
                                date_label = "Reschedule"
                        else:
                            # No reschedule date, use requested date
                            display_date = line.get("requested_del_date", "N/A")
                            date_label = "Req. Date"
                    else:
                        # ETD not enabled, use requested date
                        display_date = line.get("requested_del_date", "N/A")

                    # Bullet point with line details
                    line_text = f" Line {line.get('item', 'N/A')}  {line.get('material_code', 'N/A')}  {line.get('short_text', 'N/A')[:40]}"
                    c.drawString(0.75 * inch, y_pos, line_text)
                    y_pos -= 0.15 * inch

                    # Quantity and dates on second line
                    detail_text = f"  Qty: {line.get('requested_qty', 'N/A')} {line.get('unit', 'EA')} | "
                    detail_text += (
                        f"Value: {float(line.get('total_amount', 0)):.2f} {currency} | "
                    )
                    detail_text += f"{date_label}: {display_date}"
                    c.drawString(0.75 * inch, y_pos, detail_text)
                    y_pos -= 0.25 * inch

                    po_total += float(line.get("total_amount", 0))

                # PO Total
                c.setFont("Helvetica-Bold", 10)
                c.drawString(
                    0.75 * inch, y_pos, f"PO Total: {po_total:,.2f} {currency}"
                )
                y_pos -= 0.4 * inch

            # Calculate grand total
            grand_total = sum(data["total_value"] for _, data in orders)
            all_currency = orders[0][1]["currency"] if orders else "EUR"

            # Grand total
            if y_pos < 1.5 * inch:
                c.showPage()
                y_pos = height - 1 * inch

            y_pos -= 0.2 * inch
            c.setFont("Helvetica-Bold", 12)
            c.drawString(
                0.5 * inch, y_pos, f"GRAND TOTAL: {grand_total:,.2f} {all_currency}"
            )

            # Footer
            y_pos -= 0.6 * inch
            c.setFont("Helvetica", 10)
            buyer_email = config.get("buyer_email", "purchasing@company.com")
            footer_lines = [
                "",
                "Please reply with your confirmation for all outstanding line items.",
                f"Contact: {buyer_email}",
                "",
            ]

            if include_etd and transport_days > 0:
                footer_lines.append(
                    f"Note: ETD dates are calculated by subtracting {transport_days} working days from the reschedule date (or requested date if no reschedule)."
                )
                footer_lines.append("")

            footer_lines.append("Thank you for your cooperation.")

            for line in footer_lines:
                if y_pos > 0.5 * inch:
                    c.drawString(0.5 * inch, y_pos, line)
                    y_pos -= 0.2 * inch
                else:
                    c.showPage()
                    y_pos = height - 1 * inch
                    c.setFont("Helvetica", 10)
                    c.drawString(0.5 * inch, y_pos, line)
                    y_pos -= 0.2 * inch

            c.save()
            return True

        except Exception as e:
            self.log(f"✗ ERROR: Failed to create summary PDF: {e}")
            import traceback

            traceback.print_exc()
            return False

    def generate_reminder_pdfs(self):
        """Generate reminder PDFs for each PO (detailed table format)"""
        try:
            pdf_paths = {}
            output_folder = os.path.join(APP_DATA_FOLDER, "ReminderPDFs")
            os.makedirs(output_folder, exist_ok=True)

            for po_num, data in self.unconfirmed_data.items():
                vendor_name = data["vendor"]
                lines = data["lines"]

                # Create PDF
                safe_vendor = re.sub(r'[\\/*?:"<>|]', "_", vendor_name)
                pdf_filename = f"Reminder_{po_num}_{safe_vendor}_{datetime.now().strftime('%Y%m%d')}.pdf"
                pdf_path = os.path.join(output_folder, pdf_filename)

                if self.create_reminder_pdf(pdf_path, po_num, vendor_name, lines):
                    if vendor_name not in pdf_paths:
                        pdf_paths[vendor_name] = []
                    pdf_paths[vendor_name].append(pdf_path)
                    self.log(f"✓ SUCCESS: Generated reminder PDF for PO {po_num}")

            messagebox.showinfo("✅ Success",
                f"Generated {sum(len(pdfs) for pdfs in pdf_paths.values())} reminder PDF(s) in:\n{output_folder}",
                parent=self,
            )
            return pdf_paths

        except Exception as e:
            self.log(f"✗ ERROR: Failed to generate reminder PDFs: {e}")
            messagebox.showerror("❌ Error", f"Failed to generate PDFs:\n{str(e)}", parent=self
            )
            return {}

    def create_reminder_pdf(self, pdf_path, po_number, vendor_name, lines):
        """Create a reminder PDF for unconfirmed orders"""
        try:
            c = canvas.Canvas(pdf_path, pagesize=letter)
            width, height = letter
            config = self.dm.get_config("company_config", {})

            # Get transport days for ETD calculation
            include_etd = self.include_etd_var.get()
            transport_days = 0

            if include_etd and lines:
                # Get transport days from vendor
                vendor_data = next(
                    (
                        v
                        for v in self.dm.get_all_vendors()
                        if v["display_name"] == vendor_name
                    ),
                    None,
                )
                if vendor_data:
                    transport_days = int(vendor_data.get("transport_days", 0))

            # Header
            c.setFont("Helvetica-Bold", 16)
            c.drawString(
                0.5 * inch, height - 0.5 * inch, "PURCHASE ORDER CONFIRMATION REMINDER"
            )

            c.setFont("Helvetica", 10)
            c.drawString(
                0.5 * inch,
                height - 0.75 * inch,
                f"Date: {datetime.now().strftime('%d.%m.%Y')}",
            )
            c.drawString(0.5 * inch, height - 0.95 * inch, f"PO Number: {po_number}")
            c.drawString(0.5 * inch, height - 1.15 * inch, f"Supplier: {vendor_name}")

            if include_etd and transport_days > 0:
                c.drawString(
                    0.5 * inch,
                    height - 1.35 * inch,
                    f"Transport Days: {transport_days} working days",
                )

            # Message
            y_pos = height - 1.6 * inch if not include_etd else height - 1.8 * inch
            c.setFont("Helvetica", 10)
            message_lines = [
                "Dear Supplier,",
                "",
                "This is a reminder that we have not yet received confirmation for the following purchase order line items.",
                "Please provide delivery confirmation at your earliest convenience.",
                "",
            ]

            for line in message_lines:
                c.drawString(0.5 * inch, y_pos, line)
                y_pos -= 0.2 * inch

            # Table header
            y_pos -= 0.3 * inch
            c.setFont("Helvetica-Bold", 9)
            c.rect(
                0.5 * inch,
                y_pos - 0.25 * inch,
                7.5 * inch,
                0.25 * inch,
                stroke=1,
                fill=0,
            )

            c.drawString(0.6 * inch, y_pos - 0.15 * inch, "Line")
            c.drawString(1.2 * inch, y_pos - 0.15 * inch, "Material")
            c.drawString(2.5 * inch, y_pos - 0.15 * inch, "Description")
            c.drawString(4.8 * inch, y_pos - 0.15 * inch, "Qty")
            c.drawString(5.5 * inch, y_pos - 0.15 * inch, "Unit Price")
            c.drawString(6.5 * inch, y_pos - 0.15 * inch, "Total")

            if include_etd:
                c.drawString(7.3 * inch, y_pos - 0.15 * inch, "ETD Date")
            else:
                c.drawString(7.3 * inch, y_pos - 0.15 * inch, "Req. Date")

            # Table data
            y_pos -= 0.45 * inch
            c.setFont("Helvetica", 8)
            total = 0
            currency = lines[0].get("currency", "EUR") if lines else "EUR"

            for line in lines:
                if y_pos < 1 * inch:  # New page if needed
                    c.showPage()
                    y_pos = height - 1 * inch
                    c.setFont("Helvetica", 8)

                # Determine which date to display
                display_date = ""

                if include_etd:
                    # Check if reschedule date exists
                    reschedule_date_str = line.get("rescheduling_date")

                    if reschedule_date_str and reschedule_date_str.strip():
                        # Use reschedule date to calculate ETD
                        try:
                            reschedule_date = datetime.strptime(
                                reschedule_date_str, "%d.%m.%Y"
                            )
                            if transport_days > 0:
                                etd_date = subtract_working_days(
                                    reschedule_date, transport_days
                                )
                            else:
                                etd_date = reschedule_date
                            display_date = etd_date.strftime("%d.%m.%Y")
                        except Exception:
                            display_date = reschedule_date_str
                    else:
                        # No reschedule date, use requested date
                        display_date = line.get("requested_del_date", "N/A")
                else:
                    # ETD not enabled, use requested date
                    display_date = line.get("requested_del_date", "N/A")

                c.drawString(0.6 * inch, y_pos, str(line.get("item", "")))
                c.drawString(1.2 * inch, y_pos, str(line.get("material_code", ""))[:15])
                c.drawString(2.5 * inch, y_pos, str(line.get("short_text", ""))[:28])
                c.drawString(4.8 * inch, y_pos, str(line.get("requested_qty", "")))

                # Unit Price - show with per unit if applicable
                unit_price = float(line.get("unit_price", 0))
                price_per = int(line.get("price_per_unit", 1))
                unit = line.get("unit", "EA")
                if price_per > 1:
                    c.drawString(
                        5.5 * inch, y_pos, f"{unit_price:.2f}/ {price_per} {unit}"
                    )
                else:
                    c.drawString(5.5 * inch, y_pos, f"{unit_price:.2f}")

                c.drawString(
                    6.5 * inch, y_pos, f"{float(line.get('total_amount', 0)):.2f}"
                )
                c.drawString(7.3 * inch, y_pos, display_date)

                total += float(line.get("total_amount", 0))
                y_pos -= 0.2 * inch

            # Total
            c.line(0.5 * inch, y_pos + 0.15 * inch, 8 * inch, y_pos + 0.15 * inch)
            y_pos -= 0.2 * inch
            c.setFont("Helvetica-Bold", 10)
            c.drawRightString(6.3 * inch, y_pos, f"TOTAL ({currency}):")
            c.drawRightString(7.1 * inch, y_pos, f"{total:,.2f}")

            # Footer
            y_pos -= 0.5 * inch
            c.setFont("Helvetica", 9)
            buyer_email = config.get("buyer_email", "purchasing@company.com")
            footer_lines = [
                "",
                "Please reply with your confirmation to this email.",
                f"Contact: {buyer_email}",
                "",
            ]

            if include_etd and transport_days > 0:
                footer_lines.append(
                    f"Note: ETD dates are calculated by subtracting {transport_days} working days from the reschedule date (or requested date if no reschedule)."
                )
                footer_lines.append("")

            footer_lines.append("Thank you for your cooperation.")

            for line in footer_lines:
                if y_pos > 0.5 * inch:
                    c.drawString(0.5 * inch, y_pos, line)
                    y_pos -= 0.2 * inch

            c.save()
            return True

        except Exception as e:
            self.log(f"✗ ERROR: Failed to create reminder PDF: {e}")
            import traceback

            traceback.print_exc()
            return False

    def send_reminder_emails(self, pdf_paths):
        """Send reminder emails with attached PDFs using templates"""
        try:
            method = self.get_send_method()
            sender = EmailSender(self.log, self.dm)
            use_custom_sig = self.use_custom_sig_var.get()

            sent_count = 0
            failed_count = 0
            failed_details = []

            for vendor_name, pdfs in pdf_paths.items():
                vendor_data = next(
                    (
                        v
                        for v in self.dm.get_all_vendors()
                        if v["display_name"] == vendor_name
                    ),
                    None,
                )
                if not vendor_data or not vendor_data.get("emails"):
                    failed_count += 1
                    failed_details.append(f"{vendor_name}: No email address")
                    continue

                # Use template
                subject, body = sender.generate_reminder_email_content(vendor_name)
                to_emails = vendor_data["emails"]

                try:
                    if method == "Outlook" and OUTLOOK_AVAILABLE:
                        # Call sender method - Note: _send_outlook only takes one attachment path
                        # For simplicity, sending only the first PDF here if using Outlook
                        # If multiple PDFs per vendor, need to adjust logic or send separate emails
                        success, message = sender._send_outlook(
                            to_emails, subject, body, pdfs[0] if pdfs else None, use_custom_sig
                        )
                        if success:
                            sent_count += 1
                            self.log(f"✓ SUCCESS: Sent reminder email to {vendor_name}")
                        else:
                            failed_count += 1
                            failed_details.append(f"{vendor_name}: {message}")
                    else:
                        # Use the new multiple attachments method (with signature support)
                        success, message = sender._send_smtp_multiple_attachments(
                            to_emails, subject, body, pdfs, use_custom_sig
                        )
                        if success:
                            sent_count += 1
                            self.log(f"✓ SUCCESS: Sent reminder email to {vendor_name}")
                        else:
                            failed_count += 1
                            failed_details.append(f"{vendor_name}: {message}")

                except Exception as e:
                    failed_count += 1
                    failed_details.append(f"{vendor_name}: {str(e)}")
                    self.log(f"✗ ERROR: Failed to send reminder to {vendor_name}: {e}")

            # Show summary
            summary = f"Sent {sent_count} reminder email(s), Failed {failed_count}"
            if failed_details:
                summary += "\n\nFailures:\n" + "\n".join(failed_details)

            messagebox.showinfo("Email Results", summary, parent=self)
            self.log(f"ℹ️ INFO: Reminder email batch complete: {summary}")

        except Exception as e:
            self.log(f"✗ ERROR: Failed to send reminder emails: {e}")
            messagebox.showerror("❌ Error", f"Failed to send emails:\n{str(e)}", parent=self
            )


# ==============================================================================
# EMAIL CONFIRMATION SCANNER
# ==============================================================================


class EmailConfirmationScanner:
    """Scans emails for order confirmations and saves PDFs"""

    def __init__(
        self, db_manager, log_callback, include_keywords=None, exclude_keywords=None
    ):
        self.db = db_manager
        self.log = log_callback
        self.confirmations_folder = os.path.join(APP_DATA_FOLDER, "Confirmations")
        os.makedirs(self.confirmations_folder, exist_ok=True)

        # Default confirmation keywords (INCLUDE - increase probability)
        default_include = [
            "confirmation",
            "confirm",
            "acknowledged",
            "acknowledge",
            "order confirmation",
            "oc",
            "purchase order confirmation",
            "po confirmation",
            "accepted",
            "acceptance",
            "order accept",
            "order acknowledgement",
            "order received",
        ]

        # Custom include keywords (set via GUI or config)
        self.confirmation_keywords = (
            include_keywords if include_keywords else default_include
        )

        # Default exclude keywords (EXCLUDE - filter out non-confirmations)
        default_exclude = [
            "invoice",
            "receipt",
            "payment",
            "bill",
            "statement",
            "reminder",
            "overdue",
            "past due",
            "notice",
            "out of office",
            "automatic reply",
            "auto-reply",
            "undeliverable",
            "delivery failure",
            "returned mail",
        ]

        # Custom exclude keywords (set via GUI or config)
        self.exclude_keywords = (
            exclude_keywords if exclude_keywords else default_exclude
        )

        # PO number patterns (adjust to your format)
        self.po_patterns = [
            r"PO[:\s#-]*(\d{7,10})",  # PO: 4500123456 or PO #4500123456
            r"P\.?O\.?[:\s#-]*(\d{7,10})",  # P.O. 4500123456
            r"(?:Purchase Order|Order)[:\s#-]*(\d{7,10})",  # Purchase Order: 4500123456
            r"\b(45\d{8})\b",  # Direct 10-digit number starting with 45
            r"Order\s+(?:Number|No\.?|#)[:\s]*(\d{7,10})",  # Order Number: 4500123456
        ]

    def update_keywords(self, include_keywords=None, exclude_keywords=None):
        """Update the keyword filters"""
        if include_keywords is not None:
            self.confirmation_keywords = include_keywords
            self.log(f"ℹ️ INFO: Updated include keywords: {len(include_keywords)} terms")

        if exclude_keywords is not None:
            self.exclude_keywords = exclude_keywords
            self.log(f"ℹ️ INFO: Updated exclude keywords: {len(exclude_keywords)} terms")

    def _is_confirmation_email_filtered(self, text):
        """
        Check if email matches confirmation criteria with include/exclude filters

        Args:
            text: Combined subject + body text (lowercase)

        Returns:
            bool: True if email is likely a confirmation
        """
        # First check: Must contain at least one INCLUDE keyword
        has_include = any(
            keyword.lower() in text for keyword in self.confirmation_keywords
        )

        if not has_include:
            return False

        # Second check: Must NOT contain any EXCLUDE keywords
        has_exclude = any(keyword.lower() in text for keyword in self.exclude_keywords)

        if has_exclude:
            return False

        return True

    def scan_outlook_emails(self, days_back=7, mark_as_read=False):
        """
        Scan Outlook emails for order confirmations (Default method)

        Args:
            days_back: how many days back to scan
            mark_as_read: whether to mark scanned emails as read

        Returns:
            dict with results summary
        """
        if not OUTLOOK_AVAILABLE:
            raise Exception(
                "Outlook is not available. Please install pywin32 or use IMAP method."
            )

        try:
            self.log(f"ℹ️ INFO: Scanning Outlook emails from last {days_back} days...")
            self.log(
                f"ℹ️ INFO: Using {len(self.confirmation_keywords)} include keywords, {len(self.exclude_keywords)} exclude keywords"
            )

            outlook = win32com.client.Dispatch("outlook.application")
            namespace = outlook.GetNamespace("MAPI")
            inbox = namespace.GetDefaultFolder(6)  # 6 = Inbox
            messages = inbox.Items
            messages.Sort("[ReceivedTime]", True)  # Sort by newest first

            cutoff_date = datetime.now() - timedelta(days=days_back)

            results = {
                "scanned": 0,
                "confirmations_found": 0,
                "pdfs_saved": 0,
                "matched_pos": [],
                "unmatched_pos": [],
                "errors": [],
                "filtered_out": 0,
            }

            for message in messages:
                try:
                    received_time = message.ReceivedTime
                    if received_time:
                        try:
                            received_dt = datetime(
                                received_time.year,
                                received_time.month,
                                received_time.day,
                                received_time.hour,
                                received_time.minute,
                                received_time.second,
                            )
                        except Exception:
                            continue

                        if received_dt < cutoff_date:
                            break  # Stop scanning older emails

                    results["scanned"] += 1

                    # Check if this is a confirmation email with filtering
                    if self._is_outlook_confirmation_email(message):
                        results["confirmations_found"] += 1

                        # Extract PO numbers from subject and body
                        po_numbers = self._extract_po_numbers_from_outlook(message)

                        if po_numbers:
                            self.log(
                                f"ℹ️ INFO: Found confirmation for PO(s): {', '.join(po_numbers)}"
                            )

                            # Process attachments
                            if message.Attachments.Count > 0:
                                for attachment in message.Attachments:
                                    filename = attachment.FileName

                                    if filename.lower().endswith(".pdf"):
                                        # Save attachment temporarily to extract text
                                        temp_path = os.path.join(
                                            self.confirmations_folder,
                                            f"temp_{filename}",
                                        )
                                        attachment.SaveAsFile(temp_path)

                                        try:
                                            # Extract PO from PDF
                                            with open(temp_path, "rb") as f:
                                                pdf_data = f.read()

                                            pdf_po_numbers = self._extract_po_from_pdf(
                                                pdf_data
                                            )
                                            all_po_numbers = list(
                                                set(po_numbers + pdf_po_numbers)
                                            )

                                            for po_num in all_po_numbers:
                                                if self._po_exists_in_database(po_num):
                                                    # Save with proper name
                                                    safe_filename = f"Confirmation_{po_num}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
                                                    final_path = os.path.join(
                                                        self.confirmations_folder,
                                                        safe_filename,
                                                    )

                                                    # Rename temp file to final name
                                                    os.rename(temp_path, final_path)

                                                    results["pdfs_saved"] += 1
                                                    results["matched_pos"].append(
                                                        po_num
                                                    )

                                                    # Update database
                                                    self._update_po_confirmation_status(
                                                        po_num, final_path
                                                    )

                                                    self.log(
                                                        f"✓ SUCCESS: Saved confirmation for PO {po_num}: {safe_filename}"
                                                    )
                                                else:
                                                    results["unmatched_pos"].append(
                                                        po_num
                                                    )
                                                    self.log(
                                                        f"⚠ WARNING: PO {po_num} not found in database"
                                                    )

                                            # Clean up temp file if still exists
                                            if os.path.exists(temp_path):
                                                os.remove(temp_path)

                                        except Exception as e:
                                            self.log(
                                                f"✗ ERROR: Failed to process PDF {filename}: {e}"
                                            )
                                            if os.path.exists(temp_path):
                                                os.remove(temp_path)
                    else:
                        # Email was filtered out
                        results["filtered_out"] += 1

                    # Mark as read if requested
                    if mark_as_read:
                        message.UnRead = False

                except Exception as e:
                    error_msg = f"Error processing email: {str(e)}"
                    results["errors"].append(error_msg)
                    self.log(f"✗ ERROR: {error_msg}")

            self.log(
                f"✓ SUCCESS: Scanned {results['scanned']} emails, "
                f"Found {results['confirmations_found']} confirmations, saved {results['pdfs_saved']} PDFs, "
                f"Filtered out {results['filtered_out']} non-confirmations"
            )
            return results

        except Exception as e:
            self.log(f"✗ ERROR: Outlook scan failed: {str(e)}")
            raise

    def _is_outlook_confirmation_email(self, message):
        """Check if Outlook message is an order confirmation with filtering"""
        try:
            subject = message.Subject or ""
            body = message.Body or ""

            text = (subject + " " + body).lower()

            return self._is_confirmation_email_filtered(text)
        except Exception:
            return False

    def _extract_po_numbers_from_outlook(self, message):
        """Extract PO numbers from Outlook message"""
        po_numbers = set()

        try:
            subject = message.Subject or ""
            body = message.Body or ""

            text = subject + " " + body

            # Try each pattern
            for pattern in self.po_patterns:
                matches = re.findall(pattern, text, re.IGNORECASE)
                po_numbers.update(matches)
        except Exception:  # TODO: Add proper error handling
            pass  # TODO: Add proper error handling
        return list(po_numbers)

    def scan_emails(self, email_config, days_back=7, mark_as_read=False):
        """
        Scan IMAP emails for order confirmations with filtering

        Args:
            email_config: dict with 'email', 'password', 'imap_server', 'imap_port'
            days_back: how many days back to scan
            mark_as_read: whether to mark scanned emails as read

        Returns:
            dict with results summary
        """
        try:
            self.log(
                f"ℹ️ INFO: Connecting to email server {email_config['imap_server']}..."
            )
            self.log(
                f"ℹ️ INFO: Using {len(self.confirmation_keywords)} include keywords, {len(self.exclude_keywords)} exclude keywords"
            )

            # Connect to IMAP server
            mail = imaplib.IMAP4_SSL(
                email_config["imap_server"], email_config.get("imap_port", 993)
            )
            mail.login(email_config["email"], email_config["password"])
            mail.select("inbox")

            # Search for emails from last N days
            date_since = (datetime.now() - timedelta(days=days_back)).strftime(
                "%d-%b-%Y"
            )
            _, message_numbers = mail.search(None, f"(SINCE {date_since})")

            email_ids = message_numbers[0].split()
            total_emails = len(email_ids)

            self.log(
                f"ℹ️ INFO: Found {total_emails} emails to scan from last {days_back} days"
            )

            results = {
                "scanned": 0,
                "confirmations_found": 0,
                "pdfs_saved": 0,
                "matched_pos": [],
                "unmatched_pos": [],
                "errors": [],
                "filtered_out": 0,
            }

            for email_id in email_ids:
                try:
                    results["scanned"] += 1

                    # Fetch email
                    _, msg_data = mail.fetch(email_id, "(RFC822)")
                    email_body = msg_data[0][1]
                    email_message = email.message_from_bytes(email_body)

                    # Check if this is a confirmation email with filtering
                    if self._is_confirmation_email(email_message):
                        results["confirmations_found"] += 1

                        # Extract PO numbers and process attachments
                        po_numbers = self._extract_po_numbers(email_message)

                        if po_numbers:
                            self.log(
                                f"ℹ️ INFO: Found confirmation for PO(s): {', '.join(po_numbers)}"
                            )

                            # Process attachments
                            for part in email_message.walk():
                                if part.get_content_maintype() == "multipart":
                                    continue
                                if part.get("Content-Disposition") is None:
                                    continue

                                filename = part.get_filename()
                                if filename and filename.lower().endswith(".pdf"):
                                    # Save PDF
                                    pdf_data = part.get_payload(decode=True)

                                    # Check for PO numbers in PDF as well
                                    pdf_po_numbers = self._extract_po_from_pdf(pdf_data)
                                    all_po_numbers = list(
                                        set(po_numbers + pdf_po_numbers)
                                    )

                                    for po_num in all_po_numbers:
                                        if self._po_exists_in_database(po_num):
                                            # Save PDF with PO number
                                            safe_filename = f"Confirmation_{po_num}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
                                            save_path = os.path.join(
                                                self.confirmations_folder, safe_filename
                                            )

                                            with open(save_path, "wb") as f:
                                                f.write(pdf_data)

                                            results["pdfs_saved"] += 1
                                            results["matched_pos"].append(po_num)

                                            # Update database
                                            self._update_po_confirmation_status(
                                                po_num, save_path
                                            )

                                            self.log(
                                                f"✓ SUCCESS: Saved confirmation for PO {po_num}: {safe_filename}"
                                            )
                                        else:
                                            results["unmatched_pos"].append(po_num)
                                            self.log(
                                                f"⚠ WARNING: PO {po_num} not found in database"
                                            )
                    else:
                        results["filtered_out"] += 1

                    # Mark as read if requested
                    if mark_as_read:
                        mail.store(email_id, "+FLAGS", "\\Seen")

                except Exception as e:
                    error_msg = f"Error processing email {email_id}: {str(e)}"
                    results["errors"].append(error_msg)
                    self.log(f"✗ ERROR: {error_msg}")

            mail.close()
            mail.logout()

            self.log(
                f"✓ SUCCESS: Email scan complete. "
                f"Found {results['confirmations_found']} confirmations, saved {results['pdfs_saved']} PDFs, "
                f"Filtered out {results['filtered_out']} non-confirmations"
            )
            return results

        except Exception as e:
            self.log(f"✗ ERROR: Email scan failed: {str(e)}")
            raise

    def _is_confirmation_email(self, email_message):
        """Check if IMAP email is an order confirmation with filtering"""
        # Check subject
        subject = self._decode_header(email_message.get("Subject", ""))

        # Check body
        body = self._get_email_body(email_message)

        # Combine subject and body for keyword search
        text = (subject + " " + body).lower()

        # Apply include/exclude filtering
        return self._is_confirmation_email_filtered(text)

    def _extract_po_numbers(self, email_message):
        """Extract PO numbers from email subject and body"""
        po_numbers = set()

        # Get subject and body
        subject = self._decode_header(email_message.get("Subject", ""))
        body = self._get_email_body(email_message)

        text = subject + " " + body

        # Try each pattern
        for pattern in self.po_patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            po_numbers.update(matches)

        return list(po_numbers)

    def _extract_po_from_pdf(self, pdf_data):
        """Extract PO numbers from PDF content"""
        po_numbers = set()

        try:
            pdf_file = io.BytesIO(pdf_data)
            pdf_reader = PyPDF2.PdfReader(pdf_file)

            # Extract text from first 3 pages (usually enough)
            text = ""
            for page_num in range(min(3, len(pdf_reader.pages))):
                page = pdf_reader.pages[page_num]
                text += page.extract_text()

            # Search for PO numbers
            for pattern in self.po_patterns:
                matches = re.findall(pattern, text, re.IGNORECASE)
                po_numbers.update(matches)

        except Exception as e:
            self.log(f"⚠ WARNING: Could not extract text from PDF: {str(e)}")

        return list(po_numbers)

    def _decode_header(self, header):
        """Decode email header"""
        if header is None:
            return ""

        decoded_parts = decode_header(header)
        decoded_header = ""

        for part, encoding in decoded_parts:
            if isinstance(part, bytes):
                try:
                    decoded_header += part.decode(encoding or "utf-8", errors="ignore")
                except Exception:
                    decoded_header += part.decode("utf-8", errors="ignore")
            else:
                decoded_header += str(part)

        return decoded_header

    def _get_email_body(self, email_message):
        """Extract email body text"""
        body = ""

        if email_message.is_multipart():
            for part in email_message.walk():
                content_type = part.get_content_type()
                content_disposition = str(part.get("Content-Disposition"))

                if (
                    content_type == "text/plain"
                    and "attachment" not in content_disposition
                ):
                    try:
                        body = part.get_payload(decode=True).decode(
                            "utf-8", errors="ignore"
                        )
                        break
                    except Exception:  # TODO: Add proper error handling
                        pass  # TODO: Add proper error handling
        else:
            try:
                body = email_message.get_payload(decode=True).decode(
                    "utf-8", errors="ignore"
                )
            except Exception:  # TODO: Add proper error handling
                pass  # TODO: Add proper error handling
        return body

    def _po_exists_in_database(self, po_number):
        """Check if PO exists in database"""
        result = self.db.execute_query(
            "SELECT 1 FROM open_orders WHERE po = ? LIMIT 1",
            (po_number,),
            fetchone=True,
        )
        return result is not None

    def _update_po_confirmation_status(self, po_number, pdf_path):
        """Update PO with confirmation received status"""
        self.db.execute_query(
            """UPDATE open_orders 
               SET pdf_status = 'Confirmed', 
                   comments = COALESCE(comments || ' | ', '') || 'Confirmation received: ' || ?
               WHERE po = ?""",
            (os.path.basename(pdf_path), po_number),
            commit=True,
        )


# ==============================================================================
# GUI WINDOW FOR EMAIL CONFIRMATION SCANNER
# ==============================================================================


class EmailConfirmationScannerWindow(tk.Toplevel):
    """Window for scanning emails for order confirmations with keyword filters"""

    def __init__(self, parent, log_callback, data_manager):
        super().__init__(parent)
        self.title("Email Confirmation Scanner")
        self.geometry("950x850")
        self.log = log_callback
        self.dm = data_manager

        # Load saved filter keywords
        filter_config = self.dm.get_config("email_scanner_filters", {})
        include_keywords = filter_config.get("include", None)
        exclude_keywords = filter_config.get("exclude", None)

        # Initialize scanner with custom keywords
        self.scanner = EmailConfirmationScanner(
            data_manager.db,
            log_callback,
            include_keywords=include_keywords,
            exclude_keywords=exclude_keywords,
        )

        # Main frame
        main_frame = ttk.Frame(self, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Method selection
        method_frame = ttk.LabelFrame(main_frame, text="Scan Method", padding=10)
        method_frame.pack(fill=tk.X, pady=(0, 10))

        self.scan_method_var = tk.StringVar(value="outlook")

        if OUTLOOK_AVAILABLE:
            ttk.Radiobutton(
                method_frame,
                text=" Use Outlook (Recommended)",
                variable=self.scan_method_var,
                value="outlook",
                command=self.toggle_method,
            ).pack(anchor="w", pady=2)
        else:
            ttk.Label(
                method_frame, text=" Outlook not available", foreground="orange"
            ).pack(anchor="w", pady=2)

        ttk.Radiobutton(
            method_frame,
            text=" Use IMAP (Gmail, etc.)",
            variable=self.scan_method_var,
            value="imap",
            command=self.toggle_method,
        ).pack(anchor="w", pady=2)

        # Set default based on availability
        if not OUTLOOK_AVAILABLE:
            self.scan_method_var.set("imap")

        # IMAP Configuration section (initially hidden if Outlook available)
        self.imap_config_frame = ttk.LabelFrame(
            main_frame, text="IMAP Email Configuration", padding=10
        )

        # Email settings
        ttk.Label(self.imap_config_frame, text="Email Address:").grid(
            row=0, column=0, sticky="w", pady=5
        )
        self.email_var = tk.StringVar()
        ttk.Entry(self.imap_config_frame, textvariable=self.email_var, width=40).grid(
            row=0, column=1, sticky="ew", pady=5
        )

        ttk.Label(self.imap_config_frame, text="Password:").grid(
            row=1, column=0, sticky="w", pady=5
        )
        self.password_var = tk.StringVar()
        ttk.Entry(
            self.imap_config_frame, textvariable=self.password_var, width=40, show="*"
        ).grid(row=1, column=1, sticky="ew", pady=5)

        ttk.Label(self.imap_config_frame, text="IMAP Server:").grid(
            row=2, column=0, sticky="w", pady=5
        )
        self.imap_server_var = tk.StringVar(value="imap.gmail.com")
        ttk.Entry(
            self.imap_config_frame, textvariable=self.imap_server_var, width=40
        ).grid(row=2, column=1, sticky="ew", pady=5)

        ttk.Label(self.imap_config_frame, text="IMAP Port:").grid(
            row=3, column=0, sticky="w", pady=5
        )
        self.imap_port_var = tk.StringVar(value="993")
        ttk.Entry(
            self.imap_config_frame, textvariable=self.imap_port_var, width=40
        ).grid(row=3, column=1, sticky="ew", pady=5)

        self.imap_config_frame.columnconfigure(1, weight=1)

        # NEW: Keyword Filters Section
        filters_frame = ttk.LabelFrame(
            main_frame, text=" Keyword Filters (Advanced)", padding=10
        )
        filters_frame.pack(fill=tk.X, pady=(0, 10))

        # Include Keywords
        include_frame = ttk.Frame(filters_frame)
        include_frame.pack(fill=tk.X, pady=5)

        ttk.Label(include_frame, text=" Include Keywords (comma-separated):").pack(
            anchor="w"
        )
        ttk.Label(
            include_frame,
            text="Emails must contain at least one of these",
            font=("Helvetica", 8),
            foreground="gray",
        ).pack(anchor="w")

        self.include_keywords_text = tk.Text(include_frame, height=3, wrap=tk.WORD)
        self.include_keywords_text.pack(fill=tk.X, pady=5)

        # Set default include keywords
        default_include = ", ".join(self.scanner.confirmation_keywords)
        self.include_keywords_text.insert("1.0", default_include)

        # Exclude Keywords
        exclude_frame = ttk.Frame(filters_frame)
        exclude_frame.pack(fill=tk.X, pady=5)

        ttk.Label(exclude_frame, text=" Exclude Keywords (comma-separated):").pack(
            anchor="w"
        )
        ttk.Label(
            exclude_frame,
            text="Emails with these keywords will be filtered out",
            font=("Helvetica", 8),
            foreground="gray",
        ).pack(anchor="w")

        self.exclude_keywords_text = tk.Text(exclude_frame, height=3, wrap=tk.WORD)
        self.exclude_keywords_text.pack(fill=tk.X, pady=5)

        # Set default exclude keywords
        default_exclude = ", ".join(self.scanner.exclude_keywords)
        self.exclude_keywords_text.insert("1.0", default_exclude)

        # Buttons for filter management
        filter_btn_frame = ttk.Frame(filters_frame)
        filter_btn_frame.pack(fill=tk.X, pady=5)

        ttk.Button(
            filter_btn_frame, text=" Save Filters", command=self.save_filters
        ).pack(side=tk.LEFT, padx=5)
        ttk.Button(
            filter_btn_frame, text=" Reset to Default", command=self.reset_filters
        ).pack(side=tk.LEFT, padx=5)
        ttk.Button(filter_btn_frame, text=" Help", command=self.show_filter_help).pack(
            side=tk.LEFT, padx=5
        )

        # Scan options
        options_frame = ttk.LabelFrame(main_frame, text="Scan Options", padding=10)
        options_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(options_frame, text="Days to scan back:").pack(side=tk.LEFT, padx=5)
        self.days_back_var = tk.StringVar(value="7")
        ttk.Entry(options_frame, textvariable=self.days_back_var, width=10).pack(
            side=tk.LEFT, padx=5
        )

        self.mark_read_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            options_frame,
            text="Mark scanned emails as read",
            variable=self.mark_read_var,
        ).pack(side=tk.LEFT, padx=20)

        # Action buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(0, 10))

        self.save_config_btn = ttk.Button(
            button_frame, text=" Save IMAP Configuration", command=self.save_config
        )
        self.save_config_btn.pack(side=tk.LEFT, padx=5)

        ttk.Button(
            button_frame,
            text=" Scan Emails Now",
            command=self.scan_emails,
            style="Accent.TButton",
        ).pack(side=tk.LEFT, padx=5)
        ttk.Button(
            button_frame,
            text=" Open Confirmations Folder",
            command=self.open_confirmations_folder,
        ).pack(side=tk.LEFT, padx=5)

        # Results section
        results_frame = ttk.LabelFrame(main_frame, text="Scan Results", padding=10)
        results_frame.pack(fill=tk.BOTH, expand=True)

        self.results_text = scrolledtext.ScrolledText(
            results_frame, wrap=tk.WORD, height=15
        )
        self.results_text.pack(fill=tk.BOTH, expand=True)

        # Load saved configuration
        self.load_config()

        # Set initial visibility
        self.toggle_method()

    def toggle_method(self):
        """Show/hide configuration based on selected method"""
        if self.scan_method_var.get() == "outlook":
            self.imap_config_frame.pack_forget()
            self.save_config_btn.pack_forget()
        else:
            self.imap_config_frame.pack(
                fill=tk.X, pady=(0, 10), after=self.scan_method_var.master
            )
            # Re-add save button if not already there
            if not self.save_config_btn.winfo_ismapped():
                self.save_config_btn.pack(
                    side=tk.LEFT, padx=5, in_=self.save_config_btn.master
                )

    def save_filters(self):
        """Save keyword filters to configuration"""
        try:
            # Get keywords from text boxes
            include_text = self.include_keywords_text.get("1.0", tk.END).strip()
            exclude_text = self.exclude_keywords_text.get("1.0", tk.END).strip()

            # Parse comma-separated keywords
            include_keywords = [
                kw.strip() for kw in include_text.split(",") if kw.strip()
            ]
            exclude_keywords = [
                kw.strip() for kw in exclude_text.split(",") if kw.strip()
            ]

            # Update scanner
            self.scanner.update_keywords(include_keywords, exclude_keywords)

            # Save to config
            filter_config = {"include": include_keywords, "exclude": exclude_keywords}
            self.dm.save_config("email_scanner_filters", filter_config)

            messagebox.showinfo(
                "Filters Saved",
                f"Saved {len(include_keywords)} include keywords and "
                f"{len(exclude_keywords)} exclude keywords.",
                parent=self,
            )
            self.log(f"ℹ️ INFO: Saved email scanner filters")

        except Exception as e:
            messagebox.showerror("❌ Error", f"Failed to save filters: {str(e)}", parent=self
            )

    def reset_filters(self):
        """Reset filters to default values"""
        if messagebox.askyesno(
            "Reset Filters", "Reset keyword filters to default values?", parent=self
        ):
            # Reset to defaults
            default_scanner = EmailConfirmationScanner(self.dm.db, self.log)

            # Update text boxes
            self.include_keywords_text.delete("1.0", tk.END)
            self.include_keywords_text.insert(
                "1.0", ", ".join(default_scanner.confirmation_keywords)
            )

            self.exclude_keywords_text.delete("1.0", tk.END)
            self.exclude_keywords_text.insert(
                "1.0", ", ".join(default_scanner.exclude_keywords)
            )

            # Update scanner
            self.scanner.update_keywords(
                default_scanner.confirmation_keywords, default_scanner.exclude_keywords
            )

            self.log("ℹ️ INFO: Reset email scanner filters to defaults")

    def show_filter_help(self):
        """Show help information about keyword filters"""
        help_text = """Keyword Filters Help
        
INCLUDE KEYWORDS ():
- Emails MUST contain at least ONE of these keywords
- Increases probability of finding confirmations
- Examples: confirmation, acknowledge, accepted

EXCLUDE KEYWORDS ():
- Emails containing ANY of these will be FILTERED OUT
- Prevents false positives (invoices, receipts, etc.)
- Examples: invoice, payment, overdue

TIPS:
- Use lowercase for keywords (matching is case-insensitive)
- Separate multiple keywords with commas
- More specific keywords = better accuracy
- Test with a small date range first (1-2 days)

EXAMPLES:
 Good include: "order confirmation, po confirmation, acknowledged"
 Good exclude: "invoice, receipt, payment reminder"
"""
        messagebox.showinfo("Keyword Filters Help", help_text, parent=self)

    def save_config(self):
        """Save email configuration"""
        config = {
            "email": self.email_var.get(),
            "password": self.password_var.get(),
            "imap_server": self.imap_server_var.get(),
            "imap_port": int(self.imap_port_var.get()),
        }

        self.dm.save_config("email_scanner_config", config)
        messagebox.showinfo("✅ Success", "Email scanner configuration saved.", parent=self
        )
        self.log("ℹ️ INFO: Email scanner configuration saved")

    def load_config(self):
        """Load saved email configuration"""
        config = self.dm.get_config("email_scanner_config", {})

        self.email_var.set(config.get("email", ""))
        self.password_var.set(config.get("password", ""))
        self.imap_server_var.set(config.get("imap_server", "imap.gmail.com"))
        self.imap_port_var.set(str(config.get("imap_port", 993)))

    def scan_emails(self):
        """Start email scanning in background thread with applied filters"""

        # Apply current filters before scanning
        try:
            include_text = self.include_keywords_text.get("1.0", tk.END).strip()
            exclude_text = self.exclude_keywords_text.get("1.0", tk.END).strip()

            include_keywords = [
                kw.strip() for kw in include_text.split(",") if kw.strip()
            ]
            exclude_keywords = [
                kw.strip() for kw in exclude_text.split(",") if kw.strip()
            ]

            if not include_keywords:
                messagebox.showwarning(
                    "No Include Keywords",
                    "Please specify at least one include keyword.",
                    parent=self,
                )
                return

            # Update scanner with current filters
            self.scanner.update_keywords(include_keywords, exclude_keywords)

        except Exception as e:
            messagebox.showerror("❌ Error", f"Failed to parse keywords: {str(e)}", parent=self
            )
            return

        try:
            days_back = int(self.days_back_var.get())
        except Exception:
            days_back = 7

        mark_as_read = self.mark_read_var.get()
        scan_method = self.scan_method_var.get()

        # Validate IMAP config if using IMAP
        if scan_method == "imap":
            email_config = {
                "email": self.email_var.get(),
                "password": self.password_var.get(),
                "imap_server": self.imap_server_var.get(),
                "imap_port": int(self.imap_port_var.get()),
            }

            if not email_config["email"] or not email_config["password"]:
                messagebox.showwarning(
                    "Missing Info", "Please enter email and password.", parent=self
                )
                return

        self.results_text.insert(tk.END, f"\n{'='*60}\n")
        self.results_text.insert(
            tk.END, f"Starting email scan at {datetime.now().strftime('%H:%M:%S')}\n"
        )
        self.results_text.insert(
            tk.END, f"Method: {'Outlook' if scan_method == 'outlook' else 'IMAP'}\n"
        )
        self.results_text.insert(
            tk.END, f"Include Keywords: {len(include_keywords)} terms\n"
        )
        self.results_text.insert(
            tk.END, f"Exclude Keywords: {len(exclude_keywords)} terms\n"
        )
        self.results_text.insert(tk.END, f"{'='*60}\n")
        self.results_text.see(tk.END)

        def scan_thread():
            try:
                # Use Outlook or IMAP based on selection
                if scan_method == "outlook":
                    results = self.scanner.scan_outlook_emails(days_back, mark_as_read)
                else:
                    results = self.scanner.scan_emails(
                        email_config, days_back, mark_as_read
                    )

                # Display results with filter stats
                summary = f"""
Scan Complete!
--------------
Emails scanned: {results['scanned']}
Confirmations found: {results['confirmations_found']}
PDFs saved: {results['pdfs_saved']}
Filtered out: {results.get('filtered_out', 0)}

Matched POs: {', '.join(results['matched_pos']) if results['matched_pos'] else 'None'}
Unmatched POs: {', '.join(results['unmatched_pos']) if results['unmatched_pos'] else 'None'}

Errors: {len(results['errors'])}
"""

                self.after(
                    0, lambda summary=summary: self.results_text.insert(tk.END, summary)
                )
                self.after(0, lambda: self.results_text.see(tk.END))

                if results["errors"]:
                    error_text = (
                        "\nErrors:\n" + "\n".join(results["errors"][:10]) + "\n"
                    )
                    self.after(
                        0,
                        lambda error_text=error_text: self.results_text.insert(
                            tk.END, error_text
                        ),
                    )

                msg = (
                    f"Found {results['confirmations_found']} confirmations\n"
                    f"Saved {results['pdfs_saved']} PDFs\n"
                    f"Filtered out {results.get('filtered_out', 0)} non-confirmations"
                )
                self.after(
                    0,
                    lambda msg=msg: messagebox.showinfo(
                        "Scan Complete", msg, parent=self
                    ),
                )

            except Exception as e:
                error_msg = f"✗ ERROR: Email scan failed: {str(e)}"
                self.after(
                    0,
                    lambda error_msg=error_msg: self.results_text.insert(
                        tk.END, f"\n{error_msg}\n"
                    ),
                )
                self.after(0, lambda: self.results_text.see(tk.END))
                self.after(
                    0,
                    lambda error_msg=error_msg: messagebox.showerror(
                        "Scan Failed", error_msg, parent=self
                    ),
                )

        threading.Thread(target=scan_thread, daemon=True).start()

    def open_confirmations_folder(self):
        """Open the confirmations folder in file explorer"""
        confirmations_folder = os.path.join(APP_DATA_FOLDER, "Confirmations")

        if sys.platform == "win32":
            os.startfile(confirmations_folder)
        elif sys.platform == "darwin":  # macOS
            os.system(f'open "{confirmations_folder}"')
        else:  # linux
            os.system(f'xdg-open "{confirmations_folder}"')


# ==============================================================================
# 6. MAIN EXECUTION BLOCK
# ==============================================================================


def main():
    """Main function to initialize and run the application."""
    # 1. Initialize the database and data manager
    db_manager = DatabaseManager(DB_FILE)

    # Force migration check
    print("Checking database columns...")
    db_manager.add_missing_columns()

    data_manager = LocalDataManager(db_manager)

    # Initialize forecast manager
    forecast_manager = ForecastDataManager(db_manager)

    # 2. Set up the Tkinter root window
    root = tk.Tk()

    # Apply a modern theme
    style = ttk.Style(root)
    try:
        # 'clam' is a good, clean, cross-platform theme
        style.theme_use("clam")
    except tk.TclError:
        print("Clam theme not available, using default.")

    app = RemoteOperationsApp(root, data_manager)

    # Center the window
    root.update_idletasks()
    width = root.winfo_width()
    height = root.winfo_height()
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f"{width}x{height}+{x}+{y}")

    # 3. Start the GUI event loop
    root.mainloop()


if __name__ == "__main__":
    main()